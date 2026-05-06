"""
Çok bantlı GeoTIFF'lerden arkeolojik alan tespiti için komut satırı aracı.

Pipeline aşağıdaki adımları gerçekleştirir:
1. Tek bir raster'dan RGB, DSM ve DTM bantlarını okur ve karo bazlı iş yüklerini hazırlar.
2. DTM'den kabartma görselleştirme katmanlarını (gökyüzü görüş faktörü, açıklık, yerel 
   kabartma modeli, eğim) Relief Visualization Toolbox (rvt-py) ile türetir.
3. feature_mode'a göre R,G,B; topo5 için SVF/SLRM; topo7 için ayrıca Slope/nDSM tensörü oluşturur,
   kanalları robust 2-98 persentil ölçekleme ile normalize eder ve nodata değerlerini güvenli şekilde işler.
4. Önceden eğitilmiş derin öğrenme modeli ile karolara bölünmüş çıkarım yapar.
   Model segmentasyon ise piksel-seviyesinde, tile-classification ise tile-seviyesinde
   skor üretir; örtüşen karolar harmanlanarak kesintisiz olasılık haritası oluşturulur.
5. İsteğe bağlı yüksek obje maskeleme uygular, eşikleme ile ikili maske oluşturur ve hem
   olasılık hem de maske GeoTIFF çıktılarını sıkıştırılmış olarak, jeoreferansı koruyarak yazar.
6. İsteğe bağlı olarak tespit edilen özellikleri alan/skor öznitelikleriyle poligonlara dönüştürür 
   ve GIS iş akışları için GeoPackage formatında dışa aktarır.

Varsayımlar:
- Girdi mozaiği RGB, DSM ve DTM bantlarını aynı CRS/kapsam/piksel-ızgarada içerir.
- Önceden eğitilmiş model metadata'daki kanal şemasını bekler.
- rvt-py ve torch uyumlu hesaplama ortamı önceden kurulmuştur.

"""

from __future__ import annotations

import argparse
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
import gc
import hashlib
import json
import logging
import math
import os
import shutil
import sys
import threading
import uuid
import zipfile
from contextlib import ExitStack, nullcontext
from dataclasses import dataclass, field, fields
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Generator, Iterable, List, Optional, Sequence, TextIO, Tuple, TypeVar
from xml.sax.saxutils import escape as xml_escape

from archeo_shared.console import configure_utf8_console

configure_utf8_console()

import numpy as np
import rasterio
from rasterio.features import shapes
from rasterio.crs import CRS as RasterioCRS
from rasterio.transform import Affine
from rasterio.warp import transform as rasterio_transform
from rasterio.windows import Window
from scipy import ndimage
from scipy.ndimage import (
    gaussian_filter,
    gaussian_gradient_magnitude,
    gaussian_laplace,
    grey_closing,
    grey_opening,
    uniform_filter,
)
import torch
from tqdm import tqdm

from archeo_shared.channels import (
    LOCKED_TRAINED_ONLY_FIELDS,
    METADATA_SCHEMA_VERSION,
    MODEL_CHANNEL_NAMES,
    TOPO7_CHANNEL_NAMES,
    canonicalize_channel_names,
    channel_names_for_feature_mode,
    expected_channel_names,
    input_band_count_for_feature_mode,
    normalize_feature_mode as normalize_shared_feature_mode,
)
from archeo_shared.modeling import AttentionWrapper, CBAM, ChannelAttention, SpatialAttention

try:
    import segmentation_models_pytorch as smp
except ImportError:
    smp = None  # segmentation_models_pytorch is optional; required only for DL inference

try:
    from shapely.geometry import Polygon, mapping, shape
    from shapely.ops import transform as shapely_transform
except ImportError:
    Polygon = None  # type: ignore[assignment]
    mapping = None  # type: ignore[assignment]
    shape = None  # type: ignore[assignment]
    shapely_transform = None  # type: ignore[assignment]

try:
    from pyproj import CRS, Transformer
except ImportError:
    CRS = None  # type: ignore[assignment]
    Transformer = None  # type: ignore[assignment]

try:
    from rvt import vis as rvt_vis
except ImportError:
    rvt_vis = None  # rvt is required only when RVT-based features are used

try:
    import geopandas as gpd
except ImportError:
    gpd = None  # geopandas is optional; fallback writer is used otherwise

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    Workbook = None  # openpyxl is optional; a built-in XLSX writer is used as fallback
    ColorScaleRule = None
    Alignment = None
    Font = None
    PatternFill = None
    Table = None
    TableStyleInfo = None

try:
    import fiona
except ImportError:
    fiona = None

try:
    from osgeo import gdal
except ImportError:  # pragma: no cover - optional, used to tame GDAL warnings
    gdal = None

try:
    import yaml
except ImportError:
    yaml = None  # YAML support is optional

try:
    from ultralytics import YOLO
except ImportError:
    YOLO = None  # YOLO11 support is optional

LOGGER = logging.getLogger("archaeo_detect")
T = TypeVar("T")

SESSION_RUN_ID = datetime.now().strftime("%Y%m%d_%H%M%S")

_WARN_ONCE_KEYS: set[str] = set()
_DERIV_CACHE_THREAD_LOCAL = threading.local()


def _warn_once(key: str, message: str) -> None:
    """Emit a warning only once per Python process."""
    if key in _WARN_ONCE_KEYS:
        return
    _WARN_ONCE_KEYS.add(key)
    LOGGER.warning(message)


_KNOWN_OUTPUT_SUFFIXES = {
    ".tif",
    ".tiff",
    ".gpkg",
    ".xlsx",
    ".csv",
    ".geojson",
    ".json",
    ".shp",
    ".vrt",
}

_CONFIG_PATH_FIELDS = {
    "input",
    "out_prefix",
    "weights",
    "weights_template",
    "training_metadata",
    "yolo_weights",
    "cache_dir",
}

WORKSPACE_ROOT = Path("workspace")
WORKSPACE_OUTPUTS_DIR = WORKSPACE_ROOT / "ciktilar"
WORKSPACE_CHECKPOINTS_DIR = WORKSPACE_ROOT / "checkpoints"
WORKSPACE_ASSETS_DIR = WORKSPACE_ROOT / "assets"


def default_config_path() -> str:
    """Prefer config.local.yaml when it exists."""
    local_config = Path("config.local.yaml")
    if local_config.exists():
        return str(local_config)
    return "config.yaml"


def _normalize_config_path_value(raw: Any, base_dir: Path) -> Any:
    """Resolve config path-like values relative to the chosen base directory."""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return raw
    path = Path(text)
    if not path.is_absolute():
        path = base_dir / path
    return str(path.resolve(strict=False))


def _resolve_yolo_weights_path(raw: Optional[str]) -> Optional[str]:
    """Prefer existing local YOLO weights before letting Ultralytics download assets."""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None

    requested = Path(text)
    candidates: List[Path] = [requested]
    if not requested.is_absolute():
        candidates.append(Path.cwd() / requested)

    weight_name = requested.name
    if weight_name:
        candidates.extend(
            [
                WORKSPACE_ASSETS_DIR / weight_name,
                WORKSPACE_CHECKPOINTS_DIR / weight_name,
                WORKSPACE_ROOT / "models" / weight_name,
                Path("models") / weight_name,
            ]
        )

    seen: set[str] = set()
    for candidate in candidates:
        resolved = candidate.resolve(strict=False)
        key = str(resolved).lower()
        if key in seen:
            continue
        seen.add(key)
        if resolved.exists():
            return str(resolved)

    return text


def _normalize_yolo_name_map(raw_names: Any) -> Dict[int, str]:
    """Normalize Ultralytics class-name containers into an int->name mapping."""
    normalized: Dict[int, str] = {}
    if isinstance(raw_names, dict):
        items = raw_names.items()
    elif isinstance(raw_names, (list, tuple)):
        items = enumerate(raw_names)
    else:
        return normalized

    for raw_id, raw_name in items:
        try:
            class_id = int(raw_id)
        except Exception:
            continue
        class_name = str(raw_name).strip()
        if not class_name:
            continue
        normalized[class_id] = class_name
    return normalized


def _resolve_yolo_class_filter(raw: Optional[str], raw_names: Any) -> Optional[List[int]]:
    """Resolve CSV class names/ids into the Ultralytics `classes=` id list."""
    text = str(raw or "").strip()
    if not text or text.lower() == "all":
        return None

    name_map = _normalize_yolo_name_map(raw_names)
    if not name_map:
        raise ValueError("YOLO class filter could not be resolved because model class names are unavailable.")

    name_to_id = {name.lower(): class_id for class_id, name in name_map.items()}
    resolved_ids: List[int] = []
    seen_ids: set[int] = set()
    invalid_tokens: List[str] = []

    for token in [part.strip() for part in text.split(",") if part.strip()]:
        class_id: Optional[int] = None
        if token.isdigit():
            numeric_id = int(token)
            if numeric_id in name_map:
                class_id = numeric_id
        else:
            class_id = name_to_id.get(token.lower())

        if class_id is None:
            invalid_tokens.append(token)
            continue
        if class_id in seen_ids:
            continue
        seen_ids.add(class_id)
        resolved_ids.append(class_id)

    if invalid_tokens:
        sample_names = ", ".join(
            f"{class_id}:{name}" for class_id, name in sorted(name_map.items())[:12]
        )
        raise ValueError(
            "Unknown YOLO class filter value(s): "
            f"{', '.join(invalid_tokens)}. Available examples: {sample_names}"
        )
    if not resolved_ids:
        raise ValueError("YOLO class filter did not match any class ids.")
    return resolved_ids


def _output_base_path(path: Path) -> Path:
    """
    Strip only real output extensions from a path.

    Path.with_suffix("") truncates suffixless names that merely contain dots, such as
    "trained_th0.5_mask". Output names use dots in threshold tokens, so only known file
    extensions should be removed.
    """
    suffix = path.suffix.lower()
    if suffix and suffix in _KNOWN_OUTPUT_SUFFIXES:
        return path.with_suffix("")
    return path


def _candidate_table_base(path: Path) -> Path:
    base_path = _output_base_path(path)
    return base_path.with_name(f"{base_path.name}_gps")


def _combined_candidate_table_base(path: Path) -> Path:
    base_path = _output_base_path(path)
    return base_path.with_name(f"{base_path.name}_all_candidates")


def _vector_package_base(path: Path) -> Path:
    base_path = _output_base_path(path)
    return base_path.with_name(f"{base_path.name}_layers")


def _append_output_suffix(path: Path, suffix: str) -> Path:
    """Append a file suffix without truncating dotted base names."""
    if not suffix.startswith("."):
        raise ValueError("suffix must start with '.'")
    base_path = _output_base_path(path)
    return base_path.parent / f"{base_path.name}{suffix}"


_CANDIDATE_TABLE_FIELD_ORDER = [
    "candidate_id",
    "area_m2",
    "score_mean",
    "center_x_native",
    "center_y_native",
    "native_crs",
    "gps_lon",
    "gps_lat",
    "google_maps_url",
]

_COMBINED_CANDIDATE_TABLE_FIELD_ORDER = [
    "source_label",
    "scale_level",
    "scale_factor",
    "candidate_type",
    "review_status",
    "candidate_id",
    "area_m2",
    "score_mean",
    "center_x_native",
    "center_y_native",
    "native_crs",
    "gps_lon",
    "gps_lat",
    "google_maps_url",
]

_COMBINED_CANDIDATE_PRIORITY_FIELD_ORDER = [
    "rank",
    "priority_band",
    "source_label",
    "scale_level",
    "scale_factor",
    "candidate_type",
    "review_status",
    "candidate_id",
    "score_mean",
    "area_m2",
    "center_x_native",
    "center_y_native",
    "native_crs",
    "gps_lon",
    "gps_lat",
    "google_maps_url",
]

_COMBINED_CANDIDATE_SUMMARY_FIELD_ORDER = [
    "source_label",
    "scale_level",
    "scale_factor",
    "candidate_count",
    "high_priority_count",
    "medium_priority_count",
    "low_priority_count",
    "score_mean_avg",
    "score_mean_max",
    "area_m2_sum",
    "area_m2_max",
]

_COMBINED_CANDIDATE_CLUSTER_FIELD_ORDER = [
    "cluster_id",
    "priority_band",
    "member_count",
    "sources_seen",
    "scale_levels_seen",
    "scale_factors_seen",
    "best_source_label",
    "best_scale_level",
    "best_scale_factor",
    "best_candidate_id",
    "best_score_mean",
    "mean_score_mean",
    "total_area_m2",
    "max_area_m2",
    "center_x_native",
    "center_y_native",
    "native_crs",
    "gps_lon",
    "gps_lat",
    "google_maps_url",
]

_CANDIDATE_BOX_FIELD_ORDER = [
    "source_label",
    "scale_level",
    "scale_factor",
    "candidate_id",
    "candidate_type",
    "review_status",
    "mask_area_m2",
    "bbox_area_m2",
    "score_mean",
    "score_max",
    "pixel_count",
    "center_x_native",
    "center_y_native",
    "bbox_xmin",
    "bbox_ymin",
    "bbox_xmax",
    "bbox_ymax",
    "native_crs",
    "gps_lon",
    "gps_lat",
    "google_maps_url",
]

_CANDIDATE_BOX_GPKG_FIELD_ORDER = [
    "candidate_id",
    "source_label",
    "scale_level",
    "scale_factor",
    "candidate_type",
    "review_status",
    "mask_area_m2",
    "bbox_area_m2",
    "score_mean",
    "score_max",
    "pixel_count",
    "center_x_native",
    "center_y_native",
    "bbox_xmin",
    "bbox_ymin",
    "bbox_xmax",
    "bbox_ymax",
    "gps_lon",
    "gps_lat",
    "google_maps_url",
]

_COMBINED_CANDIDATE_RAW_FIELD_ORDER = ["input_order"] + _COMBINED_CANDIDATE_TABLE_FIELD_ORDER

_COMBINED_CANDIDATE_PRIORITY_HIGH = 0.80
_COMBINED_CANDIDATE_PRIORITY_MEDIUM = 0.60
_COMBINED_CANDIDATE_CLUSTER_DISTANCE_M = 20.0
_CANDIDATE_TYPE_THRESHOLD = "threshold_detection"
_CANDIDATE_TYPE_FALLBACK = "fallback_top_score"
_REVIEW_STATUS_DEFAULT = "Kontrol edilecek"
_REVIEW_STATUS_FALLBACK = "Esik alti aday"


def _safe_token(value: str) -> str:
    token = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in value.strip())
    return token.strip("_")


def _safe_gpkg_layer_name(value: str, max_len: int = 63) -> str:
    token = _safe_token(value.replace("-", "_"))
    if not token:
        token = "layer"
    return token[:max_len].rstrip("_") or "layer"


def _vector_gpkg_path(path: Path) -> Path:
    return _append_output_suffix(_vector_package_base(path), ".gpkg")


def _format_gpkg_target(gpkg_path: Path, layer_name: Optional[str]) -> str:
    if layer_name:
        return f"{gpkg_path} [layer={layer_name}]"
    return str(gpkg_path)


def _fmt_float(value: float, decimals: int = 2) -> str:
    formatted = f"{value:.{decimals}f}".rstrip("0").rstrip(".")
    return formatted.replace(".", "p")

# ==== Sabit Değerler ====
# Bu değerler PipelineDefaults içinde yapılandırılabilir hale getirildi.
# Eski kodlarla uyumluluk için burada referanslar var.

# Histogram ve eşikleme parametreleri (algoritmanın çalışması için sabit kalmalı)
HISTOGRAM_BINS = 256
HISTOGRAM_RANGE = (0.0, 1.0)

# Karo işleme sabitleri
TILE_SAMPLE_CROP_SIZE = 256
LABEL_CONNECTIVITY_STRUCTURE = (3, 3)

_GDAL_HANDLER_INSTALLED = False
_SUPPRESSED_GDAL_MESSAGES = (
    "TIFFReadDirectory:Sum of Photometric type-related color channels and ExtraSamples doesn't match SamplesPerPixel",
)


def available_memory_bytes() -> Optional[int]:
    """Best-effort available system memory in bytes (None if unknown)."""
    try:
        import psutil  # type: ignore
    except Exception:
        return None
    try:
        return int(psutil.virtual_memory().available)
    except Exception:
        return None


def estimate_precompute_derivatives_bytes(
    pixels: int,
    *,
    has_dsm: bool,
    enable_curvature: bool = False,
    enable_tpi: bool = False,
) -> int:
    """
    Rough upper-bound estimate of RAM needed by `precompute_derivatives`.

    The function currently materializes full-raster arrays for:
    - RGB (3), DTM (1)
    - RVT outputs: SVF (1), SLRM (1)
    """
    float_arrays = 3 + 1 + 2  # RGB + DTM + SVF + SLRM
    return int(float_arrays * pixels * np.dtype(np.float32).itemsize)


def full_raster_cache_precompute_ok(
    input_path: Path,
    band_idx: Sequence[int],
    *,
    enable_curvature: bool,
    enable_tpi: bool,
) -> Tuple[bool, str]:
    """
    Decide whether full-raster derivative precompute/cache is feasible.

    This avoids out-of-memory crashes on very large rasters by falling back to
    per-tile derivative computation.
    """
    try:
        with rasterio.open(input_path) as src:
            height = int(src.height)
            width = int(src.width)
    except Exception as e:
        return False, f"Could not read raster size ({e}); skipping full-raster derivative cache precompute."

    pixels = height * width
    if pixels <= 0:
        return False, "Invalid raster size; skipping full-raster derivative cache precompute."

    # Heuristic: above this, full-raster materialization (many float32 layers) is rarely practical.
    if pixels >= 200_000_000:
        return (
            False,
            f"Raster is very large ({width}x{height}); skipping full-raster derivative cache precompute (using per-tile derivatives).",
        )

    avail = available_memory_bytes()
    has_dsm = bool(len(band_idx) >= 4 and band_idx[3] > 0)
    est = estimate_precompute_derivatives_bytes(
        pixels,
        has_dsm=has_dsm,
        enable_curvature=enable_curvature,
        enable_tpi=enable_tpi,
    )
    est_with_overhead = int(est * 1.35)
    if avail is None or avail <= 0:
        return True, ""

    if est_with_overhead > int(avail * 0.75):
        est_gib = est_with_overhead / (1024**3)
        avail_gib = avail / (1024**3)
        return (
            False,
            f"Full-raster derivative cache precompute would likely OOM (raster {width}x{height}; estimated ~{est_gib:.1f} GiB, available ~{avail_gib:.1f} GiB). Falling back to per-tile derivatives.",
        )

    return True, ""


def make_scratch_dir(tag: str) -> Path:
    """Create a per-run scratch directory under `workspace/checkpoints/scratch/`."""
    base = WORKSPACE_CHECKPOINTS_DIR / "scratch"
    base.mkdir(parents=True, exist_ok=True)
    run_dir = base / f"{tag}_{uuid.uuid4().hex}"
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir


def alloc_array(
    shape: Tuple[int, int],
    dtype: Any,
    *,
    fill_value: Any,
    use_memmap: bool,
    scratch_dir: Optional[Path],
    name: str,
) -> Tuple[np.ndarray, Optional[Path]]:
    """Allocate an array in RAM or as a disk-backed memmap (for huge rasters)."""
    if not use_memmap:
        if fill_value == 0 or fill_value is False:
            return np.zeros(shape, dtype=dtype), None
        return np.full(shape, fill_value, dtype=dtype), None

    scratch_dir = scratch_dir or make_scratch_dir("arrays")
    path = scratch_dir / f"{name}_{uuid.uuid4().hex}.mmap"
    arr = np.memmap(path, dtype=dtype, mode="w+", shape=shape)
    # Ensure deterministic initialization for downstream logic.
    arr.fill(fill_value)
    return arr, path


def cleanup_scratch_dir(scratch_dir: Optional[Path]) -> None:
    """Best-effort cleanup for temporary memmap scratch directories."""
    if scratch_dir is None:
        return
    try:
        gc.collect()
        shutil.rmtree(scratch_dir, ignore_errors=True)
    except Exception as e:  # pragma: no cover - defensive cleanup
        LOGGER.debug("Scratch cleanup skipped for %s: %s", scratch_dir, e)


def install_gdal_warning_filter() -> None:
    """Suppress noisy but harmless GDAL warnings about mis-declared TIFF extrasamples."""
    global _GDAL_HANDLER_INSTALLED
    if _GDAL_HANDLER_INSTALLED or gdal is None:
        return

    def _handler(err_class: int, err_num: int, err_msg: str) -> None:
        if any(err_msg.startswith(prefix) for prefix in _SUPPRESSED_GDAL_MESSAGES):
            LOGGER.debug("GDAL uyarısı bastırıldı: %s", err_msg)
            return
        try:
            gdal.CPLDefaultErrorHandler(err_class, err_num, err_msg)
        except AttributeError:  # pragma: no cover - safety for stripped bindings
            LOGGER.warning("GDAL uyarısı: %s", err_msg)

    gdal.PushErrorHandler(_handler)
    _GDAL_HANDLER_INSTALLED = True


def progress_bar(iterable: Iterable[T], **tqdm_kwargs) -> tqdm:
    """Wrapper around tqdm that prefers stdout so CLI runners can render progress."""
    stream: Optional[TextIO]
    if sys.stdout:
        stream = sys.stdout
    elif sys.stderr:
        stream = sys.stderr
    else:  # pragma: no cover - exceptionally rare in CLI usage
        stream = None
    if stream is not None:
        tqdm_kwargs.setdefault("file", stream)
    tqdm_kwargs.setdefault("dynamic_ncols", True)
    tqdm_kwargs.setdefault("mininterval", 0.5)
    tqdm_kwargs.setdefault("smoothing", 0.1)
    tqdm_kwargs.setdefault("leave", False)
    return tqdm(iterable, **tqdm_kwargs)

# ==== PIPELINE AYARLARI (TÜM PARAMETRELERİ BURADAN KONTROl EDİN) ====
# Bu dataclass tüm parametreleri tek merkezde toplar ve her alanın üzerindeki açıklama 
# parametrenin ne yaptığını gösterir. İstediğiniz ayarları buradan değiştirebilirsiniz.


@dataclass
class PipelineDefaults:
    """
    Arkeolojik alan tespit pipeline'ının tüm ayarlarını içeren merkezi yapılandırma sınıfı.
    Tüm parametreleri buradan kontrol edebilir, hangi yöntemlerin çalışacağını belirleyebilirsiniz.
    """
    
    # ===== GİRDİ/ÇIKTI AYARLARI =====
    input: str = field(
        default="kesif_alani.tif",
        metadata={"help": "Çok bantlı GeoTIFF dosyasının yolu (RGB + DSM + DTM önerilir)"},
    )
    out_prefix: Optional[str] = field(
        default=None,
        metadata={"help": "Çıktı dosyaları için ön-ek veya dizin; None ise girdi adından otomatik üretilir"},
    )
    bands: str = field(
        default="1,2,3,4,5",
        metadata={"help": "Bant sırası. RGB-only için 3 giriş (örn: 1,2,3); topo5/topo7 için 5 giriş (R,G,B,DSM,DTM; örn: 1,2,3,4,5). DSM topo7 için zorunludur."},
    )
    feature_mode: str = field(
        default="auto",
        metadata={"help": "Model feature semasi: auto=checkpoint/metadata/bands; rgb3=R,G,B; topo5=R,G,B,SVF,SLRM; topo7=R,G,B,SVF,SLRM,Slope,nDSM."},
    )
    
    # ===== YÖNTEM SEÇİMİ (HANGİ YÖNTEMLERİN ÇALIŞACAĞINI BURADAN KONTROL EDİN) =====
    enable_deep_learning: bool = field(
        default=True,
        metadata={"help": "Derin öğrenme (U-Net) modelini çalıştır"},
    )
    enable_classic: bool = field(
        default=True,
        metadata={"help": "Klasik kabartma tabanlı yöntemleri (RVT, Hessian, morfoloji) çalıştır"},
    )
    enable_yolo: bool = field(
        default=False,
        metadata={"help": "YOLO11 nesne tespit/segmentasyon modelini çalıştır"},
    )
    enable_fusion: bool = field(
        default=True,
        metadata={"help": "Derin öğrenme ve klasik yöntem sonuçlarını birleştir (fusion)"},
    )
    dl_task: str = field(
        default="segmentation",
        metadata={"help": "DL gorevi: 'segmentation' veya 'tile_classification'. Tile classification, her tile icin tek skor uretip overlap ile risk haritasi olusturur."},
    )
    
    # ===== DERİN ÖĞRENME MODEL AYARLARI =====
    arch: str = field(
        default="Unet",
        metadata={"help": "Segmentation model mimarisi (örn: Unet, UnetPlusPlus, DeepLabV3Plus)"},
    )
    encoder: str = field(
        default="resnet34",
        metadata={"help": "Tek model çalışırken kullanılacak varsayılan encoder (resnet34, resnet50, efficientnet-b3)"},
    )
    encoders: str = field(
        default="none",
        metadata={"help": "Çalıştırılacak encoder listesi; 'all'=hepsi, 'none'=sadece encoder parametresi, veya 'resnet34,resnet50'"},
    )
    weights: Optional[str] = field(
        default=None,
        metadata={"help": "Eğitilmiş model ağırlıkları (.pth dosyası); None ise zero-shot modu kullanılır"},
    )
    training_metadata: Optional[str] = field(
        default=None,
        metadata={
            "help": "Egitim metadata dosyasi (.json). trained_model_only modunda tile/overlap/bands buradan kilitlenir."
        },
    )
    weights_template: Optional[str] = field(
        default=None,
        metadata={"help": "Encoder bazli agirlik sablonu (orn: models/unet_{encoder}.pth); {encoder} yerine encoder adi konur"},
    )
    zero_shot_imagenet: bool = field(
        default=True,
        metadata={"help": "Agirlik dosyasi yoksa ImageNet encoder'ini etkin kanal sayisina genisletip zero-shot calistir"},
    )
    trained_model_only: bool = field(
        default=False,
        metadata={"help": "Sadece egitilmis DL checkpointini kullan; multi/zeroshot kapatilir (classic/fusion kalabilir)."},
    )
    use_fpn_classifier: Optional[bool] = field(
        default=None,
        metadata={
            "help": "TileClassifier icin FPN-style multi-level feature aggregation. None ise checkpoint metadata/state_dict'ten otomatik cozulur; bulunamazsa False kullanilir."
        },
    )
    th: float = field(
        default=0.6,
        metadata={"help": "DL olasılık haritasını maskeye çevirmek için eşik (0-1 arası); düşük değer daha fazla tespit üretir"},
    )
    
    # ===== KLASİK YÖNTEM AYARLARI =====
    classic_modes: str = field(
        default="combo",
        metadata={"help": "Klasik yöntemler: 'combo' (hepsini çalıştır), veya 'rvtlog,hessian,morph' (seçili olanları çalıştır)"},
    )
    classic_save_intermediate: bool = field(
        default=True,
        metadata={"help": "Her klasik modu ayrı dosya olarak kaydet (karşılaştırma için kullanışlı)"},
    )
    classic_th: Optional[float] = field(
        default=None,
        metadata={"help": "Klasik yöntem için sabit maske eşiği (0-1); None ise otomatik Otsu eşiği kullanılır"},
    )
    
    # ===== KLASİK YÖNTEM ALGORİTMA PARAMETRELERİ =====
    sigma_scales: Tuple[float, ...] = field(
        default=(1.0, 2.0, 4.0, 8.0),
        metadata={"help": "Çok ölçekli filtreleme için sigma değerleri (LoG, Hessian için)"},
    )
    morphology_radii: Tuple[int, ...] = field(
        default=(3, 5, 9, 15),
        metadata={"help": "Morfolojik filtreleme için yapısal eleman yarıçapları"},
    )
    rvt_radii: Tuple[float, ...] = field(
        default=(10.0,),
        metadata={"help": "SVF hesaplaması için arama yarıçapları (metre). Tek değer önerilir."},
    )
    local_variance_window: int = field(
        default=7,
        metadata={"help": "Yerel varyans hesaplama için pencere boyutu"},
    )
    gaussian_gradient_sigma: float = field(
        default=1.5,
        metadata={"help": "Gradyan büyüklüğü hesaplama için Gaussian sigma değeri"},
    )
    gaussian_lrm_sigma: float = field(
        default=6.0,
        metadata={"help": "LRM fallback hesaplama için Gaussian sigma değeri"},
    )
    
    # ===== GELİŞMİŞ TOPOGRAFİK ANALİZ AYARLARI (YENİ KANALLAR) =====
    enable_curvature: bool = field(
        default=True,
        metadata={"help": "Plan ve Profile Curvature kanallarını hesapla (hendek/tepe ayrımı için)"},
    )
    enable_tpi: bool = field(
        default=True,
        metadata={"help": "Multi-scale TPI (Topographic Position Index) kanalını hesapla (höyük/çukur tespiti için)"},
    )
    tpi_radii: Tuple[int, ...] = field(
        default=(5, 15, 30),
        metadata={"help": "TPI hesaplama yarıçapları (piksel); farklı boyutlardaki yapılar için"},
    )
    
    # ===== ATTENTION MEKANİZMASI AYARLARI =====
    enable_attention: bool = field(
        default=True,
        metadata={"help": "CBAM (Convolutional Block Attention Module) kullan; kanal önemini dinamik öğrenir"},
    )
    attention_reduction: int = field(
        default=4,
        metadata={"help": "Attention modülü için kanal azaltma oranı (4=1/4 oranında sıkıştırma)"},
    )
    
    save_band_importance: bool = field(
        default=True,
        metadata={"help": "DL inference sirasinda band onem raporu (txt/json) kaydet."},
    )
    band_importance_max_tiles: int = field(
        default=4,
        metadata={"help": "Band onem analizi icin kullanilacak maksimum tile sayisi. 0 ise kapali."},
    )

    # ===== YOLO11 MODEL AYARLARI =====
    yolo_weights: Optional[str] = field(
        default=None,
        metadata={"help": "YOLO11 ağırlık dosyası (örn: yolo11n.pt, yolo11n-seg.pt); None ise varsayılan yolo11n-seg.pt indirilir"},
    )
    yolo_conf: float = field(
        default=0.25,
        metadata={"help": "YOLO11 tespit güven eşiği (0-1 arası)"},
    )
    yolo_iou: float = field(
        default=0.45,
        metadata={"help": "YOLO11 NMS IoU eşiği (0-1 arası)"},
    )
    yolo_tile: Optional[int] = field(
        default=None,
        metadata={"help": "YOLO11 için özel tile boyutu; None ise --tile değerini kullanır"},
    )
    yolo_imgsz: int = field(
        default=640,
        metadata={"help": "YOLO11 model girdi boyutu (piksel); model bu boyuta ölçekler"},
    )
    yolo_device: Optional[str] = field(
        default=None,
        metadata={"help": "YOLO11 için cihaz ('0', 'cpu', vb.); None ise otomatik seçilir"},
    )

    # ===== BİRLEŞTİRME (FUSION) AYARLARI =====
    alpha: float = field(
        default=0.5,
        metadata={"help": "Fusion karışım ağırlığı (0-1); 1.0=sadece DL, 0.0=sadece klasik, 0.5=eşit ağırlık"},
    )
    fuse_encoders: str = field(
        default="all",
        metadata={"help": "Füzyon yapılacak encoder listesi: 'all' veya 'resnet34,resnet50,efficientnet-b3'"},
    )
    
    # ===== KARO İŞLEME AYARLARI =====
    tile: int = field(
        default=1024,
        metadata={"help": "Karo boyutu (piksel); büyük değer daha fazla GPU belleği kullanır ama daha az karo demektir"},
    )
    overlap: int = field(
        default=128,
        metadata={"help": "Komşu karolar arası bindirme (piksel); dikiş hatlarını azaltır ama işlem süresini artırır"},
    )
    feather: bool = field(
        default=True,
        metadata={"help": "Karo sınırlarında yumuşatma uygula; dikiş izlerini azaltır"},
    )
    
    # ===== NORMALİZASYON AYARLARI =====
    global_norm: bool = field(
        default=True,
        metadata={"help": "Tüm karolar için tek bir global persentil normalizasyonu kullan (tutarlılık için önerilir)"},
    )
    norm_sample_tiles: int = field(
        default=32,
        metadata={"help": "Global normalizasyon için örneklenecek karo sayısı"},
    )
    percentile_low: float = field(
        default=2.0,
        metadata={"help": "Normalizasyon alt persentil değeri"},
    )
    percentile_high: float = field(
        default=98.0,
        metadata={"help": "Normalizasyon üst persentil değeri"},
    )
    
    # ===== MASKELEME VE FİLTRELEME AYARLARI =====
    mask_talls: Optional[float] = field(
        default=2.5,
        metadata={"help": "Bu yüksekliği (metre) aşan nDSM piksellerini maskele (yüksek bina/ağaçları eleme); None=kapalı"},
    )
    rgb_only: bool = field(
        default=False,
        metadata={"help": "If True, ignore derivative channels and run effective RGB-only inference by zero-filling SVF/Openness/LRM/Slope/nDSM."},
    )
    
    # ===== VEKTÖRLEŞTİRME AYARLARI =====
    vectorize: bool = field(
        default=True,
        metadata={"help": "Tespit maskelerini poligonlara dönüştürüp GeoPackage (.gpkg) olarak dışa aktar; dosyalama stratejisi gpkg_mode ile belirlenir."},
    )
    gpkg_mode: str = field(
        default="single",
        metadata={"help": "GeoPackage stratejisi: 'single' = tek GPKG içinde çok katman, 'split' = her çıktı için ayrı GPKG."},
    )
    export_candidate_excel: bool = field(
        default=True,
        metadata={
            "help": "Adayları merkez koordinatları ve GPS (WGS84) kolonlarıyla tek bir birlesik Excel'e (.xlsx) aktar; multi-scale aciksa scale_level/scale_factor kolonlari da eklenir. openpyxl yoksa yerlesik XLSX yazici kullan."
        },
    )
    export_candidate_boxes: bool = field(
        default=True,
        metadata={
            "help": "Mevcut mask/probability ciktisindan birlesik aday kutulari uret; vectorize aciksa GPKG'ye ayri layer, Excel'e ayri sayfa olarak ekler. Yeni egitim/etiketleme gerektirmez."
        },
    )
    fallback_candidates_enabled: bool = field(
        default=True,
        metadata={
            "help": "Esik/min_area sonrasi hic aday yoksa birlesik Excel'e en yuksek skorlu esik alti adaylari ekle."
        },
    )
    fallback_candidates_top_k: int = field(
        default=20,
        metadata={"help": "Fallback aktifken eklenecek maksimum en yuksek skorlu aday sayisi."},
    )
    fallback_candidates_min_score: Optional[float] = field(
        default=None,
        metadata={
            "help": "Fallback adaylari icin opsiyonel minimum skor. None/null ise sadece top_k kullanilir."
        },
    )
    min_area: float = field(
        default=80.0,
        metadata={"help": "Vektörleştirmede kabul edilecek minimum poligon alanı (metrekare)"},
    )
    simplify: Optional[float] = field(
        default=None,
        metadata={"help": "Poligon basitleştirme toleransı (metre); None ise basitleştirme yapılmaz"},
    )
    vector_opening_size: int = field(
        default=3,
        metadata={"help": "Vektörleştirme için morfolojik açma (opening) boyutu (piksel)"},
    )
    label_connectivity: int = field(
        default=8,
        metadata={"help": "Bileşen etiketlemede bağlanırlık (4 veya 8 komşuluk)"},
    )
    
    # ===== PERFORMANS VE TEKNİK AYARLAR =====
    half: bool = field(
        default=True,
        metadata={"help": "CUDA varsa float16 (yarı hassasiyet) kullan; bellek tasarrufu ve hız kazancı sağlar"},
    )
    seed: int = field(
        default=42,
        metadata={"help": "Rastgelelik tohum değeri (tekrarlanabilirlik için)"},
    )
    verbose: int = field(
        default=1,
        metadata={"help": "Log detay seviyesi: 0=WARNING, 1=INFO, 2=DEBUG"},
    )
    
    # ===== CACHE YÖNETİMİ (ÇOK HIZLI TEKRAR ÇALIŞTIRMALAR İÇİN) =====
    cache_derivatives: bool = field(
        default=False,
        metadata={"help": "RVT türevlerini (SVF, openness, LRM, slope) diske kaydet/oku; tekrar çalıştırmalarda çok hızlı"},
    )
    cache_derivatives_mode: str = field(
        default="auto",
        metadata={
            "help": "Cache formatı: 'auto' (küçükte .npz, büyükte blok-bazlı raster-cache), 'npz' veya 'raster'."
        },
    )
    deriv_cache_chunk: int = field(
        default=2048,
        metadata={"help": "Raster-cache üretiminde blok boyutu (piksel). Büyük değer daha hızlı ama daha çok RAM kullanır."},
    )
    deriv_cache_workers: int = field(
        default=1,
        metadata={"help": "Raster-cache hesaplamasÄ± iÃ§in worker sayÄ±sÄ±. 1=seri, 2-4 genelde daha iyi hÄ±z/RAM dengesi verir."},
    )
    deriv_cache_halo: Optional[int] = field(
        default=None,
        metadata={"help": "Raster-cache için halo (piksel). None ise rvt_radii/tpi_radii/sigma'dan otomatik hesaplanır."},
    )
    cache_dir: Optional[str] = field(
        default=None,
        metadata={"help": "Cache dosyalarının kaydedileceği dizin; None ise girdi dosyasının yanına yazar"},
    )
    recalculate_cache: bool = field(
        default=False,
        metadata={"help": "Mevcut cache'i yoksay ve RVT türevlerini yeniden hesapla"},
    )
    
    # ===== MULTI-SCALE INFERENCE =====
    enable_multiscale: bool = field(
        default=False,
        metadata={"help": "Çoklu ölçekte inference çalıştır. Her ölçekte girdi raster küçültülüp aynı model ile tahmin yapılır."},
    )
    multiscale_scales: Tuple[float, ...] = field(
        default=(1.0, 0.5, 0.25),
        metadata={"help": "Inference ölçekleri. 1.0=orijinal, 0.5=2x küçük, 0.25=4x küçük."},
    )
    multiscale_merge: str = field(
        default="max",
        metadata={"help": "Ölçek birleştirme stratejisi: 'max', 'mean', 'weighted_mean'."},
    )
    multiscale_weights: Optional[Tuple[float, ...]] = field(
        default=None,
        metadata={"help": "weighted_mean için ölçek ağırlıkları. None ise eşit ağırlık. Sırası scales ile aynı olmalı."},
    )
    multiscale_save_individual_outputs: bool = field(
        default=False,
        metadata={"help": "Her ölçek için ayri prob/mask ciktisi uret. vectorize aciksa bu olcekler tek GPKG icinde katman ya da ayri GPKG dosyalari olarak yazilir; Excel adaylari ise tek bir birlesik dosyada scale kolonlariyla toplanir."},
    )

    # ===== CİHAZ SEÇİMİ =====
    device: Optional[str] = field(
        default=None,
        metadata={"help": "Hesaplama cihazı: 'cuda', 'cpu', 'cuda:0', vb. None ise otomatik seçilir (CUDA varsa GPU)"},
    )
    
    def __post_init__(self) -> None:
        """Konfigürasyon değerlerini doğrula ve geçersiz değerlerde hata fırlat."""
        errors: List[str] = []
        
        # Eşik değerleri kontrolü (0-1 aralığı)
        if not 0.0 <= self.th <= 1.0:
            errors.append(f"th değeri 0-1 arasında olmalı, verilen: {self.th}")
        
        if self.classic_th is not None and not 0.0 <= self.classic_th <= 1.0:
            errors.append(f"classic_th değeri 0-1 arasında olmalı, verilen: {self.classic_th}")
        
        if not 0.0 <= self.alpha <= 1.0:
            errors.append(f"alpha değeri 0-1 arasında olmalı, verilen: {self.alpha}")
        
        if not 0.0 <= self.yolo_conf <= 1.0:
            errors.append(f"yolo_conf değeri 0-1 arasında olmalı, verilen: {self.yolo_conf}")
        
        if not 0.0 <= self.yolo_iou <= 1.0:
            errors.append(f"yolo_iou değeri 0-1 arasında olmalı, verilen: {self.yolo_iou}")

        # Pozitif değer kontrolleri
        if self.tile <= 0:
            errors.append(f"tile pozitif olmalı, verilen: {self.tile}")
        
        if self.overlap < 0:
            errors.append(f"overlap negatif olamaz, verilen: {self.overlap}")
        
        if self.overlap >= self.tile:
            errors.append(f"overlap ({self.overlap}) tile'dan ({self.tile}) küçük olmalı")
        
        if self.yolo_imgsz <= 0:
            errors.append(f"yolo_imgsz pozitif olmalı, verilen: {self.yolo_imgsz}")
        
        if self.yolo_tile is not None and self.yolo_tile <= 0:
            errors.append(f"yolo_tile pozitif olmalı, verilen: {self.yolo_tile}")

        if self.band_importance_max_tiles < 0:
            errors.append(
                f"band_importance_max_tiles negatif olamaz, verilen: {self.band_importance_max_tiles}"
            )
        if self.norm_sample_tiles <= 0:
            errors.append(f"norm_sample_tiles pozitif olmalı, verilen: {self.norm_sample_tiles}")
        
        if self.min_area < 0:
            errors.append(f"min_area negatif olamaz, verilen: {self.min_area}")

        if self.fallback_candidates_top_k < 0:
            errors.append(
                f"fallback_candidates_top_k negatif olamaz, verilen: {self.fallback_candidates_top_k}"
            )

        if self.fallback_candidates_min_score is not None and not 0.0 <= self.fallback_candidates_min_score <= 1.0:
            errors.append(
                "fallback_candidates_min_score 0-1 arasinda olmali veya null olmali, "
                f"verilen: {self.fallback_candidates_min_score}"
            )
        
        if self.simplify is not None and self.simplify < 0:
            errors.append(f"simplify negatif olamaz, verilen: {self.simplify}")
        
        # Persentil değerleri kontrolü
        if not 0.0 <= self.percentile_low <= 100.0:
            errors.append(f"percentile_low 0-100 arasında olmalı, verilen: {self.percentile_low}")
        
        if not 0.0 <= self.percentile_high <= 100.0:
            errors.append(f"percentile_high 0-100 arasında olmalı, verilen: {self.percentile_high}")
        
        if self.percentile_low >= self.percentile_high:
            errors.append(f"percentile_low ({self.percentile_low}) percentile_high'dan ({self.percentile_high}) küçük olmalı")
        
        # Connectivity kontrolü
        if self.label_connectivity not in (4, 8):
            errors.append(f"label_connectivity 4 veya 8 olmalı, verilen: {self.label_connectivity}")

        self.gpkg_mode = str(self.gpkg_mode).strip().lower() or "single"
        if self.gpkg_mode not in ("single", "split"):
            errors.append(f"gpkg_mode gecersiz: {self.gpkg_mode}. Gecerli degerler: 'single', 'split'")
        
        # Verbose seviyesi
        if self.verbose < 0 or self.verbose > 2:
            errors.append(f"verbose 0-2 arasında olmalı, verilen: {self.verbose}")
        
        # Cihaz kontrolü
        if self.device is not None:
            valid_devices = {"cpu", "cuda", "mps"}
            device_base = self.device.split(":")[0].lower()
            if device_base not in valid_devices and not device_base.startswith("cuda"):
                errors.append(f"device geçersiz: {self.device}. Geçerli değerler: 'cpu', 'cuda', 'cuda:0', 'mps'")

        if str(self.dl_task).strip().lower() not in {"segmentation", "tile_classification"}:
            errors.append(
                f"dl_task gecersiz: {self.dl_task}. Gecerli degerler: 'segmentation', 'tile_classification'"
            )
        feature_mode_text = str(self.feature_mode).strip().lower()
        if feature_mode_text == "auto":
            self.feature_mode = "auto"
        else:
            try:
                self.feature_mode = normalize_shared_feature_mode(feature_mode_text)
            except ValueError as exc:
                errors.append(str(exc))
        if self.use_fpn_classifier not in (None, True, False):
            errors.append("use_fpn_classifier degeri true/false veya null (None) olmalidir.")
        
        # Sigma ve radii kontrolleri
        if len(self.sigma_scales) == 0:
            errors.append("sigma_scales en az bir değer içermeli")
        if any(s <= 0 for s in self.sigma_scales):
            errors.append("sigma_scales tüm değerleri pozitif olmalı")
        
        if len(self.morphology_radii) == 0:
            errors.append("morphology_radii en az bir değer içermeli")
        if any(r <= 0 for r in self.morphology_radii):
            errors.append("morphology_radii tüm değerleri pozitif olmalı")
        
        if len(self.rvt_radii) == 0:
            errors.append("rvt_radii en az bir değer içermeli")
        if any(r <= 0 for r in self.rvt_radii):
            errors.append("rvt_radii tüm değerleri pozitif olmalı")

        # Multi-scale kontrolleri
        if self.enable_multiscale:
            if len(self.multiscale_scales) == 0:
                errors.append("multiscale_scales en az bir ölçek içermeli")
            if any(s <= 0 or s > 1.0 for s in self.multiscale_scales):
                errors.append("multiscale_scales değerleri (0, 1.0] aralığında olmalı")
            if self.multiscale_merge not in ("max", "mean", "weighted_mean"):
                errors.append(f"multiscale_merge geçersiz: {self.multiscale_merge}. max/mean/weighted_mean")
            if self.multiscale_merge == "weighted_mean" and self.multiscale_weights is not None:
                if len(self.multiscale_weights) != len(self.multiscale_scales):
                    errors.append("multiscale_weights uzunluğu multiscale_scales ile eşit olmalı")

        # Cache (NPZ vs raster) kontrolleri
        mode = str(self.cache_derivatives_mode).strip().lower()
        if mode not in ("auto", "npz", "raster"):
            errors.append(f"cache_derivatives_mode geçersiz: {self.cache_derivatives_mode} (auto/npz/raster)")
        if self.deriv_cache_chunk <= 0:
            errors.append(f"deriv_cache_chunk pozitif olmalı, verilen: {self.deriv_cache_chunk}")
        if self.deriv_cache_workers <= 0:
            errors.append(f"deriv_cache_workers pozitif olmalı, verilen: {self.deriv_cache_workers}")
        if self.deriv_cache_halo is not None and self.deriv_cache_halo < 0:
            errors.append(f"deriv_cache_halo negatif olamaz, verilen: {self.deriv_cache_halo}")

        # Single-model (encoders=none) consistency checks
        enc_mode = (self.encoders or "").strip().lower()
        single_mode = enc_mode in ("", "none", "single")
        if single_mode and self.weights_template:
            errors.append(
                "encoders=none (single-model) iken weights_template kullanilamaz; weights kullanin."
            )
        if single_mode and not str(self.encoder or "").strip():
            errors.append("Single-model mode icin encoder degeri bos olamaz.")
        if single_mode and self.enable_deep_learning and not self.weights and not self.zero_shot_imagenet:
            errors.append("Single-model mode icin weights veya zero_shot_imagenet=true gereklidir.")
        
        # Tüm hataları topla ve fırlat
        # trained_model_only: strict trained-checkpoint path
        if self.trained_model_only:
            if not self.enable_deep_learning:
                errors.append("trained_model_only icin enable_deep_learning=true olmalidir.")
            if not self.weights:
                errors.append("trained_model_only icin weights zorunludur.")

        if errors:
            error_msg = "Konfigürasyon doğrulama hataları:\n" + "\n".join(f"  - {e}" for e in errors)
            raise ValueError(error_msg)


DEFAULTS = PipelineDefaults()


def default_for(name: str) -> Any:
    """PipelineDefaults'tan verilen isimli varsayılan değeri döndür."""
    return getattr(DEFAULTS, name)


def help_for(name: str) -> str:
    for f in fields(PipelineDefaults):
        if f.name == name:
            return f.metadata.get("help", "")
    raise KeyError(name)


def cli_help(name: str, extra: Optional[str] = None) -> str:
    base = help_for(name)
    if extra:
        return f"{base} {extra}".strip()
    return base


def load_config_from_yaml(yaml_path: Path) -> Dict[str, Any]:
    """YAML config dosyasını oku ve dictionary olarak döndür."""
    if yaml is None:
        raise ImportError(
            "YAML desteği için PyYAML yükleyin: pip install pyyaml"
        )
    
    if not yaml_path.exists():
        raise FileNotFoundError(f"Config dosyası bulunamadı: {yaml_path}")
    
    with open(yaml_path, 'r', encoding='utf-8') as f:
        config_dict = yaml.safe_load(f)
    
    if config_dict is None:
        config_dict = {}
    
    return config_dict


def collect_cli_overrides(parser: argparse.ArgumentParser, argv: Sequence[str]) -> set[str]:
    """CLI üzerinden açıkça verilen argparse dest isimlerini döndür."""
    overrides: set[str] = set()
    option_actions = getattr(parser, "_option_string_actions", {})
    for raw in argv:
        if not raw or not raw.startswith("-"):
            continue
        opt = raw.split("=", 1)[0]
        action = option_actions.get(opt)
        if action is None:
            continue
        dest = getattr(action, "dest", None)
        if isinstance(dest, str) and dest:
            overrides.add(dest)
    return overrides


def build_config_from_args(
    args: argparse.Namespace, *, cli_overrides: Optional[Iterable[str]] = None
) -> PipelineDefaults:
    """Komut satırı argümanlarından ve opsiyonel YAML config'den PipelineDefaults oluştur."""
    # Önce YAML config'i yükle (varsa)
    base_config = {}
    yaml_supplied_fields: set[str] = set()
    config_dir: Optional[Path] = None
    if hasattr(args, 'config') and args.config:
        yaml_path = Path(args.config)
        try:
            base_config = load_config_from_yaml(yaml_path)
            yaml_supplied_fields = {str(k) for k in base_config.keys()}
            config_dir = yaml_path.resolve(strict=False).parent
            LOGGER.info(f"Config dosyası yüklendi: {yaml_path}")
        except Exception as e:
            LOGGER.warning(f"Config dosyası yüklenemedi ({yaml_path}): {e}")

    fallback_group = base_config.pop("fallback_candidates", None)
    if isinstance(fallback_group, dict):
        fallback_key_map = {
            "enabled": "fallback_candidates_enabled",
            "top_k": "fallback_candidates_top_k",
            "min_score": "fallback_candidates_min_score",
        }
        for nested_key, field_name in fallback_key_map.items():
            if nested_key in fallback_group and field_name not in base_config:
                base_config[field_name] = fallback_group[nested_key]
                yaml_supplied_fields.add(field_name)
        yaml_supplied_fields.discard("fallback_candidates")
    elif fallback_group is not None:
        LOGGER.warning("fallback_candidates config degeri sozluk olmali; yok sayiliyor.")
    
    # YAML'dan gelen listeleri tuple'a çevir
    list_to_tuple_fields = {'sigma_scales', 'morphology_radii', 'rvt_radii', 'tpi_radii',
                             'multiscale_scales', 'multiscale_weights'}
    for field_name in list_to_tuple_fields:
        if field_name in base_config and isinstance(base_config[field_name], list):
            base_config[field_name] = tuple(base_config[field_name])
    
    values = {f.name: default_for(f.name) for f in fields(PipelineDefaults)}
    unknown_config_keys = sorted(str(key) for key in base_config if str(key) not in values)
    if unknown_config_keys:
        LOGGER.warning(
            "Config dosyasindaki desteklenmeyen anahtarlar yok sayiliyor: %s",
            ", ".join(unknown_config_keys),
        )
        for key in unknown_config_keys:
            base_config.pop(key, None)
            yaml_supplied_fields.discard(key)
    values.update(base_config)

    # Komut satırı argümanlarıyla override et (sadece açıkça belirtilenler).
    if cli_overrides is None:
        # Geriye dönük uyumluluk: Eski davranış (yalnızca default'tan farklıysa override).
        # NOT: Bu yaklaşım, "default değerine geri dön" niyetini algılayamaz.
        for f in fields(PipelineDefaults):
            if not hasattr(args, f.name):
                continue
            arg_value = getattr(args, f.name)
            default_value = default_for(f.name)
            if arg_value != default_value:
                values[f.name] = arg_value
    else:
        for name in set(cli_overrides):
            if name not in values or not hasattr(args, name):
                continue
            values[name] = getattr(args, name)

    cwd = Path.cwd()
    cli_override_names = set(cli_overrides or ())
    for name in _CONFIG_PATH_FIELDS:
        if name not in values:
            continue
        if name in cli_override_names:
            base_dir = cwd
        elif name in yaml_supplied_fields and config_dir is not None:
            base_dir = config_dir
        else:
            base_dir = cwd
        values[name] = _normalize_config_path_value(values[name], base_dir)
    
    config = PipelineDefaults(**values)
    apply_yolo_output_defaults(config)
    return config


def apply_yolo_output_defaults(config: PipelineDefaults) -> List[str]:
    """When YOLO is enabled, keep its vector and Excel side outputs enabled."""
    messages: List[str] = []
    if not config.enable_yolo:
        return messages

    if not config.vectorize:
        config.vectorize = True
        messages.append(
            "enable_yolo aktif: YOLO GeoPackage ciktilari icin vectorize=true zorlandi."
        )
    if not config.export_candidate_excel:
        config.export_candidate_excel = True
        messages.append(
            "enable_yolo aktif: YOLO satirlari birlesik Excel'e eklensin diye export_candidate_excel=true zorlandi."
        )
    return messages



def configure_logging(verbosity: int) -> None:
    """Configure global logging level."""
    level = logging.WARNING
    if verbosity == 1:
        level = logging.INFO
    elif verbosity >= 2:
        level = logging.DEBUG
    logging.basicConfig(
        level=level,
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%H:%M:%S",
    )


def set_random_seeds(seed: int) -> None:
    """Enforce deterministic behaviour across numpy and torch."""
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False


def read_windowed(
    raster: rasterio.io.DatasetReader | str | Path,
    window: Window,
    indexes: Optional[Sequence[int]] = None,
    boundless: bool = True,
) -> np.ndarray:
    """Read a raster window from either an open dataset or a file path."""
    if hasattr(raster, "read"):
        dataset = raster
        return dataset.read(indexes=indexes, window=window, boundless=boundless)
    with rasterio.open(raster) as dataset:
        return dataset.read(indexes=indexes, window=window, boundless=boundless)


def compute_ndsm(dsm: Optional[np.ndarray], dtm: np.ndarray) -> np.ndarray:
    """Compute normalised DSM (nDSM). Returns zeros if DSM is missing."""
    ndsm = np.full_like(dtm, np.nan, dtype=np.float32)
    dtm_valid = np.isfinite(dtm)
    if dsm is None:
        ndsm[dtm_valid] = 0.0
        return ndsm
    dsm_valid = np.isfinite(dsm)
    mask = dtm_valid & dsm_valid
    ndsm[mask] = (dsm - dtm)[mask]
    return ndsm


def compute_slope(dtm: np.ndarray, pixel_size: float = 1.0) -> np.ndarray:
    """Compute finite terrain slope in degrees from a DTM."""
    dtm_filled, valid_mask = fill_nodata(dtm)
    pixel = max(float(pixel_size), 1e-6)
    gy, gx = np.gradient(dtm_filled.astype(np.float32, copy=False), pixel, pixel)
    slope = np.degrees(np.arctan(np.hypot(gx, gy))).astype(np.float32)
    slope = np.nan_to_num(slope, nan=0.0, posinf=90.0, neginf=0.0).astype(np.float32)
    slope[~valid_mask] = np.nan
    return slope


def percentile_clip(
    arr: np.ndarray, low: Optional[float] = None, high: Optional[float] = None
) -> np.ndarray:
    """Dizi değerlerini verilen persentiller arasında kırp, NaN değerlerine saygı göster."""
    if low is None:
        low = DEFAULTS.percentile_low
    if high is None:
        high = DEFAULTS.percentile_high
    if arr.size == 0:
        return arr
    valid = np.isfinite(arr)
    if not np.any(valid):
        return np.zeros_like(arr, dtype=np.float32)
    lower = np.percentile(arr[valid], low)
    upper = np.percentile(arr[valid], high)
    if upper - lower <= 1e-6:
        clipped = np.clip(arr - lower, 0.0, 1.0)
    else:
        clipped = np.clip((arr - lower) / (upper - lower), 0.0, 1.0)
    clipped[~valid] = 0.0
    return clipped.astype(np.float32)


def _norm01(arr: np.ndarray) -> np.ndarray:
    """Persentil tabanlı 0-1 ölçekleme için kolaylık sarmalayıcısı."""
    return percentile_clip(arr)


def _otsu_threshold_0to1(arr: np.ndarray, valid: np.ndarray) -> float:
    """0..1 arası veride geçerlilik maskesini dikkate alarak Otsu eşiği hesapla."""
    data = arr[valid]
    if data.size == 0:
        return 0.5
    clipped = np.clip(data, 0.0, 1.0)
    hist, bin_edges = np.histogram(clipped, bins=HISTOGRAM_BINS, range=HISTOGRAM_RANGE)
    if not np.any(hist):
        return float(np.nanmean(clipped)) if np.any(np.isfinite(clipped)) else 0.5
    hist = hist.astype(np.float64)
    prob = hist / hist.sum()
    cumulative = np.cumsum(prob)
    cumulative_mean = np.cumsum(prob * (bin_edges[:-1] + bin_edges[1:]) * 0.5)
    global_mean = cumulative_mean[-1]
    numerator = (global_mean * cumulative - cumulative_mean) ** 2
    denominator = cumulative * (1.0 - cumulative)
    with np.errstate(divide="ignore", invalid="ignore"):
        sigma_b = numerator / np.maximum(denominator, 1e-12)
    sigma_b[cumulative == 0.0] = 0.0
    sigma_b[cumulative == 1.0] = 0.0
    # If there is a plateau of equally-good thresholds (common when there is a wide empty gap),
    # pick the midpoint instead of the first bin. This yields more intuitive thresholds.
    best_val = float(np.max(sigma_b))
    best_idxs = np.flatnonzero(np.isclose(sigma_b, best_val, rtol=0.0, atol=1e-12))
    best = int(best_idxs[len(best_idxs) // 2]) if best_idxs.size else int(np.argmax(sigma_b))
    threshold = (bin_edges[best] + bin_edges[best + 1]) * 0.5
    return float(np.clip(threshold, 0.0, 1.0))


def otsu_threshold_from_hist(hist: np.ndarray, bin_edges: np.ndarray) -> float:
    """Compute Otsu threshold from a histogram (0..1 range)."""
    if hist.size == 0 or not np.any(hist):
        return 0.5
    hist_f = hist.astype(np.float64, copy=False)
    prob = hist_f / hist_f.sum()
    cumulative = np.cumsum(prob)
    bin_centers = (bin_edges[:-1] + bin_edges[1:]) * 0.5
    cumulative_mean = np.cumsum(prob * bin_centers)
    global_mean = cumulative_mean[-1]
    numerator = (global_mean * cumulative - cumulative_mean) ** 2
    denominator = cumulative * (1.0 - cumulative)
    with np.errstate(divide="ignore", invalid="ignore"):
        sigma_b = numerator / np.maximum(denominator, 1e-12)
    sigma_b[cumulative == 0.0] = 0.0
    sigma_b[cumulative == 1.0] = 0.0
    best_val = float(np.max(sigma_b))
    best_idxs = np.flatnonzero(np.isclose(sigma_b, best_val, rtol=0.0, atol=1e-12))
    best = int(best_idxs[len(best_idxs) // 2]) if best_idxs.size else int(np.argmax(sigma_b))
    threshold = (bin_edges[best] + bin_edges[best + 1]) * 0.5
    return float(np.clip(threshold, 0.0, 1.0))


def otsu_threshold_streaming(
    prob_map: np.ndarray,
    valid_global: np.ndarray,
    *,
    block_rows: int = 1024,
) -> float:
    """Compute Otsu threshold without materializing all valid pixels at once."""
    bin_edges = np.linspace(HISTOGRAM_RANGE[0], HISTOGRAM_RANGE[1], HISTOGRAM_BINS + 1)
    hist = np.zeros(HISTOGRAM_BINS, dtype=np.int64)
    height = int(prob_map.shape[0])
    for row0 in range(0, height, int(block_rows)):
        row1 = min(height, row0 + int(block_rows))
        block = prob_map[row0:row1, :]
        valid_block = valid_global[row0:row1, :] & np.isfinite(block)
        if not np.any(valid_block):
            continue
        values = block[valid_block]
        if values.size == 0:
            continue
        values = np.clip(values, HISTOGRAM_RANGE[0], HISTOGRAM_RANGE[1])
        h, _ = np.histogram(values, bins=HISTOGRAM_BINS, range=HISTOGRAM_RANGE)
        hist += h.astype(np.int64, copy=False)
    return otsu_threshold_from_hist(hist, bin_edges)


def _local_variance(arr: np.ndarray, size: Optional[int] = None) -> np.ndarray:
    """Uniform filtreleme kullanarak yerel varyans tahmini yap."""
    if size is None:
        size = DEFAULTS.local_variance_window
    if size <= 1:
        return np.zeros_like(arr, dtype=np.float32)
    arr_f = arr.astype(np.float32, copy=False)
    mean = uniform_filter(arr_f, size=size, mode="nearest")
    mean_sq = uniform_filter(arr_f * arr_f, size=size, mode="nearest")
    var = np.maximum(mean_sq - mean * mean, 0.0)
    return var.astype(np.float32)


def _hessian_response(im: np.ndarray, sigma: float) -> np.ndarray:
    """Hessian matrisinden normalize edilmiş ikinci özdeğer büyüklüğünü hesapla."""
    im_f = im.astype(np.float32, copy=False)
    Hxx = gaussian_filter(im_f, sigma=sigma, order=(2, 0), mode="nearest")
    Hxy = gaussian_filter(im_f, sigma=sigma, order=(1, 1), mode="nearest")
    Hyy = gaussian_filter(im_f, sigma=sigma, order=(0, 2), mode="nearest")
    tr = Hxx + Hyy
    det = Hxx * Hyy - Hxy * Hxy
    tmp = np.sqrt(np.maximum((tr * tr) / 4.0 - det, 0.0))
    l2 = tr / 2.0 - tmp
    return _norm01(np.abs(l2))


def robust_norm(
    arr: np.ndarray, p_low: Optional[float] = None, p_high: Optional[float] = None
) -> np.ndarray:
    """Persentil kırpma ile kanal başına robust min-max normalizasyonu."""
    if p_low is None:
        p_low = DEFAULTS.percentile_low
    if p_high is None:
        p_high = DEFAULTS.percentile_high
    if arr.ndim != 3:
        raise ValueError("robust_norm expects a (C, H, W) array.")
    normed = np.zeros_like(arr, dtype=np.float32)
    for idx in range(arr.shape[0]):
        channel = arr[idx]
        normed[idx] = percentile_clip(channel, low=p_low, high=p_high)
    return normed


def robust_norm_fixed(arr: np.ndarray, lows: np.ndarray, highs: np.ndarray) -> np.ndarray:
    """Normalize (C,H,W) with fixed per-channel lows/highs; NaNsâ†’0."""
    assert arr.ndim == 3
    out = np.zeros_like(arr, dtype=np.float32)
    C = arr.shape[0]
    for c in range(C):
        ch = arr[c]
        low = float(lows[c])
        high = float(highs[c])
        if high - low <= 1e-6:
            norm = np.clip(ch - low, 0.0, 1.0)
        else:
            norm = np.clip((ch - low) / (high - low), 0.0, 1.0)
        norm[~np.isfinite(ch)] = 0.0
        out[c] = norm.astype(np.float32)
    return out


def _raw_channel_names(value: Any) -> Optional[Tuple[str, ...]]:
    if isinstance(value, (list, tuple)) and value:
        names = tuple(str(item).strip() for item in value if str(item).strip())
        return names or None
    return None


def _extract_channel_names(source: Optional[Dict[str, Any]]) -> Optional[Tuple[str, ...]]:
    if not source:
        return None
    return _raw_channel_names(source.get("channel_names"))


def _extract_in_channels(source: Optional[Dict[str, Any]], *keys: str) -> Optional[int]:
    if not source:
        return None
    for key in keys:
        value = source.get(key)
        if isinstance(value, bool):
            continue
        if isinstance(value, int):
            return int(value)
        if isinstance(value, str):
            text = value.strip()
            if not text:
                continue
            try:
                return int(text)
            except ValueError:
                continue
    return None


def normalize_model_channel_names(channel_names: Sequence[str]) -> Tuple[str, ...]:
    names = tuple(str(name).strip() for name in channel_names if str(name).strip())
    if len(names) < 3:
        raise ValueError("channel_names en az RGB kanallarini icermelidir.")
    try:
        return canonicalize_channel_names(names)
    except ValueError as exc:
        raise ValueError(str(exc)) from exc


def band_indexes_support_dsm(band_idx: Sequence[int]) -> bool:
    return bool(len(band_idx) >= 4 and int(band_idx[3]) > 0)


def band_indexes_support_topography(band_idx: Sequence[int]) -> bool:
    return bool(len(band_idx) >= 5 and int(band_idx[4]) > 0)


def infer_channel_names_from_band_indexes(
    band_idx: Sequence[int],
    feature_mode: Optional[object] = None,
) -> Tuple[str, ...]:
    if feature_mode is not None:
        return channel_names_for_feature_mode(feature_mode)
    return expected_channel_names(5 if band_indexes_support_topography(band_idx) else 3)


def resolve_model_channel_names(
    *,
    band_idx: Optional[Sequence[int]] = None,
    checkpoint_hints: Optional[Dict[str, Any]] = None,
    metadata: Optional[Dict[str, Any]] = None,
    feature_mode: Optional[object] = None,
) -> Tuple[str, ...]:
    if feature_mode is not None and str(feature_mode).strip().lower() != "auto":
        return channel_names_for_feature_mode(feature_mode)

    for source_name, source in (
        ("checkpoint metadata", checkpoint_hints),
        ("training metadata", metadata),
    ):
        raw_names = _extract_channel_names(source)
        if raw_names is None:
            continue
        try:
            return normalize_model_channel_names(raw_names)
        except ValueError as exc:
            raise ValueError(f"{source_name} channel_names gecersiz: {exc}") from exc

    for source_name, source, keys in (
        ("checkpoint metadata", checkpoint_hints, ("in_channels", "num_channels")),
        ("training metadata", metadata, ("in_channels", "num_channels")),
    ):
        in_channels = _extract_in_channels(source, *keys)
        if in_channels is None:
            continue
        try:
            return normalize_model_channel_names(expected_channel_names(in_channels))
        except ValueError as exc:
            raise ValueError(
                f"{source_name} in_channels={in_channels} desteklenmiyor: {exc}"
            ) from exc

    if feature_mode is not None and str(feature_mode).strip().lower() != "auto":
        return channel_names_for_feature_mode(feature_mode)
    if band_idx is not None:
        return infer_channel_names_from_band_indexes(band_idx)
    return tuple(MODEL_CHANNEL_NAMES)


def channel_names_require_topography(channel_names: Sequence[str]) -> bool:
    return len(normalize_model_channel_names(channel_names)) > 3


def channel_names_require_dsm(channel_names: Sequence[str]) -> bool:
    return "nDSM" in normalize_model_channel_names(channel_names)


def derivative_band_names_for_channel_names(channel_names: Sequence[str]) -> Tuple[str, ...]:
    """Return derivative cache band names needed by a model channel schema."""
    resolved = normalize_model_channel_names(channel_names)
    names: List[str] = []
    if "SVF" in resolved:
        names.append("svf")
    if "SLRM" in resolved:
        names.append("slrm")
    if "Slope" in resolved:
        names.append("slope")
    if "nDSM" in resolved:
        names.append("ndsm")
    return tuple(names)


def _source_declared_channel_names(source: Optional[Dict[str, Any]]) -> Optional[Tuple[str, ...]]:
    """Resolve channel schema declared by checkpoint/training metadata."""
    raw_names = _extract_channel_names(source)
    if raw_names is not None:
        return normalize_model_channel_names(raw_names)
    in_channels = _extract_in_channels(source, "in_channels", "num_channels")
    if in_channels is None:
        return None
    return normalize_model_channel_names(expected_channel_names(in_channels))


def assemble_model_input(
    rgb: np.ndarray,
    *,
    channel_names: Sequence[str],
    svf: Optional[np.ndarray] = None,
    slrm: Optional[np.ndarray] = None,
    slope: Optional[np.ndarray] = None,
    ndsm: Optional[np.ndarray] = None,
) -> np.ndarray:
    resolved_names = normalize_model_channel_names(channel_names)
    rgb_arr = np.asarray(rgb, dtype=np.float32)
    if rgb_arr.ndim != 3 or rgb_arr.shape[0] != 3:
        raise ValueError("assemble_model_input expects rgb with shape (3, H, W).")

    spatial_shape = rgb_arr.shape[1:]
    channel_map: Dict[str, np.ndarray] = {
        "R": rgb_arr[0],
        "G": rgb_arr[1],
        "B": rgb_arr[2],
    }
    for name, arr in (("SVF", svf), ("SLRM", slrm), ("Slope", slope), ("nDSM", ndsm)):
        if name not in resolved_names:
            continue
        if arr is None:
            raise ValueError(f"{name} channel requested but no array was provided.")
        arr_f = np.asarray(arr, dtype=np.float32)
        if arr_f.ndim != 2:
            raise ValueError(f"{name} channel must have shape (H, W).")
        if arr_f.shape != spatial_shape:
            raise ValueError(
                f"{name} shape {arr_f.shape} does not match RGB spatial shape {spatial_shape}."
            )
        channel_map[name] = arr_f

    stacked = np.stack([channel_map[name] for name in resolved_names], axis=0)
    return stacked.astype(np.float32, copy=False)


def stack_channels(rgb: np.ndarray, svf: np.ndarray, slrm: np.ndarray) -> np.ndarray:
    """Build the canonical 5-channel model tensor [R, G, B, SVF, SLRM]."""
    return assemble_model_input(
        rgb,
        channel_names=MODEL_CHANNEL_NAMES,
        svf=svf,
        slrm=slrm,
    )


def stack_topo7_channels(
    rgb: np.ndarray,
    svf: np.ndarray,
    slrm: np.ndarray,
    slope: np.ndarray,
    ndsm: np.ndarray,
) -> np.ndarray:
    """Build the canonical topo7 tensor [R, G, B, SVF, SLRM, Slope, nDSM]."""
    return assemble_model_input(
        rgb,
        channel_names=TOPO7_CHANNEL_NAMES,
        svf=svf,
        slrm=slrm,
        slope=slope,
        ndsm=ndsm,
    )


def fill_nodata(
    arr: np.ndarray, fill_value: Optional[float] = None
) -> Tuple[np.ndarray, np.ndarray]:
    """Fill NaNs in an array with a constant (fallback median if not provided)."""
    filled = arr.astype(np.float32, copy=True)
    invalid_mask = ~np.isfinite(filled)
    if not np.any(invalid_mask):
        return filled, ~invalid_mask
    if fill_value is None:
        valid_vals = filled[~invalid_mask]
        fill_value = float(np.nanmedian(valid_vals)) if valid_vals.size else 0.0
    filled[invalid_mask] = fill_value
    return filled, ~invalid_mask


# ==============================================================================
# Topographic feature helpers
# ==============================================================================

def compute_curvatures(
    dtm: np.ndarray,
    pixel_size: float = 1.0,
) -> Tuple[np.ndarray, np.ndarray]:
    """Compute plan and profile curvature from a DTM."""
    dtm_filled, valid_mask = fill_nodata(dtm)

    fy, fx = np.gradient(dtm_filled, pixel_size)
    fyy, _fyx = np.gradient(fy, pixel_size)
    fxy, fxx = np.gradient(fx, pixel_size)

    p = fx ** 2
    q = fy ** 2
    denominator = np.where((p + q) < 1e-10, 1e-10, p + q)

    plan_curv = -((fxx * q) - (2 * fxy * fx * fy) + (fyy * p)) / (denominator ** 1.5)
    profile_curv = -((fxx * p) + (2 * fxy * fx * fy) + (fyy * q)) / (denominator ** 1.5)

    plan_curv = np.nan_to_num(plan_curv, nan=0.0, posinf=0.0, neginf=0.0)
    profile_curv = np.nan_to_num(profile_curv, nan=0.0, posinf=0.0, neginf=0.0)
    plan_curv = np.clip(plan_curv, -0.1, 0.1)
    profile_curv = np.clip(profile_curv, -0.1, 0.1)

    plan_curv[~valid_mask] = np.nan
    profile_curv[~valid_mask] = np.nan
    return plan_curv.astype(np.float32), profile_curv.astype(np.float32)


def compute_tpi_multiscale(
    dtm: np.ndarray,
    radii: Tuple[int, ...] = (5, 15, 30),
) -> np.ndarray:
    """Compute a simple multi-scale topographic position index."""
    dtm_filled, valid_mask = fill_nodata(dtm)
    tpi_stack: List[np.ndarray] = []

    for radius in radii:
        window_size = 2 * int(radius) + 1
        mean_elevation = uniform_filter(dtm_filled, size=window_size, mode="reflect")
        tpi = dtm_filled - mean_elevation
        tpi_std = np.nanstd(tpi[valid_mask]) if np.any(valid_mask) else 1.0
        if tpi_std > 0:
            tpi = tpi / (3 * tpi_std)
        tpi_stack.append(np.clip(tpi, -1.0, 1.0))

    tpi_combined = np.nanmean(np.stack(tpi_stack, axis=0), axis=0)
    tpi_combined[~valid_mask] = np.nan
    return tpi_combined.astype(np.float32)


def _rvt_as_float32(result_obj: Any, name: str) -> np.ndarray:
    """Best-effort conversion of RVT outputs to float32 ndarrays."""
    if isinstance(result_obj, np.ndarray):
        return result_obj.astype(np.float32)
    if isinstance(result_obj, np.ma.MaskedArray):
        return np.ma.filled(result_obj, np.nan).astype(np.float32)
    if isinstance(result_obj, dict):
        preferred_keys = (
            "svf",
            "SVF",
            "slrm",
            "SLRM",
            "lrm",
            "LRM",
            "positive_openness",
            "pos_open",
            "negative_openness",
            "neg_open",
            "slope",
            "Slope",
            "opns",
            "OPNS",
        )
        for key in preferred_keys:
            if key in result_obj and isinstance(result_obj[key], (np.ndarray, np.ma.MaskedArray)):
                return _rvt_as_float32(result_obj[key], name)
        for value in result_obj.values():
            try:
                return _rvt_as_float32(value, name)
            except Exception:
                continue
        raise TypeError(f"RVT '{name}' returned dict without ndarray values.")
    if isinstance(result_obj, (list, tuple)):
        for value in result_obj:
            try:
                return _rvt_as_float32(value, name)
            except Exception:
                continue
        raise TypeError(f"RVT '{name}' returned a sequence without ndarray values.")
    try:
        arr = np.asarray(result_obj)
        if isinstance(arr, np.ndarray):
            return arr.astype(np.float32)
    except Exception:
        pass
    raise TypeError(f"RVT '{name}' returned unsupported type: {type(result_obj)}")


def compute_derivatives_with_rvt(
    dtm: np.ndarray,
    pixel_size: float,
    radii: Optional[Sequence[float]] = None,
    gaussian_lrm_sigma: Optional[float] = None,
    *,
    show_progress: bool = True,
    log_steps: bool = True,
) -> Tuple[np.ndarray, np.ndarray]:
    """Compute SVF and SLRM from a DTM tile for the 5-band model."""
    del show_progress  # kept for backward-compatible call sites

    if rvt_vis is None:
        raise ImportError("Install rvt-py: pip install rvt")

    radii = DEFAULTS.rvt_radii if radii is None else radii
    dtm_filled, valid_mask = fill_nodata(dtm)
    log_fn = LOGGER.info if log_steps else LOGGER.debug

    def _radius_to_cells(radius_val: float) -> int:
        pixel = max(float(pixel_size), 1e-6)
        return max(1, int(round(float(radius_val) / pixel)))

    def _call_sky_view_factor(dem_arr: np.ndarray, radius_val: float) -> Any:
        radius_px = _radius_to_cells(radius_val)
        candidates = (
            {
                "dem": dem_arr,
                "resolution": pixel_size,
                "compute_svf": True,
                "compute_asvf": False,
                "compute_opns": False,
                "svf_r_max": radius_px,
                "svf_noise": 0,
                "no_data": None,
            },
            {
                "dem": dem_arr,
                "resolution": pixel_size,
                "compute_svf": True,
                "compute_asvf": False,
                "compute_opns": False,
                "svf_r_max": radius_px,
                "svf_noise": 0,
            },
            {
                "dem": dem_arr,
                "resolution": pixel_size,
                "compute_svf": True,
                "compute_opns": False,
                "svf_r_max": radius_px,
                "svf_noise": 0,
            },
        )
        last_err: Optional[Exception] = None
        for kwargs in candidates:
            try:
                return rvt_vis.sky_view_factor(**kwargs)
            except TypeError as exc:
                last_err = exc
        try:
            return rvt_vis.sky_view_factor(dem=dem_arr, resolution=pixel_size)
        except Exception as exc:
            if last_err is not None:
                raise last_err
            raise exc

    def _extract_named(result_obj: Any, keys: Sequence[str], name: str) -> np.ndarray:
        if isinstance(result_obj, dict):
            for key in keys:
                if key in result_obj:
                    return _rvt_as_float32(result_obj[key], name)
        return _rvt_as_float32(result_obj, name)

    log_fn("  -> SVF (Sky View Factor) hesaplanıyor...")
    svf_radius = float(radii[0]) if radii else 10.0
    log_fn("    SVF radius=%.1f m (%d px)", svf_radius, _radius_to_cells(svf_radius))
    svf_res = _call_sky_view_factor(dtm_filled, svf_radius)
    svf_avg = _extract_named(svf_res, ("svf", "SVF"), "sky_view_factor")

    log_fn("  -> SLRM (Simple Local Relief Model) hesaplanıyor...")
    from math import ceil as _ceil

    slrm_rad_cell = int(min(50, max(10, _ceil(10.0 / max(float(pixel_size), 1e-6)))))
    log_fn("    SLRM rad_cell=%d px (pixel_size=%.3f m)", slrm_rad_cell, pixel_size)

    slrm: np.ndarray
    slrm_ready = False
    if hasattr(rvt_vis, "slrm"):
        for try_kwargs in (
            {"dem": dtm_filled, "radius_cell": slrm_rad_cell, "no_data": None},
            {"dem": dtm_filled, "radius_cell": slrm_rad_cell},
        ):
            try:
                slrm = _rvt_as_float32(rvt_vis.slrm(**try_kwargs), "slrm")
                slrm_ready = True
                break
            except TypeError:
                continue
    if not slrm_ready and hasattr(rvt_vis, "local_relief_model"):
        radius_keys = ("search_radius", "radius", "r_max", "max_radius")
        for radius_key in radius_keys:
            try:
                slrm = _rvt_as_float32(
                    rvt_vis.local_relief_model(
                        dem=dtm_filled,
                        resolution=pixel_size,
                        **{radius_key: float(max(radii)), "no_data": None},
                    ),
                    "local_relief_model",
                )
                slrm_ready = True
                break
            except TypeError:
                continue
    if not slrm_ready:
        _warn_once("rvt_slrm_missing", "RVT SLRM fonksiyonu eksik; Gaussian fallback kullaniliyor.")
        sigma = (
            DEFAULTS.gaussian_lrm_sigma
            if gaussian_lrm_sigma is None
            else float(gaussian_lrm_sigma)
        )
        low_pass = gaussian_filter(dtm_filled, sigma=sigma)
        slrm = (dtm_filled - low_pass).astype(np.float32)

    svf_avg[~valid_mask] = np.nan
    slrm[~valid_mask] = np.nan
    return svf_avg.astype(np.float32, copy=False), slrm.astype(np.float32, copy=False)


def _score_rvtlog(
    dtm: np.ndarray,
    pixel_size: float,
    pre_svf: Optional[np.ndarray] = None,
    pre_neg_open: Optional[np.ndarray] = None,
    pre_lrm: Optional[np.ndarray] = None,
    pre_slope: Optional[np.ndarray] = None,
    sigmas: Optional[Sequence[float]] = None,
    gaussian_gradient_sigma: Optional[float] = None,
    local_variance_window: Optional[int] = None,
    rvt_radii: Optional[Sequence[float]] = None,
    gaussian_lrm_sigma: Optional[float] = None,
) -> np.ndarray:
    """Compute the combined RVT + LoG + gradient classical score."""
    dtm_filled, valid = fill_nodata(dtm)
    if pre_svf is None or pre_lrm is None:
        svf, slrm = compute_derivatives_with_rvt(
            dtm,
            pixel_size=pixel_size,
            radii=rvt_radii,
            gaussian_lrm_sigma=gaussian_lrm_sigma,
            show_progress=False,
            log_steps=False,
        )
        if pre_svf is None:
            pre_svf = svf
        if pre_lrm is None:
            pre_lrm = slrm
    svf = pre_svf
    lrm = pre_lrm
    neg = pre_neg_open if pre_neg_open is not None else np.zeros_like(dtm_filled, dtype=np.float32)
    resolved_sigmas = DEFAULTS.sigma_scales if sigmas is None else sigmas
    log_responses = [
        np.abs(gaussian_laplace(dtm_filled, sigma=sigma, mode="nearest"))
        for sigma in resolved_sigmas
    ]
    blob = np.maximum.reduce(log_responses)
    if pre_slope is None:
        grad_sigma = (
            DEFAULTS.gaussian_gradient_sigma
            if gaussian_gradient_sigma is None
            else float(gaussian_gradient_sigma)
        )
        grad = gaussian_gradient_magnitude(dtm_filled, sigma=grad_sigma, mode="nearest")
        slope_term = _norm01(grad)
    else:
        slope_term = _norm01(pre_slope)
    lrm_n = _norm01(lrm)
    svf_c = 1.0 - _norm01(svf)
    neg_n = _norm01(neg)
    variance_window = (
        DEFAULTS.local_variance_window
        if local_variance_window is None
        else int(local_variance_window)
    )
    var_n = _norm01(_local_variance(dtm_filled, size=variance_window))
    score = (
        0.30 * _norm01(blob)
        + 0.20 * lrm_n
        + 0.15 * _norm01(svf_c)
        + 0.15 * slope_term
        + 0.10 * neg_n
        + 0.10 * var_n
    )
    score[~valid] = np.nan
    return score.astype(np.float32)


def _score_hessian(dtm: np.ndarray, sigmas: Optional[Sequence[float]] = None) -> np.ndarray:
    """Compute a multi-scale Hessian-based ridge/valley response."""
    dtm_filled, valid = fill_nodata(dtm)
    resolved_sigmas = DEFAULTS.sigma_scales if sigmas is None else sigmas
    responses = [_hessian_response(dtm_filled, sigma=sigma) for sigma in resolved_sigmas]
    score = np.maximum.reduce(responses)
    score[~valid] = np.nan
    return score.astype(np.float32)


def _score_morph(dtm: np.ndarray, radii: Optional[Sequence[int]] = None) -> np.ndarray:
    """Compute a morphological top-hat prominence score."""
    dtm_filled, valid = fill_nodata(dtm)
    resolved_radii = DEFAULTS.morphology_radii if radii is None else radii
    wth_list: List[np.ndarray] = []
    bth_list: List[np.ndarray] = []
    for radius in resolved_radii:
        size = (int(radius), int(radius))
        opening = grey_opening(dtm_filled, size=size, mode="nearest")
        closing = grey_closing(dtm_filled, size=size, mode="nearest")
        wth_list.append(_norm01(dtm_filled - opening))
        bth_list.append(_norm01(closing - dtm_filled))
    score = np.maximum(np.maximum.reduce(wth_list), np.maximum.reduce(bth_list))
    score[~valid] = np.nan
    return score.astype(np.float32)


# ==============================================================================
# Shared attention layers live in archeo_shared.modeling.
# ==============================================================================

def get_num_channels(*_args, **_kwargs) -> int:
    """Return the resolved model channel count, keeping legacy unused args compatible."""
    channel_names = _kwargs.get("channel_names")
    band_idx = _kwargs.get("band_idx")
    checkpoint_hints = _kwargs.get("checkpoint_hints")
    metadata = _kwargs.get("metadata")
    feature_mode = _kwargs.get("feature_mode")
    if any(value is not None for value in (channel_names, band_idx, checkpoint_hints, metadata, feature_mode)):
        if channel_names is not None:
            return len(normalize_model_channel_names(channel_names))
        return len(
            resolve_model_channel_names(
                band_idx=band_idx,
                checkpoint_hints=checkpoint_hints,
                metadata=metadata,
                feature_mode=feature_mode,
            )
        )
    return len(MODEL_CHANNEL_NAMES)


class TileClassifier(torch.nn.Module):
    """SMP encoder tabanli tile-level binary classifier.

    use_fpn=True ise encoder'in tum katmanlarindan feature toplar
    (FPN-style multi-level aggregation).
    """

    def __init__(
        self,
        *,
        encoder_name: str,
        in_channels: int,
        encoder_weights: Optional[str] = "imagenet",
        dropout: float = 0.2,
        use_fpn: bool = False,
    ):
        super().__init__()
        if smp is None:
            raise ImportError("Install segmentation_models_pytorch: pip install segmentation-models-pytorch")
        self.encoder = smp.encoders.get_encoder(
            encoder_name,
            in_channels=in_channels,
            depth=5,
            weights=encoder_weights,
        )
        out_channels = list(getattr(self.encoder, "out_channels", []))
        if not out_channels:
            raise ValueError(f"Encoder cikis kanallari okunamadi: {encoder_name}")

        self.use_fpn = bool(use_fpn)
        self.pool = torch.nn.AdaptiveAvgPool2d(1)
        self.dropout = (
            torch.nn.Dropout(p=float(dropout))
            if float(dropout) > 0.0
            else torch.nn.Identity()
        )
        if self.use_fpn and len(out_channels) > 1:
            fpn_ch = sum(int(c) for c in out_channels[1:])
            self.classifier = torch.nn.Linear(fpn_ch, 1)
        else:
            self.use_fpn = False
            self.classifier = torch.nn.Linear(int(out_channels[-1]), 1)

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        feats = self.encoder(x)
        if not isinstance(feats, (list, tuple)):
            feats = [feats]
        if self.use_fpn and len(feats) > 1:
            pooled = [self.pool(f).flatten(1) for f in feats[1:]]
            x = torch.cat(pooled, dim=1)
        else:
            x = self.pool(feats[-1]).flatten(1)
        x = self.dropout(x)
        return self.classifier(x)


def build_model(
    arch: str = "Unet",
    encoder: str = "resnet34",
    in_ch: int = 5,
    enable_attention: bool = True,
    attention_reduction: int = 4,
) -> torch.nn.Module:
    """
    Segmentation model oluşturur (opsiyonel CBAM attention ile).
    
    Args:
        arch: Model mimarisi (Unet, UnetPlusPlus, DeepLabV3Plus, vb.)
        encoder: Encoder ismi (resnet34, resnet50, efficientnet-b3, vb.)
        in_ch: Giriş kanal sayısı (5: R,G,B,SVF,SLRM)
        enable_attention: CBAM attention modülünü etkinleştir
        attention_reduction: Attention için kanal azaltma oranı
        
    Returns:
        PyTorch modeli (attention sarmalayıcılı veya değil)
    """
    if smp is None:
        raise ImportError("Install segmentation_models_pytorch: pip install segmentation-models-pytorch")
    if not hasattr(smp, arch):
        raise ValueError(f"Architecture '{arch}' not found in segmentation_models_pytorch.")
    model_cls = getattr(smp, arch)
    base_model = model_cls(
        encoder_name=encoder,
        in_channels=in_ch,
        classes=1,
        activation=None,
    )
    
    # Attention modülü ekle (istenirse)
    if enable_attention:
        model = AttentionWrapper(base_model, in_ch, reduction=attention_reduction)
        LOGGER.debug(f"CBAM Attention modülü eklendi (reduction={attention_reduction})")
    else:
        model = base_model
    
    return model


def build_tile_classifier(
    encoder: str = "resnet34",
    in_ch: int = 5,
    enable_attention: bool = True,
    attention_reduction: int = 4,
    encoder_weights: Optional[str] = None,
    use_fpn: bool = False,
) -> torch.nn.Module:
    """Tile-level classification modeli olusturur."""
    base_model = TileClassifier(
        encoder_name=encoder,
        in_channels=in_ch,
        encoder_weights=encoder_weights,
        use_fpn=use_fpn,
    )
    if enable_attention:
        model = AttentionWrapper(base_model, in_ch, reduction=attention_reduction)
        LOGGER.debug(
            "CBAM Attention modulu eklendi (tile classifier, reduction=%d)",
            attention_reduction,
        )
    else:
        model = base_model
    return model


def inflate_conv1_to_n(conv_w: torch.Tensor, in_ch: int = 5, mode: str = "avg") -> torch.Tensor:
    """
    Inflate a first-conv weight tensor (out,c,kH,kW) to match the requested in_ch.
    - If channels already match, the tensor is returned unchanged.
    - If fewer channels are present, extra channels are filled using the average (or first) channel.
    - If more channels exist than requested, the tensor is truncated.
    """
    if conv_w.ndim != 4:
        raise ValueError("inflate_conv1_to_n expects a 4D convolution weight tensor.")
    out_ch, c_in, kH, kW = conv_w.shape
    if c_in == in_ch:
        return conv_w.clone()
    if in_ch < c_in:
        return conv_w[:, :in_ch].clone()

    if mode == "avg":
        base = conv_w.mean(dim=1, keepdim=True)
    elif mode == "first":
        base = conv_w[:, :1]
    else:
        raise ValueError(f"Unsupported inflate mode: {mode}")

    extra = base.repeat(1, in_ch - c_in, 1, 1)
    return torch.cat([conv_w, extra], dim=1).clone()


def build_model_with_imagenet_inflated(
    arch: str = "Unet", 
    encoder: str = "resnet34", 
    in_ch: int = 5,
    enable_attention: bool = True,
    attention_reduction: int = 4,
) -> torch.nn.Module:
    """
    Build a model that uses an ImageNet-pretrained 3-ch encoder and inflate its first conv to in_ch.
    
    ImageNet encoder'ları 3 kanal (RGB) için eğitilmiştir. Bu fonksiyon:
    1. ImageNet ağırlıklarını yükler
    2. İlk konvolüsyon katmanını 3 kanaldan in_ch kanala genişletir ("inflate")
    3. Opsiyonel olarak CBAM attention ekler
    
    Args:
        arch: Model mimarisi
        encoder: Encoder ismi
        in_ch: Hedef kanal sayısı (5: R,G,B,SVF,SLRM)
        enable_attention: CBAM attention ekle
        attention_reduction: Attention için kanal azaltma oranı
    
    Decoder weights come from the 3-ch SMP model where shapes match; others remain default-initialized.
    """
    if smp is None:
        raise ImportError("Install segmentation_models_pytorch: pip install segmentation-models-pytorch")
    if not hasattr(smp, arch):
        raise ValueError(f"Architecture '{arch}' not found in segmentation_models_pytorch.")

    model_3 = getattr(smp, arch)(
        encoder_name=encoder, encoder_weights="imagenet", in_channels=3, classes=1, activation=None
    )
    state_3 = model_3.state_dict()

    model_9 = getattr(smp, arch)(
        encoder_name=encoder, encoder_weights=None, in_channels=in_ch, classes=1, activation=None
    )
    state_9 = model_9.state_dict()

    for k, v in state_3.items():
        if k in state_9 and state_9[k].shape == v.shape:
            state_9[k] = v.clone()

    def _first_conv_key(state_dict: dict) -> Optional[str]:
        convs: List[Tuple[str, torch.Tensor]] = []
        for key, val in state_dict.items():
            if not key.endswith(".weight"):
                continue
            if getattr(val, "ndim", 0) != 4:
                continue
            convs.append((key, val))
        if not convs:
            return None

        def priority(item: Tuple[str, torch.Tensor]) -> Tuple[int, str]:
            key, tensor = item
            score = 0
            if "conv1" in key:
                score -= 20
            if "stem" in key:
                score -= 10
            if key.startswith("encoder."):
                score -= 5
            # prefer smaller input channel counts (likely RGB)
            score += int(tensor.shape[1])
            return (score, key)

        # Prefer weights with exactly three input channels, then any other.
        for target_channels in (3, 1, None):
            filtered = [
                item for item in convs if target_channels is None or item[1].shape[1] == target_channels
            ]
            if not filtered:
                continue
            filtered.sort(key=priority)
            return filtered[0][0]

        convs.sort(key=priority)
        return convs[0][0]

    conv1_key_src = _first_conv_key(state_3)
    conv1_key_tgt = _first_conv_key(state_9)
    if conv1_key_src is None or conv1_key_tgt is None:
        raise RuntimeError("Could not locate a 4D first-conv weight to inflate.")

    w3 = state_3[conv1_key_src]  # (out,3,kH,kW)
    state_9[conv1_key_tgt] = inflate_conv1_to_n(w3, in_ch=in_ch, mode="avg")

    model_9.load_state_dict(state_9, strict=False)
    
    # Attention modülü ekle (istenirse)
    if enable_attention:
        model = AttentionWrapper(model_9, in_ch, reduction=attention_reduction)
        LOGGER.debug(f"CBAM Attention modülü eklendi (ImageNet inflated, reduction={attention_reduction})")
    else:
        model = model_9
    
    return model


def safe_torch_load(weights_path: Path, map_location: Any) -> Any:
    """Load checkpoints safely when possible, with backward-compatible fallbacks."""
    try:
        return torch.load(weights_path, map_location=map_location, weights_only=True)
    except TypeError:
        # Older torch versions do not support `weights_only`.
        return torch.load(weights_path, map_location=map_location)
    except Exception as exc:
        LOGGER.debug(
            "weights_only checkpoint load failed for %s (%s). Retrying full load.",
            weights_path,
            exc,
        )
        return torch.load(weights_path, map_location=map_location, weights_only=False)


def _extract_checkpoint_state_dict(state_obj: Any) -> Optional[Dict[str, Any]]:
    """Extract the tensor state dict from raw torch.load output."""
    if not isinstance(state_obj, dict):
        return None
    for key in ("state_dict", "model_state_dict"):
        maybe_state = state_obj.get(key)
        if isinstance(maybe_state, dict):
            return maybe_state
    if any(isinstance(v, torch.Tensor) for v in state_obj.values()):
        return state_obj
    return None


def _infer_checkpoint_in_channels_from_state_dict(
    state_dict: Optional[Dict[str, Any]],
) -> Optional[int]:
    """Infer encoder input channels from the first convolution in a checkpoint."""
    if not state_dict:
        return None
    preferred_suffixes = (
        "encoder.conv1.weight",
        "base_model.encoder.conv1.weight",
        "module.encoder.conv1.weight",
        "module.base_model.encoder.conv1.weight",
    )
    for suffix in preferred_suffixes:
        for key, tensor in state_dict.items():
            if isinstance(key, str) and key.endswith(suffix):
                if isinstance(tensor, torch.Tensor) and tensor.ndim == 4:
                    return int(tensor.shape[1])
    for tensor in state_dict.values():
        if isinstance(tensor, torch.Tensor) and tensor.ndim == 4:
            channels = int(tensor.shape[1])
            if channels in (3, 5, 7):
                return channels
    return None


def _infer_checkpoint_task_type_from_state_dict(state_dict: Optional[Dict[str, Any]]) -> Optional[str]:
    """Infer checkpoint task type from state-dict keys when metadata is missing."""
    if not state_dict:
        return None
    keys = tuple(str(key) for key in state_dict.keys())
    has_segmentation_head = any("segmentation_head" in key for key in keys)
    has_classifier = any(
        key.endswith("classifier.weight")
        or key.endswith("classifier.bias")
        or ".classifier." in key
        for key in keys
    )
    if has_segmentation_head and not has_classifier:
        return "segmentation"
    if has_classifier and not has_segmentation_head:
        return "tile_classification"
    return None


def _checkpoint_classifier_in_features(state_dict: Optional[Dict[str, Any]]) -> Optional[int]:
    """Return the classifier input width from a checkpoint state dict when available."""
    if not state_dict:
        return None
    preferred_suffixes = (
        "classifier.weight",
        "base_model.classifier.weight",
        "module.classifier.weight",
        "module.base_model.classifier.weight",
    )
    for suffix in preferred_suffixes:
        for key, tensor in state_dict.items():
            if isinstance(key, str) and key.endswith(suffix):
                if isinstance(tensor, torch.Tensor) and tensor.ndim == 2:
                    return int(tensor.shape[1])
    for key, tensor in state_dict.items():
        if not isinstance(key, str) or "classifier.weight" not in key:
            continue
        if isinstance(tensor, torch.Tensor) and tensor.ndim == 2:
            return int(tensor.shape[1])
    return None


def _resolve_tile_classifier_feature_dims(
    encoder_name: str,
    in_channels: int,
) -> Optional[Tuple[int, int, int]]:
    """Resolve TileClassifier feature dims: last, current FPN, legacy all-level FPN."""
    if smp is None:
        return None
    try:
        encoder = smp.encoders.get_encoder(
            encoder_name,
            in_channels=in_channels,
            depth=5,
            weights=None,
        )
    except Exception as exc:
        LOGGER.debug(
            "Could not inspect TileClassifier feature dims for encoder=%s in_channels=%s: %s",
            encoder_name,
            in_channels,
            exc,
        )
        return None

    out_channels = list(getattr(encoder, "out_channels", []))
    if not out_channels:
        return None
    last_dim = int(out_channels[-1])
    fpn_dim = sum(int(c) for c in out_channels[1:]) if len(out_channels) > 1 else last_dim
    legacy_fpn_dim = sum(int(c) for c in out_channels)
    return last_dim, fpn_dim, legacy_fpn_dim


def _infer_tile_classifier_use_fpn_from_state_dict(
    state_dict: Optional[Dict[str, Any]],
    *,
    encoder_name: str,
    in_channels: int,
) -> Optional[bool]:
    """Infer whether a TileClassifier checkpoint used FPN aggregation."""
    classifier_in_features = _checkpoint_classifier_in_features(state_dict)
    if classifier_in_features is None:
        return None

    feature_dims = _resolve_tile_classifier_feature_dims(encoder_name, int(in_channels))
    if feature_dims is None:
        return None
    last_dim, fpn_dim, legacy_fpn_dim = feature_dims

    if classifier_in_features == last_dim:
        return False
    if classifier_in_features in {fpn_dim, legacy_fpn_dim} and classifier_in_features != last_dim:
        return True
    return None


def _checkpoint_optional_bool(config_dict: Dict[str, Any], key: str) -> Optional[bool]:
    value = config_dict.get(key)
    return value if isinstance(value, bool) else None


def resolve_tile_classifier_use_fpn(
    configured_value: Optional[bool],
    checkpoint_hints: Optional[Dict[str, Any]] = None,
) -> bool:
    """Resolve TileClassifier FPN flag from checkpoint hints, config, then fallback."""
    if checkpoint_hints:
        hint_value = checkpoint_hints.get("use_fpn_classifier")
        if isinstance(hint_value, bool):
            return hint_value
    if isinstance(configured_value, bool):
        return configured_value
    return False


def get_checkpoint_model_hints(weights_path: Path) -> Dict[str, Any]:
    """Extract optional model metadata from a checkpoint."""
    if not weights_path.exists():
        return {}
    try:
        state_obj = safe_torch_load(weights_path, map_location="cpu")
    except Exception as exc:
        LOGGER.debug("Could not inspect checkpoint metadata from %s: %s", weights_path, exc)
        return {}

    if not isinstance(state_obj, dict):
        return {}
    cfg = state_obj.get("config")

    hints: Dict[str, Any] = {}
    if isinstance(cfg, dict):
        for key in (
            "arch",
            "encoder",
            "in_channels",
            "enable_attention",
            "attention_reduction",
            "use_fpn_classifier",
            "task_type",
            "channel_names",
            "bands",
            "tile_size",
            "overlap",
            "input_file",
            "mask_file",
            "schema_version",
        ):
            if key in cfg:
                hints[key] = cfg[key]
    state_dict = _extract_checkpoint_state_dict(state_obj)
    if "in_channels" not in hints:
        inferred_in_channels = _infer_checkpoint_in_channels_from_state_dict(state_dict)
        if inferred_in_channels is not None:
            hints["in_channels"] = inferred_in_channels
    inferred_task = _infer_checkpoint_task_type_from_state_dict(state_dict)
    if inferred_task and "task_type" not in hints:
        hints["task_type"] = inferred_task
    task_type = str(hints.get("task_type", inferred_task or "")).strip().lower()
    if "use_fpn_classifier" not in hints and task_type == "tile_classification":
        encoder_name = str(hints.get("encoder", "")).strip()
        in_channels = hints.get("in_channels")
        if encoder_name and isinstance(in_channels, int):
            inferred_use_fpn = _infer_tile_classifier_use_fpn_from_state_dict(
                state_dict,
                encoder_name=encoder_name,
                in_channels=in_channels,
            )
            if inferred_use_fpn is not None:
                hints["use_fpn_classifier"] = inferred_use_fpn
    return hints


def load_training_metadata_hints(metadata_path: Path) -> Dict[str, Any]:
    """Load optional training-data metadata for consistency warnings."""
    if not metadata_path.exists():
        return {}
    try:
        raw = json.loads(metadata_path.read_text(encoding="utf-8"))
    except Exception as exc:
        LOGGER.debug("Could not read training metadata from %s: %s", metadata_path, exc)
        return {}
    return raw if isinstance(raw, dict) else {}


def _metadata_from_checkpoint_hints(checkpoint_hints: Dict[str, Any]) -> Dict[str, Any]:
    """Synthesize training metadata from checkpoint hints when JSON is unavailable."""
    if not checkpoint_hints:
        return {}
    metadata: Dict[str, Any] = {}
    for key in (
        "schema_version",
        "task_type",
        "arch",
        "encoder",
        "in_channels",
        "channel_names",
        "bands",
        "tile_size",
        "overlap",
        "input_file",
        "mask_file",
        "use_fpn_classifier",
    ):
        value = checkpoint_hints.get(key)
        if value in (None, ""):
            continue
        metadata[key] = list(value) if isinstance(value, tuple) else value
    return metadata


def _candidate_training_metadata_paths(
    configured_path: Optional[Path],
    weights_path: Optional[Path],
) -> List[Path]:
    """Return likely training metadata locations in priority order."""
    candidates: List[Path] = []
    seen: set[str] = set()

    def add_candidate(path: Optional[Path]) -> None:
        if path is None:
            return
        key = str(path)
        if key in seen:
            return
        seen.add(key)
        candidates.append(path)

    add_candidate(configured_path)
    if weights_path is not None:
        weights_dir = weights_path.parent
        add_candidate(weights_dir / "training_metadata.json")
        add_candidate(weights_dir / "active" / "training_metadata.json")
        add_candidate(weights_dir.parent / "training_metadata.json")
        add_candidate(weights_dir.parent / "active" / "training_metadata.json")
    add_candidate(WORKSPACE_CHECKPOINTS_DIR / "active" / "training_metadata.json")
    return candidates


def resolve_training_metadata(
    *,
    configured_path: Optional[Path],
    weights_path: Optional[Path],
    checkpoint_hints: Optional[Dict[str, Any]] = None,
) -> Tuple[Optional[Path], Dict[str, Any], str]:
    """Resolve metadata from file fallbacks first, then checkpoint-embedded config."""
    for candidate in _candidate_training_metadata_paths(configured_path, weights_path):
        metadata = load_training_metadata_hints(candidate)
        if metadata:
            return candidate, metadata, "file"

    synthesized = _metadata_from_checkpoint_hints(checkpoint_hints or {})
    if synthesized:
        return configured_path, synthesized, "checkpoint"

    return configured_path, {}, "missing"


def _canonical_bands_string(value: Any) -> str:
    if isinstance(value, (list, tuple)):
        return ",".join(str(part).strip() for part in value if str(part).strip())
    return ",".join(part.strip() for part in str(value).split(",") if part.strip())


def _metadata_channel_names(metadata: Dict[str, Any]) -> Optional[Tuple[str, ...]]:
    raw_channel_names = metadata.get("channel_names")
    if isinstance(raw_channel_names, list) and raw_channel_names:
        return tuple(str(x) for x in raw_channel_names)
    return None


def apply_trained_only_metadata_locks(
    *,
    config: PipelineDefaults,
    metadata: Dict[str, Any],
    cli_overrides: Iterable[str],
) -> None:
    """Lock trained-only runtime parameters to metadata values."""
    if not config.trained_model_only:
        return
    if not metadata:
        raise ValueError(
            "trained_model_only active but training metadata could not be loaded."
        )

    meta_channels = _metadata_channel_names(metadata)
    schema_version = metadata.get("schema_version")
    if schema_version not in (None, "") and int(schema_version) < METADATA_SCHEMA_VERSION:
        try:
            legacy_channels = (
                canonicalize_channel_names(meta_channels)
                if meta_channels is not None
                else expected_channel_names(int(metadata.get("in_channels", metadata.get("num_channels", 0))))
            )
        except Exception as exc:
            raise ValueError(
                f"training metadata schema_version={schema_version} eski ve kanal semasi okunamadi; "
                f"trained_model_only icin >= {METADATA_SCHEMA_VERSION} beklenir."
            ) from exc
        if legacy_channels != tuple(MODEL_CHANNEL_NAMES):
            raise ValueError(
                f"training metadata schema_version={schema_version} eski; "
                f"trained_model_only icin >= {METADATA_SCHEMA_VERSION} beklenir."
            )
        LOGGER.warning(
            "training metadata schema_version=%s eski topo5 metadata olarak kabul edildi; "
            "yeni egitimler schema_version=%s yazacaktir.",
            schema_version,
            METADATA_SCHEMA_VERSION,
        )

    locked_sources = {
        "tile": ("tile_size", metadata.get("tile_size")),
        "overlap": ("overlap", metadata.get("overlap")),
        "bands": ("bands", metadata.get("bands")),
    }
    for field_name in LOCKED_TRAINED_ONLY_FIELDS:
        metadata_key, source_value = locked_sources[field_name]
        if source_value in (None, ""):
            raise ValueError(
                f"trained_model_only icin training metadata '{metadata_key}' bilgisini icermelidir."
            )
        current_value = getattr(config, field_name)
        if field_name in cli_overrides:
            current_text = (
                _canonical_bands_string(current_value)
                if field_name == "bands"
                else str(int(current_value))
            )
            source_text = (
                _canonical_bands_string(source_value)
                if field_name == "bands"
                else str(int(source_value))
            )
            if current_text != source_text:
                raise ValueError(
                    f"trained_model_only icin --{field_name}={current_text} kullanilamaz; "
                    f"training metadata {field_name}={source_text} bekliyor."
                )
        if field_name == "bands":
            setattr(config, field_name, _canonical_bands_string(source_value))
        else:
            setattr(config, field_name, int(source_value))

    if meta_channels is not None and canonicalize_channel_names(meta_channels) != expected_channel_names(len(meta_channels)):
        raise ValueError(
            "training metadata channel_names guncel kanal semasi ile uyusmuyor; "
            "legacy/curvature tabanli metadata trained_model_only modunda desteklenmiyor."
        )

    LOGGER.info(
        "trained_model_only active: tile/overlap/bands values locked to training metadata"
    )
    LOGGER.info(
        "Effective tile/overlap/bands: tile=%d overlap=%d bands=%s",
        int(config.tile),
        int(config.overlap),
        str(config.bands),
    )


def validate_checkpoint_metadata_consistency(
    *, checkpoint_hints: Dict[str, Any], metadata: Dict[str, Any]
) -> None:
    """Fail closed when checkpoint hints and metadata disagree on critical fields."""
    if not checkpoint_hints or not metadata:
        return

    checks = (
        ("task_type", "task_type"),
        ("arch", "arch"),
        ("encoder", "encoder"),
        ("in_channels", "in_channels"),
        ("bands", "bands"),
        ("tile_size", "tile_size"),
        ("overlap", "overlap"),
        ("schema_version", "schema_version"),
    )
    for ck_key, meta_key in checks:
        ck_val = checkpoint_hints.get(ck_key)
        meta_val = metadata.get(meta_key)
        if ck_val in (None, "") or meta_val in (None, ""):
            continue
        if ck_key in {"task_type", "arch", "encoder"}:
            left = str(ck_val).strip().lower()
            right = str(meta_val).strip().lower()
        elif ck_key == "bands":
            left = _canonical_bands_string(ck_val)
            right = _canonical_bands_string(meta_val)
        else:
            left = int(ck_val)
            right = int(meta_val)
        if left != right:
            raise ValueError(
                f"Checkpoint {ck_key}={ck_val} ile training metadata {meta_key}={meta_val} uyusmuyor."
            )

    ck_channels = checkpoint_hints.get("channel_names")
    meta_channels = _metadata_channel_names(metadata)
    ck_channel_tuple = _raw_channel_names(ck_channels)
    if ck_channel_tuple is not None and meta_channels is not None:
        try:
            ck_norm = normalize_model_channel_names(ck_channel_tuple)
            meta_norm = normalize_model_channel_names(meta_channels)
        except ValueError as exc:
            raise ValueError(f"Checkpoint/training metadata channel_names gecersiz: {exc}") from exc
    else:
        ck_norm = None
        meta_norm = None
    if ck_norm is not None and meta_norm is not None and ck_norm != meta_norm:
        raise ValueError(
            "Checkpoint channel_names ile training metadata channel_names uyusmuyor."
        )


def warn_if_training_inference_mismatch(
    *,
    metadata: Dict[str, Any],
    input_path: Path,
    dl_task: str,
) -> None:
    """Emit soft warnings for non-fatal training/inference differences."""
    if not metadata or str(dl_task).strip().lower() != "tile_classification":
        return

    meta_input_name = Path(str(metadata.get("input_file", "")).strip()).name
    if meta_input_name and meta_input_name != input_path.name:
        LOGGER.warning(
            "Training metadata input '%s' ile inference input '%s' farkli.",
            meta_input_name,
            input_path.name,
        )


def load_weights(
    model: torch.nn.Module,
    weights_path: Path,
    map_location: torch.device,
) -> None:
    """Load model weights with helpful diagnostics on mismatch."""
    def _resolve_expected_in_channels(target_model: torch.nn.Module) -> Optional[int]:
        queue: List[torch.nn.Module] = [target_model]
        visited: set[int] = set()
        while queue:
            current = queue.pop(0)
            obj_id = id(current)
            if obj_id in visited:
                continue
            visited.add(obj_id)

            encoder = getattr(current, "encoder", None)
            conv1 = getattr(encoder, "conv1", None) if encoder is not None else None
            weight = getattr(conv1, "weight", None) if conv1 is not None else None
            if isinstance(weight, torch.Tensor) and weight.ndim == 4:
                return int(weight.shape[1])

            for attr in ("base_model", "module"):
                nested = getattr(current, attr, None)
                if isinstance(nested, torch.nn.Module):
                    queue.append(nested)
        return None

    def _state_first_conv_in_channels(state_dict: Dict[str, Any]) -> Optional[int]:
        preferred_suffixes = (
            "encoder.conv1.weight",
            "base_model.encoder.conv1.weight",
            "module.encoder.conv1.weight",
            "module.base_model.encoder.conv1.weight",
        )
        for suffix in preferred_suffixes:
            for key, tensor in state_dict.items():
                if isinstance(key, str) and key.endswith(suffix):
                    if isinstance(tensor, torch.Tensor) and tensor.ndim == 4:
                        return int(tensor.shape[1])
        for tensor in state_dict.values():
            if isinstance(tensor, torch.Tensor) and tensor.ndim == 4:
                return int(tensor.shape[1])
        return None

    def _resolve_model_encoder_name(target_model: torch.nn.Module) -> Optional[str]:
        queue: List[torch.nn.Module] = [target_model]
        visited: set[int] = set()
        while queue:
            current = queue.pop(0)
            obj_id = id(current)
            if obj_id in visited:
                continue
            visited.add(obj_id)

            encoder = getattr(current, "encoder", None)
            name = getattr(encoder, "name", None) if encoder is not None else None
            if isinstance(name, str) and name:
                return name

            for attr in ("base_model", "module"):
                nested = getattr(current, attr, None)
                if isinstance(nested, torch.nn.Module):
                    queue.append(nested)
        return None

    def _resolve_model_arch_name(target_model: torch.nn.Module) -> str:
        base_model = getattr(target_model, "base_model", target_model)
        return base_model.__class__.__name__

    def _resolve_model_use_fpn(target_model: torch.nn.Module) -> Optional[bool]:
        queue: List[torch.nn.Module] = [target_model]
        visited: set[int] = set()
        while queue:
            current = queue.pop(0)
            obj_id = id(current)
            if obj_id in visited:
                continue
            visited.add(obj_id)

            use_fpn = getattr(current, "use_fpn", None)
            if isinstance(use_fpn, bool):
                return use_fpn

            for attr in ("base_model", "module"):
                nested = getattr(current, attr, None)
                if isinstance(nested, torch.nn.Module):
                    queue.append(nested)
        return None

    def _checkpoint_cfg(obj: Any) -> Dict[str, Any]:
        if not isinstance(obj, dict):
            return {}
        cfg = obj.get("config")
        return cfg if isinstance(cfg, dict) else {}

    if not weights_path.exists():
        raise FileNotFoundError(f"Weights not found: {weights_path}")
    state_obj = safe_torch_load(weights_path, map_location=map_location)
    state = state_obj
    if isinstance(state_obj, dict):
        for key in ("state_dict", "model_state_dict"):
            maybe_state = state_obj.get(key)
            if isinstance(maybe_state, dict):
                state = maybe_state
                break
    if not isinstance(state, dict):
        raise RuntimeError("Weights file did not contain a valid state dict.")
    if not any(isinstance(v, torch.Tensor) for v in state.values()):
        raise RuntimeError(
            "Weights file does not contain tensor weights. "
            "Expected a state dict or a checkpoint with 'state_dict' / 'model_state_dict'."
        )

    expected_in_ch = _resolve_expected_in_channels(model)
    if expected_in_ch is None:
        raise RuntimeError(
            "Could not determine encoder input channels from model. "
            "Expected an encoder.conv1 weight tensor."
        )

    in_channels = _state_first_conv_in_channels(state)
    if in_channels is not None:
        if in_channels != expected_in_ch:
            raise RuntimeError(
                f"Encoder expects {expected_in_ch} channels but weights provide {in_channels}. "
                "Ensure training and inference use the same channel configuration."
            )

    candidate_states: List[Dict[str, Any]] = [state]
    stripped_state: Dict[str, Any] = {}
    stripped_changed = False
    for key, value in state.items():
        if isinstance(key, str) and key.startswith("module."):
            stripped_state[key.replace("module.", "", 1)] = value
            stripped_changed = True
        else:
            stripped_state[key] = value
    if stripped_changed:
        candidate_states.append(stripped_state)

    last_error: Optional[RuntimeError] = None
    for candidate in candidate_states:
        try:
            model.load_state_dict(candidate, strict=True)
            return
        except RuntimeError as exc:
            last_error = exc

    if last_error is not None:
        details: List[str] = []
        ck_cfg = _checkpoint_cfg(state_obj)
        ck_encoder = str(ck_cfg.get("encoder", "")).strip()
        ck_arch = str(ck_cfg.get("arch", "")).strip()
        ck_in_ch = ck_cfg.get("in_channels")
        model_encoder = _resolve_model_encoder_name(model)
        model_arch = _resolve_model_arch_name(model)
        ck_use_fpn = _checkpoint_optional_bool(ck_cfg, "use_fpn_classifier")
        if ck_use_fpn is None and model_encoder:
            ck_use_fpn = _infer_tile_classifier_use_fpn_from_state_dict(
                state,
                encoder_name=model_encoder,
                in_channels=expected_in_ch,
            )
        model_use_fpn = _resolve_model_use_fpn(model)

        if ck_encoder and model_encoder and ck_encoder != model_encoder:
            details.append(f"encoder mismatch (checkpoint={ck_encoder}, model={model_encoder})")
        if ck_arch and model_arch and ck_arch != model_arch:
            details.append(f"arch mismatch (checkpoint={ck_arch}, model={model_arch})")
        if isinstance(ck_in_ch, int) and ck_in_ch != expected_in_ch:
            details.append(f"in_channels mismatch (checkpoint={ck_in_ch}, model={expected_in_ch})")
        if isinstance(ck_use_fpn, bool) and isinstance(model_use_fpn, bool) and ck_use_fpn != model_use_fpn:
            details.append(
                f"use_fpn_classifier mismatch (checkpoint={ck_use_fpn}, model={model_use_fpn})"
            )

        error_head = str(last_error).splitlines()[0].strip()
        if details:
            raise RuntimeError(
                f"Unable to load weights from {weights_path}: {error_head}. "
                f"{'; '.join(details)}. Check --encoder/--arch/channel/use_fpn_classifier settings."
            ) from last_error
        raise RuntimeError(f"Unable to load weights from {weights_path}: {error_head}") from last_error


def generate_windows(
    width: int, height: int, tile: int, overlap: int
) -> Generator[Tuple[Window, int, int], None, None]:
    """Yield raster windows with the desired tile size and overlap."""
    if overlap >= tile:
        raise ValueError("--overlap must be smaller than --tile.")
    stride = tile - overlap
    if stride <= 0:
        raise ValueError("Stride must be positive; reduce overlap or increase tile size.")
    for row in range(0, height, stride):
        for col in range(0, width, stride):
            win_height = min(tile, height - row)
            win_width = min(tile, width - col)
            yield Window(col, row, win_width, win_height), row, col


def make_feather_weights(h: int, w: int, tile: int, overlap: int) -> np.ndarray:
    """Cosine feathering weights for a tile of size (h,w)."""
    if overlap <= 0:
        return np.ones((h, w), dtype=np.float32)
    ramp = max(1, min(overlap, tile // 2))

    def _ramp(n: int) -> np.ndarray:
        win = np.ones(n, dtype=np.float32)
        # Border tiles can be smaller; clamp ramp length
        local_ramp = max(1, min(ramp, n // 2 if n > 1 else 1))
        # half-cosine ramps
        t = np.linspace(0, np.pi, local_ramp, endpoint=False, dtype=np.float32)
        up = (1 - np.cos(t)) * 0.5  # 0â†’1
        win[:local_ramp] = up
        win[-local_ramp:] = up[::-1]
        return win

    wr = _ramp(h)
    wc = _ramp(w)
    return np.outer(wr, wc).astype(np.float32)


@dataclass
class InferenceOutputs:
    prob_path: Path
    mask_path: Path
    prob_map: Optional[np.ndarray]
    mask: Optional[np.ndarray]
    transform: Affine
    crs: Optional[RasterioCRS]
    band_importance_txt: Optional[Path] = None
    band_importance_json: Optional[Path] = None


@dataclass
class ClassicModeOutput:
    prob_path: Path
    mask_path: Path
    prob_map: Optional[np.ndarray]
    mask: Optional[np.ndarray]
    threshold: float


@dataclass
class ClassicOutputs:
    prob_path: Path
    mask_path: Path
    prob_map: Optional[np.ndarray]
    mask: Optional[np.ndarray]
    transform: Affine
    crs: Optional[RasterioCRS]
    threshold: float
    per_mode: Dict[str, ClassicModeOutput]


@dataclass
class FusionOutputs:
    prob_path: Path
    mask_path: Path
    prob_map: Optional[np.ndarray]
    mask: Optional[np.ndarray]
    transform: Affine
    crs: Optional[RasterioCRS]
    threshold: float


@dataclass
class YoloOutputs:
    prob_path: Path
    mask_path: Path
    prob_map: Optional[np.ndarray]
    mask: Optional[np.ndarray]
    transform: Affine
    crs: Optional[RasterioCRS]
    threshold: float
    label_records: Optional[List[Dict[str, Any]]] = None
    labels_out_base: Optional[Path] = None


@dataclass
class MultiscaleSavedOutput:
    scale_token: str
    scale_factor: float
    scale_level: int
    outputs: InferenceOutputs


@dataclass
class VectorExportJob:
    label: str
    mask: Optional[np.ndarray]
    prob_map: Optional[np.ndarray]
    transform: Affine
    crs: Optional[RasterioCRS]
    out_base: Path
    prob_path: Path
    mask_path: Path
    scale_level: Optional[int] = None
    scale_factor: Optional[float] = None
    gpkg_layer_name: Optional[str] = None


def _resolve_importance_channel_names(
    channel_names: Optional[Sequence[str]],
    in_channels: int,
) -> List[str]:
    names = [str(name) for name in (channel_names or ())]
    if len(names) < in_channels:
        names.extend(f"ch_{idx+1}" for idx in range(len(names), in_channels))
    return names[:in_channels]


def _find_first_channel_attention_module(model: torch.nn.Module) -> Optional[ChannelAttention]:
    for module in model.modules():
        if isinstance(module, ChannelAttention):
            return module
    return None


def _estimate_attention_channel_scores(
    attention_module: ChannelAttention,
    tensor: torch.Tensor,
) -> Optional[np.ndarray]:
    if tensor.ndim != 4:
        return None
    with torch.no_grad():
        avg_out = attention_module.fc(attention_module.avg_pool(tensor))
        max_out = attention_module.fc(attention_module.max_pool(tensor))
        attention = torch.sigmoid(avg_out + max_out)
        scores = attention.detach().float().mean(dim=(0, 2, 3)).cpu().numpy()
    return scores


def _estimate_gradient_channel_scores(
    model: torch.nn.Module,
    tensor: torch.Tensor,
    task_key: str,
) -> Tuple[torch.Tensor, Optional[np.ndarray]]:
    tensor_grad = tensor.detach().clone().requires_grad_(True)
    with torch.enable_grad():
        logits = model(tensor_grad)
        target = logits.view(-1).mean() if task_key == "tile_classification" else logits.mean()
        model.zero_grad(set_to_none=True)
        target.backward()
        grad = tensor_grad.grad
        model.zero_grad(set_to_none=True)
    if grad is None:
        return logits.detach(), None
    with torch.no_grad():
        scores = (
            grad.detach().float().abs() * tensor_grad.detach().float().abs()
        ).mean(dim=(0, 2, 3)).cpu().numpy()
    return logits.detach(), scores


def _build_band_importance_summary(
    *,
    scores_sum: Optional[np.ndarray],
    sample_count: int,
    mode: str,
    channel_names: Sequence[str],
) -> Optional[Dict[str, Any]]:
    if scores_sum is None or sample_count <= 0:
        return None
    avg_scores = np.asarray(scores_sum, dtype=np.float64) / float(sample_count)
    if mode == "gradient":
        denom = float(np.sum(avg_scores))
        if denom > 0:
            avg_scores = avg_scores / denom

    names = _resolve_importance_channel_names(channel_names, int(avg_scores.shape[0]))
    ranking = sorted(
        [{"channel": name, "score": float(score)} for name, score in zip(names, avg_scores)],
        key=lambda x: x["score"],
        reverse=True,
    )
    top_item = ranking[0] if ranking else {"channel": "n/a", "score": 0.0}
    return {
        "mode": mode,
        "samples": int(sample_count),
        "scores": {item["channel"]: float(item["score"]) for item in ranking},
        "ranking": ranking,
        "top_channel": str(top_item["channel"]),
        "top_score": float(top_item["score"]),
    }


def _write_band_importance_report(
    *,
    out_dir: Path,
    filename_stem: str,
    summary: Optional[Dict[str, Any]],
) -> Tuple[Optional[Path], Optional[Path]]:
    if summary is None:
        return None, None
    out_dir.mkdir(parents=True, exist_ok=True)
    txt_path = out_dir / f"{filename_stem}_band_importance.txt"
    json_path = out_dir / f"{filename_stem}_band_importance.json"

    lines = [
        "# Band Importance Report",
        f"mode: {summary.get('mode', 'n/a')}",
        f"samples: {summary.get('samples', 0)}",
        f"top_channel: {summary.get('top_channel', 'n/a')}",
        f"top_score: {float(summary.get('top_score', 0.0)):.6f}",
        "",
        "[ranking]",
    ]
    for idx, item in enumerate(summary.get("ranking", []), start=1):
        lines.append(
            f"{idx:02d}. {item.get('channel', 'n/a')}: {float(item.get('score', 0.0)):.6f}"
        )
    txt_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return txt_path, json_path


def _downsample_raster(input_path: Path, scale: float, output_path: Path) -> Path:
    """Raster'ı verilen ölçeğe küçültüp geçici dosyaya yaz."""
    with rasterio.open(input_path) as src:
        new_h = max(1, int(round(src.height * scale)))
        new_w = max(1, int(round(src.width * scale)))
        transform = src.transform * src.transform.scale(
            src.width / new_w,
            src.height / new_h,
        )
        profile = src.profile.copy()
        profile.update(height=new_h, width=new_w, transform=transform)
        data = src.read(
            out_shape=(src.count, new_h, new_w),
            resampling=rasterio.enums.Resampling.bilinear,
        )
        with rasterio.open(output_path, "w", **profile) as dst:
            dst.write(data)
    LOGGER.info("  Ölçek %.2fx: %dx%d → %dx%d  (%s)", scale, src.height, src.width, new_h, new_w, output_path.name)
    return output_path


def _upsample_prob_map(prob_map: np.ndarray, target_h: int, target_w: int) -> np.ndarray:
    """Olasılık haritasını orijinal boyuta bilinear interpolasyon ile büyüt."""
    from scipy.ndimage import zoom as _zoom
    if prob_map.shape[0] == target_h and prob_map.shape[1] == target_w:
        return prob_map
    zoom_h = target_h / max(prob_map.shape[0], 1)
    zoom_w = target_w / max(prob_map.shape[1], 1)
    return _zoom(prob_map.astype(np.float32), (zoom_h, zoom_w), order=1).astype(np.float32)


def _merge_prob_maps(
    prob_maps: List[np.ndarray],
    merge: str,
    weights: Optional[Sequence[float]] = None,
) -> np.ndarray:
    """Birden fazla olasılık haritasını birleştir."""
    if len(prob_maps) == 1:
        return prob_maps[0]
    stack = np.stack(prob_maps, axis=0)
    if merge == "max":
        return np.nanmax(stack, axis=0)
    if merge == "mean":
        return np.nanmean(stack, axis=0)
    if merge == "weighted_mean":
        w = np.array(weights if weights else [1.0] * len(prob_maps), dtype=np.float32)
        w = w / w.sum()
        return np.nansum(stack * w[:, None, None], axis=0)
    raise ValueError(f"Bilinmeyen merge stratejisi: {merge}")


def infer_tiled(
    model: torch.nn.Module,
    input_path: Path,
    band_idx: Sequence[int],
    task_type: str,
    tile: int,
    overlap: int,
    device: torch.device,
    use_half: bool,
    threshold: float,
    mask_talls: Optional[float],
    out_prefix: Path,
    global_norm: bool,
    norm_sample_tiles: int,
    feather: bool,
    precomputed_deriv: Optional[PrecomputedDerivatives] = None,
    derivative_cache_tif: Optional[Path] = None,
    derivative_cache_meta: Optional[Path] = None,
    enable_curvature: bool = True,
    enable_tpi: bool = True,
    encoder: Optional[str] = None,
    min_area: Optional[float] = None,
    percentile_low: Optional[float] = None,
    percentile_high: Optional[float] = None,
    rvt_radii: Optional[Sequence[float]] = None,
    gaussian_lrm_sigma: Optional[float] = None,
    rgb_only: bool = False,
    tpi_radii: Tuple[int, ...] = (5, 15, 30),
    arch: Optional[str] = None,
    weight_type: Optional[str] = None,
    channel_names: Optional[Sequence[str]] = None,
    save_band_importance: bool = False,
    band_importance_max_tiles: int = 0,
) -> InferenceOutputs:
    """
    Run tiled inference and save outputs.
    
    Args:
        precomputed_deriv: Önceden hesaplanmış RVT türevleri (cache kullanımı için).
                          None ise her tile için RVT yeniden hesaplanır.
    """
    task_key = str(task_type).strip().lower()
    if task_key not in {"segmentation", "tile_classification"}:
        raise ValueError(f"Unsupported task_type: {task_type}")
    model.eval()
    model.to(device)
    resolved_channel_names = (
        normalize_model_channel_names(channel_names)
        if channel_names is not None
        else infer_channel_names_from_band_indexes(band_idx)
    )
    requires_topography = channel_names_require_topography(resolved_channel_names)
    requires_dsm = channel_names_require_dsm(resolved_channel_names)
    derivative_band_names = derivative_band_names_for_channel_names(resolved_channel_names)
    has_dsm = band_indexes_support_dsm(band_idx)
    has_topography = band_indexes_support_topography(band_idx)

    rgb_only_log_emitted = False

    def log_rgb_only_once() -> None:
        nonlocal rgb_only_log_emitted
        if rgb_only and requires_topography and not rgb_only_log_emitted:
            LOGGER.info("RGB-only modu aktif: türev kanalları sıfırla dolduruldu; efektif girdi sadece RGB.")
            rgb_only_log_emitted = True

    if requires_topography and not rgb_only and not has_topography:
        raise ValueError(
            f"Model kanal semasi {resolved_channel_names} topografya gerektiriyor "
            "ama --bands yalnizca RGB girisi sagliyor."
        )
    if requires_dsm and not rgb_only and not has_dsm:
        raise ValueError(
            f"Model kanal semasi {resolved_channel_names} nDSM gerektiriyor "
            "ama --bands DSM bandi saglamiyor."
        )

    if mask_talls is not None and not has_dsm:
        LOGGER.warning("Yüksek obje maskeleme istendi ama DSM bandı eksik; devre dışı bırakılıyor.")
        mask_talls = None

    prob_map_out: Optional[np.ndarray] = None
    mask_out: Optional[np.ndarray] = None
    band_importance_txt_path: Optional[Path] = None
    band_importance_json_path: Optional[Path] = None
    band_importance_sum: Optional[np.ndarray] = None
    band_importance_samples = 0
    max_importance_tiles = max(0, int(band_importance_max_tiles))
    band_importance_enabled = bool(save_band_importance) and max_importance_tiles > 0
    band_importance_mode = "disabled"
    attention_module: Optional[ChannelAttention] = None
    if band_importance_enabled:
        attention_module = _find_first_channel_attention_module(model)
        if attention_module is not None:
            band_importance_mode = "attention"
        else:
            band_importance_mode = "gradient"
        LOGGER.info(
            "Band onem analizi aktif (mode=%s, max_tiles=%d).",
            band_importance_mode,
            max_importance_tiles,
        )

    with ExitStack() as stack:
        src = stack.enter_context(rasterio.open(input_path))
        height, width = src.height, src.width
        transform = src.transform
        crs = src.crs
        pixel_size = float((abs(transform.a) + abs(transform.e)) / 2.0)

        deriv_ds: Optional[rasterio.io.DatasetReader] = None
        deriv_band_map: Optional[Dict[str, int]] = None
        if (
            precomputed_deriv is None
            and derivative_cache_tif is not None
            and derivative_cache_meta is not None
            and requires_topography
            and not rgb_only
        ):
            info = load_derivative_raster_cache_info(
                derivative_cache_tif,
                derivative_cache_meta,
                input_path,
                band_idx,
                rvt_radii=rvt_radii,
                gaussian_lrm_sigma=gaussian_lrm_sigma,
                enable_curvature=enable_curvature,
                enable_tpi=enable_tpi,
                tpi_radii=tpi_radii,
            )
            if info is not None:
                deriv_band_map = {str(k): int(v) for k, v in dict(info.get("band_map", {})).items()}
                missing_cache_bands = [
                    name for name in derivative_band_names if name not in deriv_band_map
                ]
                if missing_cache_bands:
                    LOGGER.warning(
                        "Derivatives raster-cache gerekli bandlari icermiyor (%s); tile-bazli hesaplama kullanilacak.",
                        ", ".join(missing_cache_bands),
                    )
                    deriv_band_map = None
                else:
                    deriv_ds = stack.enter_context(rasterio.open(derivative_cache_tif))
                LOGGER.info("Derivatives raster-cache kullanılıyor: %s", derivative_cache_tif)

        pixels = int(height) * int(width)
        one_float_bytes = pixels * np.dtype(np.float32).itemsize
        est_bytes = one_float_bytes * (3 if mask_talls is not None else 2) + pixels
        avail = available_memory_bytes()
        use_memmap = False
        if avail is not None and avail > 0:
            # Prefer RAM if it comfortably fits; fall back to disk when it likely risks OOM.
            use_memmap = est_bytes > int(avail * 0.75)
        else:
            # If we cannot detect available RAM, be conservative only for very large accumulators.
            use_memmap = est_bytes >= int(24 * 1024**3)
        scratch_dir: Optional[Path] = make_scratch_dir("infer") if use_memmap else None
        if scratch_dir is not None:
            stack.callback(cleanup_scratch_dir, scratch_dir)
        if use_memmap:
            LOGGER.info(
                "Large raster detected (%dx%d). Using disk-backed accumulators under %s",
                width,
                height,
                scratch_dir,
            )

        prob_acc, _prob_acc_path = alloc_array(
            (height, width),
            np.float32,
            fill_value=0.0,
            use_memmap=use_memmap,
            scratch_dir=scratch_dir,
            name="dl_prob_acc",
        )
        weight_acc, _weight_acc_path = alloc_array(
            (height, width),
            np.float32,
            fill_value=0.0,
            use_memmap=use_memmap,
            scratch_dir=scratch_dir,
            name="dl_weight_acc",
        )
        ndsm_max: Optional[np.ndarray] = None
        _ndsm_max_path: Optional[Path] = None
        if mask_talls is not None:
            ndsm_max, _ndsm_max_path = alloc_array(
                (height, width),
                np.float32,
                fill_value=-np.inf,
                use_memmap=use_memmap,
                scratch_dir=scratch_dir,
                name="dl_ndsm_max",
            )
        valid_global, _valid_global_path = alloc_array(
            (height, width),
            bool,
            fill_value=False,
            use_memmap=use_memmap,
            scratch_dir=scratch_dir,
            name="dl_valid_global",
        )

        total_tiles = math.ceil(height / max(tile - overlap, 1)) * math.ceil(
            width / max(tile - overlap, 1)
        )

        autocast_ctx = (
            torch.autocast(device_type=device.type, dtype=torch.float16)
            if use_half and device.type == "cuda"
            else nullcontext()
        )

        # ---- Global normalization pre-pass (optional) ----
        fixed_lows = None
        fixed_highs = None
        if global_norm:
            lows_list = []
            highs_list = []
            sampled = 0
            for window, row, col in generate_windows(width, height, tile, overlap):
                if precomputed_deriv is not None:
                    # Use precomputed full-raster derivatives; slice to window
                    row_start = int(window.row_off)
                    col_start = int(window.col_off)
                    row_end = row_start + int(window.height)
                    col_end = col_start + int(window.width)

                    rgb_s = precomputed_deriv.rgb[:, row_start:row_end, col_start:col_end].copy()
                    svf_s: Optional[np.ndarray] = None
                    slrm_s: Optional[np.ndarray] = None
                    slope_s: Optional[np.ndarray] = None
                    ndsm_s: Optional[np.ndarray] = None
                    if requires_topography:
                        if rgb_only:
                            log_rgb_only_once()
                            Hs, Ws = rgb_s.shape[1], rgb_s.shape[2]
                            Z = np.zeros((Hs, Ws), dtype=np.float32)
                            svf_s = Z
                            slrm_s = Z
                            slope_s = Z
                            ndsm_s = Z
                        else:
                            svf_s = precomputed_deriv.svf[row_start:row_end, col_start:col_end].copy()
                            slrm_s = precomputed_deriv.slrm[row_start:row_end, col_start:col_end].copy()
                            if "Slope" in resolved_channel_names:
                                slope_s = precomputed_deriv.slope[row_start:row_end, col_start:col_end].copy()
                            if "nDSM" in resolved_channel_names:
                                ndsm_s = precomputed_deriv.ndsm[row_start:row_end, col_start:col_end].copy()
                    stack_s = assemble_model_input(
                        rgb_s,
                        channel_names=resolved_channel_names,
                        svf=svf_s,
                        slrm=slrm_s,
                        slope=slope_s,
                        ndsm=ndsm_s,
                    )
                elif deriv_ds is not None and deriv_band_map is not None:
                    def read_band(idx: int) -> Optional[np.ndarray]:
                        if idx <= 0:
                            return None
                        data = src.read(idx, window=window, boundless=True, masked=True)
                        return np.ma.filled(data.astype(np.float32), np.nan)

                    rgb_s = np.stack([read_band(band_idx[i]) for i in range(3)], axis=0)
                    svf_s = None
                    slrm_s = None
                    slope_s = None
                    ndsm_s = None
                    if requires_topography:
                        if rgb_only:
                            log_rgb_only_once()
                            Hs, Ws = rgb_s.shape[1], rgb_s.shape[2]
                            Z = np.zeros((Hs, Ws), dtype=np.float32)
                            svf_s = Z
                            slrm_s = Z
                            slope_s = Z
                            ndsm_s = Z
                        else:
                            band_names = list(derivative_band_names)
                            indexes = [int(deriv_band_map[name]) for name in band_names]
                            deriv_stack = deriv_ds.read(indexes=indexes, window=window, boundless=True, masked=True)
                            deriv_stack = np.ma.filled(deriv_stack.astype(np.float32), np.nan)
                            deriv_by_name = {
                                name: deriv_stack[idx] for idx, name in enumerate(band_names)
                            }
                            svf_s = deriv_by_name.get("svf")
                            slrm_s = deriv_by_name.get("slrm")
                            slope_s = deriv_by_name.get("slope")
                            ndsm_s = deriv_by_name.get("ndsm")

                    stack_s = assemble_model_input(
                        rgb_s,
                        channel_names=resolved_channel_names,
                        svf=svf_s,
                        slrm=slrm_s,
                        slope=slope_s,
                        ndsm=ndsm_s,
                    )
                else:
                    def read_band(idx: int) -> Optional[np.ndarray]:
                        if idx <= 0:
                            return None
                        data = src.read(idx, window=window, boundless=True, masked=True)
                        return np.ma.filled(data.astype(np.float32), np.nan)

                    rgb_s = np.stack([read_band(band_idx[i]) for i in range(3)], axis=0)
                    svf_s = None
                    slrm_s = None
                    slope_s = None
                    ndsm_s = None
                    if requires_topography:
                        if rgb_only:
                            log_rgb_only_once()
                            Hs, Ws = rgb_s.shape[1], rgb_s.shape[2]
                            Z = np.zeros((Hs, Ws), dtype=np.float32)
                            svf_s = Z
                            slrm_s = Z
                            slope_s = Z
                            ndsm_s = Z
                        else:
                            dtm_s = read_band(band_idx[4])
                            if dtm_s is None:
                                raise ValueError("DTM band gerekli; topografik kanal hesabi yapilamadi.")
                            svf_s, slrm_s = compute_derivatives_with_rvt(
                                dtm_s,
                                pixel_size=pixel_size,
                                radii=rvt_radii,
                                gaussian_lrm_sigma=gaussian_lrm_sigma,
                                show_progress=False,
                                log_steps=False,
                            )
                            if "Slope" in resolved_channel_names:
                                slope_s = compute_slope(dtm_s, pixel_size=pixel_size)
                            if "nDSM" in resolved_channel_names:
                                dsm_s = read_band(band_idx[3]) if has_dsm else None
                                ndsm_s = compute_ndsm(dsm_s, dtm_s)
                    stack_s = assemble_model_input(
                        rgb_s,
                        channel_names=resolved_channel_names,
                        svf=svf_s,
                        slrm=slrm_s,
                        slope=slope_s,
                        ndsm=ndsm_s,
                    )

                Hs, Ws = stack_s.shape[1], stack_s.shape[2]
                ch = min(TILE_SAMPLE_CROP_SIZE, Hs)
                cw = min(TILE_SAMPLE_CROP_SIZE, Ws)
                r0 = max(0, (Hs - ch) // 2)
                c0 = max(0, (Ws - cw) // 2)
                crop = stack_s[:, r0 : r0 + ch, c0 : c0 + cw]

                Cc = crop.shape[0]
                lows = np.zeros(Cc, dtype=np.float32)
                highs = np.zeros(Cc, dtype=np.float32)
                for ci in range(Cc):
                    v = crop[ci].ravel()
                    v = v[np.isfinite(v)]
                    if v.size == 0:
                        lows[ci] = 0.0
                        highs[ci] = 1.0
                    else:
                        _p_low = percentile_low if percentile_low is not None else DEFAULTS.percentile_low
                        _p_high = percentile_high if percentile_high is not None else DEFAULTS.percentile_high
                        lows[ci] = float(np.percentile(v, _p_low))
                        highs[ci] = float(np.percentile(v, _p_high))
                lows_list.append(lows)
                highs_list.append(highs)

                sampled += 1
                if sampled >= norm_sample_tiles:
                    break

            if lows_list and highs_list:
                fixed_lows = np.median(np.stack(lows_list, axis=0), axis=0)
                fixed_highs = np.median(np.stack(highs_list, axis=0), axis=0)
            if global_norm and (fixed_lows is None or fixed_highs is None):
                LOGGER.info("Global normalizasyon atlandı; karo başına normalizasyona geçiliyor.")

        for window, row, col in progress_bar(
            generate_windows(width, height, tile, overlap),
            total=total_tiles,
            desc="Inference",
            unit="tile",
        ):
            win_height = int(window.height)
            win_width = int(window.width)
            pad_h = max(0, tile - win_height)
            pad_w = max(0, tile - win_width)
            
            # Precomputed derivatives kullan (cache modunda)
            dsm: Optional[np.ndarray] = None
            dtm: Optional[np.ndarray] = None
            if precomputed_deriv is not None:
                row_start = int(window.row_off)
                col_start = int(window.col_off)
                row_end = row_start + win_height
                col_end = col_start + win_width
                
                rgb = precomputed_deriv.rgb[:, row_start:row_end, col_start:col_end].copy()
                svf: Optional[np.ndarray] = None
                slrm: Optional[np.ndarray] = None
                slope: Optional[np.ndarray] = None
                ndsm: Optional[np.ndarray] = None
                if requires_topography and not rgb_only:
                    svf = precomputed_deriv.svf[row_start:row_end, col_start:col_end].copy()
                    slrm = precomputed_deriv.slrm[row_start:row_end, col_start:col_end].copy()
                    if "Slope" in resolved_channel_names:
                        slope = precomputed_deriv.slope[row_start:row_end, col_start:col_end].copy()
                    if "nDSM" in resolved_channel_names:
                        ndsm = precomputed_deriv.ndsm[row_start:row_end, col_start:col_end].copy()
                
                if pad_h or pad_w:
                    rgb = np.pad(rgb, ((0, 0), (0, pad_h), (0, pad_w)), mode="constant", constant_values=np.nan)
                    if svf is not None:
                        svf = np.pad(svf, ((0, pad_h), (0, pad_w)), mode="constant", constant_values=np.nan)
                    if slrm is not None:
                        slrm = np.pad(slrm, ((0, pad_h), (0, pad_w)), mode="constant", constant_values=np.nan)
                    if slope is not None:
                        slope = np.pad(slope, ((0, pad_h), (0, pad_w)), mode="constant", constant_values=np.nan)
                    if ndsm is not None:
                        ndsm = np.pad(ndsm, ((0, pad_h), (0, pad_w)), mode="constant", constant_values=np.nan)

                if requires_topography and rgb_only:
                    log_rgb_only_once()
                    Ht, Wt = rgb.shape[1], rgb.shape[2]
                    Z = np.zeros((Ht, Wt), dtype=np.float32)
                    svf = Z
                    slrm = Z
                    slope = Z
                    ndsm = Z
            elif deriv_ds is not None and deriv_band_map is not None:
                def read_band(idx: int) -> Optional[np.ndarray]:
                    if idx <= 0:
                        return None
                    data = src.read(idx, window=window, boundless=True, masked=True)
                    return np.ma.filled(data.astype(np.float32), np.nan)

                rgb = np.stack([read_band(band_idx[i]) for i in range(3)], axis=0)
                svf = None
                slrm = None
                slope = None
                ndsm = None
                if requires_topography and not rgb_only:
                    band_names = list(derivative_band_names)
                    indexes = [int(deriv_band_map[name]) for name in band_names]
                    deriv_stack = deriv_ds.read(indexes=indexes, window=window, boundless=True, masked=True)
                    deriv_stack = np.ma.filled(deriv_stack.astype(np.float32), np.nan)
                    deriv_by_name = {
                        name: deriv_stack[idx] for idx, name in enumerate(band_names)
                    }
                    svf = deriv_by_name.get("svf")
                    slrm = deriv_by_name.get("slrm")
                    slope = deriv_by_name.get("slope")
                    ndsm = deriv_by_name.get("ndsm")

                if pad_h or pad_w:
                    rgb = np.pad(rgb, ((0, 0), (0, pad_h), (0, pad_w)), mode="constant", constant_values=np.nan)
                    if svf is not None:
                        svf = np.pad(svf, ((0, pad_h), (0, pad_w)), mode="constant", constant_values=np.nan)
                    if slrm is not None:
                        slrm = np.pad(slrm, ((0, pad_h), (0, pad_w)), mode="constant", constant_values=np.nan)
                    if slope is not None:
                        slope = np.pad(slope, ((0, pad_h), (0, pad_w)), mode="constant", constant_values=np.nan)
                    if ndsm is not None:
                        ndsm = np.pad(ndsm, ((0, pad_h), (0, pad_w)), mode="constant", constant_values=np.nan)

                if requires_topography and rgb_only:
                    log_rgb_only_once()
                    Ht, Wt = rgb.shape[1], rgb.shape[2]
                    Z = np.zeros((Ht, Wt), dtype=np.float32)
                    svf = Z
                    slrm = Z
                    slope = Z
                    ndsm = Z
            else:
                # Normal mod - Her tile için RVT hesapla
                def read_band(idx: int) -> Optional[np.ndarray]:
                    if idx <= 0:
                        return None
                    data = src.read(idx, window=window, boundless=True, masked=True)
                    return np.ma.filled(data.astype(np.float32), np.nan)

                rgb = np.stack([read_band(band_idx[i]) for i in range(3)], axis=0)
                svf = None
                slrm = None
                slope = None
                ndsm = None
                if has_dsm:
                    dsm = read_band(band_idx[3])
                if requires_topography or mask_talls is not None:
                    dtm = read_band(band_idx[4]) if has_topography else None
                    if requires_topography and not rgb_only and dtm is None:
                        raise ValueError("DTM band gerekli; topografik kanal hesabi yapilamadi.")

                if pad_h or pad_w:
                    rgb = np.pad(
                        rgb,
                        ((0, 0), (0, pad_h), (0, pad_w)),
                        mode="constant",
                        constant_values=np.nan,
                    )
                    if dtm is not None:
                        dtm = np.pad(
                            dtm,
                            ((0, pad_h), (0, pad_w)),
                            mode="constant",
                            constant_values=np.nan,
                        )
                    if dsm is not None:
                        dsm = np.pad(
                            dsm,
                            ((0, pad_h), (0, pad_w)),
                            mode="constant",
                            constant_values=np.nan,
                        )

                if requires_topography and rgb_only:
                    log_rgb_only_once()
                    Ht, Wt = rgb.shape[1], rgb.shape[2]
                    Z = np.zeros((Ht, Wt), dtype=np.float32)
                    svf = Z
                    slrm = Z
                    slope = Z
                    ndsm = Z
                elif requires_topography:
                    svf, slrm = compute_derivatives_with_rvt(
                        dtm,
                        pixel_size=pixel_size,
                        radii=rvt_radii,
                        gaussian_lrm_sigma=gaussian_lrm_sigma,
                        show_progress=False,
                        log_steps=False,
                    )
                    if "Slope" in resolved_channel_names:
                        slope = compute_slope(dtm, pixel_size=pixel_size)
                    if "nDSM" in resolved_channel_names:
                        ndsm = compute_ndsm(dsm, dtm)

            stacked = assemble_model_input(
                rgb,
                channel_names=resolved_channel_names,
                svf=svf,
                slrm=slrm,
                slope=slope,
                ndsm=ndsm,
            )
            valid_mask = np.all(np.isfinite(stacked), axis=0)
            if global_norm and fixed_lows is not None and fixed_highs is not None:
                normed = robust_norm_fixed(stacked, fixed_lows, fixed_highs)
            else:
                normed = robust_norm(
                    stacked,
                    p_low=percentile_low if percentile_low is not None else DEFAULTS.percentile_low,
                    p_high=percentile_high if percentile_high is not None else DEFAULTS.percentile_high,
                )

            tensor = torch.from_numpy(normed).unsqueeze(0).to(device)
            collect_band_importance = (
                band_importance_enabled
                and band_importance_samples < max_importance_tiles
                and np.any(valid_mask)
            )

            scores: Optional[np.ndarray] = None
            if collect_band_importance and band_importance_mode == "gradient":
                logits, scores = _estimate_gradient_channel_scores(model, tensor, task_key)
            else:
                if collect_band_importance and band_importance_mode == "attention" and attention_module is not None:
                    scores = _estimate_attention_channel_scores(attention_module, tensor)
                with torch.no_grad(), autocast_ctx:
                    logits = model(tensor)

            if task_key == "tile_classification":
                tile_prob_scalar = float(torch.sigmoid(logits).view(-1)[0].item())
                probs = np.full((tile, tile), tile_prob_scalar, dtype=np.float32)
            else:
                probs = torch.sigmoid(logits).cpu().numpy()[0, 0]

            if collect_band_importance and scores is not None:
                if band_importance_sum is None:
                    band_importance_sum = np.zeros_like(scores, dtype=np.float64)
                band_importance_sum += scores.astype(np.float64)
                band_importance_samples += 1

            row_slice = slice(row, row + win_height)
            col_slice = slice(col, col + win_width)
            tile_rows = row_slice.stop - row_slice.start
            tile_cols = col_slice.stop - col_slice.start
            prob_tile = probs.astype(np.float32)[:tile_rows, :tile_cols]
            valid_tile = valid_mask[:tile_rows, :tile_cols]

            if feather:
                W = make_feather_weights(tile_rows, tile_cols, tile, overlap)
            else:
                W = np.ones((tile_rows, tile_cols), dtype=np.float32)

            prob_acc[row_slice, col_slice] += prob_tile * W * valid_tile
            weight_acc[row_slice, col_slice] += W * valid_tile.astype(np.float32)
            valid_global[row_slice, col_slice] |= valid_tile

            if mask_talls is not None and ndsm_max is not None:
                if precomputed_deriv is None and deriv_ds is None:
                    ndsm_tile = compute_ndsm(dsm, dtm)
                else:
                    _dsm_mt = src.read(band_idx[3], window=window, boundless=True, masked=True) if has_dsm else None
                    _dtm_mt = src.read(band_idx[4], window=window, boundless=True, masked=True)
                    _dsm_f = np.ma.filled(_dsm_mt.astype(np.float32), np.nan) if _dsm_mt is not None else None
                    _dtm_f = np.ma.filled(_dtm_mt.astype(np.float32), np.nan)
                    ndsm_tile = compute_ndsm(_dsm_f, _dtm_f)
                    if pad_h or pad_w:
                        ndsm_tile = np.pad(ndsm_tile, ((0, pad_h), (0, pad_w)), mode="constant", constant_values=np.nan)
                ndsm_chunk = ndsm_tile.astype(np.float32)[:tile_rows, :tile_cols]
                existing = ndsm_max[row_slice, col_slice]
                ndsm_chunk[~valid_tile] = existing[~valid_tile]
                ndsm_max[row_slice, col_slice] = np.maximum(existing, ndsm_chunk)

        binary_mask, _binary_mask_path = alloc_array(
            (height, width),
            np.uint8,
            fill_value=0,
            use_memmap=use_memmap,
            scratch_dir=scratch_dir,
            name="dl_binary_mask",
        )

        # Finalize in blocks to avoid huge temporary allocations (especially on large rasters).
        block_rows = 1024 if use_memmap else min(8192, int(height))
        total_blocks = math.ceil(int(height) / block_rows)
        for idx in progress_bar(
            range(total_blocks),
            desc="DL finalize",
            unit="block",
            total=total_blocks,
        ):
            row0 = idx * block_rows
            row1 = min(int(height), row0 + block_rows)
            acc_block = prob_acc[row0:row1, :]
            w_block = weight_acc[row0:row1, :]
            valid_block = valid_global[row0:row1, :]

            with np.errstate(divide="ignore", invalid="ignore"):
                np.divide(acc_block, w_block, out=acc_block, where=w_block > 0)
            acc_block[~valid_block] = np.nan

            if mask_talls is not None:
                tall_block = ndsm_max[row0:row1, :] > float(mask_talls)
                acc_block[tall_block] = 0.0

            binary_mask[row0:row1, :] = ((acc_block >= threshold) & valid_block).astype(np.uint8)

        prob_map = prob_acc

        filename, prob_path, mask_path = build_dl_output_paths(
            out_prefix=out_prefix,
            task_type=task_key,
            arch=arch,
            encoder=encoder,
            weight_type=weight_type,
            threshold=threshold,
            tile=tile,
            min_area=min_area,
        )

        if band_importance_enabled:
            summary = _build_band_importance_summary(
                scores_sum=band_importance_sum,
                sample_count=band_importance_samples,
                mode=band_importance_mode,
                channel_names=_resolve_importance_channel_names(
                    resolved_channel_names,
                    int(band_importance_sum.shape[0]) if band_importance_sum is not None else 0,
                ),
            )
            band_importance_txt_path, band_importance_json_path = _write_band_importance_report(
                out_dir=_output_base_path(out_prefix).parent,
                filename_stem=filename,
                summary=summary,
            )
            if band_importance_txt_path is not None:
                LOGGER.info("Band onem raporu yazildi: %s", band_importance_txt_path)
        
        write_prob_and_mask_rasters(
            prob_map=prob_map,
            mask=binary_mask,
            transform=transform,
            crs=crs,
            prob_path=prob_path,
            mask_path=mask_path,
        )

        if use_memmap:
            prob_map_out = None
            mask_out = None
            del prob_map, binary_mask, prob_acc, weight_acc, valid_global
            if ndsm_max is not None:
                del ndsm_max
            gc.collect()
        else:
            prob_map_out = prob_map
            mask_out = binary_mask

    # GPU belleğini temizle (bellek sızıntısını önle)
    if device.type == "cuda":
        torch.cuda.empty_cache()
        LOGGER.debug("GPU belleği temizlendi")

    return InferenceOutputs(
        prob_path=prob_path,
        mask_path=mask_path,
        prob_map=prob_map_out,
        mask=mask_out,
        transform=transform,
        crs=crs,
        band_importance_txt=band_importance_txt_path,
        band_importance_json=band_importance_json_path,
    )


def stretch_to_uint8(rgb: np.ndarray, low: float = 2.0, high: float = 98.0) -> np.ndarray:
    """Stretch float/int bands to uint8 using percentile bounds."""
    stretched = np.zeros_like(rgb, dtype=np.uint8)
    for idx in range(rgb.shape[0]):
        band = rgb[idx].astype(np.float32)
        valid = np.isfinite(band)
        if not np.any(valid):
            continue
        lo = float(np.nanpercentile(band, low))
        hi = float(np.nanpercentile(band, high))
        if not np.isfinite(lo) or not np.isfinite(hi) or hi <= lo:
            lo = float(np.nanmin(band[valid]))
            hi = float(np.nanmax(band[valid]))
        if not np.isfinite(lo) or not np.isfinite(hi) or hi <= lo:
            continue
        scaled = np.zeros_like(band, dtype=np.float32)
        scaled[valid] = (band[valid] - lo) / (hi - lo)
        np.clip(scaled, 0.0, 1.0, out=scaled)
        stretched[idx] = (scaled * 255.0).astype(np.uint8)
    return stretched


def infer_yolo_tiled(
    input_path: Path,
    band_idx: Sequence[int],
    tile: int,
    overlap: int,
    out_prefix: Path,
    yolo_weights: Optional[str] = None,
    conf_threshold: float = 0.25,
    iou_threshold: float = 0.45,
    imgsz: int = 640,
    device: Optional[str] = None,
    min_area: Optional[float] = None,
    precomputed_deriv: Optional[PrecomputedDerivatives] = None,
    percentile_low: float = 2.0,
    percentile_high: float = 98.0,
    save_labels: bool = True,
) -> YoloOutputs:
    """
    Run YOLO11 segmentation inference on RGB tiles.
    
    Args:
        input_path: Path to multi-band GeoTIFF
        band_idx: Band indices (RGB bands used: band_idx[0:3])
        tile: Tile size for sliding window
        overlap: Overlap between tiles
        out_prefix: Output file prefix
        yolo_weights: Path to YOLO weights file (e.g. 'yolo11n-seg.pt')
        conf_threshold: Detection confidence threshold
        iou_threshold: NMS IoU threshold
        imgsz: YOLO model input size
        device: Device for inference ('0', 'cpu', etc.)
        min_area: Minimum polygon area filter
        precomputed_deriv: Precomputed derivatives (for RGB extraction)
        percentile_low: Lower percentile for RGB stretch
        percentile_high: Upper percentile for RGB stretch
    
    Returns:
        YoloOutputs with probability map, mask, and metadata
    """
    if YOLO is None:
        raise ImportError(
            "Ultralytics YOLO gerekli. Yüklemek için: pip install ultralytics>=8.1.0"
        )
    
    # Load YOLO model
    if yolo_weights is None:
        yolo_weights = "yolo11n-seg.pt"
        LOGGER.info("YOLO ağırlık dosyası belirtilmedi, varsayılan kullanılıyor: %s", yolo_weights)
        LOGGER.info("Not: Daha güçlü model için yolo_weights='yolo11s-seg.pt' kullanabilirsiniz.")

    resolved_yolo_weights = _resolve_yolo_weights_path(yolo_weights)
    if resolved_yolo_weights and str(resolved_yolo_weights) != str(yolo_weights):
        LOGGER.info("YOLO agirlik dosyasi yerelde bulundu: %s", resolved_yolo_weights)
    yolo_weights = resolved_yolo_weights or yolo_weights
    
    LOGGER.info("YOLO modeli yükleniyor: %s", yolo_weights)
    
    # Kuş bakışı (nadir) görüntüler için uyarı
    # Varsayılan COCO modelleri (yolov8, yolo11) için uyar, özel nadir modeller için uyarma
    weights_str = str(yolo_weights).lower()
    is_default_coco = ("yolov8" in weights_str or "yolo11" in weights_str) and \
                      "nadir" not in weights_str and \
                      "aerial" not in weights_str and \
                      "drone" not in weights_str and \
                      "visdrone" not in weights_str
    
    if is_default_coco:
        model_version = "YOLOv8" if "yolov8" in weights_str else "YOLO11"
        LOGGER.warning("")
        LOGGER.warning("=" * 70)
        LOGGER.warning("⚠️  KUŞ BAKIŞI (NADIR) GÖRÜNTÜ UYARISI")
        LOGGER.warning("=" * 70)
        LOGGER.warning("Varsayılan %s COCO modeli YATAY perspektiften eğitilmiştir.", model_version)
        LOGGER.warning("Arkeolojik LiDAR/İHA görüntüleri KUŞ BAKIŞI perspektiftedir.")
        LOGGER.warning("")
        LOGGER.warning("ÖNERİ:")
        LOGGER.warning("  1. VisDrone ile %s'i fine-tune edin (1-2 gün)", model_version)
        LOGGER.warning("     yolo segment train data=visdrone_yolo/data.yaml model=%s epochs=100", 
                      "yolov8s-seg.pt" if "yolov8" in weights_str else "yolo11s-seg.pt")
        LOGGER.warning("  2. Detaylar için: YOLO11_NADIR_TRAINING.md dosyasına bakın")
        LOGGER.warning("  3. Hazır nadir model: yolo_weights: 'models/yolov8_nadir_visdrone.pt'")
        LOGGER.warning("")
        LOGGER.warning("ŞUANKI SONUÇLAR: Genel envanter amaçlı, düşük doğruluk beklenir")
        LOGGER.warning("=" * 70)
        LOGGER.warning("")
    
    model = YOLO(yolo_weights)
    
    with rasterio.open(input_path) as src:
        height, width = src.height, src.width
        transform = src.transform
        crs = src.crs
        
        pixels = int(height) * int(width)
        one_float_bytes = pixels * np.dtype(np.float32).itemsize
        est_bytes = one_float_bytes * 2 + pixels  # prob_acc + weight_acc + valid_global(bool)
        avail = available_memory_bytes()
        use_memmap = False
        if avail is not None and avail > 0:
            use_memmap = est_bytes > int(avail * 0.75)
        else:
            use_memmap = est_bytes >= int(16 * 1024**3)
        scratch_dir: Optional[Path] = make_scratch_dir("yolo") if use_memmap else None
        if use_memmap:
            LOGGER.info(
                "Large raster detected (%dx%d). Using disk-backed accumulators under %s",
                width,
                height,
                scratch_dir,
            )

        # Initialize accumulator arrays
        prob_acc, _prob_acc_path = alloc_array(
            (height, width),
            np.float32,
            fill_value=0.0,
            use_memmap=use_memmap,
            scratch_dir=scratch_dir,
            name="yolo_prob_acc",
        )
        weight_acc, _weight_acc_path = alloc_array(
            (height, width),
            np.float32,
            fill_value=0.0,
            use_memmap=use_memmap,
            scratch_dir=scratch_dir,
            name="yolo_weight_acc",
        )
        valid_global, _valid_global_path = alloc_array(
            (height, width),
            bool,
            fill_value=False,
            use_memmap=use_memmap,
            scratch_dir=scratch_dir,
            name="yolo_valid_global",
        )
        
        # Initialize label detection storage
        all_detections: List[Dict[str, Any]] = []
        detection_id = 0
        
        total_tiles = math.ceil(height / max(tile - overlap, 1)) * math.ceil(
            width / max(tile - overlap, 1)
        )
        
        LOGGER.info("YOLO11 inference başlıyor: %dx%d piksel, %d tile", width, height, total_tiles)
        
        for window, row, col in progress_bar(
            generate_windows(width, height, tile, overlap),
            total=total_tiles,
            desc="YOLO11",
            unit="tile",
        ):
            win_height = int(window.height)
            win_width = int(window.width)
            
            # Read RGB bands
            if precomputed_deriv is not None:
                row_start = int(window.row_off)
                col_start = int(window.col_off)
                row_end = row_start + win_height
                col_end = col_start + win_width
                rgb = precomputed_deriv.rgb[:, row_start:row_end, col_start:col_end].copy()
            else:
                def read_band(idx: int) -> Optional[np.ndarray]:
                    if idx <= 0:
                        return None
                    data = src.read(idx, window=window, boundless=True, masked=True)
                    return np.ma.filled(data.astype(np.float32), np.nan)
                
                rgb = np.stack([read_band(band_idx[i]) for i in range(3)], axis=0)
            
            # Convert to uint8 for YOLO
            rgb_uint8 = stretch_to_uint8(rgb, low=percentile_low, high=percentile_high)
            
            # Transpose to HWC format and convert to BGR
            image_rgb = np.transpose(rgb_uint8, (1, 2, 0))
            image_bgr = np.ascontiguousarray(image_rgb[:, :, ::-1])
            
            # Check if tile is valid (not all NaN/zero)
            valid_mask = np.any(rgb_uint8 > 0, axis=0)
            if not np.any(valid_mask):
                continue
            
            # Run YOLO inference
            predict_kwargs = {
                'conf': conf_threshold,
                'iou': iou_threshold,
                'imgsz': imgsz,
                'verbose': False,
                'save': False,
            }
            if device is not None:
                predict_kwargs['device'] = device
            
            try:
                results = model.predict(image_bgr, **predict_kwargs)
                result = results[0]
                
                # Create probability map from masks
                tile_prob = np.zeros((win_height, win_width), dtype=np.float32)
                
                # Get class names from model
                class_names = result.names if hasattr(result, 'names') else {}
                
                if hasattr(result, 'masks') and result.masks is not None:
                    # Segmentation masks available
                    masks_data = result.masks.data.cpu().numpy()  # (N, H, W)
                    boxes = result.boxes.xyxy.cpu().numpy() if result.boxes is not None else None
                    confidences = result.boxes.conf.cpu().numpy() if result.boxes is not None else None
                    class_ids = result.boxes.cls.cpu().numpy().astype(int) if result.boxes is not None else None
                    
                    for i, mask in enumerate(masks_data):
                        # Resize mask to tile size if needed
                        if mask.shape != (win_height, win_width):
                            from scipy.ndimage import zoom
                            zy = win_height / mask.shape[0]
                            zx = win_width / mask.shape[1]
                            mask = zoom(mask, (zy, zx), order=1)
                        
                        # Weight by confidence
                        conf = float(confidences[i]) if confidences is not None and i < len(confidences) else 1.0
                        tile_prob = np.maximum(tile_prob, mask * conf)
                        
                        # Store detection info for label export
                        if save_labels and boxes is not None and i < len(boxes):
                            bbox = boxes[i]
                            xmin_tile, ymin_tile, xmax_tile, ymax_tile = bbox
                            
                            # Convert to global pixel coordinates
                            xmin_global = float(xmin_tile + col)
                            xmax_global = float(xmax_tile + col)
                            ymin_global = float(ymin_tile + row)
                            ymax_global = float(ymax_tile + row)
                            
                            # Get class info
                            class_id = int(class_ids[i]) if class_ids is not None and i < len(class_ids) else -1
                            class_name = class_names.get(class_id, f"class_{class_id}")
                            
                            # Create polygon from mask in geographic coordinates
                            mask_binary = (mask > 0.5).astype(np.uint8)
                            if np.any(mask_binary):
                                # Apply proper transformation: tile offset + main transform
                                # Window transform for this tile
                                tile_transform = transform * Affine.translation(col, row)
                                
                                # Convert mask to polygon using rasterio.features.shapes
                                geoms = list(shapes(mask_binary, mask=mask_binary, transform=tile_transform))
                                if geoms:
                                    geom_dict, _ = geoms[0]
                                    from shapely.geometry import shape as shapely_shape
                                    polygon = shapely_shape(geom_dict)
                                    
                                    # Calculate center in geographic coordinates from polygon
                                    center_geom = polygon.centroid
                                    center_x = float(center_geom.x)
                                    center_y = float(center_geom.y)
                                    
                                    all_detections.append({
                                        'id': detection_id,
                                        'class_id': class_id,
                                        'class_name': class_name,
                                        'confidence': conf,
                                        'bbox_xmin': xmin_global,
                                        'bbox_ymin': ymin_global,
                                        'bbox_xmax': xmax_global,
                                        'bbox_ymax': ymax_global,
                                        'center_x': center_x,
                                        'center_y': center_y,
                                        'tile_row': int(row),
                                        'tile_col': int(col),
                                        'geometry': polygon,
                                    })
                                    detection_id += 1
                
                elif hasattr(result, 'boxes') and result.boxes is not None:
                    # Only bounding boxes available - create masks from boxes
                    boxes = result.boxes.xyxy.cpu().numpy()
                    confidences = result.boxes.conf.cpu().numpy()
                    class_ids = result.boxes.cls.cpu().numpy().astype(int)
                    
                    for i, (bbox, conf, class_id) in enumerate(zip(boxes, confidences, class_ids)):
                        xmin, ymin, xmax, ymax = bbox.astype(int)
                        xmin = max(0, min(xmin, win_width - 1))
                        xmax = max(0, min(xmax, win_width))
                        ymin = max(0, min(ymin, win_height - 1))
                        ymax = max(0, min(ymax, win_height))
                        
                        if xmax > xmin and ymax > ymin:
                            tile_prob[ymin:ymax, xmin:xmax] = np.maximum(
                                tile_prob[ymin:ymax, xmin:xmax],
                                float(conf)
                            )
                            
                            # Store detection info for label export
                            if save_labels:
                                # Convert to global pixel coordinates
                                xmin_global = float(xmin + col)
                                xmax_global = float(xmax + col)
                                ymin_global = float(ymin + row)
                                ymax_global = float(ymax + row)
                                
                                # Get class name
                                class_name = class_names.get(int(class_id), f"class_{class_id}")
                                
                                # Create bounding box polygon in geographic coordinates
                                # Use Affine transform to convert pixel coords to geo coords
                                from shapely.geometry import Polygon
                                
                                # Get corners of bounding box in geo coordinates
                                # Affine transform * (col, row) -> (x_geo, y_geo)
                                corners = []
                                for px, py in [(xmin_global, ymin_global), 
                                              (xmax_global, ymin_global),
                                              (xmax_global, ymax_global),
                                              (xmin_global, ymax_global),
                                              (xmin_global, ymin_global)]:
                                    geo_x, geo_y = transform * (px, py)
                                    corners.append((geo_x, geo_y))
                                
                                polygon = Polygon(corners)
                                
                                # Calculate center from polygon
                                center_geom = polygon.centroid
                                center_x = float(center_geom.x)
                                center_y = float(center_geom.y)
                                
                                all_detections.append({
                                    'id': detection_id,
                                    'class_id': int(class_id),
                                    'class_name': class_name,
                                    'confidence': float(conf),
                                    'bbox_xmin': xmin_global,
                                    'bbox_ymin': ymin_global,
                                    'bbox_xmax': xmax_global,
                                    'bbox_ymax': ymax_global,
                                    'center_x': center_x,
                                    'center_y': center_y,
                                    'tile_row': int(row),
                                    'tile_col': int(col),
                                    'geometry': polygon,
                                })
                                detection_id += 1
                
                # Accumulate results
                row_slice = slice(row, row + win_height)
                col_slice = slice(col, col + win_width)
                
                valid_tile = valid_mask & np.isfinite(tile_prob)
                
                # Apply feathering weights
                weights = make_feather_weights(win_height, win_width, tile, overlap)
                
                prob_acc[row_slice, col_slice] += tile_prob * weights * valid_tile
                weight_acc[row_slice, col_slice] += weights * valid_tile.astype(np.float32)
                valid_global[row_slice, col_slice] |= valid_tile
                
            except Exception as e:
                LOGGER.warning("YOLO inference başarısız oldu (satır=%d, sütun=%d): %s", row, col, e)
                continue
        
        binary_mask, _binary_mask_path = alloc_array(
            (height, width),
            np.uint8,
            fill_value=0,
            use_memmap=use_memmap,
            scratch_dir=scratch_dir,
            name="yolo_binary_mask",
        )

        # Normalize accumulated probabilities in blocks to avoid huge temporaries.
        block_rows = 1024 if use_memmap else min(8192, int(height))
        total_blocks = math.ceil(int(height) / block_rows)
        for idx in progress_bar(
            range(total_blocks),
            desc="YOLO finalize",
            unit="block",
            total=total_blocks,
        ):
            row0 = idx * block_rows
            row1 = min(int(height), row0 + block_rows)
            acc_block = prob_acc[row0:row1, :]
            w_block = weight_acc[row0:row1, :]
            valid_block = valid_global[row0:row1, :]
            with np.errstate(divide="ignore", invalid="ignore"):
                np.divide(acc_block, w_block, out=acc_block, where=w_block > 0)
            acc_block[~valid_block] = np.nan
            binary_mask[row0:row1, :] = ((acc_block >= conf_threshold) & valid_block).astype(np.uint8)

        prob_map = prob_acc
        
        # Build output paths
        base_prefix = _output_base_path(out_prefix)
        base_prefix.parent.mkdir(parents=True, exist_ok=True)
        
        filename = build_filename_with_params(
            base_name=base_prefix.name,
            mode_suffix="yolo11",
            yolo_model=yolo_weights,
            threshold=conf_threshold,
            tile=tile,
            min_area=min_area,
        )
        
        prob_path = base_prefix.parent / f"{filename}_prob.tif"
        mask_path = base_prefix.parent / f"{filename}_mask.tif"
        
        # Write outputs
        write_prob_and_mask_rasters(
            prob_map=prob_map,
            mask=binary_mask,
            transform=transform,
            crs=crs,
            prob_path=prob_path,
            mask_path=mask_path,
        )
        
        # Collect labeled detections for downstream GeoPackage export
        labels_out_base = None
        if save_labels and all_detections:
            labels_out_base = base_prefix.parent / f"{filename}_labels"
            LOGGER.info("YOLO11 etiketli tespitler hazirlandi: %d nesne", len(all_detections))
        elif save_labels and not all_detections:
            LOGGER.info("YOLO11 hiç tespit bulamadı, etiketli çıktı yok")
        
        # GPU belleğini temizle (bellek sızıntısını önle)
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
            LOGGER.debug("GPU belleği temizlendi (YOLO)")

        if use_memmap:
            prob_map_out = None
            mask_out = None
            del prob_map, binary_mask, prob_acc, weight_acc, valid_global
            gc.collect()
        else:
            prob_map_out = prob_map
            mask_out = binary_mask

        cleanup_scratch_dir(scratch_dir)

        return YoloOutputs(
            prob_path=prob_path,
            mask_path=mask_path,
            prob_map=prob_map_out,
            mask=mask_out,
            transform=transform,
            crs=crs,
            threshold=conf_threshold,
            label_records=all_detections if save_labels else None,
            labels_out_base=labels_out_base,
        )


def infer_classic_tiled(
    input_path: Path,
    band_idx: Sequence[int],
    tile: int,
    overlap: int,
    out_prefix: Path,
    classic_th: Optional[float] = None,
    modes: Sequence[str] = ("rvtlog",),
    feather: bool = True,
    save_intermediate: bool = False,
    min_area: Optional[float] = None,
    precomputed_deriv: Optional[PrecomputedDerivatives] = None,
    derivative_cache_tif: Optional[Path] = None,
    derivative_cache_meta: Optional[Path] = None,
    enable_curvature: bool = True,
    enable_tpi: bool = True,
    tpi_radii: Tuple[int, ...] = (5, 15, 30),
    sigma_scales: Optional[Sequence[float]] = None,
    morphology_radii: Optional[Sequence[int]] = None,
    rvt_radii: Optional[Sequence[float]] = None,
    gaussian_gradient_sigma: Optional[float] = None,
    local_variance_window: Optional[int] = None,
    gaussian_lrm_sigma: Optional[float] = None,
) -> ClassicOutputs:
    """Run classical raster scoring methods in a tiled fashion."""
    if len(band_idx) < 5 or band_idx[4] <= 0:
        raise ValueError("DTM band index must be provided for classic inference.")
    base_modes = []
    for mode in modes:
        key = mode.strip().lower()
        if key not in base_modes:
            base_modes.append(key)
    if not base_modes:
        raise ValueError("At least one classic mode must be specified.")
    allowed = {"rvtlog", "hessian", "morph"}
    unknown = [m for m in base_modes if m not in allowed]
    if unknown:
        raise ValueError(f"Unsupported classic mode(s): {', '.join(unknown)}")
    base_prefix = _output_base_path(out_prefix)
    base_prefix.parent.mkdir(parents=True, exist_ok=True)

    with ExitStack() as stack:
        src = stack.enter_context(rasterio.open(input_path))
        height, width = src.height, src.width
        transform = src.transform
        crs = src.crs
        pixel_size = float((abs(transform.a) + abs(transform.e)) / 2.0)

        deriv_ds: Optional[rasterio.io.DatasetReader] = None
        deriv_band_map: Optional[Dict[str, int]] = None
        if precomputed_deriv is None and derivative_cache_tif is not None and derivative_cache_meta is not None:
            info = load_derivative_raster_cache_info(
                derivative_cache_tif,
                derivative_cache_meta,
                input_path,
                band_idx,
                rvt_radii=rvt_radii,
                gaussian_lrm_sigma=gaussian_lrm_sigma,
                enable_curvature=enable_curvature,
                enable_tpi=enable_tpi,
                tpi_radii=tpi_radii,
            )
            if info is not None:
                deriv_band_map = {str(k): int(v) for k, v in dict(info.get("band_map", {})).items()}
                deriv_ds = stack.enter_context(rasterio.open(derivative_cache_tif))
                LOGGER.info("Classic iin derivatives raster-cache kullanlyor: %s", derivative_cache_tif)

        # Large-raster fast path: avoid per-mode full-size buffers + global percentile ops.
        pixels = int(height) * int(width)
        one_float_bytes = pixels * np.dtype(np.float32).itemsize
        est_bytes = one_float_bytes * 2 + pixels  # combined prob + weights + valid_global(bool)
        avail = available_memory_bytes()
        use_memmap = False
        if avail is not None and avail > 0:
            use_memmap = est_bytes > int(avail * 0.75)
        else:
            use_memmap = est_bytes >= int(16 * 1024**3)
        large_raster = pixels >= 200_000_000 or use_memmap

        if large_raster:
            scratch_dir: Optional[Path] = make_scratch_dir("classic") if use_memmap else None
            if scratch_dir is not None:
                stack.callback(cleanup_scratch_dir, scratch_dir)
            if use_memmap:
                LOGGER.info(
                    "Large raster detected (%dx%d). Using disk-backed classic accumulators under %s",
                    width,
                    height,
                    scratch_dir,
                )
            elif pixels >= 200_000_000:
                LOGGER.info(
                    "Large raster detected (%dx%d). Using streaming classic aggregation path.",
                    width,
                    height,
                )

            prob_acc_big, _prob_acc_path = alloc_array(
                (height, width),
                np.float32,
                fill_value=0.0,
                use_memmap=use_memmap,
                scratch_dir=scratch_dir,
                name="classic_prob_acc",
            )
            weight_acc_big, _weight_acc_path = alloc_array(
                (height, width),
                np.float32,
                fill_value=0.0,
                use_memmap=use_memmap,
                scratch_dir=scratch_dir,
                name="classic_weight_acc",
            )
            valid_global_big, _valid_global_path = alloc_array(
                (height, width),
                bool,
                fill_value=False,
                use_memmap=use_memmap,
                scratch_dir=scratch_dir,
                name="classic_valid_global",
            )

            dtm_idx = band_idx[4]
            total_tiles = math.ceil(height / max(tile - overlap, 1)) * math.ceil(
                width / max(tile - overlap, 1)
            )

            for window, row, col in progress_bar(
                generate_windows(width, height, tile, overlap),
                total=total_tiles,
                desc="Classic",
                unit="tile",
            ):
                dtm_window = src.read(dtm_idx, window=window, boundless=True, masked=True)
                dtm_tile = np.ma.filled(dtm_window.astype(np.float32), np.nan)

                win_height = int(window.height)
                win_width = int(window.width)
                row_slice = slice(row, row + win_height)
                col_slice = slice(col, col + win_width)
                valid_tile = np.isfinite(dtm_tile)
                valid_global_big[row_slice, col_slice] |= valid_tile
                if not np.any(valid_tile):
                    continue

                if feather:
                    weights = make_feather_weights(win_height, win_width, tile, overlap)
                else:
                    weights = np.ones((win_height, win_width), dtype=np.float32)

                prob_tiles: List[np.ndarray] = []
                for mode in base_modes:
                    if mode == "rvtlog":
                        if precomputed_deriv is not None:
                            row_start = int(window.row_off)
                            col_start = int(window.col_off)
                            row_end = row_start + win_height
                            col_end = col_start + win_width
                            svf_tile = precomputed_deriv.svf[row_start:row_end, col_start:col_end]
                            slrm_tile = precomputed_deriv.slrm[row_start:row_end, col_start:col_end]
                            score_tile = _score_rvtlog(
                                dtm_tile,
                                pixel_size=pixel_size,
                                pre_svf=svf_tile,
                                pre_lrm=slrm_tile,
                                sigmas=sigma_scales,
                                gaussian_gradient_sigma=gaussian_gradient_sigma,
                                local_variance_window=local_variance_window,
                                rvt_radii=rvt_radii,
                                gaussian_lrm_sigma=gaussian_lrm_sigma,
                            )
                        elif deriv_ds is not None and deriv_band_map is not None:
                            idxs = [
                                int(deriv_band_map["svf"]),
                                int(deriv_band_map["slrm"]),
                            ]
                            deriv_tile = deriv_ds.read(indexes=idxs, window=window, boundless=True, masked=True)
                            deriv_tile = np.ma.filled(deriv_tile.astype(np.float32), np.nan)
                            svf_tile, slrm_tile = deriv_tile[0], deriv_tile[1]
                            score_tile = _score_rvtlog(
                                dtm_tile,
                                pixel_size=pixel_size,
                                pre_svf=svf_tile,
                                pre_lrm=slrm_tile,
                                sigmas=sigma_scales,
                                gaussian_gradient_sigma=gaussian_gradient_sigma,
                                local_variance_window=local_variance_window,
                                rvt_radii=rvt_radii,
                                gaussian_lrm_sigma=gaussian_lrm_sigma,
                            )
                        else:
                            score_tile = _score_rvtlog(
                                dtm_tile,
                                pixel_size=pixel_size,
                                sigmas=sigma_scales,
                                gaussian_gradient_sigma=gaussian_gradient_sigma,
                                local_variance_window=local_variance_window,
                                rvt_radii=rvt_radii,
                                gaussian_lrm_sigma=gaussian_lrm_sigma,
                            )
                    elif mode == "hessian":
                        score_tile = _score_hessian(dtm_tile, sigmas=sigma_scales)
                    elif mode == "morph":
                        score_tile = _score_morph(dtm_tile, radii=morphology_radii)
                    else:  # pragma: no cover - guarded above
                        continue

                    valid_mode = valid_tile & np.isfinite(score_tile)
                    if not np.any(valid_mode):
                        continue
                    score_tile_masked = np.where(valid_mode, score_tile, np.nan).astype(np.float32)
                    prob_tile = _norm01(score_tile_masked)
                    prob_tile[~valid_mode] = np.nan
                    prob_tiles.append(prob_tile.astype(np.float32, copy=False))

                if not prob_tiles:
                    continue

                if len(prob_tiles) == 1:
                    mean_prob = prob_tiles[0]
                else:
                    with np.errstate(all="ignore"):
                        mean_prob = np.nanmean(np.stack(prob_tiles, axis=0), axis=0).astype(np.float32)
                mean_prob[~valid_tile] = np.nan

                valid_combined = np.isfinite(mean_prob)
                if not np.any(valid_combined):
                    continue

                weight_tile = weights * valid_combined.astype(np.float32)
                prob_fill = np.nan_to_num(mean_prob, nan=0.0).astype(np.float32, copy=False)

                prob_acc_big[row_slice, col_slice] += prob_fill * weight_tile
                weight_acc_big[row_slice, col_slice] += weight_tile

            combined_prob = prob_acc_big
            combined_mask, _combined_mask_path = alloc_array(
                (height, width),
                np.uint8,
                fill_value=0,
                use_memmap=use_memmap,
                scratch_dir=scratch_dir,
                name="classic_binary_mask",
            )

            block_rows = 1024 if use_memmap else min(8192, int(height))
            total_blocks = math.ceil(int(height) / block_rows)
            for idx in progress_bar(
                range(total_blocks),
                desc="Classic finalize",
                unit="block",
                total=total_blocks,
            ):
                row0 = idx * block_rows
                row1 = min(int(height), row0 + block_rows)
                acc_block = combined_prob[row0:row1, :]
                w_block = weight_acc_big[row0:row1, :]
                valid_block = valid_global_big[row0:row1, :]
                with np.errstate(divide="ignore", invalid="ignore"):
                    np.divide(acc_block, w_block, out=acc_block, where=w_block > 0)
                acc_block[(w_block <= 0) | (~valid_block)] = np.nan
                np.clip(acc_block, 0.0, 1.0, out=acc_block)

            if classic_th is not None:
                combined_threshold = float(classic_th)
            else:
                combined_threshold = otsu_threshold_streaming(
                    combined_prob,
                    valid_global_big,
                    block_rows=block_rows,
                )

            for idx in progress_bar(
                range(total_blocks),
                desc="Classic mask",
                unit="block",
                total=total_blocks,
            ):
                row0 = idx * block_rows
                row1 = min(int(height), row0 + block_rows)
                prob_block = combined_prob[row0:row1, :]
                valid_block = np.isfinite(prob_block)
                combined_mask[row0:row1, :] = (valid_block & (prob_block >= combined_threshold)).astype(
                    np.uint8
                )

            # Mode listesini stringe çevir (combo veya virgülle ayrılmış)
            modes_str = "combo" if len(modes) > 1 else modes[0] if modes else "combo"
            # Otsu mu manuel eşik mi?
            th_type = "otsu" if classic_th is None else None
            
            classic_filename = build_filename_with_params(
                base_name=base_prefix.name,
                mode_suffix="classic",
                classic_modes=modes_str,
                threshold=combined_threshold,
                threshold_type=th_type,
                tile=tile,
                min_area=min_area,
            )
            classic_prob_path = base_prefix.parent / f"{classic_filename}_prob.tif"
            classic_mask_path = base_prefix.parent / f"{classic_filename}_mask.tif"

            write_prob_and_mask_rasters(
                prob_map=combined_prob,
                mask=combined_mask,
                transform=transform,
                crs=crs,
                prob_path=classic_prob_path,
                mask_path=classic_mask_path,
            )

            if use_memmap:
                prob_map_out = None
                mask_out = None
                del combined_prob, combined_mask, prob_acc_big, weight_acc_big, valid_global_big
                gc.collect()
            else:
                prob_map_out = combined_prob
                mask_out = combined_mask

            return ClassicOutputs(
                prob_path=classic_prob_path,
                mask_path=classic_mask_path,
                prob_map=prob_map_out,
                mask=mask_out,
                transform=transform,
                crs=crs,
                threshold=combined_threshold,
                per_mode={},
            )

        prob_acc: Dict[str, np.ndarray] = {
            mode: np.zeros((height, width), dtype=np.float32) for mode in base_modes
        }
        weight_acc: Dict[str, np.ndarray] = {
            mode: np.zeros((height, width), dtype=np.float32) for mode in base_modes
        }
        valid_global = np.zeros((height, width), dtype=bool)
        dtm_idx = band_idx[4]

        total_tiles = math.ceil(height / max(tile - overlap, 1)) * math.ceil(
            width / max(tile - overlap, 1)
        )

        for window, row, col in progress_bar(
            generate_windows(width, height, tile, overlap),
            total=total_tiles,
            desc="Classic",
            unit="tile",
        ):
            dtm_window = src.read(dtm_idx, window=window, boundless=True, masked=True)
            dtm_tile = np.ma.filled(dtm_window.astype(np.float32), np.nan)

            win_height = int(window.height)
            win_width = int(window.width)
            row_slice = slice(row, row + win_height)
            col_slice = slice(col, col + win_width)
            valid_tile = np.isfinite(dtm_tile)
            valid_global[row_slice, col_slice] |= valid_tile
            if not np.any(valid_tile):
                continue

            if feather:
                weights = make_feather_weights(win_height, win_width, tile, overlap)
            else:
                weights = np.ones((win_height, win_width), dtype=np.float32)

            for mode in base_modes:
                if mode == "rvtlog":
                    if precomputed_deriv is not None:
                        row_start = int(window.row_off)
                        col_start = int(window.col_off)
                        row_end = row_start + win_height
                        col_end = col_start + win_width
                        svf_tile = precomputed_deriv.svf[row_start:row_end, col_start:col_end]
                        slrm_tile = precomputed_deriv.slrm[row_start:row_end, col_start:col_end]
                        score_tile = _score_rvtlog(
                            dtm_tile, pixel_size=pixel_size,
                            pre_svf=svf_tile,
                            pre_lrm=slrm_tile,
                            sigmas=sigma_scales,
                            gaussian_gradient_sigma=gaussian_gradient_sigma,
                            local_variance_window=local_variance_window,
                            rvt_radii=rvt_radii,
                            gaussian_lrm_sigma=gaussian_lrm_sigma,
                        )
                    elif deriv_ds is not None and deriv_band_map is not None:
                        idxs = [
                            int(deriv_band_map["svf"]),
                            int(deriv_band_map["slrm"]),
                        ]
                        deriv_tile = deriv_ds.read(indexes=idxs, window=window, boundless=True, masked=True)
                        deriv_tile = np.ma.filled(deriv_tile.astype(np.float32), np.nan)
                        svf_tile, slrm_tile = deriv_tile[0], deriv_tile[1]
                        score_tile = _score_rvtlog(
                            dtm_tile,
                            pixel_size=pixel_size,
                            pre_svf=svf_tile,
                            pre_lrm=slrm_tile,
                            sigmas=sigma_scales,
                            gaussian_gradient_sigma=gaussian_gradient_sigma,
                            local_variance_window=local_variance_window,
                            rvt_radii=rvt_radii,
                            gaussian_lrm_sigma=gaussian_lrm_sigma,
                        )
                    else:
                        score_tile = _score_rvtlog(
                            dtm_tile, pixel_size=pixel_size,
                            sigmas=sigma_scales,
                            gaussian_gradient_sigma=gaussian_gradient_sigma,
                            local_variance_window=local_variance_window,
                            rvt_radii=rvt_radii,
                            gaussian_lrm_sigma=gaussian_lrm_sigma,
                        )
                elif mode == "hessian":
                    score_tile = _score_hessian(dtm_tile, sigmas=sigma_scales)
                elif mode == "morph":
                    score_tile = _score_morph(dtm_tile, radii=morphology_radii)
                else:  # pragma: no cover - guarded above
                    continue

                valid_mode = valid_tile & np.isfinite(score_tile)
                if not np.any(valid_mode):
                    continue
                score_tile_masked = np.where(valid_mode, score_tile, np.nan).astype(np.float32)
                prob_tile = _norm01(score_tile_masked)
                prob_tile[~valid_mode] = np.nan
                weight_tile = weights * valid_mode.astype(np.float32)
                prob_fill = np.nan_to_num(prob_tile, nan=0.0)

                prob_acc[mode][row_slice, col_slice] += prob_fill * weight_tile
                weight_acc[mode][row_slice, col_slice] += weight_tile

    per_mode_prob: Dict[str, np.ndarray] = {}
    per_mode_mask: Dict[str, np.ndarray] = {}
    per_mode_thresholds: Dict[str, float] = {}

    for mode in base_modes:
        acc = prob_acc[mode]
        weights = weight_acc[mode]
        with np.errstate(divide="ignore", invalid="ignore"):
            prob_map = np.divide(acc, weights, out=np.zeros_like(acc), where=weights > 0)
        prob_map = prob_map.astype(np.float32)
        prob_map[weights <= 0] = np.nan
        prob_map[~valid_global] = np.nan
        valid_mode = np.isfinite(prob_map)
        if np.any(valid_mode):
            prob_norm = _norm01(np.where(valid_mode, prob_map, np.nan))
            prob_norm[~valid_mode] = np.nan
        else:
            prob_norm = np.full_like(prob_map, np.nan, dtype=np.float32)
        if classic_th is not None:
            threshold = float(classic_th)
        else:
            threshold = _otsu_threshold_0to1(prob_norm, valid_mode)
        mask = np.zeros_like(prob_norm, dtype=np.uint8)
        mask[valid_mode & (prob_norm >= threshold)] = 1
        per_mode_prob[mode] = prob_norm.astype(np.float32)
        per_mode_mask[mode] = mask
        per_mode_thresholds[mode] = threshold

    if per_mode_prob:
        stack = np.stack([per_mode_prob[m] for m in base_modes], axis=0)
        combined_prob = np.nanmean(stack, axis=0).astype(np.float32)
    else:
        combined_prob = np.full((height, width), np.nan, dtype=np.float32)
    combined_prob[~valid_global] = np.nan
    combined_valid = np.isfinite(combined_prob)
    combined_prob = np.clip(combined_prob, 0.0, 1.0, out=combined_prob)

    if classic_th is not None:
        combined_threshold = float(classic_th)
    else:
        combined_threshold = _otsu_threshold_0to1(combined_prob, combined_valid)
    combined_mask = np.zeros_like(combined_prob, dtype=np.uint8)
    combined_mask[combined_valid & (combined_prob >= combined_threshold)] = 1

    # Mode listesini stringe çevir (combo veya virgülle ayrılmış)
    modes_str = "combo" if len(base_modes) > 1 else base_modes[0] if base_modes else "combo"
    # Otsu mu manuel eşik mi?
    th_type = "otsu" if classic_th is None else None
    
    # Parametreli dosya adı oluştur (classic için)
    classic_filename = build_filename_with_params(
        base_name=base_prefix.name,
        mode_suffix="classic",
        classic_modes=modes_str,
        threshold=combined_threshold,
        threshold_type=th_type,
        tile=tile,
        min_area=min_area,
    )
    
    classic_prob_path = base_prefix.parent / f"{classic_filename}_prob.tif"
    classic_mask_path = base_prefix.parent / f"{classic_filename}_mask.tif"
    
    write_prob_and_mask_rasters(
        prob_map=combined_prob,
        mask=combined_mask,
        transform=transform,
        crs=crs,
        prob_path=classic_prob_path,
        mask_path=classic_mask_path,
    )

    per_mode_outputs: Dict[str, ClassicModeOutput] = {}
    write_individual = save_intermediate or len(base_modes) == 1
    if write_individual:
        for mode in base_modes:
            prob_map = per_mode_prob.get(mode)
            mask_map = per_mode_mask.get(mode)
            if prob_map is None or mask_map is None:
                continue
            
            # Her mode için ayrı parametreli dosya adı
            mode_filename = build_filename_with_params(
                base_name=base_prefix.name,
                mode_suffix=f"classic_{mode}",
                threshold=combined_threshold,
                threshold_type=th_type,
                tile=tile,
                min_area=min_area,
            )
            
            mode_prob_path = base_prefix.parent / f"{mode_filename}_prob.tif"
            mode_mask_path = base_prefix.parent / f"{mode_filename}_mask.tif"
            
            write_prob_and_mask_rasters(
                prob_map=prob_map,
                mask=mask_map,
                transform=transform,
                crs=crs,
                prob_path=mode_prob_path,
                mask_path=mode_mask_path,
            )
            
            per_mode_outputs[mode] = ClassicModeOutput(
                prob_path=mode_prob_path,
                mask_path=mode_mask_path,
                prob_map=prob_map,
                mask=mask_map,
                threshold=per_mode_thresholds.get(mode, combined_threshold),
            )

    return ClassicOutputs(
        prob_path=classic_prob_path,
        mask_path=classic_mask_path,
        prob_map=combined_prob,
        mask=combined_mask,
        transform=transform,
        crs=crs,
        threshold=combined_threshold,
        per_mode=per_mode_outputs,
    )


def _format_crs_label(crs: Optional[RasterioCRS]) -> str:
    """Return a compact CRS label for tabular exports."""
    if crs is None:
        return "unknown"
    try:
        epsg = crs.to_epsg()
        if epsg is not None:
            return f"EPSG:{epsg}"
    except Exception:
        pass
    try:
        text = crs.to_string()
        if text:
            return str(text)
    except Exception:
        pass
    return "unknown"


def _resolve_area_transformers(
    crs: Optional[RasterioCRS],
) -> Tuple[Optional[CRS], Optional[Transformer], Optional[Transformer]]:
    crs_obj: Optional[CRS] = None
    if crs:
        try:
            crs_obj = CRS.from_wkt(crs.to_wkt())
        except Exception:  # pragma: no cover - defensive
            LOGGER.warning("Unable to parse CRS; assuming projected coordinates for area.")

    to_area: Optional[Transformer] = None
    to_native: Optional[Transformer] = None
    if crs_obj and not crs_obj.is_projected:
        area_crs = CRS.from_epsg(6933)
        to_area = Transformer.from_crs(crs_obj, area_crs, always_xy=True)
        to_native = Transformer.from_crs(area_crs, crs_obj, always_xy=True)
    elif not crs_obj:
        LOGGER.warning(
            "Input CRS unknown; polygon areas assume coordinate units are meters."
        )
    return crs_obj, to_area, to_native


def _normalize_gpkg_property_value(value: Any) -> Any:
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float, str)) or value is None:
        return value
    return str(value)


def _ordered_record_property_names(
    records: Sequence[Dict[str, Any]],
    column_order: Optional[Sequence[str]] = None,
) -> List[str]:
    available = {str(key) for rec in records for key in rec.keys() if str(key) != "geometry"}
    ordered: List[str] = []
    if column_order:
        ordered.extend([str(key) for key in column_order if str(key) in available and str(key) != "geometry"])
    extras = sorted(key for key in available if key not in ordered)
    ordered.extend(extras)
    return ordered


def _infer_fiona_property_type(values: Sequence[Any]) -> str:
    for value in values:
        if value is None:
            continue
        if isinstance(value, np.generic):
            value = value.item()
        if isinstance(value, bool):
            return "int"
        if isinstance(value, int):
            return "int"
        if isinstance(value, float):
            return "float"
        return "str"
    return "str"


def _resolve_fiona_geometry_type(records: Sequence[Dict[str, Any]]) -> str:
    geom_types = {
        str(getattr(rec.get("geometry"), "geom_type", "")).strip()
        for rec in records
        if rec.get("geometry") is not None and not getattr(rec.get("geometry"), "is_empty", False)
    }
    geom_types.discard("")
    if not geom_types:
        raise ValueError("No valid geometries available for GPKG export.")

    if geom_types <= {"Polygon", "MultiPolygon"}:
        return "MultiPolygon" if "MultiPolygon" in geom_types else "Polygon"
    if geom_types <= {"LineString", "MultiLineString"}:
        return "MultiLineString" if "MultiLineString" in geom_types else "LineString"
    if geom_types <= {"Point", "MultiPoint"}:
        return "MultiPoint" if "MultiPoint" in geom_types else "Point"
    if len(geom_types) == 1:
        return next(iter(geom_types))
    raise ValueError(f"Unsupported mixed geometry types for GPKG export: {sorted(geom_types)}")


def _write_records_to_gpkg(
    *,
    records: Sequence[Dict[str, Any]],
    crs: Optional[RasterioCRS],
    gpkg_path: Path,
    layer_name: str,
    column_order: Optional[Sequence[str]] = None,
) -> Path:
    if not records:
        raise ValueError("No records available for GPKG export.")

    safe_layer_name = _safe_gpkg_layer_name(layer_name)
    property_names = _ordered_record_property_names(records, column_order=column_order)
    gpkg_path.parent.mkdir(parents=True, exist_ok=True)

    if gpd is not None:
        gdf = gpd.GeoDataFrame(records, geometry="geometry", crs=crs)
        ordered_columns = [name for name in property_names if name in gdf.columns]
        if "geometry" in gdf.columns:
            ordered_columns.append("geometry")
        gdf = gdf[ordered_columns]
        gdf.to_file(gpkg_path, driver="GPKG", layer=safe_layer_name)
        return gpkg_path

    if fiona is None:
        raise RuntimeError("Neither geopandas nor fiona is available for GPKG export.")

    schema = {
        "geometry": _resolve_fiona_geometry_type(records),
        "properties": {
            name: _infer_fiona_property_type([rec.get(name) for rec in records])
            for name in property_names
        },
    }
    crs_wkt = crs.to_wkt() if crs else None
    with fiona.open(
        gpkg_path,
        mode="w",
        driver="GPKG",
        layer=safe_layer_name,
        schema=schema,
        crs_wkt=crs_wkt,
    ) as dst:
        for rec in records:
            geom = rec.get("geometry")
            if geom is None or getattr(geom, "is_empty", False):
                continue
            dst.write(
                {
                    "geometry": mapping(geom),
                    "properties": {
                        name: _normalize_gpkg_property_value(rec.get(name))
                        for name in property_names
                    },
                }
            )
    return gpkg_path


def _resolve_vector_gpkg_target(
    *,
    out_base: Path,
    layer_name: str,
    gpkg_mode: str,
    single_gpkg_path: Optional[Path],
) -> Tuple[Path, str]:
    safe_layer_name = _safe_gpkg_layer_name(layer_name)
    if str(gpkg_mode).strip().lower() == "single":
        if single_gpkg_path is None:
            raise ValueError("single_gpkg_path must be provided when gpkg_mode='single'.")
        return single_gpkg_path, safe_layer_name
    gpkg_path = _append_output_suffix(_output_base_path(out_base), ".gpkg")
    return gpkg_path, safe_layer_name


def _prepare_yolo_detection_records(
    records: Sequence[Dict[str, Any]],
    crs: Optional[RasterioCRS],
) -> List[Dict[str, Any]]:
    if not records:
        return []

    _crs_obj, to_area, _to_native = _resolve_area_transformers(crs)
    prepared: List[Dict[str, Any]] = []
    for rec in records:
        geom = rec.get("geometry")
        if geom is None or getattr(geom, "is_empty", False):
            continue
        if to_area:
            area_m2 = float(shapely_transform(to_area.transform, geom).area)
        else:
            area_m2 = float(geom.area)
        item = dict(rec)
        item["area_m2"] = area_m2
        prepared.append(item)
    return prepared


def _can_vectorize_predictions() -> Tuple[bool, Optional[str]]:
    """Return whether polygon vectorization dependencies are available."""
    if fiona is None and gpd is None:
        return False, "geopandas veya fiona kurulu degil"
    if mapping is None or shape is None or shapely_transform is None:
        return False, "shapely kurulu degil"
    if CRS is None or Transformer is None:
        return False, "pyproj kurulu degil"
    return True, None


def _transform_points_to_wgs84(
    xs: Sequence[float],
    ys: Sequence[float],
    crs: Optional[RasterioCRS],
) -> Tuple[List[Optional[float]], List[Optional[float]]]:
    """Best-effort transform of native coordinates to WGS84 lon/lat."""
    if not xs or not ys or crs is None:
        return [None] * len(xs), [None] * len(ys)

    try:
        lon_vals, lat_vals = rasterio_transform(crs, "EPSG:4326", list(xs), list(ys))
    except Exception:
        return [None] * len(xs), [None] * len(ys)

    lons: List[Optional[float]] = []
    lats: List[Optional[float]] = []
    for lon, lat in zip(lon_vals, lat_vals):
        try:
            lons.append(float(lon))
            lats.append(float(lat))
        except Exception:
            lons.append(None)
            lats.append(None)
    return lons, lats


def _xlsx_column_name(index: int) -> str:
    """Convert 1-based column index to Excel column name."""
    if index <= 0:
        raise ValueError("Excel column index must be positive.")
    letters: List[str] = []
    while index > 0:
        index, rem = divmod(index - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def _build_inline_xlsx_sheet_xml(
    *,
    rows: Sequence[Sequence[Any]],
    sheet_name: str,
) -> str:
    """Build a minimal worksheet XML using inline strings/numbers."""
    if not rows:
        rows = [()]

    max_cols = max(len(row) for row in rows)
    max_cols = max(1, max_cols)
    last_ref = f"{_xlsx_column_name(max_cols)}{max(1, len(rows))}"

    row_xml: List[str] = []
    for row_idx, row in enumerate(rows, start=1):
        cell_xml: List[str] = []
        for col_idx, value in enumerate(row, start=1):
            cell_ref = f"{_xlsx_column_name(col_idx)}{row_idx}"
            if value is None or value == "":
                cell_xml.append(f'<c r="{cell_ref}" t="inlineStr"><is><t></t></is></c>')
                continue
            if isinstance(value, (int, float, np.integer, np.floating)) and np.isfinite(value):
                cell_xml.append(f'<c r="{cell_ref}"><v>{value}</v></c>')
                continue
            text = xml_escape(str(value))
            cell_xml.append(
                f'<c r="{cell_ref}" t="inlineStr"><is><t xml:space="preserve">{text}</t></is></c>'
            )
        row_xml.append(f'<row r="{row_idx}">{"".join(cell_xml)}</row>')

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<dimension ref="A1:{last_ref}"/>'
        '<sheetViews><sheetView workbookViewId="0">'
        '<pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>'
        "</sheetView></sheetViews>"
        '<sheetFormatPr defaultRowHeight="15"/>'
        f'<sheetData>{"".join(row_xml)}</sheetData>'
        "</worksheet>"
    )


def _write_builtin_xlsx(
    *,
    rows: Sequence[Dict[str, Any]],
    field_order: Sequence[str],
    xlsx_path: Path,
) -> Path:
    """Write a minimal valid .xlsx file without external Excel libraries."""
    sheet_rows: List[List[Any]] = [list(field_order)]
    for row in rows:
        sheet_rows.append([row.get(col) for col in field_order])

    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="candidate_locations" sheetId="1" r:id="rId1"/></sheets>'
        "</workbook>"
    )
    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
        'Target="styles.xml"/>'
        "</Relationships>"
    )
    package_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/styles.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        "</Types>"
    )
    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="1"><font><sz val="11"/><name val="Calibri"/><family val="2"/></font></fonts>'
        '<fills count="2"><fill><patternFill patternType="none"/></fill>'
        '<fill><patternFill patternType="gray125"/></fill></fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        "</styleSheet>"
    )
    sheet_xml = _build_inline_xlsx_sheet_xml(rows=sheet_rows, sheet_name="candidate_locations")

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(xlsx_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types_xml)
        zf.writestr("_rels/.rels", package_rels_xml)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
        zf.writestr("xl/styles.xml", styles_xml)
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return xlsx_path


def _write_candidate_location_rows_xlsx(
    *,
    rows: Sequence[Dict[str, Any]],
    field_order: Sequence[str],
    out_base: Path,
) -> Optional[Path]:
    """Write candidate rows to XLSX, using built-in fallback if openpyxl is missing."""
    xlsx_path = _append_output_suffix(out_base, ".xlsx")
    if Workbook is not None:
        wb = Workbook()
        ws = wb.active
        ws.title = "candidate_locations"
        ws.append(list(field_order))
        if rows:
            google_maps_col = list(field_order).index("google_maps_url") + 1
            for row_idx, row in enumerate(rows, start=2):
                ws.append([row.get(col) for col in field_order])
                maps_url = str(row.get("google_maps_url") or "").strip()
                if maps_url:
                    cell = ws.cell(row=row_idx, column=google_maps_col)
                    cell.hyperlink = maps_url
                    cell.style = "Hyperlink"
        ws.freeze_panes = "A2"
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(xlsx_path)
        return xlsx_path

    _warn_once(
        "builtin_xlsx_writer",
        "openpyxl kurulu degil; aday tablosu yerlesik XLSX yazicisi ile olusturulacak.",
    )
    return _write_builtin_xlsx(rows=rows, field_order=field_order, xlsx_path=xlsx_path)


def _normalize_excel_cell_value(value: Any) -> Any:
    """Convert numpy scalars and non-finite floats into Excel-safe primitives."""
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, float):
        if not np.isfinite(value):
            return None
    return value


def _candidate_row_number(row: Dict[str, Any], key: str) -> Optional[float]:
    value = _normalize_excel_cell_value(row.get(key))
    if value in (None, ""):
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if not np.isfinite(numeric):
        return None
    return numeric


def _candidate_row_int(row: Dict[str, Any], key: str) -> Optional[int]:
    value = _candidate_row_number(row, key)
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _combined_candidate_priority_band(score: Optional[float]) -> str:
    if score is None:
        return "Bilinmiyor"
    if score >= _COMBINED_CANDIDATE_PRIORITY_HIGH:
        return "Yuksek"
    if score >= _COMBINED_CANDIDATE_PRIORITY_MEDIUM:
        return "Orta"
    return "Dusuk"


def _combined_candidate_sort_key(row: Dict[str, Any]) -> Tuple[float, float, str, int, int]:
    score = _candidate_row_number(row, "score_mean")
    area = _candidate_row_number(row, "area_m2")
    source_label = str(row.get("source_label") or "")
    scale_level = _candidate_row_int(row, "scale_level")
    candidate_id = _candidate_row_int(row, "candidate_id")
    return (
        -(score if score is not None else -1.0),
        -(area if area is not None else -1.0),
        source_label,
        scale_level if scale_level is not None else 999_999,
        candidate_id if candidate_id is not None else 999_999,
    )


def _build_combined_priority_rows(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    ordered_rows = sorted((dict(row) for row in rows), key=_combined_candidate_sort_key)
    priority_rows: List[Dict[str, Any]] = []
    for rank, row in enumerate(ordered_rows, start=1):
        review_row = dict(row)
        review_row["rank"] = rank
        review_row["priority_band"] = _combined_candidate_priority_band(
            _candidate_row_number(row, "score_mean")
        )
        priority_rows.append(review_row)
    return priority_rows


def _build_combined_summary_rows(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[str, Optional[int], Optional[float]], Dict[str, Any]] = {}
    for row in rows:
        source_label = str(row.get("source_label") or "")
        scale_level = _candidate_row_int(row, "scale_level")
        scale_factor = _candidate_row_number(row, "scale_factor")
        key = (source_label, scale_level, scale_factor)
        bucket = grouped.setdefault(
            key,
            {
                "source_label": source_label,
                "scale_level": scale_level,
                "scale_factor": scale_factor,
                "candidate_count": 0,
                "high_priority_count": 0,
                "medium_priority_count": 0,
                "low_priority_count": 0,
                "score_sum": 0.0,
                "score_count": 0,
                "score_mean_max": None,
                "area_m2_sum": 0.0,
                "area_m2_max": None,
            },
        )

        bucket["candidate_count"] += 1

        score = _candidate_row_number(row, "score_mean")
        if score is not None:
            bucket["score_sum"] += score
            bucket["score_count"] += 1
            prev_max = bucket["score_mean_max"]
            if prev_max is None or score > float(prev_max):
                bucket["score_mean_max"] = score

        area_m2 = _candidate_row_number(row, "area_m2")
        if area_m2 is not None:
            bucket["area_m2_sum"] += area_m2
            prev_area_max = bucket["area_m2_max"]
            if prev_area_max is None or area_m2 > float(prev_area_max):
                bucket["area_m2_max"] = area_m2

        priority_band = _combined_candidate_priority_band(score)
        if priority_band == "Yuksek":
            bucket["high_priority_count"] += 1
        elif priority_band == "Orta":
            bucket["medium_priority_count"] += 1
        elif priority_band == "Dusuk":
            bucket["low_priority_count"] += 1

    summary_rows: List[Dict[str, Any]] = []
    for bucket in grouped.values():
        score_count = int(bucket.pop("score_count"))
        score_sum = float(bucket.pop("score_sum"))
        bucket["score_mean_avg"] = (score_sum / score_count) if score_count > 0 else None
        summary_rows.append(bucket)

    summary_rows.sort(
        key=lambda row: (
            -(_candidate_row_number(row, "score_mean_max") or -1.0),
            -float(row.get("candidate_count") or 0),
            str(row.get("source_label") or ""),
            _candidate_row_int(row, "scale_level") if _candidate_row_int(row, "scale_level") is not None else 999_999,
        )
    )
    return summary_rows


def _cluster_combined_candidate_members(
    rows: Sequence[Dict[str, Any]],
    distance_threshold_m: float,
) -> List[List[Dict[str, Any]]]:
    clusters: List[List[Dict[str, Any]]] = []
    grouped_by_crs: Dict[str, List[Tuple[Dict[str, Any], float, float]]] = {}

    for row in rows:
        x = _candidate_row_number(row, "center_x_native")
        y = _candidate_row_number(row, "center_y_native")
        native_crs = str(row.get("native_crs") or "").strip()
        if x is None or y is None or not native_crs or native_crs.lower() == "unknown":
            clusters.append([row])
            continue
        grouped_by_crs.setdefault(native_crs, []).append((row, x, y))

    max_distance_sq = float(distance_threshold_m) ** 2
    for items in grouped_by_crs.values():
        item_count = len(items)
        parents = list(range(item_count))

        def _find(index: int) -> int:
            while parents[index] != index:
                parents[index] = parents[parents[index]]
                index = parents[index]
            return index

        def _union(left: int, right: int) -> None:
            left_root = _find(left)
            right_root = _find(right)
            if left_root != right_root:
                parents[right_root] = left_root

        for left_idx in range(item_count):
            _, left_x, left_y = items[left_idx]
            for right_idx in range(left_idx + 1, item_count):
                _, right_x, right_y = items[right_idx]
                dx = left_x - right_x
                dy = left_y - right_y
                if (dx * dx) + (dy * dy) <= max_distance_sq:
                    _union(left_idx, right_idx)

        grouped_members: Dict[int, List[Dict[str, Any]]] = {}
        for item_idx, (row, _, _) in enumerate(items):
            grouped_members.setdefault(_find(item_idx), []).append(row)
        clusters.extend(grouped_members.values())

    return clusters


def _format_scale_factor_token(scale_factor: Optional[float]) -> str:
    if scale_factor is None:
        return "-"
    return f"{scale_factor:.2f}x"


def _build_combined_cluster_rows(
    rows: Sequence[Dict[str, Any]],
    distance_threshold_m: float = _COMBINED_CANDIDATE_CLUSTER_DISTANCE_M,
) -> List[Dict[str, Any]]:
    cluster_members = _cluster_combined_candidate_members(rows, distance_threshold_m)
    cluster_rows: List[Dict[str, Any]] = []

    for members in cluster_members:
        ordered_members = sorted(members, key=_combined_candidate_sort_key)
        best_row = ordered_members[0]

        score_values = [
            score for score in (_candidate_row_number(member, "score_mean") for member in members) if score is not None
        ]
        area_values = [
            area for area in (_candidate_row_number(member, "area_m2") for member in members) if area is not None
        ]
        center_points = [
            (
                _candidate_row_number(member, "center_x_native"),
                _candidate_row_number(member, "center_y_native"),
            )
            for member in members
        ]
        center_points = [
            (center_x, center_y)
            for center_x, center_y in center_points
            if center_x is not None and center_y is not None
        ]
        gps_points = [
            (_candidate_row_number(member, "gps_lon"), _candidate_row_number(member, "gps_lat"))
            for member in members
        ]
        gps_points = [(lon, lat) for lon, lat in gps_points if lon is not None and lat is not None]

        source_labels = sorted(
            {str(member.get("source_label") or "").strip() for member in members if str(member.get("source_label") or "").strip()}
        )
        scale_levels = sorted(
            {
                scale_level
                for scale_level in (_candidate_row_int(member, "scale_level") for member in members)
                if scale_level is not None
            }
        )
        scale_factors = sorted(
            {
                scale_factor
                for scale_factor in (_candidate_row_number(member, "scale_factor") for member in members)
                if scale_factor is not None
            }
        )

        cluster_center_x = (
            float(np.mean([point[0] for point in center_points]))
            if center_points
            else _candidate_row_number(best_row, "center_x_native")
        )
        cluster_center_y = (
            float(np.mean([point[1] for point in center_points]))
            if center_points
            else _candidate_row_number(best_row, "center_y_native")
        )
        gps_lon = (
            float(np.mean([point[0] for point in gps_points]))
            if gps_points
            else _candidate_row_number(best_row, "gps_lon")
        )
        gps_lat = (
            float(np.mean([point[1] for point in gps_points]))
            if gps_points
            else _candidate_row_number(best_row, "gps_lat")
        )

        google_maps_url = ""
        if gps_lat is not None and gps_lon is not None:
            google_maps_url = f"https://maps.google.com/?q={gps_lat:.8f},{gps_lon:.8f}"
        elif best_row.get("google_maps_url"):
            google_maps_url = str(best_row.get("google_maps_url"))

        best_score = _candidate_row_number(best_row, "score_mean")
        cluster_rows.append(
            {
                "cluster_id": 0,
                "priority_band": _combined_candidate_priority_band(best_score),
                "member_count": len(members),
                "sources_seen": ", ".join(source_labels) if source_labels else str(best_row.get("source_label") or ""),
                "scale_levels_seen": ", ".join(str(level) for level in scale_levels) if scale_levels else "-",
                "scale_factors_seen": ", ".join(_format_scale_factor_token(scale) for scale in scale_factors)
                if scale_factors
                else "-",
                "best_source_label": str(best_row.get("source_label") or ""),
                "best_scale_level": _candidate_row_int(best_row, "scale_level"),
                "best_scale_factor": _candidate_row_number(best_row, "scale_factor"),
                "best_candidate_id": _candidate_row_int(best_row, "candidate_id"),
                "best_score_mean": best_score,
                "mean_score_mean": (float(np.mean(score_values)) if score_values else None),
                "total_area_m2": float(sum(area_values)) if area_values else None,
                "max_area_m2": max(area_values) if area_values else None,
                "center_x_native": cluster_center_x,
                "center_y_native": cluster_center_y,
                "native_crs": str(best_row.get("native_crs") or "unknown"),
                "gps_lon": gps_lon,
                "gps_lat": gps_lat,
                "google_maps_url": google_maps_url,
            }
        )

    cluster_rows.sort(
        key=lambda row: (
            -(_candidate_row_number(row, "best_score_mean") or -1.0),
            -float(row.get("member_count") or 0),
            -(_candidate_row_number(row, "total_area_m2") or -1.0),
            str(row.get("best_source_label") or ""),
        )
    )
    for cluster_id, row in enumerate(cluster_rows, start=1):
        row["cluster_id"] = cluster_id
    return cluster_rows


def _build_combined_summary_metrics(
    rows: Sequence[Dict[str, Any]],
    cluster_rows: Sequence[Dict[str, Any]],
    distance_threshold_m: float,
) -> List[Dict[str, Any]]:
    scores = [score for score in (_candidate_row_number(row, "score_mean") for row in rows) if score is not None]
    areas = [area for area in (_candidate_row_number(row, "area_m2") for row in rows) if area is not None]
    source_labels = {
        str(row.get("source_label") or "").strip()
        for row in rows
        if str(row.get("source_label") or "").strip()
    }
    scale_factors = {
        scale_factor
        for scale_factor in (_candidate_row_number(row, "scale_factor") for row in rows)
        if scale_factor is not None
    }

    priority_counts = {"Yuksek": 0, "Orta": 0, "Dusuk": 0}
    fallback_count = 0
    for row in rows:
        band = _combined_candidate_priority_band(_candidate_row_number(row, "score_mean"))
        if band in priority_counts:
            priority_counts[band] += 1
        if str(row.get("candidate_type") or "") == _CANDIDATE_TYPE_FALLBACK:
            fallback_count += 1

    return [
        {"metric": "toplam_aday_sayisi", "value": len(rows)},
        {"metric": "fallback_aday_sayisi", "value": fallback_count},
        {"metric": "toplam_kume_sayisi", "value": len(cluster_rows)},
        {"metric": "benzersiz_kaynak_sayisi", "value": len(source_labels)},
        {"metric": "benzersiz_scale_sayisi", "value": len(scale_factors)},
        {"metric": "yuksek_oncelik_aday_sayisi", "value": priority_counts["Yuksek"]},
        {"metric": "orta_oncelik_aday_sayisi", "value": priority_counts["Orta"]},
        {"metric": "dusuk_oncelik_aday_sayisi", "value": priority_counts["Dusuk"]},
        {"metric": "score_mean_avg", "value": float(np.mean(scores)) if scores else None},
        {"metric": "score_mean_max", "value": max(scores) if scores else None},
        {"metric": "area_m2_sum", "value": float(sum(areas)) if areas else None},
        {"metric": "cluster_distance_m", "value": float(distance_threshold_m)},
    ]


def _excel_number_format_for_field(field_name: str) -> Optional[str]:
    if field_name in {
        "rank",
        "scale_level",
        "candidate_id",
        "cluster_id",
        "member_count",
        "best_scale_level",
        "best_candidate_id",
        "input_order",
        "candidate_count",
        "high_priority_count",
        "medium_priority_count",
        "low_priority_count",
        "pixel_count",
    }:
        return "0"
    if field_name in {
        "score_mean",
        "score_mean_avg",
        "score_mean_max",
        "best_score_mean",
        "mean_score_mean",
        "score_max",
    }:
        return "0.0%"
    if field_name in {"scale_factor", "best_scale_factor"}:
        return "0.00x"
    if field_name in {
        "area_m2",
        "area_m2_sum",
        "area_m2_max",
        "total_area_m2",
        "max_area_m2",
        "mask_area_m2",
        "bbox_area_m2",
    }:
        return "#,##0.00"
    if field_name in {
        "center_x_native",
        "center_y_native",
        "bbox_xmin",
        "bbox_ymin",
        "bbox_xmax",
        "bbox_ymax",
        "gps_lon",
        "gps_lat",
    }:
        return "#,##0.000000"
    return None


def _style_excel_header_row(ws, *, row_idx: int, field_order: Sequence[str]) -> None:
    for col_idx, field_name in enumerate(field_order, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=field_name)
        if Font is not None:
            cell.font = Font(bold=True, color="FFFFFF")
        if PatternFill is not None:
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        if Alignment is not None:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_idx].height = 20


def _autosize_excel_columns(
    ws,
    *,
    field_order: Sequence[str],
    start_row: int,
    end_row: int,
) -> None:
    for col_idx, field_name in enumerate(field_order, start=1):
        max_len = len(str(field_name))
        for row_idx in range(start_row + 1, end_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is None:
                continue
            max_len = max(max_len, len(str(cell_value)))
        max_len = max(max_len + 2, 10)
        if field_name in {"google_maps_url"}:
            max_len = max(max_len, 16)
        if field_name in {"sources_seen", "scale_factors_seen", "scale_levels_seen"}:
            max_len = min(max_len, 36)
        else:
            max_len = min(max_len, 24)
        ws.column_dimensions[_xlsx_column_name(col_idx)].width = max_len


def _add_excel_color_scale(
    ws,
    *,
    field_order: Sequence[str],
    start_row: int,
    end_row: int,
) -> None:
    if ColorScaleRule is None or end_row <= start_row:
        return
    score_fields = {
        "score_mean",
        "score_mean_avg",
        "score_mean_max",
        "best_score_mean",
        "mean_score_mean",
        "score_max",
    }
    for col_idx, field_name in enumerate(field_order, start=1):
        if field_name not in score_fields:
            continue
        col_letter = _xlsx_column_name(col_idx)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row + 1}:{col_letter}{end_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="F8696B",
                mid_type="num",
                mid_value=_COMBINED_CANDIDATE_PRIORITY_MEDIUM,
                mid_color="FFEB84",
                end_type="num",
                end_value=1,
                end_color="63BE7B",
            ),
        )


def _write_excel_table(
    ws,
    *,
    rows: Sequence[Dict[str, Any]],
    field_order: Sequence[str],
    start_row: int,
    table_name: Optional[str] = None,
    hyperlink_display_text: Optional[str] = None,
) -> int:
    _style_excel_header_row(ws, row_idx=start_row, field_order=field_order)

    for row_offset, row in enumerate(rows, start=1):
        excel_row = start_row + row_offset
        for col_idx, field_name in enumerate(field_order, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            value = _normalize_excel_cell_value(row.get(field_name))
            if field_name == "google_maps_url":
                maps_url = str(value or "").strip()
                if maps_url:
                    cell.value = hyperlink_display_text or maps_url
                    cell.hyperlink = maps_url
                    cell.style = "Hyperlink"
                else:
                    cell.value = None
            else:
                cell.value = value

            number_format = _excel_number_format_for_field(field_name)
            if number_format:
                cell.number_format = number_format
            if Alignment is not None:
                wrap_text = field_name in {"sources_seen", "scale_levels_seen", "scale_factors_seen"}
                cell.alignment = Alignment(vertical="top", wrap_text=wrap_text)

    end_row = start_row + max(len(rows), 0)
    last_col = _xlsx_column_name(len(field_order))
    data_ref = f"A{start_row}:{last_col}{max(start_row, end_row)}"

    if table_name and Table is not None and TableStyleInfo is not None and len(rows) > 0:
        excel_table = Table(displayName=table_name, ref=data_ref)
        excel_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(excel_table)
    else:
        ws.auto_filter.ref = data_ref

    _add_excel_color_scale(ws, field_order=field_order, start_row=start_row, end_row=end_row)
    _autosize_excel_columns(ws, field_order=field_order, start_row=start_row, end_row=end_row)
    return end_row


def _apply_summary_metric_formats(
    ws,
    *,
    metrics: Sequence[Dict[str, Any]],
    start_row: int,
) -> None:
    for row_offset, metric_row in enumerate(metrics, start=1):
        metric_name = str(metric_row.get("metric") or "")
        value_cell = ws.cell(row=start_row + row_offset, column=2)
        if metric_name in {"score_mean_avg", "score_mean_max"}:
            value_cell.number_format = "0.0%"
        elif metric_name in {"area_m2_sum", "cluster_distance_m"}:
            value_cell.number_format = "#,##0.00"
        else:
            value_cell.number_format = "0"


def _write_combined_candidate_review_workbook(
    *,
    rows: Sequence[Dict[str, Any]],
    out_base: Path,
    candidate_box_rows: Optional[Sequence[Dict[str, Any]]] = None,
) -> Optional[Path]:
    xlsx_path = _append_output_suffix(out_base, ".xlsx")
    if Workbook is None:
        _warn_once(
            "combined_candidate_excel_fallback",
            "openpyxl kurulu degil; birlesik aday workbook'u tek sayfali fallback XLSX olarak yazilacak.",
        )
        return _write_builtin_xlsx(
            rows=rows,
            field_order=_COMBINED_CANDIDATE_TABLE_FIELD_ORDER,
            xlsx_path=xlsx_path,
        )

    priority_rows = _build_combined_priority_rows(rows)
    summary_rows = _build_combined_summary_rows(priority_rows)
    cluster_rows = _build_combined_cluster_rows(
        priority_rows,
        distance_threshold_m=_COMBINED_CANDIDATE_CLUSTER_DISTANCE_M,
    )
    summary_metrics = _build_combined_summary_metrics(
        priority_rows,
        cluster_rows,
        _COMBINED_CANDIDATE_CLUSTER_DISTANCE_M,
    )
    box_rows = list(candidate_box_rows) if candidate_box_rows is not None else None
    if box_rows is not None:
        summary_metrics.append({"metric": "aday_kutu_sayisi", "value": len(box_rows)})
    raw_rows: List[Dict[str, Any]] = []
    for input_order, row in enumerate(rows, start=1):
        raw_row = dict(row)
        raw_row["input_order"] = input_order
        raw_rows.append(raw_row)

    wb = Workbook()

    ws_priority = wb.active
    ws_priority.title = "01_Oncelikli_Adaylar"
    _write_excel_table(
        ws_priority,
        rows=priority_rows,
        field_order=_COMBINED_CANDIDATE_PRIORITY_FIELD_ORDER,
        start_row=1,
        table_name="tblOncelikliAdaylar",
        hyperlink_display_text="Haritada Ac",
    )
    ws_priority.freeze_panes = "A2"

    ws_summary = wb.create_sheet("02_Ozet")
    _write_excel_table(
        ws_summary,
        rows=summary_metrics,
        field_order=("metric", "value"),
        start_row=1,
        table_name="tblGenelOzet",
    )
    _apply_summary_metric_formats(ws_summary, metrics=summary_metrics, start_row=1)
    summary_start_row = len(summary_metrics) + 4
    ws_summary.cell(row=summary_start_row - 1, column=1, value="kaynak_ozeti")
    if Font is not None:
        ws_summary.cell(row=summary_start_row - 1, column=1).font = Font(bold=True)
    _write_excel_table(
        ws_summary,
        rows=summary_rows,
        field_order=_COMBINED_CANDIDATE_SUMMARY_FIELD_ORDER,
        start_row=summary_start_row,
        table_name="tblKaynakOzet",
    )
    ws_summary.freeze_panes = "A2"

    ws_cluster = wb.create_sheet("03_Kumelenmis_Adaylar")
    _write_excel_table(
        ws_cluster,
        rows=cluster_rows,
        field_order=_COMBINED_CANDIDATE_CLUSTER_FIELD_ORDER,
        start_row=1,
        table_name="tblKumelenmisAdaylar",
        hyperlink_display_text="Haritada Ac",
    )
    ws_cluster.freeze_panes = "A2"

    if box_rows is not None:
        ws_boxes = wb.create_sheet("04_Aday_Kutulari")
        _write_excel_table(
            ws_boxes,
            rows=box_rows,
            field_order=_CANDIDATE_BOX_FIELD_ORDER,
            start_row=1,
            table_name="tblAdayKutulari",
            hyperlink_display_text="Haritada Ac",
        )
        ws_boxes.freeze_panes = "A2"

    ws_raw = wb.create_sheet("99_Ham_Veriler")
    _write_excel_table(
        ws_raw,
        rows=raw_rows,
        field_order=_COMBINED_CANDIDATE_RAW_FIELD_ORDER,
        start_row=1,
        table_name="tblHamVeriler",
    )
    ws_raw.freeze_panes = "A2"

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return xlsx_path


def _build_candidate_location_rows(
    records: Sequence[Dict[str, Any]],
    crs: Optional[RasterioCRS],
    extra_fields: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    """Build candidate location rows with native and optional WGS84 coordinates."""
    if not records:
        return []

    native_crs = _format_crs_label(crs)
    centers_x: List[float] = []
    centers_y: List[float] = []
    for rec in records:
        geom = rec.get("geometry")
        if geom is None or getattr(geom, "is_empty", False):
            continue

        center = geom.centroid
        centers_x.append(float(center.x))
        centers_y.append(float(center.y))

    gps_lons, gps_lats = _transform_points_to_wgs84(centers_x, centers_y, crs)

    rows: List[Dict[str, Any]] = []
    center_idx = 0
    for rec in records:
        geom = rec.get("geometry")
        if geom is None or getattr(geom, "is_empty", False):
            continue

        center_x = centers_x[center_idx]
        center_y = centers_y[center_idx]
        gps_lon = gps_lons[center_idx]
        gps_lat = gps_lats[center_idx]
        center_idx += 1

        google_maps_url = ""
        if gps_lat is not None and gps_lon is not None:
            google_maps_url = f"https://maps.google.com/?q={gps_lat:.8f},{gps_lon:.8f}"

        row = {
            "candidate_id": int(rec.get("id", len(rows) + 1)),
            "area_m2": float(rec.get("area_m2", 0.0)),
            "score_mean": float(rec.get("score_mean", 0.0)),
            "center_x_native": center_x,
            "center_y_native": center_y,
            "native_crs": native_crs,
            "gps_lon": gps_lon,
            "gps_lat": gps_lat,
            "google_maps_url": google_maps_url,
        }
        if extra_fields:
            row.update(extra_fields)
        rows.append(row)
    return rows


def export_candidate_locations_table(
    records: Sequence[Dict[str, Any]],
    crs: Optional[RasterioCRS],
    out_base: Path,
) -> Optional[Path]:
    """Write candidate locations as XLSX."""
    rows = _build_candidate_location_rows(records=records, crs=crs)
    return _write_candidate_location_rows_xlsx(
        rows=rows,
        field_order=_CANDIDATE_TABLE_FIELD_ORDER,
        out_base=out_base,
    )


def write_combined_candidate_locations_table(
    *,
    rows: Sequence[Dict[str, Any]],
    out_base: Path,
    candidate_box_rows: Optional[Sequence[Dict[str, Any]]] = None,
) -> Optional[Path]:
    """Write the consolidated candidate workbook with review, summary and raw sheets."""
    return _write_combined_candidate_review_workbook(
        rows=rows,
        out_base=out_base,
        candidate_box_rows=candidate_box_rows,
    )


def _build_combined_candidate_extra_fields(
    *,
    source_label: str,
    scale_level: Optional[int] = None,
    scale_factor: Optional[float] = None,
    candidate_type: str = _CANDIDATE_TYPE_THRESHOLD,
    review_status: str = _REVIEW_STATUS_DEFAULT,
) -> Dict[str, Any]:
    return {
        "source_label": str(source_label),
        "scale_level": scale_level,
        "scale_factor": scale_factor,
        "candidate_type": str(candidate_type),
        "review_status": str(review_status),
    }


def write_empty_candidate_locations_table(
    *,
    out_base: Path,
) -> Optional[Path]:
    """Always write an empty candidate XLSX with headers only."""
    return _write_candidate_location_rows_xlsx(
        rows=[],
        field_order=_CANDIDATE_TABLE_FIELD_ORDER,
        out_base=out_base,
    )


def ensure_candidate_locations_table_exists(
    *,
    out_base: Path,
    log_prefix: str = "",
    reason: Optional[str] = None,
) -> Optional[Path]:
    """Ensure a candidate XLSX exists; create an empty one if needed."""
    prefix = f"{log_prefix} " if log_prefix else ""
    xlsx_path = _append_output_suffix(out_base, ".xlsx")
    if xlsx_path.exists():
        return xlsx_path
    if reason:
        LOGGER.warning("%s%s", prefix, reason)
    table_path = write_empty_candidate_locations_table(out_base=out_base)
    if table_path:
        LOGGER.info("%sBos Excel aday tablosu yazildi: %s", prefix, table_path)
    return table_path


def export_candidate_locations_from_prediction(
    *,
    mask: np.ndarray,
    prob_map: np.ndarray,
    transform: Affine,
    crs: Optional[RasterioCRS],
    out_base: Path,
    min_area: float,
    opening_size: int,
    label_connectivity: int,
) -> Optional[Path]:
    """Export candidate table directly from raster predictions without polygon libraries."""
    rows = build_candidate_location_rows_from_prediction(
        mask=mask,
        prob_map=prob_map,
        transform=transform,
        crs=crs,
        min_area=min_area,
        opening_size=opening_size,
        label_connectivity=label_connectivity,
    )
    if not rows:
        return write_empty_candidate_locations_table(out_base=out_base)
    return _write_candidate_location_rows_xlsx(
        rows=rows,
        field_order=_CANDIDATE_TABLE_FIELD_ORDER,
        out_base=out_base,
    )


def build_candidate_location_rows_from_prediction(
    *,
    mask: np.ndarray,
    prob_map: np.ndarray,
    transform: Affine,
    crs: Optional[RasterioCRS],
    min_area: float,
    opening_size: int,
    label_connectivity: int,
    extra_fields: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    """Build candidate rows directly from raster predictions without writing files."""
    if mask.size == 0 or prob_map.size == 0:
        LOGGER.warning("Bos mask/prob rasteri; aday satirlari bos donuyor.")
        return []

    k = max(1, int(opening_size))
    cleaned_mask = grey_opening(mask.astype(np.uint8), size=(k, k))
    if int(label_connectivity) == 4:
        structure = np.array([[0, 1, 0], [1, 1, 1], [0, 1, 0]], dtype=int)
    else:
        structure = np.ones((3, 3), dtype=int)

    labels, num_features = ndimage.label(cleaned_mask.astype(bool), structure=structure)
    if num_features == 0:
        LOGGER.warning("Aday tablo icin esik ustunde bilesen bulunamadi; aday satirlari bos donuyor.")
        return []

    label_ids = np.arange(1, num_features + 1)
    pixel_counts = np.asarray(
        ndimage.sum(cleaned_mask.astype(np.uint8), labels, index=label_ids),
        dtype=np.float64,
    )
    prob_sums = np.asarray(
        ndimage.sum(prob_map.astype(np.float32), labels, index=label_ids),
        dtype=np.float64,
    )
    centers = ndimage.center_of_mass(cleaned_mask.astype(np.float32), labels, index=label_ids)

    pixel_width = abs(transform[0])
    pixel_height = abs(transform[4])
    pixel_area = float(pixel_width * pixel_height)
    native_crs = _format_crs_label(crs)

    native_xs: List[float] = []
    native_ys: List[float] = []
    temp_rows: List[Tuple[int, float, float, float, float]] = []
    next_id = 1
    for idx, center in enumerate(centers):
        pixels = float(pixel_counts[idx])
        if pixels <= 0:
            continue
        area_m2 = pixels * pixel_area
        if area_m2 < float(min_area):
            continue
        row_c, col_c = center
        center_x, center_y = transform * (float(col_c) + 0.5, float(row_c) + 0.5)
        native_xs.append(float(center_x))
        native_ys.append(float(center_y))
        mean_score = float(prob_sums[idx]) / pixels
        temp_rows.append((next_id, area_m2, mean_score, float(center_x), float(center_y)))
        next_id += 1

    if not temp_rows:
        LOGGER.warning("Aday tablo icin minimum alan filtresini gecen bilesen bulunamadi; aday satirlari bos donuyor.")
        return []

    gps_lons, gps_lats = _transform_points_to_wgs84(native_xs, native_ys, crs)
    rows: List[Dict[str, Any]] = []
    for idx, (candidate_id, area_m2, mean_score, center_x, center_y) in enumerate(temp_rows):
        gps_lon = gps_lons[idx]
        gps_lat = gps_lats[idx]
        google_maps_url = ""
        if gps_lat is not None and gps_lon is not None:
            google_maps_url = f"https://maps.google.com/?q={gps_lat:.8f},{gps_lon:.8f}"
        row = {
            "candidate_id": candidate_id,
            "area_m2": area_m2,
            "score_mean": mean_score,
            "center_x_native": center_x,
            "center_y_native": center_y,
            "native_crs": native_crs,
            "gps_lon": gps_lon,
            "gps_lat": gps_lat,
            "google_maps_url": google_maps_url,
        }
        if extra_fields:
            row.update(extra_fields)
        rows.append(row)
    return rows


def _fallback_candidate_min_separation_pixels(transform: Affine, min_area: float) -> int:
    pixel_width = abs(float(transform[0]))
    pixel_height = abs(float(transform[4]))
    pixel_area = pixel_width * pixel_height
    if not np.isfinite(pixel_area) or pixel_area <= 0:
        return 8
    if float(min_area) <= 0:
        return 8
    diameter_px = math.sqrt(max(float(min_area), pixel_area) / pixel_area)
    return max(1, min(256, int(round(diameter_px * 0.5))))


def _select_top_score_pixels(
    prob_map: np.ndarray,
    *,
    top_k: int,
    min_score: Optional[float],
    min_separation_pixels: int,
) -> List[Tuple[int, int, float]]:
    if top_k <= 0 or prob_map.size == 0:
        return []

    scores = np.array(prob_map, dtype=np.float32, copy=True)
    valid = np.isfinite(scores)
    if min_score is not None:
        valid &= scores >= float(min_score)
    scores[~valid] = -np.inf

    selected: List[Tuple[int, int, float]] = []
    radius = max(0, int(min_separation_pixels))
    height, width = scores.shape
    for _ in range(int(top_k)):
        flat_idx = int(np.argmax(scores))
        score = float(scores.flat[flat_idx])
        if not np.isfinite(score):
            break
        row, col = (int(v) for v in np.unravel_index(flat_idx, scores.shape))
        selected.append((row, col, score))
        if radius <= 0:
            scores[row, col] = -np.inf
            continue
        r0 = max(0, row - radius)
        r1 = min(height, row + radius + 1)
        c0 = max(0, col - radius)
        c1 = min(width, col + radius + 1)
        scores[r0:r1, c0:c1] = -np.inf
    return selected


def build_fallback_candidate_location_rows_from_prediction(
    *,
    prob_map: np.ndarray,
    transform: Affine,
    crs: Optional[RasterioCRS],
    min_area: float,
    top_k: int,
    min_score: Optional[float],
    extra_fields: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    """Build low-threshold review rows from the highest probability pixels."""
    selected = _select_top_score_pixels(
        prob_map,
        top_k=top_k,
        min_score=min_score,
        min_separation_pixels=_fallback_candidate_min_separation_pixels(transform, min_area),
    )
    if not selected:
        return []

    pixel_area = float(abs(transform[0]) * abs(transform[4]))
    native_crs = _format_crs_label(crs)
    native_xs: List[float] = []
    native_ys: List[float] = []
    for row, col, _score in selected:
        center_x, center_y = transform * (float(col) + 0.5, float(row) + 0.5)
        native_xs.append(float(center_x))
        native_ys.append(float(center_y))

    gps_lons, gps_lats = _transform_points_to_wgs84(native_xs, native_ys, crs)
    rows: List[Dict[str, Any]] = []
    for idx, (_row, _col, score) in enumerate(selected):
        gps_lon = gps_lons[idx]
        gps_lat = gps_lats[idx]
        google_maps_url = ""
        if gps_lat is not None and gps_lon is not None:
            google_maps_url = f"https://maps.google.com/?q={gps_lat:.8f},{gps_lon:.8f}"
        row = {
            "candidate_id": idx + 1,
            "candidate_type": _CANDIDATE_TYPE_FALLBACK,
            "review_status": _REVIEW_STATUS_FALLBACK,
            "area_m2": pixel_area,
            "score_mean": float(score),
            "center_x_native": native_xs[idx],
            "center_y_native": native_ys[idx],
            "native_crs": native_crs,
            "gps_lon": gps_lon,
            "gps_lat": gps_lat,
            "google_maps_url": google_maps_url,
        }
        if extra_fields:
            row.update(extra_fields)
        row["candidate_type"] = _CANDIDATE_TYPE_FALLBACK
        row["review_status"] = _REVIEW_STATUS_FALLBACK
        rows.append(row)
    return rows


def _fallback_rows_for_empty_candidate_result(
    *,
    normal_rows: Sequence[Dict[str, Any]],
    enabled: bool,
    prob_map: np.ndarray,
    transform: Affine,
    crs: Optional[RasterioCRS],
    min_area: float,
    top_k: int,
    min_score: Optional[float],
    extra_fields: Optional[Dict[str, Any]],
    label: str,
) -> List[Dict[str, Any]]:
    if normal_rows or not enabled or top_k <= 0:
        return []
    fallback_rows = build_fallback_candidate_location_rows_from_prediction(
        prob_map=prob_map,
        transform=transform,
        crs=crs,
        min_area=min_area,
        top_k=top_k,
        min_score=min_score,
        extra_fields=extra_fields,
    )
    if fallback_rows:
        LOGGER.info(
            "    Esik ustu aday yok (%s); fallback top-score adaylari hazirlandi: %d",
            label,
            len(fallback_rows),
        )
    return fallback_rows


def _use_fallback_rows_if_no_candidates(
    *,
    rows: Sequence[Dict[str, Any]],
    fallback_rows: Sequence[Dict[str, Any]],
    top_k: int,
) -> List[Dict[str, Any]]:
    if rows:
        return list(rows)
    if top_k <= 0 or not fallback_rows:
        return list(rows)
    selected = sorted((dict(row) for row in fallback_rows), key=_combined_candidate_sort_key)[: int(top_k)]
    LOGGER.info(
        "Esik/min_area sonrasi hic aday bulunamadi; birlesik Excel'e en yuksek skorlu %d fallback aday eklendi.",
        len(selected),
    )
    return selected


def _candidate_label_structure(label_connectivity: int) -> np.ndarray:
    if int(label_connectivity) == 4:
        return np.array([[0, 1, 0], [1, 1, 1], [0, 1, 0]], dtype=int)
    return np.ones((3, 3), dtype=int)


def _component_bbox_polygon(
    *,
    transform: Affine,
    row_start: int,
    row_stop: int,
    col_start: int,
    col_stop: int,
):
    if Polygon is None:
        return None
    corners = [
        transform * (float(col_start), float(row_start)),
        transform * (float(col_stop), float(row_start)),
        transform * (float(col_stop), float(row_stop)),
        transform * (float(col_start), float(row_stop)),
    ]
    geom = Polygon(corners)
    if not geom.is_valid:
        geom = geom.buffer(0)
    if geom.is_empty:
        return None
    return geom


def _build_candidate_box_records_from_prediction(
    *,
    mask: np.ndarray,
    prob_map: np.ndarray,
    transform: Affine,
    crs: Optional[RasterioCRS],
    min_area: float,
    opening_size: int,
    label_connectivity: int,
    extra_fields: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    """Build object-like candidate boxes from connected prediction components.

    This is a post-processing view of the existing tile/polygon output. It does
    not require bbox labels, a detector model, or an additional segmentation
    training pass.
    """
    if mask.size == 0 or prob_map.size == 0:
        return []
    if Polygon is None or shapely_transform is None or CRS is None or Transformer is None:
        return []

    k = max(1, int(opening_size))
    cleaned_mask = grey_opening(mask.astype(np.uint8), size=(k, k))
    labels, num_features = ndimage.label(
        cleaned_mask.astype(bool),
        structure=_candidate_label_structure(label_connectivity),
    )
    if num_features == 0:
        return []

    label_ids = np.arange(1, num_features + 1)
    pixel_counts = np.asarray(
        ndimage.sum(cleaned_mask.astype(np.uint8), labels, index=label_ids),
        dtype=np.float64,
    )
    prob_sums = np.asarray(
        ndimage.sum(prob_map.astype(np.float32), labels, index=label_ids),
        dtype=np.float64,
    )
    prob_maxes = np.asarray(
        ndimage.maximum(prob_map.astype(np.float32), labels, index=label_ids),
        dtype=np.float64,
    )
    centers = ndimage.center_of_mass(cleaned_mask.astype(np.float32), labels, index=label_ids)
    slices = ndimage.find_objects(labels)

    pixel_width = abs(transform[0])
    pixel_height = abs(transform[4])
    pixel_area = float(pixel_width * pixel_height)
    native_crs = _format_crs_label(crs)
    _crs_obj, to_area, _to_native = _resolve_area_transformers(crs)

    temp_records: List[Dict[str, Any]] = []
    center_xs: List[float] = []
    center_ys: List[float] = []
    next_id = 1

    for idx in range(num_features):
        pixels = float(pixel_counts[idx])
        if pixels <= 0:
            continue
        mask_area_m2 = pixels * pixel_area
        if mask_area_m2 < float(min_area):
            continue
        component_slice = slices[idx] if idx < len(slices) else None
        if component_slice is None:
            continue
        row_slice, col_slice = component_slice
        geom = _component_bbox_polygon(
            transform=transform,
            row_start=int(row_slice.start),
            row_stop=int(row_slice.stop),
            col_start=int(col_slice.start),
            col_stop=int(col_slice.stop),
        )
        if geom is None:
            continue
        if to_area:
            bbox_area_m2 = float(shapely_transform(to_area.transform, geom).area)
        else:
            bbox_area_m2 = float(geom.area)

        row_c, col_c = centers[idx]
        center_x, center_y = transform * (float(col_c) + 0.5, float(row_c) + 0.5)
        center_x = float(center_x)
        center_y = float(center_y)
        center_xs.append(center_x)
        center_ys.append(center_y)

        minx, miny, maxx, maxy = geom.bounds
        mean_score = float(prob_sums[idx]) / pixels
        rec: Dict[str, Any] = {
            "id": next_id,
            "candidate_id": next_id,
            "candidate_type": "connected_component_box",
            "review_status": "Kontrol edilecek",
            "area_m2": mask_area_m2,
            "mask_area_m2": mask_area_m2,
            "bbox_area_m2": bbox_area_m2,
            "score_mean": mean_score,
            "score_max": float(prob_maxes[idx]),
            "pixel_count": int(round(pixels)),
            "center_x_native": center_x,
            "center_y_native": center_y,
            "bbox_xmin": float(minx),
            "bbox_ymin": float(miny),
            "bbox_xmax": float(maxx),
            "bbox_ymax": float(maxy),
            "native_crs": native_crs,
            "gps_lon": None,
            "gps_lat": None,
            "google_maps_url": "",
            "geometry": geom,
        }
        if extra_fields:
            rec.update(extra_fields)
        temp_records.append(rec)
        next_id += 1

    if not temp_records:
        return []

    gps_lons, gps_lats = _transform_points_to_wgs84(center_xs, center_ys, crs)
    for idx, rec in enumerate(temp_records):
        gps_lon = gps_lons[idx]
        gps_lat = gps_lats[idx]
        rec["gps_lon"] = gps_lon
        rec["gps_lat"] = gps_lat
        if gps_lat is not None and gps_lon is not None:
            rec["google_maps_url"] = f"https://maps.google.com/?q={gps_lat:.8f},{gps_lon:.8f}"
    return temp_records


def _candidate_box_records_to_rows(records: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {field_name: rec.get(field_name) for field_name in _CANDIDATE_BOX_FIELD_ORDER}
        for rec in records
    ]


def write_candidate_boxes_layer_from_prediction(
    *,
    mask: np.ndarray,
    prob_map: np.ndarray,
    transform: Affine,
    crs: Optional[RasterioCRS],
    gpkg_path: Path,
    layer_name: str,
    min_area: float,
    opening_size: int,
    label_connectivity: int,
    extra_fields: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    records = _build_candidate_box_records_from_prediction(
        mask=mask,
        prob_map=prob_map,
        transform=transform,
        crs=crs,
        min_area=min_area,
        opening_size=opening_size,
        label_connectivity=label_connectivity,
        extra_fields=extra_fields,
    )
    if not records:
        return []
    _write_records_to_gpkg(
        records=records,
        crs=crs,
        gpkg_path=gpkg_path,
        layer_name=layer_name,
        column_order=_CANDIDATE_BOX_GPKG_FIELD_ORDER,
    )
    return records


def load_prediction_arrays(
    *,
    prob_path: Path,
    mask_path: Path,
) -> Tuple[np.ndarray, np.ndarray]:
    """Load prediction rasters from disk for downstream vector/table export."""
    with rasterio.open(prob_path) as src_prob:
        prob_arr = src_prob.read(1).astype(np.float32, copy=False)
    with rasterio.open(mask_path) as src_mask:
        mask_arr = src_mask.read(1).astype(np.uint8, copy=False)
    return mask_arr, prob_arr


def export_candidate_locations_from_prediction_safe(
    *,
    mask: Optional[np.ndarray],
    prob_map: Optional[np.ndarray],
    transform: Affine,
    crs: Optional[RasterioCRS],
    out_base: Path,
    min_area: float,
    opening_size: int,
    label_connectivity: int,
    log_prefix: str = "",
) -> Optional[Path]:
    """Best-effort candidate XLSX export; fall back to an empty workbook on any failure."""
    prefix = f"{log_prefix} " if log_prefix else ""
    try:
        if mask is None or prob_map is None:
            LOGGER.warning("%sRaster verisi eksik; bos Excel yaziliyor.", prefix)
            return write_empty_candidate_locations_table(out_base=out_base)
        return export_candidate_locations_from_prediction(
            mask=mask,
            prob_map=prob_map,
            transform=transform,
            crs=crs,
            out_base=out_base,
            min_area=min_area,
            opening_size=opening_size,
            label_connectivity=label_connectivity,
        )
    except Exception as exc:
        LOGGER.warning("%sAday Excel aktarimi basarisiz (%s); bos Excel yaziliyor.", prefix, exc)
        return write_empty_candidate_locations_table(out_base=out_base)


def vectorize_predictions(
    mask: np.ndarray,
    prob_map: np.ndarray,
    transform: Affine,
    crs: Optional[RasterioCRS],
    out_path: Path,
    min_area: float,
    simplify_tol: Optional[float],
    opening_size: int,
    label_connectivity: int,
    export_candidate_excel: bool = True,
    gpkg_path: Optional[Path] = None,
    gpkg_layer_name: Optional[str] = None,
) -> Optional[Path]:
    """Convert binary mask into polygons and write to GeoPackage."""
    table_base = _candidate_table_base(out_path)
    if fiona is None and gpd is None:
        LOGGER.warning("Vektör çıktısı atlandı; vektörleştirme için geopandas veya fiona yükleyin.")
        return None
    if mapping is None or shape is None or shapely_transform is None or CRS is None or Transformer is None:
        LOGGER.warning("Vektör çıktısı atlandı; shapely/pyproj kurulu değil.")
        return None

    # Küçük gürültüleri temizlemek için binary opening (hızlı!)
    LOGGER.info("Küçük gürültüler temizleniyor...")
    k = max(1, int(opening_size))
    cleaned_mask = grey_opening(mask.astype(np.uint8), size=(k, k))
    
    LOGGER.info("Etiketleme yapılıyor...")
    # Connectivity: 4-neighbour (cross) or 8-neighbour (3x3 ones)
    if int(label_connectivity) == 4:
        structure = np.array([[0, 1, 0], [1, 1, 1], [0, 1, 0]], dtype=int)
    else:
        structure = np.ones((3, 3), dtype=int)
    labels, num_features = ndimage.label(cleaned_mask.astype(bool), structure=structure)
    if num_features == 0:
        LOGGER.info("Eşik üstünde özellik bulunamadı; vektörleştirme atlanıyor.")
        return None

    LOGGER.info("Tespit edilen özellik sayısı: %d", num_features)
    
    # Performans optimizasyonu: Çok küçük poligonları erken filtrelemek için piksel sayılarını hesapla
    label_ids = np.arange(1, num_features + 1)
    pixel_counts = ndimage.sum(mask.astype(np.uint8), labels, index=label_ids)
    prob_sums = ndimage.sum(prob_map.astype(np.float32), labels, index=label_ids)
    
    # Piksel tabanlı ön filtreleme: min_area'yı piksel sayısına çevir
    # Transform'dan piksel boyutunu hesapla (m²)
    pixel_width = abs(transform[0])  # x yönünde piksel boyutu
    pixel_height = abs(transform[4])  # y yönünde piksel boyutu
    pixel_area = pixel_width * pixel_height  # piksel alanı (m²)
    
    from math import ceil
    min_pixels = max(1, int(ceil(min_area / pixel_area)))
    valid_labels = pixel_counts >= min_pixels
    
    if not valid_labels.any():
        LOGGER.info("Tüm poligonlar minimum piksel sayısının altında; vektörleştirme atlanıyor.")
        return None
    
    filtered_count = valid_labels.sum()
    LOGGER.info("Piksel filtresinden sonra kalan poligon sayısı: %d (elenen: %d)", 
                filtered_count, num_features - filtered_count)
    
    # YENİ ETİKETLEME: Geçerli label'ları ardışık numaralarla yeniden etiketle
    # Bu, shapes() fonksiyonunun çok daha hızlı çalışmasını sağlar
    LOGGER.info("Label'lar yeniden numaralandırılıyor...")
    filtered_labels = np.zeros_like(labels)
    new_id = 1
    filtered_pixel_counts = []
    filtered_prob_sums = []
    
    for old_id in range(1, num_features + 1):
        if valid_labels[old_id - 1]:
            filtered_labels[labels == old_id] = new_id
            filtered_pixel_counts.append(pixel_counts[old_id - 1])
            filtered_prob_sums.append(prob_sums[old_id - 1])
            new_id += 1
    
    # Dizileri numpy array'e çevir
    pixel_counts = np.array(filtered_pixel_counts)
    prob_sums = np.array(filtered_prob_sums)

    _crs_obj, to_area, to_native = _resolve_area_transformers(crs)

    records = []
    LOGGER.info("Poligonlar oluşturuluyor...")
    # Filtrelenmiş label'lardan poligon oluştur (artık ardışık numaralı)
    shape_generator = shapes(filtered_labels.astype(np.int32), mask=None, transform=transform)
    
    # Progress bar ile poligon işleme
    for geom, value in progress_bar(
        shape_generator,
        total=filtered_count + 1,  # +1 for background (0)
        desc="Vektörleştirme",
        unit="poligon"
    ):
        new_label_id = int(value)
        if new_label_id == 0:
            continue
        geom_shape = shape(geom)
        if to_area:
            geom_area_space = shapely_transform(to_area.transform, geom_shape)
            area_m2 = geom_area_space.area
        else:
            geom_area_space = geom_shape
            area_m2 = geom_shape.area
        if area_m2 < float(min_area):
            continue
        pixels = float(pixel_counts[new_label_id - 1])
        if pixels <= 0:
            continue
        mean_score = float(prob_sums[new_label_id - 1]) / pixels
        if simplify_tol and simplify_tol > 0:
            if to_area and to_native:
                simplified_area_geom = geom_area_space.simplify(
                    simplify_tol, preserve_topology=True
                )
                geom_shape = shapely_transform(to_native.transform, simplified_area_geom)
            else:
                geom_shape = geom_shape.simplify(simplify_tol, preserve_topology=True)
        records.append(
            {
                "id": int(new_label_id),
                "area_m2": float(area_m2),
                "score_mean": float(mean_score),
                "geometry": geom_shape,
            }
        )

    if not records:
        LOGGER.info("No polygons passed the min-area filter; skipping vector output.")
        if export_candidate_excel:
            empty_base = _candidate_table_base(out_path)
            empty_table = export_candidate_locations_table(records=[], crs=crs, out_base=empty_base)
            if empty_table:
                LOGGER.info("Bo? aday konum tablosu yaz?ld?: %s", empty_table)
        return None

    base_out_path = _output_base_path(out_path)
    resolved_gpkg_path = gpkg_path or _append_output_suffix(base_out_path, ".gpkg")
    resolved_layer_name = gpkg_layer_name or base_out_path.name
    LOGGER.info("GeoPackage dosyası yazılıyor (%d poligon)...", len(records))
    _write_records_to_gpkg(
        records=records,
        crs=crs,
        gpkg_path=resolved_gpkg_path,
        layer_name=resolved_layer_name,
        column_order=("id", "area_m2", "score_mean"),
    )
    if export_candidate_excel:
        table_base = _candidate_table_base(base_out_path)
        table_path = export_candidate_locations_table(records=records, crs=crs, out_base=table_base)
        if table_path:
            LOGGER.info("Aday konum tablosu yazıldı: %s", table_path)

    return resolved_gpkg_path



def parse_band_indexes(band_string: str) -> Tuple[int, ...]:
    """Parse CSV band specification to a 3-band RGB or 5-band RGB+DSM+DTM tuple."""
    if str(band_string or "").strip().lower() == "auto":
        raise argparse.ArgumentTypeError("bands='auto' once input raster'a gore cozulmelidir.")
    parts = [int(val.strip()) for val in band_string.split(",")]
    if len(parts) not in (3, 5):
        raise argparse.ArgumentTypeError("--bands must specify either 3 or 5 entries.")
    if any(idx <= 0 for idx in parts[:3]):
        raise argparse.ArgumentTypeError("RGB band indices must be positive (1-based).")
    if len(parts) == 5 and parts[4] <= 0:
        raise argparse.ArgumentTypeError("DTM band index must be provided (>=1).")
    return tuple(parts)


def _bands_auto_requested(band_string: object) -> bool:
    return str(band_string or "").strip().lower() == "auto"


def infer_auto_band_indexes(input_path: Path) -> Tuple[int, ...]:
    """Infer conventional band indexes from raster band count for bands='auto'."""
    try:
        with rasterio.open(input_path) as src:
            band_count = int(src.count)
    except Exception as exc:
        raise argparse.ArgumentTypeError(f"bands='auto' icin raster bant sayisi okunamadi: {exc}") from exc

    if band_count == 1:
        raise argparse.ArgumentTypeError(
            "bands='auto' tek bant raster icin desteklenmez; classic/DL/YOLO akislarinda RGB veya RGB+DSM+DTM gerekir. "
            "Tek bant raster icin bu pipeline yerine uygun ayri araci kullanin."
        )
    if band_count >= 5:
        return (1, 2, 3, 4, 5)
    if band_count >= 3:
        return (1, 2, 3)
    raise argparse.ArgumentTypeError(f"bands='auto' icin desteklenmeyen bant sayisi: {band_count}")


def resolve_band_indexes_for_config(input_path: Path, config: PipelineDefaults) -> Tuple[int, ...]:
    """Resolve configured band indexes, including bands='auto'."""
    if not _bands_auto_requested(config.bands):
        return parse_band_indexes(config.bands)
    return infer_auto_band_indexes(input_path)


def create_raster_metadata(
    height: int,
    width: int,
    transform: Affine,
    crs: Optional[RasterioCRS],
    dtype: str,
    nodata: Optional[float] = None,
) -> Dict[str, Any]:
    """Create standardized raster metadata dictionary."""
    # Choose predictor: 3 for floating point, 2 otherwise
    _dtype = str(dtype).lower()
    predictor_val = 3 if _dtype.startswith("float") else 2
    meta = {
        "driver": "GTiff",
        "height": height,
        "width": width,
        "count": 1,
        "transform": transform,
        "crs": crs,
        "dtype": dtype,
        "compress": "deflate",
        "tiled": True,
        "blockxsize": 256,
        "blockysize": 256,
        "predictor": predictor_val,
    }
    if nodata is not None:
        meta["nodata"] = nodata
    return meta


def write_prob_and_mask_rasters(
    prob_map: np.ndarray,
    mask: np.ndarray,
    transform: Affine,
    crs: Optional[RasterioCRS],
    prob_path: Path,
    mask_path: Path,
) -> None:
    """Write probability and mask rasters with standard metadata."""
    height, width = prob_map.shape
    
    # Write probability raster
    prob_meta = create_raster_metadata(
        height=height,
        width=width,
        transform=transform,
        crs=crs,
        dtype="float32",
        nodata=np.nan,
    )
    with rasterio.open(prob_path, "w", **prob_meta) as dst:
        dst.write(prob_map[np.newaxis, :, :])
    
    # Write mask raster
    mask_meta = create_raster_metadata(
        height=height,
        width=width,
        transform=transform,
        crs=crs,
        dtype="uint8",
        nodata=0,
    )
    with rasterio.open(mask_path, "w", **mask_meta) as dst:
        dst.write(mask[np.newaxis, :, :])


def compute_fused_probability(
    dl_prob: np.ndarray,
    classic_prob: np.ndarray,
    alpha: float,
    threshold: float,
) -> Tuple[np.ndarray, np.ndarray]:
    """
    Compute fused probability map and mask from DL and classic outputs.
    
    Args:
        dl_prob: Deep learning probability map
        classic_prob: Classical method probability map
        alpha: Mixing weight (1.0 = pure DL, 0.0 = pure classical)
        threshold: Probability threshold for binary mask
    
    Returns:
        Tuple of (fused_prob, fused_mask)
    """
    dl_filled = np.nan_to_num(dl_prob, nan=0.0)
    classic_filled = np.nan_to_num(classic_prob, nan=0.0)
    fused_prob = alpha * dl_filled + (1.0 - alpha) * classic_filled
    fused_prob = fused_prob.astype(np.float32)
    fused_valid = np.isfinite(dl_prob) | np.isfinite(classic_prob)
    fused_prob = np.clip(fused_prob, 0.0, 1.0, out=fused_prob)
    fused_prob[~fused_valid] = np.nan
    fused_mask = np.zeros_like(fused_prob, dtype=np.uint8)
    fused_mask[fused_valid & (fused_prob >= threshold)] = 1
    return fused_prob, fused_mask


def write_fused_probability_rasters_from_paths(
    dl_prob_path: Path,
    classic_prob_path: Path,
    *,
    alpha: float,
    threshold: float,
    out_prob_path: Path,
    out_mask_path: Path,
) -> Tuple[Affine, Optional[RasterioCRS]]:
    """Fuse two probability rasters in a streaming fashion (memory-safe for huge rasters)."""
    with rasterio.open(dl_prob_path) as dl_ds, rasterio.open(classic_prob_path) as cl_ds:
        if dl_ds.width != cl_ds.width or dl_ds.height != cl_ds.height:
            raise ValueError(
                f"Fusion raster sizes differ: dl={dl_ds.width}x{dl_ds.height}, classic={cl_ds.width}x{cl_ds.height}"
            )
        height = int(dl_ds.height)
        width = int(dl_ds.width)
        transform = dl_ds.transform
        crs = dl_ds.crs

        out_prob_path.parent.mkdir(parents=True, exist_ok=True)
        prob_meta = create_raster_metadata(
            height=height,
            width=width,
            transform=transform,
            crs=crs,
            dtype="float32",
            nodata=np.nan,
        )
        mask_meta = create_raster_metadata(
            height=height,
            width=width,
            transform=transform,
            crs=crs,
            dtype="uint8",
            nodata=0,
        )

        # Estimate block count for progress reporting.
        block_h = int(dl_ds.block_shapes[0][0]) if getattr(dl_ds, "block_shapes", None) else 256
        block_w = int(dl_ds.block_shapes[0][1]) if getattr(dl_ds, "block_shapes", None) else 256
        total_blocks = math.ceil(height / block_h) * math.ceil(width / block_w)

        with rasterio.open(out_prob_path, "w", **prob_meta) as out_prob, rasterio.open(
            out_mask_path, "w", **mask_meta
        ) as out_mask:
            for _, window in progress_bar(
                dl_ds.block_windows(1),
                total=total_blocks,
                desc="Fusion",
                unit="block",
            ):
                dl_block = dl_ds.read(1, window=window, masked=False).astype(np.float32, copy=False)
                cl_block = cl_ds.read(1, window=window, masked=False).astype(np.float32, copy=False)

                valid = np.isfinite(dl_block) | np.isfinite(cl_block)
                dl_block = np.nan_to_num(dl_block, nan=0.0, copy=False)
                cl_block = np.nan_to_num(cl_block, nan=0.0, copy=False)

                fused = alpha * dl_block + (1.0 - alpha) * cl_block
                fused = fused.astype(np.float32, copy=False)
                np.clip(fused, 0.0, 1.0, out=fused)
                fused[~valid] = np.nan

                mask = np.zeros(fused.shape, dtype=np.uint8)
                mask[valid & (fused >= threshold)] = 1

                out_prob.write(fused[np.newaxis, :, :], window=window)
                out_mask.write(mask[np.newaxis, :, :], window=window)

    return transform, crs


def build_session_folder_name(input_path: Path, config: PipelineDefaults) -> str:
    """Build a compact, timestamped output subfolder name."""

    def _short_token(text: Optional[str], max_len: int) -> str:
        token = _safe_token(str(text or "").strip())
        if not token:
            return "na"
        return token[:max_len].rstrip("_-") or "na"

    stem_short = _short_token(input_path.stem, 20)

    method_flags: List[str] = []
    if config.enable_deep_learning:
        method_flags.append("dl")
    if config.enable_classic:
        method_flags.append("cl")
    if config.enable_yolo:
        method_flags.append("yo")
    if config.enable_fusion:
        method_flags.append("fu")
    method_summary = "+".join(method_flags) if method_flags else "run"

    folder_tile = config.tile
    folder_overlap = config.overlap
    tile_part = f"t{int(folder_tile)}" if folder_tile else ""
    overlap_part = f"o{int(folder_overlap)}" if folder_overlap else ""
    tile_overlap = f"{tile_part}{overlap_part}" if (tile_part or overlap_part) else ""

    model_part = ""
    if config.enable_deep_learning:
        if config.weights:
            model_part = f"m-{_short_token(Path(config.weights).stem, 18)}"
        elif config.zero_shot_imagenet:
            model_part = "m-zs"
        elif config.encoder:
            model_part = f"m-{_short_token(config.encoder, 12)}"

    parts: List[str] = [SESSION_RUN_ID, stem_short, method_summary]
    if tile_overlap:
        parts.append(tile_overlap)
    if model_part:
        parts.append(model_part)
    return "_".join(parts)


def resolve_out_prefix(input_path: Path, prefix: Optional[str], config: PipelineDefaults) -> Path:
    """Resolve output prefix path."""
    fallback_name = _output_base_path(input_path).name
    if prefix:
        out_path = Path(prefix)
        if out_path.is_dir():
            out_path = out_path / input_path.stem
        out_name = _output_base_path(out_path).name or fallback_name
    else:
        out_name = fallback_name

    # Route all outputs under repo-local 'workspace/ciktilar/<session>/' regardless of input location.
    session_folder = build_session_folder_name(input_path, config)
    return WORKSPACE_OUTPUTS_DIR / session_folder / out_name


def _normalize_param_value(value: Any) -> Any:
    """Convert values into JSON/text-friendly primitives."""
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, torch.device):
        return str(value)
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, (list, tuple, set)):
        return [_normalize_param_value(v) for v in value]
    if isinstance(value, dict):
        return {str(k): _normalize_param_value(v) for k, v in value.items()}
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def write_run_params_txt(
    *,
    out_prefix: Path,
    config: PipelineDefaults,
    argv_list: Sequence[str],
    cli_overrides: Iterable[str],
    input_path: Path,
    bands: Sequence[int],
    device: torch.device,
) -> Optional[Path]:
    """Write effective run parameters into the session output folder as a txt file."""
    try:
        base_prefix = _output_base_path(out_prefix)
        session_dir = base_prefix.parent
        session_dir.mkdir(parents=True, exist_ok=True)
        txt_path = session_dir / "run_params.txt"

        lines: List[str] = []
        lines.append("# Archaeo Detect Run Parameters")
        lines.append(f"created_at: {datetime.now().isoformat(timespec='seconds')}")
        lines.append(f"session_run_id: {SESSION_RUN_ID}")
        lines.append(f"input_path: {input_path}")
        lines.append(f"output_session_dir: {session_dir}")
        lines.append(f"output_base_prefix: {base_prefix}")
        lines.append(f"device: {device}")
        lines.append(f"bands_parsed: {list(int(v) for v in bands)}")
        lines.append(f"dl_task_resolved: {str(config.dl_task).strip().lower()}")
        lines.append(f"cli_overrides: {json.dumps(sorted(set(cli_overrides)), ensure_ascii=False)}")
        lines.append(f"argv: {json.dumps(list(argv_list), ensure_ascii=False)}")
        lines.append("")
        lines.append("[config_effective]")

        for f in fields(PipelineDefaults):
            name = f.name
            value = getattr(config, name)
            normalized = _normalize_param_value(value)
            lines.append(f"{name}: {json.dumps(normalized, ensure_ascii=False)}")

        txt_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        LOGGER.info("Çalışma parametreleri kaydedildi: %s", txt_path)
        return txt_path
    except Exception as exc:
        LOGGER.warning("Çalışma parametreleri txt dosyasına yazılamadı: %s", exc)
        return None


def build_filename_with_params(
    base_name: str,
    encoder: Optional[str] = None,
    threshold: Optional[float] = None,
    tile: Optional[int] = None,
    alpha: Optional[float] = None,
    min_area: Optional[float] = None,
    mode_suffix: Optional[str] = None,
    arch: Optional[str] = None,
    weight_type: Optional[str] = None,  # "imagenet", "trained", veya model adı
    classic_modes: Optional[str] = None,  # "combo", "rvtlog", vb.
    yolo_model: Optional[str] = None,  # "yolo11n-seg", "yolo11s-seg", vb.
    threshold_type: Optional[str] = None,  # "otsu" veya None (manuel)
) -> str:
    """
    Parametreleri içeren dosya adı oluştur.
    
    Format: {base}_{mode}_{arch}_{encoder}_{wt}_{modes}_th{th}_tile{tile}_alpha{alpha}_minarea{min}
    
    Örnekler:
      - DL: kesif_alani_dl_Unet_resnet34_imagenet_th0.5_tile1024_minarea80
      - Classic: kesif_alani_classic_combo_otsu_tile1024_minarea80
      - Fused: kesif_alani_fused_Unet_resnet34_th0.5_tile1024_alpha0.7_minarea80
      - YOLO: kesif_alani_yolo11_seg_conf0.3_tile1024_minarea50
    """
    parts = [base_name]
    
    # 1. Mode suffix (dl, classic, fused, yolo11)
    if mode_suffix:
        parts.append(mode_suffix)
    
    # 2. Architecture (Unet, DeepLabV3Plus vb.) - DL ve fusion için
    if arch:
        parts.append(arch)
    
    # 3. Encoder (resnet34, efficientnet-b3 vb.)
    if encoder:
        parts.append(encoder)
    
    # 4. Weight type (imagenet, trained, veya model checkpoint adı)
    if weight_type:
        # Eğer bir dosya yolu ise sadece dosya adını al (uzantısız)
        if "/" in weight_type or "\\" in weight_type:
            weight_type = Path(weight_type).stem
        # Çok uzunsa kısalt (max 20 karakter)
        if len(weight_type) > 20:
            weight_type = weight_type[:20]
        parts.append(weight_type)
    
    # 5. Classic modes (combo, rvtlog, hessian, morph)
    if classic_modes:
        parts.append(classic_modes)
    
    # 6. YOLO model tipi (seg, det, veya model adı)
    if yolo_model:
        # Model adından tipi çıkar (yolo11n-seg.pt -> seg)
        if "-seg" in yolo_model:
            parts.append("seg")
        elif "-det" in yolo_model or yolo_model.endswith(".pt") and "-seg" not in yolo_model:
            parts.append("det")
        else:
            parts.append(yolo_model.replace(".pt", ""))
    
    # 7. Threshold (th0.5 veya otsu)
    if threshold_type == "otsu":
        parts.append("otsu")
    elif threshold is not None:
        parts.append(f"th{threshold:.2f}".rstrip('0').rstrip('.'))
    
    # 8. Tile size
    if tile is not None:
        parts.append(f"tile{tile}")
    
    # 9. Alpha değeri (fusion için)
    if alpha is not None:
        parts.append(f"alpha{alpha:.2f}".rstrip('0').rstrip('.'))
    
    # 10. Min area
    if min_area is not None and min_area > 0:
        parts.append(f"minarea{int(min_area)}")
    
    return "_".join(parts)


def _format_multiscale_scale_token(scale: float) -> str:
    """Create a stable filename token for a multi-scale factor."""
    scale_text = f"{float(scale):.6f}".rstrip("0").rstrip(".")
    return f"scale{scale_text.replace('.', 'p')}"


def _append_output_token(path: Path, token: str) -> Path:
    """Append a safe token to an output prefix without stripping existing dotted parts."""
    base_path = _output_base_path(path)
    safe_token = _safe_token(token)
    if not safe_token:
        return base_path
    return base_path.parent / f"{base_path.name}_{safe_token}"


def build_dl_output_paths(
    *,
    out_prefix: Path,
    task_type: str,
    arch: Optional[str],
    encoder: Optional[str],
    weight_type: Optional[str],
    threshold: float,
    tile: int,
    min_area: Optional[float],
) -> Tuple[str, Path, Path]:
    """Build standard DL probability/mask output paths."""
    base_prefix = _output_base_path(out_prefix)
    base_prefix.parent.mkdir(parents=True, exist_ok=True)
    task_key = str(task_type).strip().lower()
    filename = build_filename_with_params(
        base_name=base_prefix.name,
        mode_suffix="dlcls" if task_key == "tile_classification" else "dl",
        arch="TileClassifier" if task_key == "tile_classification" else arch,
        encoder=encoder,
        weight_type=weight_type,
        threshold=threshold,
        tile=tile,
        min_area=min_area,
    )
    prob_path = base_prefix.parent / f"{filename}_prob.tif"
    mask_path = base_prefix.parent / f"{filename}_mask.tif"
    return filename, prob_path, mask_path


def _stat_mtime_ns(stat_result: os.stat_result) -> int:
    raw = getattr(stat_result, "st_mtime_ns", None)
    if raw is not None:
        return int(raw)
    return int(round(float(stat_result.st_mtime) * 1_000_000_000.0))


def _input_cache_identity(input_path: Path) -> Dict[str, Any]:
    path = Path(input_path)
    stat_result = path.stat()
    identity: Dict[str, Any] = {
        "name": path.name,
        "size": int(stat_result.st_size),
        "mtime_ns": _stat_mtime_ns(stat_result),
    }
    try:
        with rasterio.open(path) as src:
            identity.update(
                {
                    "width": int(src.width),
                    "height": int(src.height),
                    "count": int(src.count),
                    "crs": src.crs.to_string() if src.crs is not None else None,
                    "transform": [
                        round(float(src.transform.a), 12),
                        round(float(src.transform.b), 12),
                        round(float(src.transform.c), 12),
                        round(float(src.transform.d), 12),
                        round(float(src.transform.e), 12),
                        round(float(src.transform.f), 12),
                    ],
                }
            )
    except Exception as exc:
        LOGGER.debug("Cache kimligi icin raster metadata okunamadi (%s): %s", path, exc)
    return identity


def build_derivative_cache_key(
    input_path: Path,
    band_idx: Sequence[int],
    *,
    rvt_radii: Optional[Sequence[float]] = None,
    gaussian_lrm_sigma: Optional[float] = None,
    enable_curvature: bool = True,
    enable_tpi: bool = True,
    tpi_radii: Optional[Sequence[int]] = None,
) -> str:
    """Cache dosya adlari icin dosya+parametre duyarli kisa anahtar uret."""
    resolved_radii = DEFAULTS.rvt_radii if rvt_radii is None else rvt_radii
    resolved_sigma = float(
        DEFAULTS.gaussian_lrm_sigma if gaussian_lrm_sigma is None else gaussian_lrm_sigma
    )
    resolved_tpi_radii = DEFAULTS.tpi_radii if tpi_radii is None else tpi_radii
    payload = {
        "schema": 2,
        "input": _input_cache_identity(Path(input_path)),
        "bands": [int(v) for v in band_idx],
        "rvt_radii": [round(float(v), 6) for v in resolved_radii],
        "gaussian_lrm_sigma": round(float(resolved_sigma), 6),
        "enable_curvature": bool(enable_curvature),
        "enable_tpi": bool(enable_tpi),
        "tpi_radii": [int(v) for v in resolved_tpi_radii],
    }
    raw = json.dumps(payload, ensure_ascii=True, sort_keys=True, separators=(",", ":"))
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def get_cache_path(
    input_path: Path,
    cache_dir: Optional[str] = None,
    *,
    band_idx: Sequence[int] = (),
    rvt_radii: Optional[Sequence[float]] = None,
    gaussian_lrm_sigma: Optional[float] = None,
    enable_curvature: bool = True,
    enable_tpi: bool = True,
    tpi_radii: Optional[Sequence[int]] = None,
) -> Path:
    """RVT türevleri için cache dosya yolunu oluştur."""
    cache_key = build_derivative_cache_key(
        input_path,
        band_idx,
        rvt_radii=rvt_radii,
        gaussian_lrm_sigma=gaussian_lrm_sigma,
        enable_curvature=enable_curvature,
        enable_tpi=enable_tpi,
        tpi_radii=tpi_radii,
    )
    filename = f"{input_path.stem}.{cache_key}.derivatives.npz"
    if cache_dir:
        cache_path = Path(cache_dir)
        cache_path.mkdir(parents=True, exist_ok=True)
        # Dosya adını uzantısız al ve .derivatives.npz ekle
        # Örn: kesif_alani.tif -> kesif_alani.derivatives.npz
        return cache_path / filename
    return input_path.with_name(filename)


def get_derivative_raster_cache_paths(
    input_path: Path,
    cache_dir: Optional[str] = None,
    *,
    band_idx: Sequence[int] = (),
    rvt_radii: Optional[Sequence[float]] = None,
    gaussian_lrm_sigma: Optional[float] = None,
    enable_curvature: bool = True,
    enable_tpi: bool = True,
    tpi_radii: Optional[Sequence[int]] = None,
) -> Tuple[Path, Path]:
    """Blok bazl RVT trev raster-cache (GeoTIFF) yolu + metadata JSON yolu."""
    if cache_dir:
        cache_root = Path(cache_dir)
        cache_root.mkdir(parents=True, exist_ok=True)
        tif_path = cache_root / (
            f"{input_path.stem}."
            f"{build_derivative_cache_key(input_path, band_idx, rvt_radii=rvt_radii, gaussian_lrm_sigma=gaussian_lrm_sigma, enable_curvature=enable_curvature, enable_tpi=enable_tpi, tpi_radii=tpi_radii)}"
            ".derivatives_raster.tif"
        )
        meta_path = cache_root / (
            f"{input_path.stem}."
            f"{build_derivative_cache_key(input_path, band_idx, rvt_radii=rvt_radii, gaussian_lrm_sigma=gaussian_lrm_sigma, enable_curvature=enable_curvature, enable_tpi=enable_tpi, tpi_radii=tpi_radii)}"
            ".derivatives_raster.json"
        )
        return tif_path, meta_path
    return (
        input_path.with_name(
            f"{input_path.stem}."
            f"{build_derivative_cache_key(input_path, band_idx, rvt_radii=rvt_radii, gaussian_lrm_sigma=gaussian_lrm_sigma, enable_curvature=enable_curvature, enable_tpi=enable_tpi, tpi_radii=tpi_radii)}"
            ".derivatives_raster.tif"
        ),
        input_path.with_name(
            f"{input_path.stem}."
            f"{build_derivative_cache_key(input_path, band_idx, rvt_radii=rvt_radii, gaussian_lrm_sigma=gaussian_lrm_sigma, enable_curvature=enable_curvature, enable_tpi=enable_tpi, tpi_radii=tpi_radii)}"
            ".derivatives_raster.json"
        ),
    )


def _json_dump(path: Path, data: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, sort_keys=True)


def _json_load(path: Path) -> Optional[Dict[str, Any]]:
    try:
        with open(path, "r", encoding="utf-8") as f:
            obj = json.load(f)
        if isinstance(obj, dict):
            return obj
        return None
    except FileNotFoundError:
        return None
    except Exception as e:
        LOGGER.warning("Raster-cache metadata okunamad (%s): %s", path, e)
        return None


def validate_derivative_raster_cache_metadata(
    metadata: Dict[str, Any],
    input_path: Path,
    band_idx: Sequence[int],
    *,
    rvt_radii: Optional[Sequence[float]],
    gaussian_lrm_sigma: Optional[float],
    enable_curvature: bool,
    enable_tpi: bool,
    tpi_radii: Sequence[int],
) -> bool:
    """Raster-cache metadata dogrula (girdi + parametre uyumu)."""
    required_keys = [
        "complete",
        "input_name",
        "input_mtime",
        "input_size",
        "bands",
        "shape",
        "pixel_size",
        "rvt_radii",
        "gaussian_lrm_sigma",
        "enable_curvature",
        "enable_tpi",
        "tpi_radii",
        "band_map",
    ]
    if not all(k in metadata for k in required_keys):
        LOGGER.warning("Raster-cache metadata eksik anahtarlar var; cache gecersiz")
        return False

    expected_cache_key = build_derivative_cache_key(
        input_path,
        band_idx,
        rvt_radii=rvt_radii,
        gaussian_lrm_sigma=gaussian_lrm_sigma,
        enable_curvature=enable_curvature,
        enable_tpi=enable_tpi,
        tpi_radii=tpi_radii,
    )
    cached_cache_key = str(metadata.get("cache_key", "")).strip()
    if cached_cache_key and cached_cache_key != expected_cache_key:
        LOGGER.warning("Raster-cache anahtari uyusmuyor; cache gecersiz")
        return False

    if not bool(metadata.get("complete", False)):
        LOGGER.warning("Raster-cache tamamlanmamis gorunuyor; cache gecersiz")
        return False

    if str(metadata.get("input_name", "")) != input_path.name:
        LOGGER.warning(
            "Raster-cache dosya adi uyusmuyor: cache'deki=%s, mevcut=%s",
            metadata.get("input_name"),
            input_path.name,
        )
        return False

    try:
        stat = input_path.stat()
        current_mtime = float(stat.st_mtime)
        current_size = int(stat.st_size)
    except (OSError, FileNotFoundError) as e:
        LOGGER.warning("Girdi dosyasi kontrol edilemedi: %s", e)
        return False

    cached_mtime = float(metadata.get("input_mtime", -1))
    cached_size = int(metadata.get("input_size", -1))
    if cached_size != current_size:
        LOGGER.warning("Girdi dosyasi boyutu degismis; cache gecersiz")
        return False
    if abs(cached_mtime - current_mtime) > 2.0:
        LOGGER.warning("Girdi dosyasi mtime degismis; cache gecersiz")
        return False

    if list(metadata.get("bands", [])) != list(band_idx):
        LOGGER.warning("Raster-cache bant sirasi degismis; cache gecersiz")
        return False

    wanted_radii = list(DEFAULTS.rvt_radii if rvt_radii is None else rvt_radii)
    cached_radii = list(metadata.get("rvt_radii", []))
    if len(cached_radii) != len(wanted_radii) or not np.allclose(
        np.array(cached_radii, dtype=np.float32),
        np.array(wanted_radii, dtype=np.float32),
        rtol=0.0,
        atol=1e-6,
        equal_nan=True,
    ):
        LOGGER.warning("Raster-cache RVT radii parametreleri degismis; cache gecersiz")
        return False

    wanted_sigma = float(DEFAULTS.gaussian_lrm_sigma if gaussian_lrm_sigma is None else gaussian_lrm_sigma)
    cached_sigma = float(metadata.get("gaussian_lrm_sigma", float("nan")))
    if not math.isfinite(cached_sigma) or abs(cached_sigma - wanted_sigma) > 1e-6:
        LOGGER.warning("Raster-cache gaussian_lrm_sigma degismis; cache gecersiz")
        return False

    if bool(metadata.get("enable_curvature", False)) != bool(enable_curvature):
        LOGGER.warning("Raster-cache enable_curvature degismis; cache gecersiz")
        return False
    if bool(metadata.get("enable_tpi", False)) != bool(enable_tpi):
        LOGGER.warning("Raster-cache enable_tpi degismis; cache gecersiz")
        return False

    wanted_tpi = [int(x) for x in (tpi_radii if enable_tpi else [])]
    cached_tpi = [int(x) for x in metadata.get("tpi_radii", [])]
    if wanted_tpi != cached_tpi:
        LOGGER.warning("Raster-cache tpi_radii degismis; cache gecersiz")
        return False

    band_map = metadata.get("band_map")
    if not isinstance(band_map, dict) or not band_map:
        LOGGER.warning("Raster-cache band_map eksik; cache gecersiz")
        return False

    required_bands = ["svf", "slrm"]
    missing = [b for b in required_bands if b not in band_map]
    if missing:
        LOGGER.warning("Raster-cache band_map eksik kanallar: %s", ", ".join(missing))
        return False

    return True


def load_derivative_raster_cache_info(
    cache_tif_path: Path,
    cache_meta_path: Path,
    input_path: Path,
    band_idx: Sequence[int],
    *,
    rvt_radii: Optional[Sequence[float]],
    gaussian_lrm_sigma: Optional[float],
    enable_curvature: bool,
    enable_tpi: bool,
    tpi_radii: Sequence[int],
) -> Optional[Dict[str, Any]]:
    """Raster-cache mevcut ve uyumluysa metadata d”ndr."""
    if not cache_tif_path.exists() or not cache_meta_path.exists():
        return None
    metadata = _json_load(cache_meta_path)
    if metadata is None:
        return None
    if not validate_derivative_raster_cache_metadata(
        metadata,
        input_path,
        band_idx,
        rvt_radii=rvt_radii,
        gaussian_lrm_sigma=gaussian_lrm_sigma,
        enable_curvature=enable_curvature,
        enable_tpi=enable_tpi,
        tpi_radii=tpi_radii,
    ):
        return None

    # Quick sanity-check of the GeoTIFF container (shape + band count).
    try:
        with rasterio.open(cache_tif_path) as ds:
            shape = metadata.get("shape")
            if isinstance(shape, (list, tuple)) and len(shape) == 2:
                exp_h = int(shape[0])
                exp_w = int(shape[1])
                if int(ds.height) != exp_h or int(ds.width) != exp_w:
                    LOGGER.warning("Raster-cache boyutu uyumsuz; cache gecersiz")
                    return None
            band_map = metadata.get("band_map", {})
            if isinstance(band_map, dict) and band_map:
                max_band = max(int(v) for v in band_map.values())
                if int(ds.count) < max_band:
                    LOGGER.warning("Raster-cache band says yetersiz; cache gecersiz")
                    return None
    except Exception as e:
        LOGGER.warning("Raster-cache GeoTIFF a‡lamad (%s): %s", cache_tif_path, e)
        return None
    return metadata


def _estimate_derivative_cache_halo_px(
    pixel_size: float,
    *,
    rvt_radii: Optional[Sequence[float]],
    gaussian_lrm_sigma: Optional[float],
    enable_curvature: bool,
    enable_tpi: bool,
    tpi_radii: Sequence[int],
) -> int:
    # RVT radii are in meters; convert to pixels using resolution.
    radii_m = DEFAULTS.rvt_radii if rvt_radii is None else rvt_radii
    max_radius_m = float(max(radii_m)) if radii_m else 0.0
    px = max(pixel_size, 1e-9)
    halo_rvt = int(math.ceil(max_radius_m / px)) if max_radius_m > 0 else 0

    # Gaussian fallback uses sigma in pixels; use ~4*sigma support.
    sigma = float(DEFAULTS.gaussian_lrm_sigma if gaussian_lrm_sigma is None else gaussian_lrm_sigma)
    halo_gauss = int(math.ceil(4.0 * sigma)) if sigma > 0 else 0

    halo_curv = 2 if enable_curvature else 0
    halo_tpi = int(max(tpi_radii)) if (enable_tpi and tpi_radii) else 0

    return max(2, halo_rvt, halo_gauss, halo_curv, halo_tpi)


def _get_derivative_cache_thread_reader(input_path: str) -> Any:
    """Reuse one raster reader per worker thread to avoid reopening on every block."""
    reader = getattr(_DERIV_CACHE_THREAD_LOCAL, "reader", None)
    reader_path = getattr(_DERIV_CACHE_THREAD_LOCAL, "reader_path", None)
    if reader is None or reader_path != input_path:
        if reader is not None:
            try:
                reader.close()
            except Exception:
                pass
        reader = rasterio.open(input_path)
        _DERIV_CACHE_THREAD_LOCAL.reader = reader
        _DERIV_CACHE_THREAD_LOCAL.reader_path = input_path
    return reader


def _compute_derivative_cache_block_from_source(
    src: Any,
    *,
    width: int,
    height: int,
    row: int,
    col: int,
    win_h: int,
    win_w: int,
    halo: int,
    dtm_idx: int,
    dsm_idx: int,
    pixel_size: float,
    rvt_radii: Sequence[float],
    gaussian_lrm_sigma: Optional[float],
    enable_curvature: bool,
    enable_tpi: bool,
    tpi_radii: Sequence[int],
) -> np.ndarray:
    row0 = max(0, int(row) - halo)
    col0 = max(0, int(col) - halo)
    row1 = min(height, int(row) + win_h + halo)
    col1 = min(width, int(col) + win_w + halo)
    padded = Window(col0, row0, col1 - col0, row1 - row0)

    dtm_ma = src.read(dtm_idx, window=padded, boundless=False, masked=True)
    dtm = np.ma.filled(dtm_ma.astype(np.float32), np.nan)
    dsm: Optional[np.ndarray] = None
    if int(dsm_idx) > 0:
        dsm_ma = src.read(int(dsm_idx), window=padded, boundless=False, masked=True)
        dsm = np.ma.filled(dsm_ma.astype(np.float32), np.nan)

    svf, slrm = compute_derivatives_with_rvt(
        dtm,
        pixel_size=pixel_size,
        radii=rvt_radii,
        gaussian_lrm_sigma=gaussian_lrm_sigma,
        show_progress=False,
        log_steps=False,
    )

    roff = int(row) - row0
    coff = int(col) - col0
    rs = slice(roff, roff + win_h)
    cs = slice(coff, coff + win_w)
    slope = compute_slope(dtm, pixel_size=pixel_size)
    ndsm = compute_ndsm(dsm, dtm)

    layers: List[np.ndarray] = [
        svf[rs, cs],
        slrm[rs, cs],
        slope[rs, cs],
        ndsm[rs, cs],
    ]

    return np.stack(layers, axis=0).astype(np.float32, copy=False)


def _compute_derivative_cache_block(
    input_path: str,
    *,
    width: int,
    height: int,
    row: int,
    col: int,
    win_h: int,
    win_w: int,
    halo: int,
    dtm_idx: int,
    dsm_idx: int,
    pixel_size: float,
    rvt_radii: Sequence[float],
    gaussian_lrm_sigma: Optional[float],
    enable_curvature: bool,
    enable_tpi: bool,
    tpi_radii: Sequence[int],
) -> np.ndarray:
    src = _get_derivative_cache_thread_reader(input_path)
    return _compute_derivative_cache_block_from_source(
        src,
        width=width,
        height=height,
        row=row,
        col=col,
        win_h=win_h,
        win_w=win_w,
        halo=halo,
        dtm_idx=dtm_idx,
        dsm_idx=dsm_idx,
        pixel_size=pixel_size,
        rvt_radii=rvt_radii,
        gaussian_lrm_sigma=gaussian_lrm_sigma,
        enable_curvature=enable_curvature,
        enable_tpi=enable_tpi,
        tpi_radii=tpi_radii,
    )


def build_derivative_raster_cache(
    *,
    input_path: Path,
    band_idx: Sequence[int],
    cache_tif_path: Path,
    cache_meta_path: Path,
    recalculate: bool,
    rvt_radii: Optional[Sequence[float]],
    gaussian_lrm_sigma: Optional[float],
    enable_curvature: bool,
    enable_tpi: bool,
    tpi_radii: Sequence[int],
    chunk_size: int = 2048,
    worker_count: int = 1,
    halo_px: Optional[int] = None,
) -> Dict[str, Any]:
    """RVT trevlerini blok blok hesaplayp GeoTIFF olarak cache'ler."""
    if chunk_size <= 0:
        raise ValueError("chunk_size pozitif olmal")
    if worker_count <= 0:
        raise ValueError("worker_count pozitif olmalı")
    if halo_px is not None and halo_px < 0:
        raise ValueError("halo_px negatif olamaz")

    if recalculate:
        for p in (cache_tif_path, cache_meta_path):
            try:
                if p.exists():
                    p.unlink()
            except Exception:
                pass

    cache_tif_path.parent.mkdir(parents=True, exist_ok=True)
    cache_meta_path.parent.mkdir(parents=True, exist_ok=True)
    temp_tif_path = cache_tif_path.with_name(f"{cache_tif_path.stem}.building{cache_tif_path.suffix}")
    temp_meta_path = cache_meta_path.with_name(f"{cache_meta_path.stem}.building{cache_meta_path.suffix}")
    for p in (temp_tif_path, temp_meta_path):
        try:
            if p.exists():
                p.unlink()
        except Exception:
            pass

    with rasterio.open(input_path) as src:
        height = int(src.height)
        width = int(src.width)
        transform = src.transform
        crs = src.crs
        pixel_size = float((abs(transform.a) + abs(transform.e)) / 2.0)

        dtm_idx = int(band_idx[4])
        dsm_idx = int(band_idx[3]) if len(band_idx) >= 4 else 0
        if dtm_idx <= 0:
            raise ValueError("DTM band gerekli (band_idx[4])")

        halo = (
            int(halo_px)
            if halo_px is not None
            else _estimate_derivative_cache_halo_px(
                pixel_size,
                rvt_radii=rvt_radii,
                gaussian_lrm_sigma=gaussian_lrm_sigma,
                enable_curvature=enable_curvature,
                enable_tpi=enable_tpi,
                tpi_radii=tpi_radii,
            )
        )

        radii_used = list(DEFAULTS.rvt_radii if rvt_radii is None else rvt_radii)
        sigma_used = float(DEFAULTS.gaussian_lrm_sigma if gaussian_lrm_sigma is None else gaussian_lrm_sigma)

        band_map: Dict[str, int] = {
            "svf": 1,
            "slrm": 2,
            "slope": 3,
            "ndsm": 4,
        }
        band_count = len(band_map)

        stat = input_path.stat()
        metadata: Dict[str, Any] = {
            "version": 1,
            "cache_key": build_derivative_cache_key(
                input_path,
                band_idx,
                rvt_radii=rvt_radii,
                gaussian_lrm_sigma=gaussian_lrm_sigma,
                enable_curvature=enable_curvature,
                enable_tpi=enable_tpi,
                tpi_radii=tpi_radii,
            ),
            "complete": False,
            "input_path": str(input_path),
            "input_name": input_path.name,
            "input_mtime": float(stat.st_mtime),
            "input_size": int(stat.st_size),
            "bands": list(band_idx),
            "shape": [height, width],
            "pixel_size": pixel_size,
            "rvt_radii": [float(x) for x in radii_used],
            "gaussian_lrm_sigma": float(sigma_used),
            "enable_curvature": bool(enable_curvature),
            "enable_tpi": bool(enable_tpi),
            "tpi_radii": [int(x) for x in (tpi_radii if enable_tpi else [])],
            "chunk_size": int(chunk_size),
            "worker_count": int(worker_count),
            "halo_px": int(halo),
            "band_map": {k: int(v) for k, v in band_map.items()},
        }

        _json_dump(temp_meta_path, metadata)

        out_meta = create_raster_metadata(
            height=height,
            width=width,
            transform=transform,
            crs=crs,
            dtype="float32",
            nodata=np.nan,
        )
        out_meta["count"] = int(band_count)
        out_meta["BIGTIFF"] = "IF_SAFER"
        out_meta["interleave"] = "pixel"

        total_blocks = math.ceil(height / chunk_size) * math.ceil(width / chunk_size)
        effective_workers = max(1, min(int(worker_count), total_blocks, os.cpu_count() or int(worker_count)))
        block_radii = tuple(float(x) for x in radii_used)
        block_tpi_radii = tuple(int(x) for x in tpi_radii)
        LOGGER.info(
            "Raster-cache oluŸturuluyor: %s (%d band, block=%d, halo=%d px)",
            cache_tif_path,
            band_count,
            chunk_size,
            halo,
        )

        with rasterio.open(cache_tif_path, "w", **out_meta) as dst:
            if effective_workers == 1:
                for window, row, col in progress_bar(
                    generate_windows(width, height, chunk_size, overlap=0),
                    total=total_blocks,
                    desc="DerivCache",
                    unit="block",
                ):
                    data = _compute_derivative_cache_block_from_source(
                        src,
                        width=width,
                        height=height,
                        row=int(row),
                        col=int(col),
                        win_h=int(window.height),
                        win_w=int(window.width),
                        halo=halo,
                        dtm_idx=dtm_idx,
                        dsm_idx=dsm_idx,
                        pixel_size=pixel_size,
                        rvt_radii=block_radii,
                        gaussian_lrm_sigma=sigma_used,
                        enable_curvature=enable_curvature,
                        enable_tpi=enable_tpi,
                        tpi_radii=block_tpi_radii,
                    )
                    dst.write(data, window=window)
            else:
                LOGGER.info("Raster-cache paralel hesaplama etkin: %d worker", effective_workers)
                window_iter = iter(generate_windows(width, height, chunk_size, overlap=0))
                pending: Dict[Any, Tuple[Window, int, int]] = {}
                progress = progress_bar(
                    range(total_blocks),
                    total=total_blocks,
                    desc="DerivCache",
                    unit="block",
                )

                def _submit_next(executor: ThreadPoolExecutor) -> bool:
                    try:
                        window, row, col = next(window_iter)
                    except StopIteration:
                        return False
                    future = executor.submit(
                        _compute_derivative_cache_block,
                        str(input_path),
                        width=width,
                        height=height,
                        row=int(row),
                        col=int(col),
                        win_h=int(window.height),
                        win_w=int(window.width),
                        halo=halo,
                        dtm_idx=dtm_idx,
                        dsm_idx=dsm_idx,
                        pixel_size=pixel_size,
                        rvt_radii=block_radii,
                        gaussian_lrm_sigma=sigma_used,
                        enable_curvature=enable_curvature,
                        enable_tpi=enable_tpi,
                        tpi_radii=block_tpi_radii,
                    )
                    pending[future] = (window, int(row), int(col))
                    return True

                try:
                    with ThreadPoolExecutor(
                        max_workers=effective_workers,
                        thread_name_prefix="deriv-cache",
                    ) as executor:
                        for _ in range(effective_workers):
                            if not _submit_next(executor):
                                break

                        while pending:
                            done, _ = wait(tuple(pending.keys()), return_when=FIRST_COMPLETED)
                            for future in done:
                                window, _row, _col = pending.pop(future)
                                data = future.result()
                                dst.write(data, window=window)
                                progress.update(1)
                                _submit_next(executor)
                finally:
                    progress.close()

        metadata["complete"] = True
        _json_dump(cache_meta_path, metadata)

        LOGGER.info("Raster-cache hazr: %s", cache_tif_path)
        return metadata


def save_derivatives_cache(
    cache_path: Path,
    derivatives_data: Dict[str, Any],
    metadata: Dict[str, Any],
) -> None:
    """RVT türevlerini compressed numpy dosyası olarak kaydet."""
    LOGGER.info(f"RVT türevleri kaydediliyor: {cache_path}")
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Tüm verileri birleştir
    save_dict = {**derivatives_data, **{"_metadata": metadata}}
    
    # Compressed format ile kaydet
    np.savez_compressed(cache_path, **save_dict)
    
    # Boyut bilgisi
    size_mb = cache_path.stat().st_size / (1024 * 1024)
    LOGGER.info(f"Cache kaydedildi: {size_mb:.1f} MB")


def load_derivatives_cache(cache_path: Path) -> Optional[Tuple[Dict[str, Any], Dict[str, Any]]]:
    """RVT türevlerini cache'den yükle."""
    if not cache_path.exists():
        return None
    
    try:
        size_mb = cache_path.stat().st_size / (1024 * 1024)
        LOGGER.info(f"Cache yükleniyor: {cache_path} ({size_mb:.1f} MB)")
        
        # Dosyayı aç
        with np.load(cache_path, allow_pickle=True) as data:
            # Metadata'yı ayır
            if "_metadata" not in data:
                LOGGER.warning("Cache dosyası geçersiz (metadata eksik)")
                return None
            
            metadata = data["_metadata"].item()
            
            # Diğer verileri progress bar ile yükle
            derivatives_data = {}
            layer_names = [key for key in data.files if key != "_metadata"]
            
            for key in progress_bar(
                layer_names,
                desc="Cache katmanları yükleniyor",
                unit="katman",
                total=len(layer_names)
            ):
                derivatives_data[key] = data[key].copy()
        
        LOGGER.info(f"✓ Cache yüklendi: {len(derivatives_data)} katman")
        
        return derivatives_data, metadata
        
    except Exception as e:
        LOGGER.error(f"Cache yüklenirken hata: {e}")
        return None


def validate_cache(
    metadata: Dict[str, Any],
    input_path: Path,
    bands: Sequence[int],
    *,
    rvt_radii: Optional[Sequence[float]] = None,
    gaussian_lrm_sigma: Optional[float] = None,
    enable_curvature: bool = True,
    enable_tpi: bool = True,
    tpi_radii: Optional[Sequence[int]] = None,
) -> bool:
    """
    Cache'in mevcut parametrelerle uyumlu olup olmadini kontrol et.
    
    NOT: tile ve overlap parametreleri cache gecerliligini ETKILEMEZ!
    RVT turevleri tum raster icin bir kez hesaplanir, tile/overlap sadece
    model inference sirasinda kullanilir.
    
    NOT: Dosya yolu degisse bile, dosya adi ve mtime ayniysa cache gecerlidir.
    Bu, dosyalarin farkli dizinlere tasinmasi durumunda cache'in kullanilabilmesini saglar.
    
    Args:
        metadata: Cache metadata sozlugu
        input_path: Girdi dosyasi yolu
        bands: Bant indeksleri
        rvt_radii: RVT radii parametreleri
        gaussian_lrm_sigma: LRM sigma parametresi
        enable_curvature: Curvature kanallari aktif mi
        enable_tpi: TPI kanali aktif mi
        tpi_radii: TPI radii parametreleri (enable_tpi=True ise kontrol edilir)
    
    Returns:
        bool: Cache gecerli mi
    """
    required_keys = ["input_path", "input_mtime", "input_size", "bands", "shape", "rvt_radii", "gaussian_lrm_sigma"]
    
    if not all(key in metadata for key in required_keys):
        LOGGER.warning("Cache metadata eksik")
        return False

    expected_cache_key = build_derivative_cache_key(
        input_path,
        bands,
        rvt_radii=rvt_radii,
        gaussian_lrm_sigma=gaussian_lrm_sigma,
        enable_curvature=enable_curvature,
        enable_tpi=enable_tpi,
        tpi_radii=tpi_radii,
    )
    cached_cache_key = str(metadata.get("cache_key", "")).strip()
    if cached_cache_key and cached_cache_key != expected_cache_key:
        LOGGER.warning("Cache anahtari uyusmuyor, cache gecersiz")
        return False
    
    # Dosya adi kontrolu (yol degisse bile dosya adi ayni olmali)
    cached_path = Path(metadata["input_path"])
    if cached_path.name != input_path.name:
        LOGGER.warning(f"Cache dosya adi uyusmuyor: cache'deki={cached_path.name}, mevcut={input_path.name}")
        return False
    
    # Dosya degismis mi? (mtime kontrolu - dosya tasinsa bile mtime genelde ayni kalir)
    try:
        stat = input_path.stat()
        current_mtime = stat.st_mtime
        current_size = stat.st_size
        cached_mtime = metadata["input_mtime"]
        cached_size = metadata["input_size"]
        # 2 saniye tolerans (dosya sistemi zamanlama farklari icin)
        if abs(cached_mtime - current_mtime) > 2.0:
            LOGGER.warning(f"Girdi dosyasi degismis, cache gecersiz (mtime farki: {abs(cached_mtime - current_mtime):.1f} saniye)")
            return False
        if int(cached_size) != int(current_size):
            LOGGER.warning("Girdi dosyasi boyutu degismis, cache gecersiz")
            return False
    except (OSError, FileNotFoundError) as e:
        LOGGER.warning(f"Girdi dosyasi kontrol edilemedi: {e}")
        return False

    # Raster boyutu (H,W) kontrolu
    try:
        with rasterio.open(input_path) as src:
            h = int(src.height)
            w = int(src.width)
        cached_shape = metadata["shape"]
        if tuple(cached_shape) != (h, w):
            LOGGER.warning("Raster boyutu degismis, cache gecersiz")
            return False
    except Exception as e:
        LOGGER.warning(f"Raster boyutu kontrol edilemedi: {e}")
        return False
    
    # Bant sirasi ayni mi?
    if metadata["bands"] != list(bands):
        LOGGER.warning(f"Bant sirasi degismis, cache gecersiz (cache: {metadata['bands']}, mevcut: {list(bands)})")
        return False

    # RVT parametreleri (radii/sigma) ayni mi?
    wanted_radii = list(DEFAULTS.rvt_radii if rvt_radii is None else rvt_radii)
    cached_radii = list(metadata.get("rvt_radii", []))
    if len(wanted_radii) != len(cached_radii) or not np.allclose(
        np.array(wanted_radii, dtype=np.float32),
        np.array(cached_radii, dtype=np.float32),
        rtol=0.0,
        atol=1e-6,
        equal_nan=True,
    ):
        LOGGER.warning("RVT radii degismis, cache gecersiz")
        return False
    wanted_sigma = float(DEFAULTS.gaussian_lrm_sigma if gaussian_lrm_sigma is None else gaussian_lrm_sigma)
    cached_sigma = float(metadata.get("gaussian_lrm_sigma", float("nan")))
    if not math.isfinite(cached_sigma) or abs(cached_sigma - wanted_sigma) > 1e-6:
        LOGGER.warning("gaussian_lrm_sigma degismis, cache gecersiz")
        return False
    
    # enable_curvature parametresi kontrolu
    cached_enable_curvature = metadata.get("enable_curvature", False)
    if enable_curvature != cached_enable_curvature:
        LOGGER.warning(f"enable_curvature degismis (cache: {cached_enable_curvature}, mevcut: {enable_curvature}), cache gecersiz")
        return False
    
    # enable_tpi parametresi kontrolu
    cached_enable_tpi = metadata.get("enable_tpi", False)
    if enable_tpi != cached_enable_tpi:
        LOGGER.warning(f"enable_tpi degismis (cache: {cached_enable_tpi}, mevcut: {enable_tpi}), cache gecersiz")
        return False
    
    # TPI radii kontrolu (sadece TPI aktifse)
    if enable_tpi:
        wanted_tpi_radii = list(DEFAULTS.tpi_radii if tpi_radii is None else tpi_radii)
        cached_tpi_radii = list(metadata.get("tpi_radii", []))
        if len(wanted_tpi_radii) != len(cached_tpi_radii) or wanted_tpi_radii != cached_tpi_radii:
            LOGGER.warning(f"tpi_radii degismis (cache: {cached_tpi_radii}, mevcut: {wanted_tpi_radii}), cache gecersiz")
            return False
    
    LOGGER.info("Cache gecerli")
    return True


@dataclass
class PrecomputedDerivatives:
    """Precomputed derivatives for topo5/topo7 schemas."""
    rgb: np.ndarray          # (3, H, W)
    svf: np.ndarray          # (H, W)
    slrm: np.ndarray         # (H, W)
    slope: np.ndarray        # (H, W)
    ndsm: np.ndarray         # (H, W)
    # Metadata
    transform: Affine = field(default=None)  # type: ignore
    crs: Optional[RasterioCRS] = None
    pixel_size: float = 1.0
    

def precompute_derivatives(
    input_path: Path,
    band_idx: Sequence[int],
    use_cache: bool = False,
    cache_path: Optional[Path] = None,
    recalculate: bool = False,
    rvt_radii: Optional[Sequence[float]] = None,
    gaussian_lrm_sigma: Optional[float] = None,
    enable_curvature: bool = True,
    enable_tpi: bool = True,
    tpi_radii: Tuple[int, ...] = (5, 15, 30),
) -> Optional[PrecomputedDerivatives]:
    """
    Tüm raster için topo türevlerini önceden hesapla.
    
    Bu fonksiyon:
    1. Cache varsa ve geçerliyse yükler
    2. Yoksa tüm raster'ı okur ve türevleri hesaplar
    3. İstenirse cache'e kaydeder
    
    Hesaplanan türevler: SVF, SLRM, Slope ve nDSM. Topo5 yalnızca SVF/SLRM,
    topo7 tümünü kullanır.
    
    Args:
        input_path: GeoTIFF dosya yolu
        band_idx: Bant indeksleri [R, G, B, DSM, DTM]
        use_cache: Cache kullan
        cache_path: Cache dosya yolu
        recalculate: Cache'i yoksay, yeniden hesapla
        rvt_radii: RVT yarıçapları (metre)
        gaussian_lrm_sigma: LRM için Gaussian sigma
        enable_curvature: Unused (kept for backward compatibility)
        enable_tpi: Unused (kept for backward compatibility)
        tpi_radii: Unused (kept for backward compatibility)
    
    Returns:
        PrecomputedDerivatives veya None
    """
    stage_pbar: Optional[tqdm] = None
    if use_cache and cache_path is not None:
        # High-level cache progress (in addition to per-band progress bars).
        stage_pbar = tqdm(
            total=5,
            desc="Cache",
            unit="adım",
            dynamic_ncols=True,
            mininterval=0.5,
            smoothing=0.1,
            leave=False,
        )

    with rasterio.open(input_path) as src:
        height, width = src.height, src.width
        transform = src.transform
        crs = src.crs
        pixel_size = float((abs(transform.a) + abs(transform.e)) / 2.0)
        
        # Cache kullanılacak mı kontrol et
        if use_cache and cache_path and not recalculate:
            cache_result = load_derivatives_cache(cache_path)
            if cache_result:
                derivatives_data, metadata = cache_result
                if validate_cache(
                    metadata,
                    input_path,
                    band_idx,
                    rvt_radii=rvt_radii,
                    gaussian_lrm_sigma=gaussian_lrm_sigma,
                    enable_curvature=enable_curvature,
                    enable_tpi=enable_tpi,
                    tpi_radii=tpi_radii,
                ):
                    if stage_pbar is not None:
                        stage_pbar.update(stage_pbar.total - stage_pbar.n)
                        stage_pbar.close()
                    cached_slope = derivatives_data.get("slope")
                    cached_ndsm = derivatives_data.get("ndsm")
                    if cached_slope is None or cached_ndsm is None:
                        dtm_for_missing = np.ma.filled(
                            src.read(int(band_idx[4]), masked=True).astype(np.float32),
                            np.nan,
                        )
                        if cached_slope is None:
                            cached_slope = compute_slope(dtm_for_missing, pixel_size=pixel_size)
                        if cached_ndsm is None:
                            dsm_for_missing = None
                            if len(band_idx) >= 4 and int(band_idx[3]) > 0:
                                dsm_for_missing = np.ma.filled(
                                    src.read(int(band_idx[3]), masked=True).astype(np.float32),
                                    np.nan,
                                )
                            cached_ndsm = compute_ndsm(dsm_for_missing, dtm_for_missing)
                    return PrecomputedDerivatives(
                        rgb=derivatives_data["rgb"],
                        svf=derivatives_data["svf"],
                        slrm=derivatives_data["slrm"],
                        slope=cached_slope,
                        ndsm=cached_ndsm,
                        transform=transform,
                        crs=crs,
                        pixel_size=pixel_size,
                    )
                else:
                    LOGGER.info("Cache geçersiz, yeniden hesaplanacak")
        
        # Cache kullanılmıyor veya geçersiz - yeniden hesapla
        LOGGER.info(f"RVT türevleri hesaplanıyor: {width}x{height} piksel")
        LOGGER.info("Bu işlem birkaç dakika sürebilir...")
        
        # Tüm raster'ı oku
        LOGGER.info("Raster bandları okunuyor...")
        def read_band(idx: int) -> Optional[np.ndarray]:
            if idx <= 0:
                return None
            data = src.read(idx, masked=True)
            return np.ma.filled(data.astype(np.float32), np.nan)
        
        bands_to_read = [
            ("RGB-R", band_idx[0]),
            ("RGB-G", band_idx[1]),
            ("RGB-B", band_idx[2]),
            ("DTM", band_idx[4]),
        ]
        if len(band_idx) >= 4 and int(band_idx[3]) > 0:
            bands_to_read.append(("DSM", band_idx[3]))
        band_data = {}
        for band_name, band_id in progress_bar(
            bands_to_read,
            desc="Band okuma",
            unit="band",
            total=len(bands_to_read)
        ):
            if band_id > 0:
                band_data[band_name] = read_band(band_id)
        if stage_pbar is not None:
            stage_pbar.update(1)
        
        rgb = np.stack([band_data["RGB-R"], band_data["RGB-G"], band_data["RGB-B"]], axis=0)
        dtm = band_data["DTM"]
        dsm = band_data.get("DSM")
        
        if dtm is None:
            raise ValueError("DTM band gerekli")
        
        if stage_pbar is not None:
            stage_pbar.update(1)
        
        # RVT türevlerini hesapla (SVF + SLRM)
        LOGGER.info("RVT türevleri hesaplanıyor (SVF, SLRM)...")
        svf, slrm = compute_derivatives_with_rvt(
            dtm, pixel_size=pixel_size, radii=rvt_radii, gaussian_lrm_sigma=gaussian_lrm_sigma
        )
        slope = compute_slope(dtm, pixel_size=pixel_size)
        ndsm = compute_ndsm(dsm, dtm)
        if stage_pbar is not None:
            stage_pbar.update(1)

        if stage_pbar is not None:
            stage_pbar.update(1)
        
        derivatives = PrecomputedDerivatives(
            rgb=rgb,
            svf=svf,
            slrm=slrm,
            slope=slope,
            ndsm=ndsm,
            transform=transform,
            crs=crs,
            pixel_size=pixel_size,
        )
        
        # Cache'e kaydet (istenirse)
        if use_cache and cache_path:
            derivatives_data = {
                "rgb": rgb,
                "svf": svf,
                "slrm": slrm,
                "slope": slope,
                "ndsm": ndsm,
            }
            metadata = {
                "cache_key": build_derivative_cache_key(
                    input_path,
                    band_idx,
                    rvt_radii=rvt_radii,
                    gaussian_lrm_sigma=gaussian_lrm_sigma,
                    enable_curvature=enable_curvature,
                    enable_tpi=enable_tpi,
                    tpi_radii=tpi_radii,
                ),
                "input_path": str(input_path),
                "input_mtime": input_path.stat().st_mtime,
                "input_size": input_path.stat().st_size,
                "bands": list(band_idx),
                "shape": (height, width),
                "pixel_size": pixel_size,
                "rvt_radii": list(DEFAULTS.rvt_radii if rvt_radii is None else rvt_radii),
                "gaussian_lrm_sigma": float(
                    DEFAULTS.gaussian_lrm_sigma if gaussian_lrm_sigma is None else gaussian_lrm_sigma
                ),
                "enable_curvature": enable_curvature,
                "enable_tpi": enable_tpi,
                "tpi_radii": list(tpi_radii) if enable_tpi else [],
                # NOT: tile ve overlap cache'e KAYDEDİLMEZ!
                # RVT türevleri tüm raster için hesaplanır, tile/overlap
                # sadece model inference sırasında kullanılır.
            }
            save_derivatives_cache(cache_path, derivatives_data, metadata)
            if stage_pbar is not None:
                stage_pbar.update(1)
                stage_pbar.close()
                stage_pbar = None
        
        LOGGER.info("RVT türevleri hazır ✓")
        return derivatives


def main(argv: Optional[Sequence[str]] = None) -> None:
    parser = argparse.ArgumentParser(
        description="Arkeolojik alan tespiti için derin öğrenme / klasik yöntem CLI aracı.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--config",
        type=str,
        default=default_config_path(),
        help="YAML konfigürasyon dosyası yolu. Tüm ayarları buradan kontrol edebilirsiniz.",
    )
    parser.add_argument(
        "--input",
        default=default_for("input"),
        help=cli_help("input", "(PipelineDefaults.input veya config.yaml'dan değiştirin)."),
    )
    parser.add_argument(
        "--weights",
        default=default_for("weights"),
        help=cli_help("weights"),
    )
    parser.add_argument(
        "--training-metadata",
        default=default_for("training_metadata"),
        dest="training_metadata",
        help=cli_help("training_metadata"),
    )
    parser.add_argument(
        "--bands",
        default=default_for("bands"),
        help=cli_help("bands"),
    )
    parser.add_argument(
        "--feature-mode",
        choices=("auto", "rgb3", "topo5", "topo7"),
        default=default_for("feature_mode"),
        dest="feature_mode",
        help=cli_help("feature_mode"),
    )
    parser.add_argument(
        "--tile",
        type=int,
        default=default_for("tile"),
        help=cli_help("tile"),
    )
    parser.add_argument(
        "--overlap",
        type=int,
        default=default_for("overlap"),
        help=cli_help("overlap"),
    )
    parser.add_argument(
        "--th",
        type=float,
        default=default_for("th"),
        help=cli_help("th"),
    )
    parser.add_argument(
        "--half",
        action=argparse.BooleanOptionalAction,
        default=default_for("half"),
        help=cli_help("half", "(use --no-half to force float32 even when CUDA is present)."),
    )
    parser.add_argument(
        "--device",
        type=str,
        default=default_for("device"),
        help=cli_help("device", "Örnek: 'cuda', 'cpu', 'cuda:0'. None ise otomatik seçilir."),
    )
    parser.add_argument(
        "--global-norm",
        action=argparse.BooleanOptionalAction,
        default=default_for("global_norm"),
        help=cli_help("global_norm", "(toggle with --no-global-norm for per-tile scaling)."),
    )
    parser.add_argument(
        "--norm-sample-tiles",
        type=int,
        default=default_for("norm_sample_tiles"),
        help=cli_help("norm_sample_tiles"),
    )
    parser.add_argument(
        "--feather",
        action=argparse.BooleanOptionalAction,
        default=default_for("feather"),
        help=cli_help("feather", "(switch off with --no-feather for raw tile edges)."),
    )
    parser.add_argument(
        "--enable-deep-learning",
        action=argparse.BooleanOptionalAction,
        default=default_for("enable_deep_learning"),
        dest="enable_deep_learning",
        help=cli_help("enable_deep_learning", "(--no-enable-deep-learning ile sadece klasik yöntemleri çalıştırabilirsiniz)."),
    )
    parser.add_argument(
        "--dl-task",
        default=default_for("dl_task"),
        dest="dl_task",
        help=cli_help("dl_task"),
    )
    parser.add_argument(
        "--enable-classic",
        action=argparse.BooleanOptionalAction,
        default=default_for("enable_classic"),
        dest="enable_classic",
        help=cli_help("enable_classic", "(--no-enable-classic ile sadece DL çalıştırabilirsiniz)."),
    )
    parser.add_argument(
        "--enable-fusion",
        action=argparse.BooleanOptionalAction,
        default=default_for("enable_fusion"),
        dest="enable_fusion",
        help=cli_help("enable_fusion", "(--no-enable-fusion ile DL ve klasik sonuçları ayrı tutabilirsiniz)."),
    )
    parser.add_argument(
        "--classic-modes",
        default=default_for("classic_modes"),
        help=cli_help("classic_modes"),
    )
    parser.add_argument(
        "--classic-save-intermediate",
        action=argparse.BooleanOptionalAction,
        default=default_for("classic_save_intermediate"),
        help=cli_help("classic_save_intermediate", "(use --no-classic-save-intermediate to skip per-mode rasters)."),
    )
    parser.add_argument(
        "--classic-th",
        type=float,
        default=default_for("classic_th"),
        help=cli_help("classic_th"),
    )
    parser.add_argument(
        "--alpha",
        type=float,
        default=default_for("alpha"),
        help=cli_help("alpha"),
    )
    parser.add_argument(
        "--fuse-encoders",
        default=default_for("fuse_encoders"),
        help=cli_help("fuse_encoders", "('all' or CSV of encoder suffixes)")
    )
    parser.add_argument(
        "--mask-talls",
        type=float,
        default=default_for("mask_talls"),
        help=cli_help("mask_talls"),
    )
    parser.add_argument(
        "--rgb-only",
        action=argparse.BooleanOptionalAction,
        default=default_for("rgb_only"),
        dest="rgb_only",
        help=cli_help("rgb_only", "Zero out SVF/SLRM so the model effectively runs on RGB only."),
    )
    parser.add_argument(
        "--min-area",
        type=float,
        default=default_for("min_area"),
        help=cli_help("min_area"),
    )
    parser.add_argument(
        "--simplify",
        type=float,
        default=default_for("simplify"),
        help=cli_help("simplify"),
    )
    parser.add_argument(
        "--vectorize",
        action=argparse.BooleanOptionalAction,
        default=default_for("vectorize"),
        help=cli_help("vectorize", "(set --no-vectorize to keep only raster outputs)."),
    )
    parser.add_argument(
        "--gpkg-mode",
        default=default_for("gpkg_mode"),
        dest="gpkg_mode",
        help=cli_help("gpkg_mode", "('single' = tek GPKG cok katman, 'split' = her cikti icin ayri GPKG)."),
    )
    parser.add_argument(
        "--export-candidate-excel",
        action=argparse.BooleanOptionalAction,
        default=default_for("export_candidate_excel"),
        dest="export_candidate_excel",
        help=cli_help(
            "export_candidate_excel",
            "(set --no-export-candidate-excel to disable GPS candidate table output).",
        ),
    )
    parser.add_argument(
        "--export-candidate-boxes",
        action=argparse.BooleanOptionalAction,
        default=default_for("export_candidate_boxes"),
        dest="export_candidate_boxes",
        help=cli_help(
            "export_candidate_boxes",
            "(set --no-export-candidate-boxes to disable object-like candidate boxes).",
        ),
    )
    parser.add_argument(
        "--fallback-candidates",
        action=argparse.BooleanOptionalAction,
        default=default_for("fallback_candidates_enabled"),
        dest="fallback_candidates_enabled",
        help=cli_help(
            "fallback_candidates_enabled",
            "(set --no-fallback-candidates to leave the combined Excel empty when no candidate passes filters).",
        ),
    )
    parser.add_argument(
        "--fallback-candidates-top-k",
        type=int,
        default=default_for("fallback_candidates_top_k"),
        dest="fallback_candidates_top_k",
        help=cli_help("fallback_candidates_top_k"),
    )
    parser.add_argument(
        "--fallback-candidates-min-score",
        type=float,
        default=default_for("fallback_candidates_min_score"),
        dest="fallback_candidates_min_score",
        help=cli_help("fallback_candidates_min_score"),
    )
    parser.add_argument(
        "--arch",
        default=default_for("arch"),
        help=cli_help("arch"),
    )
    parser.add_argument(
        "--encoder",
        default=default_for("encoder"),
        help=cli_help("encoder"),
    )
    parser.add_argument(
        "--encoders",
        default=default_for("encoders"),
        help=cli_help(
            "encoders",
            "(examples: 'resnet34,resnet50', 'all', 'none').",
        ),
    )
    parser.add_argument(
        "--use-fpn-classifier",
        action=argparse.BooleanOptionalAction,
        default=default_for("use_fpn_classifier"),
        dest="use_fpn_classifier",
        help=cli_help(
            "use_fpn_classifier",
            "(TileClassifier icin; belirtilmezse checkpoint metadata/state_dict'ten otomatik cozulur).",
        ),
    )
    parser.add_argument(
        "--weights-template",
        default=default_for("weights_template"),
        help=cli_help(
            "weights_template",
            "(uses {encoder} placeholder; ignored if file missing).",
        ),
    )
    parser.add_argument(
        "--out-prefix",
        default=default_for("out_prefix"),
        help=cli_help("out_prefix"),
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=default_for("seed"),
        help=cli_help("seed"),
    )
    parser.add_argument(
        "--zero-shot-imagenet",
        action=argparse.BooleanOptionalAction,
        default=default_for("zero_shot_imagenet"),
        help=cli_help("zero_shot_imagenet", "(toggle with --no-zero-shot-imagenet when supplying weights)."),
    )
    parser.add_argument(
        "--trained-model-only",
        action=argparse.BooleanOptionalAction,
        default=default_for("trained_model_only"),
        dest="trained_model_only",
        help=cli_help("trained_model_only", "(force single trained checkpoint path; disable zero-shot/multi fallbacks)."),
    )
    parser.add_argument(
        "--save-band-importance",
        action=argparse.BooleanOptionalAction,
        default=default_for("save_band_importance"),
        dest="save_band_importance",
        help=cli_help("save_band_importance"),
    )
    parser.add_argument(
        "--band-importance-max-tiles",
        type=int,
        default=default_for("band_importance_max_tiles"),
        dest="band_importance_max_tiles",
        help=cli_help("band_importance_max_tiles"),
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="count",
        default=default_for("verbose"),
        help=cli_help("verbose"),
    )
    parser.add_argument(
        "--cache-derivatives",
        action=argparse.BooleanOptionalAction,
        default=default_for("cache_derivatives"),
        dest="cache_derivatives",
        help=cli_help("cache_derivatives", "(tekrar çalıştırmalarda çok hızlı!)"),
    )
    parser.add_argument(
        "--cache-derivatives-mode",
        type=str,
        default=default_for("cache_derivatives_mode"),
        dest="cache_derivatives_mode",
        help=cli_help("cache_derivatives_mode"),
    )
    parser.add_argument(
        "--deriv-cache-chunk",
        type=int,
        default=default_for("deriv_cache_chunk"),
        dest="deriv_cache_chunk",
        help=cli_help("deriv_cache_chunk"),
    )
    parser.add_argument(
        "--deriv-cache-workers",
        type=int,
        default=default_for("deriv_cache_workers"),
        dest="deriv_cache_workers",
        help=cli_help("deriv_cache_workers"),
    )
    parser.add_argument(
        "--deriv-cache-halo",
        type=int,
        default=default_for("deriv_cache_halo"),
        dest="deriv_cache_halo",
        help=cli_help("deriv_cache_halo"),
    )
    parser.add_argument(
        "--cache-dir",
        type=str,
        default=default_for("cache_dir"),
        dest="cache_dir",
        help=cli_help("cache_dir"),
    )
    parser.add_argument(
        "--recalculate-cache",
        action=argparse.BooleanOptionalAction,
        default=default_for("recalculate_cache"),
        dest="recalculate_cache",
        help=cli_help("recalculate_cache"),
    )
    parser.add_argument(
        "--enable-yolo",
        action=argparse.BooleanOptionalAction,
        default=default_for("enable_yolo"),
        dest="enable_yolo",
        help=cli_help("enable_yolo", "(YOLO11 nesne tespit/segmentasyon modelini etkinleştir)"),
    )
    parser.add_argument(
        "--yolo-weights",
        type=str,
        default=default_for("yolo_weights"),
        dest="yolo_weights",
        help=cli_help("yolo_weights"),
    )
    parser.add_argument(
        "--yolo-conf",
        type=float,
        default=default_for("yolo_conf"),
        dest="yolo_conf",
        help=cli_help("yolo_conf"),
    )
    parser.add_argument(
        "--yolo-iou",
        type=float,
        default=default_for("yolo_iou"),
        dest="yolo_iou",
        help=cli_help("yolo_iou"),
    )
    parser.add_argument(
        "--yolo-tile",
        type=int,
        default=default_for("yolo_tile"),
        dest="yolo_tile",
        help=cli_help("yolo_tile"),
    )
    parser.add_argument(
        "--yolo-imgsz",
        type=int,
        default=default_for("yolo_imgsz"),
        dest="yolo_imgsz",
        help=cli_help("yolo_imgsz"),
    )
    parser.add_argument(
        "--yolo-device",
        type=str,
        default=default_for("yolo_device"),
        dest="yolo_device",
        help=cli_help("yolo_device"),
    )
    argv_list = list(argv) if argv is not None else sys.argv[1:]
    args = parser.parse_args(argv)
    cli_overrides = collect_cli_overrides(parser, argv_list)
    config = build_config_from_args(args, cli_overrides=cli_overrides)

    ENCODER_ALIASES = {
        "efficientnet-b3": "timm-efficientnet-b3",
        "effnet-b3": "timm-efficientnet-b3",
        "resnet34": "resnet34",
        "resnet50": "resnet50",
        "densenet121": "densenet121",
    }

    def normalize_encoder(name: str) -> Tuple[str, str]:
        key = (name or "").strip().lower()
        smp_name = ENCODER_ALIASES.get(key, key)
        suffix = "efficientnet-b3" if smp_name == "timm-efficientnet-b3" else smp_name
        return smp_name, suffix

    def available_encoders_list() -> List[str]:
        return ["resnet34", "resnet50", "efficientnet-b3", "densenet121"]

    if config.classic_th is not None and not (0.0 <= config.classic_th <= 1.0):
        parser.error("--classic-th must be between 0 and 1.")
    if not (0.0 <= config.alpha <= 1.0):
        parser.error("--alpha must be between 0 and 1.")

    configure_logging(config.verbose)
    install_gdal_warning_filter()
    set_random_seeds(config.seed)

    if config.rgb_only:
        LOGGER.info("RGB-only bayragi aktif.")

    input_path = Path(config.input)
    if not input_path.exists():
        parser.error(f"Input raster not found: {input_path}")

    if config.trained_model_only:
        if not config.enable_deep_learning:
            parser.error("--trained-model-only requires --enable-deep-learning.")
        if not config.weights:
            parser.error("--trained-model-only requires --weights.")
        if config.zero_shot_imagenet:
            LOGGER.info("trained_model_only active: forcing zero_shot_imagenet=false.")
            config.zero_shot_imagenet = False
        if config.weights_template:
            LOGGER.info("trained_model_only active: ignoring weights_template.")
            config.weights_template = None
        raw_enc_mode = (config.encoders or "").strip().lower()
        if raw_enc_mode not in ("", "none", "single"):
            LOGGER.info("trained_model_only active: forcing encoders=none (single-model).")
            config.encoders = "none"

    enc_mode = (config.encoders or "").strip().lower()
    if enc_mode == "single":
        LOGGER.info("'single' encoder modu, tek-model modu olarak yorumlandı (encoders=none).")
        enc_mode = "none"

    checkpoint_hints: Dict[str, Any] = {}
    if config.enable_deep_learning and enc_mode in ("", "none") and config.weights:
        probe_weights_path = Path(config.weights)
        if probe_weights_path.exists():
            checkpoint_hints = get_checkpoint_model_hints(probe_weights_path)
            hint_encoder_raw = str(checkpoint_hints.get("encoder", "")).strip()
            if hint_encoder_raw:
                hint_encoder, _ = normalize_encoder(hint_encoder_raw)
                if hint_encoder != config.encoder:
                    if config.trained_model_only and "encoder" in cli_overrides:
                        parser.error(
                            f"trained_model_only checkpoint encoder '{hint_encoder}' bekliyor; "
                            f"CLI encoder '{config.encoder}' kullanilamaz."
                        )
                    else:
                        LOGGER.info(
                            "Checkpoint metadata: encoder '%s' bulundu, encoder '%s' -> '%s' guncellendi.",
                            hint_encoder,
                            config.encoder,
                            hint_encoder,
                        )
                        config.encoder = hint_encoder

            hint_arch = str(checkpoint_hints.get("arch", "")).strip()
            if hint_arch and hint_arch != config.arch:
                if config.trained_model_only and "arch" in cli_overrides:
                    parser.error(
                        f"trained_model_only checkpoint arch '{hint_arch}' bekliyor; "
                        f"CLI arch '{config.arch}' kullanilamaz."
                    )
                else:
                    LOGGER.info(
                        "Checkpoint metadata: arch '%s' bulundu, arch '%s' -> '%s' guncellendi.",
                        hint_arch,
                        config.arch,
                        hint_arch,
                    )
                    config.arch = hint_arch

            hint_use_fpn = checkpoint_hints.get("use_fpn_classifier")
            if isinstance(hint_use_fpn, bool) and hint_use_fpn != config.use_fpn_classifier:
                if config.trained_model_only and "use_fpn_classifier" in cli_overrides:
                    parser.error(
                        f"trained_model_only checkpoint use_fpn_classifier='{hint_use_fpn}' bekliyor; "
                        f"CLI use_fpn_classifier='{config.use_fpn_classifier}' kullanilamaz."
                    )
                else:
                    LOGGER.info(
                        "Checkpoint metadata: use_fpn_classifier '%s' bulundu, use_fpn_classifier '%s' -> '%s' guncellendi.",
                        hint_use_fpn,
                        config.use_fpn_classifier,
                        hint_use_fpn,
                    )
                    config.use_fpn_classifier = hint_use_fpn

            hint_task = str(checkpoint_hints.get("task_type", "")).strip().lower()
            if hint_task and hint_task != str(config.dl_task).strip().lower():
                if config.trained_model_only:
                    parser.error(
                        f"trained_model_only checkpoint task '{hint_task}' ancak config/CLI dl_task "
                        f"'{config.dl_task}'. Yanlis checkpoint kullaniliyor."
                    )
                else:
                    LOGGER.info(
                        "Checkpoint metadata: task '%s' bulundu, dl_task '%s' -> '%s' guncellendi.",
                        hint_task,
                        config.dl_task,
                        hint_task,
                    )
                    config.dl_task = hint_task

    ran_multi = enc_mode not in ("", "none")
    if config.enable_deep_learning and enc_mode in ("", "none"):
        if config.weights_template:
            parser.error("--weights-template cannot be used in single-model mode (encoders=none).")
        if not str(config.encoder or "").strip():
            parser.error("Single-model mode requires a non-empty --encoder value.")
        single_encoder_name, _ = normalize_encoder(config.encoder)
        if smp is not None:
            try:
                _ = smp.encoders.get_encoder(single_encoder_name, in_channels=3)
            except Exception:
                avail = ", ".join(available_encoders_list())
                parser.error(f"Unknown encoder '{config.encoder}'. Supported: {avail}")
        config.encoder = single_encoder_name
        if not config.zero_shot_imagenet and not config.weights:
            parser.error("Single-model mode requires --weights or --zero-shot-imagenet.")

    weights_path: Optional[Path] = None
    if config.weights:
        weights_path = Path(config.weights)
        if not weights_path.exists():
            parser.error(f"Weights file not found: {weights_path}")

    training_metadata_path: Optional[Path] = None
    training_metadata: Dict[str, Any] = {}
    configured_training_metadata_path = (
        Path(config.training_metadata) if config.training_metadata else None
    )
    if configured_training_metadata_path is not None or weights_path is not None or checkpoint_hints:
        (
            training_metadata_path,
            training_metadata,
            training_metadata_source,
        ) = resolve_training_metadata(
            configured_path=configured_training_metadata_path,
            weights_path=weights_path,
            checkpoint_hints=checkpoint_hints,
        )
        if training_metadata_source == "file" and training_metadata_path is not None:
            if (
                configured_training_metadata_path is not None
                and training_metadata_path != configured_training_metadata_path
            ):
                LOGGER.info(
                    "Training metadata fallback bulundu: %s (config path: %s)",
                    training_metadata_path,
                    configured_training_metadata_path,
                )
        elif training_metadata_source == "checkpoint":
            LOGGER.warning(
                "Training metadata JSON yuklenemedi; checkpoint icindeki config ipuclari kullanilacak."
            )
        elif config.trained_model_only:
            attempted_paths = ", ".join(
                str(path)
                for path in _candidate_training_metadata_paths(
                    configured_training_metadata_path,
                    weights_path,
                )
            )
            parser.error(
                "Training metadata could not be resolved. "
                f"Configured path: {configured_training_metadata_path}. "
                f"Searched: {attempted_paths}. "
                "Checkpoint also lacks the required tile/overlap/bands metadata."
            )

    if config.trained_model_only:
        try:
            apply_trained_only_metadata_locks(
                config=config,
                metadata=training_metadata,
                cli_overrides=cli_overrides,
            )
            validate_checkpoint_metadata_consistency(
                checkpoint_hints=checkpoint_hints,
                metadata=training_metadata,
            )
        except ValueError as exc:
            parser.error(str(exc))

    dl_task = str(config.dl_task).strip().lower()

    try:
        bands = resolve_band_indexes_for_config(input_path, config)
    except argparse.ArgumentTypeError as exc:
        parser.error(str(exc))
    has_dsm = band_indexes_support_dsm(bands)
    has_topography = band_indexes_support_topography(bands)
    try:
        resolved_dl_channel_names = resolve_model_channel_names(
            band_idx=bands,
            checkpoint_hints=checkpoint_hints,
            metadata=training_metadata,
            feature_mode=config.feature_mode,
        )
    except ValueError as exc:
        parser.error(str(exc))
    if str(config.feature_mode).strip().lower() != "auto":
        for source_label, source in (
            ("training metadata", training_metadata),
            ("checkpoint metadata", checkpoint_hints),
        ):
            try:
                declared_channels = _source_declared_channel_names(source)
            except ValueError as exc:
                parser.error(f"{source_label} channel schema gecersiz: {exc}")
            if declared_channels is not None and declared_channels != resolved_dl_channel_names:
                parser.error(
                    f"{source_label} kanal semasi {declared_channels} fakat "
                    f"feature_mode={config.feature_mode} {resolved_dl_channel_names} bekliyor. "
                    "Dogru checkpoint/metadata veya feature_mode kullanin."
                )
    dl_requires_topography = channel_names_require_topography(resolved_dl_channel_names)
    if config.enable_classic and not has_topography:
        parser.error("Klasik yontemler DTM gerektirir; --bands RGB-only iken classic inference calismaz.")
    if config.enable_deep_learning and dl_requires_topography and not config.rgb_only and not has_topography:
        parser.error(
            f"Etkin derin ogrenme modeli kanal semasi {resolved_dl_channel_names} topografya gerektiriyor "
            "ama --bands yalnizca RGB sagliyor. RGB-trained checkpoint kullanin veya DTM ekleyin."
        )
    if config.enable_deep_learning and channel_names_require_dsm(resolved_dl_channel_names) and not config.rgb_only and not has_dsm:
        parser.error(
            f"Etkin derin ogrenme modeli kanal semasi {resolved_dl_channel_names} nDSM gerektiriyor "
            "ama --bands DSM bandi saglamiyor. topo7 icin R,G,B,DSM,DTM bandlari gerekir."
        )
    if config.enable_deep_learning and config.rgb_only and not dl_requires_topography:
        LOGGER.info("RGB-only bayragi acik ama etkin model zaten RGB3; topografik kanallar kullanilmayacak.")
    warn_if_training_inference_mismatch(
        metadata=training_metadata,
        input_path=input_path,
        dl_task=dl_task,
    )
    out_prefix = resolve_out_prefix(input_path, config.out_prefix, config)
    single_vector_gpkg_path: Optional[Path] = None
    if config.gpkg_mode == "single" and (config.vectorize or config.enable_yolo):
        single_vector_gpkg_path = _vector_gpkg_path(out_prefix)
        try:
            single_vector_gpkg_path.unlink(missing_ok=True)
        except Exception as exc:
            LOGGER.warning("Tekil GPKG kapsayicisi sifirlanamadi (%s): %s", single_vector_gpkg_path, exc)

    # Heuristic: treat very large rasters differently to avoid OOM in downstream steps (fusion/vectorization).
    raster_pixels = 0
    raster_height = 0
    raster_width = 0
    try:
        with rasterio.open(input_path) as _src:
            raster_height = int(_src.height)
            raster_width = int(_src.width)
            raster_pixels = raster_height * raster_width
    except Exception as e:
        LOGGER.warning("Could not read raster size for large-raster heuristics: %s", e)
    one_float_bytes = raster_pixels * np.dtype(np.float32).itemsize
    is_large_raster = raster_pixels >= 200_000_000 or one_float_bytes >= int(1.5 * 1024**3)
    if is_large_raster and config.vectorize:
        LOGGER.warning(
            "Very large raster detected (%dx%d). Vectorization/candidate export will continue and may use substantial RAM.",
            raster_width,
            raster_height,
        )

    # Yöntem etkinleştirme kontrolleri
    if (
        not config.enable_deep_learning
        and not config.enable_classic
        and not config.enable_yolo
    ):
        parser.error("En az bir yöntem etkin olmalı (deep learning, classic veya YOLO11).")
    
    classic_outputs: Optional[ClassicOutputs] = None
    yolo_outputs: Optional[YoloOutputs] = None
    fusion_outputs: Optional[FusionOutputs] = None
    fusion_encoder_label: Optional[str] = None
    dl_runs: List[Tuple[str, Path]] = []
    multi_fused_results: List[Tuple[str, FusionOutputs]] = []
    resolved_classic_modes: Optional[Tuple[str, ...]] = None
    if config.enable_classic:
        raw_modes = [mode.strip() for mode in config.classic_modes.split(",") if mode.strip()]
        if not raw_modes:
            parser.error("--classic-modes en az bir mod içermelidir.")
        valid_modes = {"rvtlog", "hessian", "morph", "combo"}
        invalid = [m for m in raw_modes if m.lower() not in valid_modes]
        if invalid:
            parser.error(f"Desteklenmeyen klasik mod(lar): {', '.join(invalid)}")
        if len(raw_modes) == 1 and raw_modes[0].lower() == "combo":
            resolved_classic_modes = ("rvtlog", "hessian", "morph")
        elif any(m.lower() == "combo" for m in raw_modes):
            parser.error("'combo' diğer klasik modlarla birleştirilemez.")
        else:
            resolved_classic_modes = tuple(m.lower() for m in raw_modes)

    # Cihaz seçimi: CLI/config'den alınır veya otomatik belirlenir
    if config.device:
        device = torch.device(config.device)
        if device.type == "cuda" and not torch.cuda.is_available():
            LOGGER.warning("CUDA istendi ama kullanılamıyor; CPU'ya geçiliyor.")
            device = torch.device("cpu")
    else:
        device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    LOGGER.info("Kullanılan cihaz: %s", device)
    if config.half and device.type != "cuda":
        LOGGER.warning("--half istendi ama CUDA kullanılamıyor; float32 ile çalışılacak.")

    # Oturum klasörüne etkin parametreleri yaz.
    write_run_params_txt(
        out_prefix=out_prefix,
        config=config,
        argv_list=argv_list,
        cli_overrides=cli_overrides,
        input_path=input_path,
        bands=bands,
        device=device,
    )

    # Cache yönetimi - RVT türevlerini bir kez hesaplayıp tekrar kullan
    precomputed_deriv: Optional[PrecomputedDerivatives] = None
    derivative_cache_tif: Optional[Path] = None
    derivative_cache_meta: Optional[Path] = None

    # Debug: Cache ayarlarını logla
    LOGGER.info("=" * 70)
    LOGGER.info("CACHE AYARLARI")
    LOGGER.info("=" * 70)
    LOGGER.info(f"  cache_derivatives: {config.cache_derivatives}")
    LOGGER.info(f"  cache_derivatives_mode: {config.cache_derivatives_mode}")
    LOGGER.info(f"  cache_dir: {config.cache_dir}")
    LOGGER.info(f"  enable_deep_learning: {config.enable_deep_learning}")
    LOGGER.info(f"  dl_task: {dl_task}")
    LOGGER.info(f"  enable_classic: {config.enable_classic}")

    cache_mode = str(getattr(config, "cache_derivatives_mode", "auto")).strip().lower()
    if cache_mode not in ("auto", "npz", "raster"):
        LOGGER.warning("cache_derivatives_mode geçersiz (%s); 'auto' kullanılacak.", cache_mode)
        cache_mode = "auto"

    needs_derivative_features = bool(
        config.enable_classic
        or (config.enable_deep_learning and dl_requires_topography and not config.rgb_only)
    )
    required_derivative_band_names = derivative_band_names_for_channel_names(resolved_dl_channel_names)
    cache_precompute_ok = True
    cache_precompute_msg = ""
    if needs_derivative_features and config.cache_derivatives and cache_mode in ("auto", "npz"):
        cache_precompute_ok, cache_precompute_msg = full_raster_cache_precompute_ok(
            input_path,
            bands,
            enable_curvature=config.enable_curvature,
            enable_tpi=config.enable_tpi,
        )
        LOGGER.info(f"  cache_precompute_ok: {cache_precompute_ok}")
        if not cache_precompute_ok and cache_precompute_msg:
            LOGGER.warning(cache_precompute_msg)
    else:
        LOGGER.info(
            "  NPZ cache kontrolü atlandı (cache_derivatives=%s, cache_mode=%s, needs_derivative_features=%s)",
            config.cache_derivatives,
            cache_mode,
            needs_derivative_features,
        )

    if needs_derivative_features and config.cache_derivatives:
        if cache_mode in ("auto", "npz") and cache_precompute_ok:
            # Full-raster NPZ cache (küçük/orta rasterlar için en hızlı okuma)
            LOGGER.info("=" * 70)
            LOGGER.info("CACHE MODU (npz): RVT türevleri bir kez hesaplanacak ve diskten okunacak")
            LOGGER.info("=" * 70)
            cache_path = get_cache_path(
                input_path,
                config.cache_dir,
                band_idx=bands,
                rvt_radii=config.rvt_radii,
                gaussian_lrm_sigma=config.gaussian_lrm_sigma,
                enable_curvature=config.enable_curvature,
                enable_tpi=config.enable_tpi,
                tpi_radii=config.tpi_radii,
            )
            precomputed_deriv = precompute_derivatives(
                input_path=input_path,
                band_idx=bands,
                use_cache=True,
                cache_path=cache_path,
                recalculate=config.recalculate_cache,
                rvt_radii=config.rvt_radii,
                gaussian_lrm_sigma=config.gaussian_lrm_sigma,
                enable_curvature=config.enable_curvature,
                enable_tpi=config.enable_tpi,
                tpi_radii=config.tpi_radii,
            )
            if precomputed_deriv:
                num_ch = len(resolved_dl_channel_names)
                LOGGER.info(f"✓ Türevler hazır ({num_ch} kanal) - Her encoder çok daha hızlı çalışacak!")
            else:
                LOGGER.warning("RVT türevleri hesaplanamadı, normal moda devam ediliyor")
        else:
            # Block-based raster cache (büyük rasterlar için bellek güvenli)
            if cache_mode == "npz" and not cache_precompute_ok:
                LOGGER.warning("Tam-raster NPZ cache bu veri için uygun değil; raster-cache kullanılacak.")

            LOGGER.info("=" * 70)
            LOGGER.info("CACHE MODU (raster): RVT türevleri blok blok diske yazılacak ve tekrar çalıştırmalarda buradan okunacak")
            LOGGER.info("=" * 70)

            derivative_cache_tif, derivative_cache_meta = get_derivative_raster_cache_paths(
                input_path,
                config.cache_dir,
                band_idx=bands,
                rvt_radii=config.rvt_radii,
                gaussian_lrm_sigma=config.gaussian_lrm_sigma,
                enable_curvature=config.enable_curvature,
                enable_tpi=config.enable_tpi,
                tpi_radii=config.tpi_radii,
            )
            raster_info = None
            if not config.recalculate_cache:
                raster_info = load_derivative_raster_cache_info(
                    derivative_cache_tif,
                    derivative_cache_meta,
                    input_path,
                    bands,
                    rvt_radii=config.rvt_radii,
                    gaussian_lrm_sigma=config.gaussian_lrm_sigma,
                    enable_curvature=config.enable_curvature,
                    enable_tpi=config.enable_tpi,
                    tpi_radii=config.tpi_radii,
                )
                if raster_info is not None:
                    band_map = {str(k): int(v) for k, v in dict(raster_info.get("band_map", {})).items()}
                    missing_bands = [
                        name for name in required_derivative_band_names if name not in band_map
                    ]
                    if missing_bands:
                        LOGGER.warning(
                            "Raster-cache topo semasi eksik (%s); yeniden olusturulacak.",
                            ", ".join(missing_bands),
                        )
                        raster_info = None

            if raster_info is None:
                chunk_size = int(getattr(config, "deriv_cache_chunk", 2048))
                worker_count = int(getattr(config, "deriv_cache_workers", 1))
                halo_px = getattr(config, "deriv_cache_halo", None)
                if halo_px is not None:
                    halo_px = int(halo_px)
                build_derivative_raster_cache(
                    input_path=input_path,
                    band_idx=bands,
                    cache_tif_path=derivative_cache_tif,
                    cache_meta_path=derivative_cache_meta,
                    recalculate=config.recalculate_cache,
                    rvt_radii=config.rvt_radii,
                    gaussian_lrm_sigma=config.gaussian_lrm_sigma,
                    enable_curvature=config.enable_curvature,
                    enable_tpi=config.enable_tpi,
                    tpi_radii=config.tpi_radii,
                    chunk_size=chunk_size,
                    worker_count=worker_count,
                    halo_px=halo_px,
                )
                raster_info = load_derivative_raster_cache_info(
                    derivative_cache_tif,
                    derivative_cache_meta,
                    input_path,
                    bands,
                    rvt_radii=config.rvt_radii,
                    gaussian_lrm_sigma=config.gaussian_lrm_sigma,
                    enable_curvature=config.enable_curvature,
                    enable_tpi=config.enable_tpi,
                    tpi_radii=config.tpi_radii,
                )

            if raster_info is None:
                LOGGER.warning("Raster-cache doğrulanamadı; normal moda devam ediliyor")
                derivative_cache_tif = None
                derivative_cache_meta = None
            else:
                deriv_band_count = len(dict(raster_info.get("band_map", {})))
                model_ch = len(resolved_dl_channel_names)
                LOGGER.info("✓ Raster-cache hazır (deriv=%d band, model=%d kanal): %s", deriv_band_count, model_ch, derivative_cache_tif)

    combined_candidate_rows: List[Dict[str, Any]] = []
    combined_candidate_box_rows: List[Dict[str, Any]] = []
    combined_candidate_fallback_rows: List[Dict[str, Any]] = []

    # Derin öğrenme etkinse modelleri çalıştır
    if config.enable_deep_learning and ran_multi:
        enc_list = (
            available_encoders_list()
            if enc_mode == "all"
            else [enc.strip() for enc in config.encoders.split(",") if enc.strip()]
        )
        single_weight_fallback_path: Optional[Path] = None
        single_weight_fallback_suffix: Optional[str] = None
        if config.weights:
            single_weight_fallback_path = Path(config.weights)
            _, single_weight_fallback_suffix = normalize_encoder(config.encoder)
            if not config.weights_template:
                if single_weight_fallback_suffix and single_weight_fallback_suffix in {
                    normalize_encoder(enc)[1] for enc in enc_list
                }:
                    LOGGER.info(
                        "Multi-encoder mode: --weights will be used only for '%s'; "
                        "other encoders require --weights-template or will fallback.",
                        single_weight_fallback_suffix,
                    )
                else:
                    LOGGER.warning(
                        "Multi-encoder mode: --weights is set but does not match requested encoders; "
                        "provide --weights-template to avoid fallback runs."
                    )
        
        # Cache kullanılıyorsa ilerleme mesajı
        if precomputed_deriv is not None or derivative_cache_tif is not None:
            LOGGER.info(f"Toplam {len(enc_list)} encoder çalıştırılacak (RVT türevleri cache'ten kullanılacak)")
        else:
            LOGGER.info(f"Toplam {len(enc_list)} encoder çalıştırılacak (her biri için RVT hesaplanacak)")
        
        for idx, enc in enumerate(enc_list, 1):
            LOGGER.info("=" * 70)
            LOGGER.info(f"ENCODER {idx}/{len(enc_list)}: {enc.upper()}")
            LOGGER.info("=" * 70)
            
            smp_name, suffix = normalize_encoder(enc)
            try:
                _ = smp.encoders.get_encoder(smp_name, in_channels=3)
            except Exception:
                avail = ", ".join(available_encoders_list())
                LOGGER.error("Unknown encoder '%s'. Supported: %s", enc, avail)
                continue

            per_weights: Optional[Path] = None
            if config.weights_template:
                cand = Path(config.weights_template.format(encoder=suffix))
                if cand.exists():
                    per_weights = cand
                else:
                    LOGGER.info("[%s] No weights at %s; falling back to zero-shot.", suffix, cand)
            if (
                per_weights is None
                and single_weight_fallback_path is not None
                and single_weight_fallback_suffix is not None
                and suffix == single_weight_fallback_suffix
                and single_weight_fallback_path.exists()
            ):
                per_weights = single_weight_fallback_path
                LOGGER.info("[%s] Using --weights fallback: %s", suffix, per_weights)

            # Checkpoint/band metadata'dan etkin kanal şemasını çöz.
            per_hints: Dict[str, Any] = {}
            if per_weights is not None:
                per_hints = get_checkpoint_model_hints(per_weights)
            try:
                run_channel_names = resolve_model_channel_names(
                    band_idx=bands,
                    checkpoint_hints=per_hints if per_weights is not None else None,
                    metadata=training_metadata,
                    feature_mode=config.feature_mode,
                )
            except ValueError as exc:
                parser.error(f"[{suffix}] {exc}")
            num_channels = len(run_channel_names)
            if str(config.feature_mode).strip().lower() != "auto" and per_hints:
                try:
                    declared_channels = _source_declared_channel_names(per_hints)
                except ValueError as exc:
                    parser.error(f"[{suffix}] checkpoint channel schema gecersiz: {exc}")
                if declared_channels is not None and declared_channels != run_channel_names:
                    parser.error(
                        f"[{suffix}] checkpoint kanal semasi {declared_channels} fakat "
                        f"feature_mode={config.feature_mode} {run_channel_names} bekliyor."
                    )
            if channel_names_require_topography(run_channel_names) and not config.rgb_only and not has_topography:
                parser.error(
                    f"[{suffix}] Model kanal semasi {run_channel_names} topografya gerektiriyor "
                    "ama mevcut --bands yalnizca RGB girisi sagliyor."
                )
            if channel_names_require_dsm(run_channel_names) and not config.rgb_only and not has_dsm:
                parser.error(
                    f"[{suffix}] Model kanal semasi {run_channel_names} nDSM gerektiriyor "
                    "ama mevcut --bands DSM bandi saglamiyor."
                )
            per_use_fpn = resolve_tile_classifier_use_fpn(
                config.use_fpn_classifier,
                per_hints if per_weights is not None else None,
            )
            
            if dl_task == "tile_classification":
                if per_weights is not None:
                    LOGGER.info("[%s] Loading trained tile-classification weights: %s", suffix, per_weights)
                    model = build_tile_classifier(
                        encoder=smp_name,
                        in_ch=num_channels,
                        enable_attention=config.enable_attention,
                        attention_reduction=config.attention_reduction,
                        encoder_weights=None,
                        use_fpn=per_use_fpn,
                    )
                    load_weights(model, per_weights, map_location=device)
                    wt = "trained"
                elif config.zero_shot_imagenet:
                    LOGGER.info("[%s] Starting in zero-shot tile-classification mode (ImageNet encoder)", suffix)
                    model = build_tile_classifier(
                        encoder=smp_name,
                        in_ch=num_channels,
                        enable_attention=config.enable_attention,
                        attention_reduction=config.attention_reduction,
                        encoder_weights="imagenet",
                        use_fpn=per_use_fpn,
                    )
                    wt = "imagenet"
                else:
                    LOGGER.warning(
                        "[%s] No weights found and zero_shot_imagenet is disabled; using random-init tile classifier.",
                        suffix,
                    )
                    model = build_tile_classifier(
                        encoder=smp_name,
                        in_ch=num_channels,
                        enable_attention=config.enable_attention,
                        attention_reduction=config.attention_reduction,
                        encoder_weights=None,
                        use_fpn=per_use_fpn,
                    )
                    wt = "random"
            else:
                if per_weights is not None:
                    LOGGER.info("[%s] Loading trained weights: %s", suffix, per_weights)
                    model = build_model(
                        arch=config.arch,
                        encoder=smp_name,
                        in_ch=num_channels,
                        enable_attention=config.enable_attention,
                        attention_reduction=config.attention_reduction,
                    )
                    load_weights(model, per_weights, map_location=device)
                    wt = "trained"
                elif config.zero_shot_imagenet:
                    LOGGER.info("[%s] Starting in zero-shot mode (ImageNet 3->%d inflate)", suffix, num_channels)
                    model = build_model_with_imagenet_inflated(
                        arch=config.arch,
                        encoder=smp_name,
                        in_ch=num_channels,
                        enable_attention=config.enable_attention,
                        attention_reduction=config.attention_reduction,
                    )
                    wt = "imagenet"
                else:
                    LOGGER.warning(
                        "[%s] No weights found and zero_shot_imagenet is disabled; using random init model.",
                        suffix,
                    )
                    model = build_model(
                        arch=config.arch,
                        encoder=smp_name,
                        in_ch=num_channels,
                        enable_attention=config.enable_attention,
                        attention_reduction=config.attention_reduction,
                    )
                    wt = "random"

            enc_prefix = _output_base_path(out_prefix)
            enc_prefix = enc_prefix.parent / f"{enc_prefix.name}_{suffix}"

            outputs = infer_tiled(
                model=model,
                input_path=input_path,
                band_idx=bands,
                task_type=dl_task,
                tile=config.tile,
                overlap=config.overlap,
                device=device,
                use_half=config.half,
                threshold=config.th,
                mask_talls=config.mask_talls,
                out_prefix=enc_prefix,
                global_norm=config.global_norm,
                norm_sample_tiles=config.norm_sample_tiles,
                feather=config.feather,
                precomputed_deriv=precomputed_deriv,  # NPZ cache (uygunsa)
                derivative_cache_tif=derivative_cache_tif,  # Raster cache (büyük rasterlarda)
                derivative_cache_meta=derivative_cache_meta,
                enable_curvature=config.enable_curvature,
                enable_tpi=config.enable_tpi,
                # encoder adı base prefix'te zaten var; tekrar eklemeyelim
                encoder=None,
                min_area=config.min_area,
                percentile_low=config.percentile_low,
                percentile_high=config.percentile_high,
                rvt_radii=config.rvt_radii,
                gaussian_lrm_sigma=config.gaussian_lrm_sigma,
                rgb_only=config.rgb_only,
                tpi_radii=config.tpi_radii,
                arch="TileClassifier" if dl_task == "tile_classification" else config.arch,
                weight_type=wt,
                channel_names=run_channel_names,
                save_band_importance=config.save_band_importance,
                band_importance_max_tiles=config.band_importance_max_tiles,
            )
            LOGGER.info("[%s] ✓ Olasılık haritası: %s", suffix, outputs.prob_path)
            LOGGER.info("[%s] ✓ İkili maske: %s", suffix, outputs.mask_path)

            if outputs.band_importance_txt is not None:
                LOGGER.info("[%s] Band onem raporu: %s", suffix, outputs.band_importance_txt)

            # Store last encoder label for fusion filenames
            fusion_encoder_label = suffix
            dl_runs.append((suffix, outputs.prob_path))
            if config.vectorize or config.export_candidate_excel:
                if config.vectorize:
                    LOGGER.info("[%s] Vekt?rle?tirme/Excel aktar?m? ba?lat?l?yor...", suffix)
                else:
                    LOGGER.info("[%s] Excel aday tablosu olu?turuluyor (vekt?r kapal?).", suffix)
                vector_base = _output_base_path(outputs.mask_path)
                vector_ok, vector_reason = _can_vectorize_predictions()
                try:
                    mask_arr = outputs.mask
                    prob_arr = outputs.prob_map
                    if mask_arr is None or prob_arr is None:
                        mask_arr, prob_arr = load_prediction_arrays(
                            prob_path=outputs.prob_path,
                            mask_path=outputs.mask_path,
                        )
                except Exception as e:
                    LOGGER.warning("[%s] Rasterlar vekt?rle?tirme/tablo i?in y?klenemedi: %s", suffix, e)
                    mask_arr = None
                    prob_arr = None

                if mask_arr is None or prob_arr is None:
                    LOGGER.warning("[%s] Vekt?r/Excel ??kt?s? atland?; rasterlar okunamad?.", suffix)
                elif config.vectorize and vector_ok:
                    try:
                        gpkg_path, gpkg_layer_name = _resolve_vector_gpkg_target(
                            out_base=vector_base,
                            layer_name=f"dl_{suffix}",
                            gpkg_mode=config.gpkg_mode,
                            single_gpkg_path=single_vector_gpkg_path,
                        )
                        gpkg = vectorize_predictions(
                            mask=mask_arr,
                            prob_map=prob_arr,
                            transform=outputs.transform,
                            crs=outputs.crs,
                            out_path=vector_base,
                            min_area=config.min_area,
                            simplify_tol=config.simplify,
                            opening_size=config.vector_opening_size,
                            label_connectivity=config.label_connectivity,
                            export_candidate_excel=False,
                            gpkg_path=gpkg_path,
                            gpkg_layer_name=gpkg_layer_name,
                        )
                    except Exception as e:
                        LOGGER.warning("[%s] Vekt?r aktarimi basarisiz: %s", suffix, e)
                        gpkg = None
                    if gpkg:
                        LOGGER.info("[%s] ? Vekt?r dosyas?: %s", suffix, _format_gpkg_target(gpkg, gpkg_layer_name))
                else:
                    LOGGER.warning("[%s] Vekt?r ??kt?s? atland? (%s).", suffix, vector_reason)

                if mask_arr is not None and prob_arr is not None and config.export_candidate_excel:
                    try:
                        job_rows = build_candidate_location_rows_from_prediction(
                            mask=mask_arr,
                            prob_map=prob_arr,
                            transform=outputs.transform,
                            crs=outputs.crs,
                            min_area=config.min_area,
                            opening_size=config.vector_opening_size,
                            label_connectivity=config.label_connectivity,
                            extra_fields=_build_combined_candidate_extra_fields(
                                source_label=f"dl_{suffix}",
                            ),
                        )
                        combined_candidate_rows.extend(job_rows)
                        LOGGER.info("[%s] Tek Excel listesine eklenen aday sayisi: %d", suffix, len(job_rows))
                        combined_candidate_fallback_rows.extend(
                            _fallback_rows_for_empty_candidate_result(
                                normal_rows=job_rows,
                                enabled=config.fallback_candidates_enabled,
                                prob_map=prob_arr,
                                transform=outputs.transform,
                                crs=outputs.crs,
                                min_area=config.min_area,
                                top_k=config.fallback_candidates_top_k,
                                min_score=config.fallback_candidates_min_score,
                                extra_fields=_build_combined_candidate_extra_fields(
                                    source_label=f"dl_{suffix}",
                                ),
                                label=f"dl_{suffix}",
                            )
                        )
                    except Exception as e:
                        LOGGER.warning("[%s] Tek Excel satirlari hazirlanamadi: %s", suffix, e)

                if mask_arr is not None and prob_arr is not None and config.export_candidate_boxes:
                    try:
                        box_extra = _build_combined_candidate_extra_fields(
                            source_label=f"dl_{suffix}",
                        )
                        if config.vectorize and vector_ok:
                            boxes_gpkg_path, boxes_layer_name = _resolve_vector_gpkg_target(
                                out_base=vector_base,
                                layer_name=f"dl_{suffix}_candidate_boxes",
                                gpkg_mode=config.gpkg_mode,
                                single_gpkg_path=single_vector_gpkg_path,
                            )
                            box_records = write_candidate_boxes_layer_from_prediction(
                                mask=mask_arr,
                                prob_map=prob_arr,
                                transform=outputs.transform,
                                crs=outputs.crs,
                                gpkg_path=boxes_gpkg_path,
                                layer_name=boxes_layer_name,
                                min_area=config.min_area,
                                opening_size=config.vector_opening_size,
                                label_connectivity=config.label_connectivity,
                                extra_fields=box_extra,
                            )
                            LOGGER.info(
                                "[%s] Aday kutu katmani: %s (%d kutu)",
                                suffix,
                                _format_gpkg_target(boxes_gpkg_path, boxes_layer_name),
                                len(box_records),
                            )
                        else:
                            box_records = _build_candidate_box_records_from_prediction(
                                mask=mask_arr,
                                prob_map=prob_arr,
                                transform=outputs.transform,
                                crs=outputs.crs,
                                min_area=config.min_area,
                                opening_size=config.vector_opening_size,
                                label_connectivity=config.label_connectivity,
                                extra_fields=box_extra,
                            )
                        if config.export_candidate_excel:
                            combined_candidate_box_rows.extend(_candidate_box_records_to_rows(box_records))
                    except Exception as e:
                        LOGGER.warning("[%s] Aday kutulari hazirlanamadi: %s", suffix, e)

        if not dl_runs:
            parser.error(
                "Hiç geçerli encoder çalıştırılamadı. --encoders değerlerini kontrol edin veya encoders=none kullanın."
            )
        
        LOGGER.info("")
        LOGGER.info("=" * 70)
        LOGGER.info("TÜM ENCODERLAR TAMAMLANDI!")
        LOGGER.info("=" * 70)
        
        # Multi-encoder loop tamamlandı
        # Klasik yöntemler kapalıysa burada çık, açıksa devam et
        if not config.enable_classic and not config.enable_yolo:
            if config.export_candidate_excel:
                table_rows = _use_fallback_rows_if_no_candidates(
                    rows=combined_candidate_rows,
                    fallback_rows=combined_candidate_fallback_rows,
                    top_k=config.fallback_candidates_top_k,
                )
                combined_table_path = write_combined_candidate_locations_table(
                    rows=table_rows,
                    candidate_box_rows=combined_candidate_box_rows if config.export_candidate_boxes else None,
                    out_base=_combined_candidate_table_base(out_prefix),
                )
                if combined_table_path:
                    LOGGER.info("Tek birlesik Excel aday tablosu yazildi: %s", combined_table_path)
            LOGGER.info("Klasik yöntemler kapalı, işlem tamamlandı.")
            return
    
    # Tek model çalıştır (sadece encoders boş/none ise)
    # Tek model için etkin kanal şemasını kullan.
    single_channel_names: Tuple[str, ...] = tuple(resolved_dl_channel_names)
    num_channels = len(single_channel_names)
    single_use_fpn = resolve_tile_classifier_use_fpn(
        config.use_fpn_classifier,
        checkpoint_hints if weights_path is not None else None,
    )
    
    if enc_mode in ("", "none") and config.enable_deep_learning and weights_path is not None:
        ckpt_in_ch = checkpoint_hints.get("in_channels")
        if isinstance(ckpt_in_ch, int) and ckpt_in_ch != num_channels:
            parser.error(
                f"Checkpoint in_channels={ckpt_in_ch} fakat etkin kanal semasi "
                f"{single_channel_names} ({num_channels} kanal)."
            )
        ckpt_channel_names = checkpoint_hints.get("channel_names")
        raw_ckpt_channels = _raw_channel_names(ckpt_channel_names)
        if raw_ckpt_channels is not None:
            try:
                ckpt_channels = normalize_model_channel_names(raw_ckpt_channels)
            except ValueError as exc:
                parser.error(f"Checkpoint channel_names gecersiz: {exc}")
            if ckpt_channels != single_channel_names:
                parser.error(
                    f"Checkpoint channel_names={ckpt_channels} fakat etkin kanal semasi {single_channel_names}."
                )
        elif ckpt_in_ch == num_channels:
            LOGGER.warning(
                "Checkpoint channel_names bilgisi icermiyor. Etkin kanal semasi %s.",
                list(single_channel_names),
            )
        LOGGER.info("Loading trained weights in single-encoder mode: %s", weights_path)
        if dl_task == "tile_classification":
            model = build_tile_classifier(
                encoder=config.encoder,
                in_ch=num_channels,
                enable_attention=config.enable_attention,
                attention_reduction=config.attention_reduction,
                encoder_weights=None,
                use_fpn=single_use_fpn,
            )
        else:
            model = build_model(
                arch=config.arch,
                encoder=config.encoder,
                in_ch=num_channels,
                enable_attention=config.enable_attention,
                attention_reduction=config.attention_reduction,
            )
        load_weights(model, weights_path, map_location=device)
    elif enc_mode in ("", "none") and config.enable_deep_learning and config.zero_shot_imagenet:
        if dl_task == "tile_classification":
            LOGGER.info("Zero-shot tile-classification mode: using ImageNet-pretrained encoder.")
            model = build_tile_classifier(
                encoder=config.encoder,
                in_ch=num_channels,
                enable_attention=config.enable_attention,
                attention_reduction=config.attention_reduction,
                encoder_weights="imagenet",
                use_fpn=single_use_fpn,
            )
        else:
            LOGGER.info(f"Zero-shot mode: using ImageNet-pretrained encoder inflated to {num_channels} channels.")
            if config.enable_attention:
                LOGGER.info("CBAM Attention module is active.")
            model = build_model_with_imagenet_inflated(
                arch=config.arch,
                encoder=config.encoder,
                in_ch=num_channels,
                enable_attention=config.enable_attention,
                attention_reduction=config.attention_reduction,
            )
    elif enc_mode in ("", "none") and config.enable_deep_learning:
        if dl_task == "tile_classification":
            LOGGER.warning("No weights provided and zero_shot_imagenet is disabled; using random-init tile classifier.")
            model = build_tile_classifier(
                encoder=config.encoder,
                in_ch=num_channels,
                enable_attention=config.enable_attention,
                attention_reduction=config.attention_reduction,
                encoder_weights=None,
                use_fpn=single_use_fpn,
            )
        else:
            LOGGER.warning("No weights provided and zero_shot_imagenet is disabled; using random init model.")
            model = build_model(
                arch=config.arch,
                encoder=config.encoder,
                in_ch=num_channels,
                enable_attention=config.enable_attention,
                attention_reduction=config.attention_reduction,
            )
    else:
        model = None

    if config.enable_deep_learning and config.mask_talls is not None and not has_dsm:
        LOGGER.warning(
            "DSM band not provided; nDSM channel will be zero and tall-object masking will be disabled."
        )

    # Derin öğrenme çıkarımını çalıştır (etkinse)
    outputs = None
    run_multiscale = False
    ms_scales: Tuple[float, ...] = (1.0,)
    multiscale_saved_outputs: List[MultiscaleSavedOutput] = []
    if enc_mode in ("", "none") and config.enable_deep_learning and model is not None:
        # Weight type belirleme (eğitilmiş mi, ImageNet mi)
        single_wt = "trained" if weights_path is not None else ("imagenet" if config.zero_shot_imagenet else "random")

        _infer_kwargs = dict(
            model=model,
            band_idx=bands,
            task_type=dl_task,
            tile=config.tile,
            overlap=config.overlap,
            device=device,
            use_half=config.half,
            threshold=config.th,
            mask_talls=config.mask_talls,
            out_prefix=out_prefix,
            global_norm=config.global_norm,
            norm_sample_tiles=config.norm_sample_tiles,
            feather=config.feather,
            enable_curvature=config.enable_curvature,
            enable_tpi=config.enable_tpi,
            encoder=config.encoder,
            min_area=config.min_area,
            percentile_low=config.percentile_low,
            percentile_high=config.percentile_high,
            rvt_radii=config.rvt_radii,
            gaussian_lrm_sigma=config.gaussian_lrm_sigma,
            rgb_only=config.rgb_only,
            tpi_radii=config.tpi_radii,
            arch="TileClassifier" if dl_task == "tile_classification" else config.arch,
            weight_type=single_wt,
            channel_names=single_channel_names,
            save_band_importance=config.save_band_importance,
            band_importance_max_tiles=config.band_importance_max_tiles,
        )

        ms_scales = tuple(float(s) for s in config.multiscale_scales) if config.enable_multiscale else (1.0,)
        run_multiscale = config.enable_multiscale and len(ms_scales) > 1

        if not run_multiscale:
            # --- Tek ölçek (mevcut davranış) ---
            outputs = infer_tiled(
                input_path=input_path,
                precomputed_deriv=precomputed_deriv,
                derivative_cache_tif=derivative_cache_tif,
                derivative_cache_meta=derivative_cache_meta,
                **_infer_kwargs,
            )
        else:
            # --- Multi-scale inference ---
            import tempfile
            LOGGER.info("")
            LOGGER.info("=" * 70)
            LOGGER.info("MULTI-SCALE INFERENCE: ölçekler=%s  birleştirme=%s",
                        ms_scales, config.multiscale_merge)
            LOGGER.info("=" * 70)

            with rasterio.open(input_path) as _src:
                orig_h, orig_w = _src.height, _src.width
                orig_transform = _src.transform
                orig_crs = _src.crs

            scale_prob_maps: List[np.ndarray] = []
            reference_outputs: Optional[InferenceOutputs] = None
            tmp_dir = Path(tempfile.mkdtemp(prefix="multiscale_"))

            for si, scale in enumerate(ms_scales):
                LOGGER.info("")
                LOGGER.info("--- Ölçek %d/%d: %.2fx ---", si + 1, len(ms_scales), scale)
                scale_token = _format_multiscale_scale_token(scale)
                scale_infer_kwargs = dict(_infer_kwargs)
                if config.multiscale_save_individual_outputs:
                    scale_infer_kwargs["out_prefix"] = _append_output_token(out_prefix, scale_token)

                if abs(scale - 1.0) < 1e-6:
                    scale_outputs = infer_tiled(
                        input_path=input_path,
                        precomputed_deriv=precomputed_deriv,
                        derivative_cache_tif=derivative_cache_tif,
                        derivative_cache_meta=derivative_cache_meta,
                        **scale_infer_kwargs,
                    )
                else:
                    scaled_path = tmp_dir / f"scaled_{scale:.2f}_{input_path.name}"
                    _downsample_raster(input_path, scale, scaled_path)
                    scale_outputs = infer_tiled(
                        input_path=scaled_path,
                        precomputed_deriv=None,
                        derivative_cache_tif=None,
                        derivative_cache_meta=None,
                        **scale_infer_kwargs,
                    )
                    try:
                        scaled_path.unlink(missing_ok=True)
                    except Exception:
                        pass

                if reference_outputs is None:
                    reference_outputs = scale_outputs
                if config.multiscale_save_individual_outputs:
                    multiscale_saved_outputs.append(
                        MultiscaleSavedOutput(
                            scale_token=scale_token,
                            scale_factor=float(scale),
                            scale_level=si + 1,
                            outputs=scale_outputs,
                        )
                    )
                    LOGGER.info(
                        "Ölçek %.2fx ayrı raster çıktılarını yazdı: %s, %s",
                        scale,
                        scale_outputs.prob_path,
                        scale_outputs.mask_path,
                    )
                elif config.export_candidate_excel:
                    try:
                        scale_mask_for_rows = scale_outputs.mask
                        scale_prob_for_rows = scale_outputs.prob_map
                        if scale_mask_for_rows is None or scale_prob_for_rows is None:
                            scale_mask_for_rows, scale_prob_for_rows = load_prediction_arrays(
                                prob_path=scale_outputs.prob_path,
                                mask_path=scale_outputs.mask_path,
                            )
                        scale_rows = build_candidate_location_rows_from_prediction(
                            mask=scale_mask_for_rows,
                            prob_map=scale_prob_for_rows,
                            transform=scale_outputs.transform,
                            crs=scale_outputs.crs,
                            min_area=config.min_area,
                            opening_size=config.vector_opening_size,
                            label_connectivity=config.label_connectivity,
                            extra_fields=_build_combined_candidate_extra_fields(
                                source_label=f"dl_{scale_token}",
                                scale_level=si + 1,
                                scale_factor=float(scale),
                            ),
                        )
                        combined_candidate_rows.extend(scale_rows)
                        if config.export_candidate_boxes:
                            scale_box_records = _build_candidate_box_records_from_prediction(
                                mask=scale_mask_for_rows,
                                prob_map=scale_prob_for_rows,
                                transform=scale_outputs.transform,
                                crs=scale_outputs.crs,
                                min_area=config.min_area,
                                opening_size=config.vector_opening_size,
                                label_connectivity=config.label_connectivity,
                                extra_fields=_build_combined_candidate_extra_fields(
                                    source_label=f"dl_{scale_token}",
                                    scale_level=si + 1,
                                    scale_factor=float(scale),
                                ),
                            )
                            combined_candidate_box_rows.extend(
                                _candidate_box_records_to_rows(scale_box_records)
                            )
                            LOGGER.info(
                                "Olcek %.2fx aday kutulari tek Excel listesine eklendi: %d satir",
                                scale,
                                len(scale_box_records),
                            )
                        LOGGER.info(
                            "Ölçek %.2fx adayları tek Excel listesine eklendi: %d satır",
                            scale,
                            len(scale_rows),
                        )
                        combined_candidate_fallback_rows.extend(
                            _fallback_rows_for_empty_candidate_result(
                                normal_rows=scale_rows,
                                enabled=config.fallback_candidates_enabled,
                                prob_map=scale_prob_for_rows,
                                transform=scale_outputs.transform,
                                crs=scale_outputs.crs,
                                min_area=config.min_area,
                                top_k=config.fallback_candidates_top_k,
                                min_score=config.fallback_candidates_min_score,
                                extra_fields=_build_combined_candidate_extra_fields(
                                    source_label=f"dl_{scale_token}",
                                    scale_level=si + 1,
                                    scale_factor=float(scale),
                                ),
                                label=f"dl_{scale_token}",
                            )
                        )
                    except Exception as exc:
                        LOGGER.warning(
                            "Ölçek %.2fx için birlesik Excel satirlari hazirlanamadi: %s",
                            scale,
                            exc,
                        )

                if scale_outputs.prob_map is not None:
                    native_prob = scale_outputs.prob_map
                else:
                    with rasterio.open(scale_outputs.prob_path) as _ps:
                        native_prob = _ps.read(1).astype(np.float32)

                if abs(scale - 1.0) < 1e-6:
                    scale_prob_maps.append(native_prob)
                else:
                    upsampled = _upsample_prob_map(native_prob, orig_h, orig_w)
                    scale_prob_maps.append(upsampled)

            LOGGER.info("")
            LOGGER.info("Ölçekler birleştiriliyor: %s ...", config.multiscale_merge)
            ms_weights = (
                tuple(float(w) for w in config.multiscale_weights)
                if config.multiscale_weights is not None
                else None
            )
            merged_prob = _merge_prob_maps(scale_prob_maps, config.multiscale_merge, ms_weights)
            merged_prob = np.clip(merged_prob, 0.0, 1.0)

            merged_mask = (merged_prob >= config.th).astype(np.uint8)

            assert reference_outputs is not None
            _, merged_prob_path, merged_mask_path = build_dl_output_paths(
                out_prefix=out_prefix,
                task_type=dl_task,
                arch="TileClassifier" if dl_task == "tile_classification" else config.arch,
                encoder=config.encoder,
                weight_type=single_wt,
                threshold=config.th,
                tile=config.tile,
                min_area=config.min_area,
            )
            write_prob_and_mask_rasters(
                prob_map=merged_prob.astype(np.float32),
                mask=merged_mask,
                transform=orig_transform,
                crs=orig_crs,
                prob_path=merged_prob_path,
                mask_path=merged_mask_path,
            )

            outputs = InferenceOutputs(
                prob_path=merged_prob_path,
                mask_path=merged_mask_path,
                prob_map=merged_prob,
                mask=merged_mask,
                transform=orig_transform,
                crs=orig_crs,
                band_importance_txt=reference_outputs.band_importance_txt,
                band_importance_json=reference_outputs.band_importance_json,
            )
            LOGGER.info("Multi-scale birleştirme tamamlandı. %d ölçek → merged prob map.", len(ms_scales))
            LOGGER.info("Birleşik raster çıktıları: %s, %s", merged_prob_path, merged_mask_path)

            try:
                import shutil
                shutil.rmtree(tmp_dir, ignore_errors=True)
            except Exception:
                pass

        if outputs.band_importance_txt is not None:
            LOGGER.info("Band onem raporu: %s", outputs.band_importance_txt)

        # Record encoder for fusion filenames (single-encoder path)
        fusion_encoder_label = config.encoder

        LOGGER.info("Olasılık haritası yazıldı: %s", outputs.prob_path)
        LOGGER.info("İkili maske yazıldı: %s", outputs.mask_path)

    fuse_enabled = config.enable_classic and config.enable_fusion and config.enable_deep_learning
    if config.enable_fusion and not (config.enable_classic and config.enable_deep_learning):
        LOGGER.warning("Fusion için hem deep learning hem de classic yöntemlerin etkin olması gerekir.")

    # Klasik pipeline'ı çalıştır (etkinse)
    if config.enable_classic and resolved_classic_modes is not None:
        LOGGER.info("")
        LOGGER.info("=" * 70)
        LOGGER.info("KLASİK YÖNTEMLER BAŞLATILIYOR")
        LOGGER.info(f"Modlar: {', '.join(resolved_classic_modes)}")
        LOGGER.info("=" * 70)
        
        classic_outputs = infer_classic_tiled(
            input_path=input_path,
            band_idx=bands,
            tile=config.tile,
            overlap=config.overlap,
            out_prefix=out_prefix,
            classic_th=config.classic_th,
            modes=resolved_classic_modes,
            feather=config.feather,
            save_intermediate=config.classic_save_intermediate,
            min_area=config.min_area,
            precomputed_deriv=precomputed_deriv,
            derivative_cache_tif=derivative_cache_tif,
            derivative_cache_meta=derivative_cache_meta,
            enable_curvature=config.enable_curvature,
            enable_tpi=config.enable_tpi,
            tpi_radii=config.tpi_radii,
            sigma_scales=config.sigma_scales,
            morphology_radii=config.morphology_radii,
            rvt_radii=config.rvt_radii,
            gaussian_gradient_sigma=config.gaussian_gradient_sigma,
            local_variance_window=config.local_variance_window,
            gaussian_lrm_sigma=config.gaussian_lrm_sigma,
        )
        LOGGER.info("✓ Klasik yöntem birleşik çıktı: %s, %s", classic_outputs.prob_path, classic_outputs.mask_path)
        if classic_outputs.per_mode:
            for mode_name, mode_out in classic_outputs.per_mode.items():
                LOGGER.info(
                    "  ✓ Klasik mod '%s': %s, %s",
                    mode_name,
                    mode_out.prob_path,
                    mode_out.mask_path,
                )

    # YOLO11 pipeline'ı çalıştır (etkinse)
    if config.enable_yolo:
        if YOLO is None:
            LOGGER.error("YOLO11 etkinleştirildi ama ultralytics yüklü değil. Yüklemek için: pip install ultralytics>=8.1.0")
        else:
            LOGGER.info("")
            LOGGER.info("=" * 70)
            LOGGER.info("YOLO11 BAŞLATILIYOR")
            LOGGER.info("=" * 70)
            
            yolo_tile_size = config.yolo_tile if config.yolo_tile is not None else config.tile
            
            yolo_outputs = infer_yolo_tiled(
                input_path=input_path,
                band_idx=bands,
                tile=yolo_tile_size,
                overlap=config.overlap,
                out_prefix=out_prefix,
                yolo_weights=config.yolo_weights,
                conf_threshold=config.yolo_conf,
                iou_threshold=config.yolo_iou,
                imgsz=config.yolo_imgsz,
                device=config.yolo_device,
                min_area=config.min_area,
                precomputed_deriv=precomputed_deriv,
                percentile_low=config.percentile_low,
                percentile_high=config.percentile_high,
                save_labels=True,  # Etiketli tespitleri GeoPackage olarak kaydet
            )
            
            LOGGER.info("✓ YOLO11 olasılık haritası: %s", yolo_outputs.prob_path)
            LOGGER.info("✓ YOLO11 ikili maske: %s", yolo_outputs.mask_path)
            if yolo_outputs.label_records:
                try:
                    detection_records = _prepare_yolo_detection_records(
                        yolo_outputs.label_records,
                        yolo_outputs.crs,
                    )
                    if detection_records:
                        labels_out_base = yolo_outputs.labels_out_base or _output_base_path(yolo_outputs.mask_path)
                        labels_gpkg_path, labels_layer_name = _resolve_vector_gpkg_target(
                            out_base=labels_out_base,
                            layer_name="yolo11_detections",
                            gpkg_mode=config.gpkg_mode,
                            single_gpkg_path=single_vector_gpkg_path,
                        )
                        _write_records_to_gpkg(
                            records=detection_records,
                            crs=yolo_outputs.crs,
                            gpkg_path=labels_gpkg_path,
                            layer_name=labels_layer_name,
                            column_order=(
                                "id",
                                "class_id",
                                "class_name",
                                "confidence",
                                "area_m2",
                                "center_x",
                                "center_y",
                                "bbox_xmin",
                                "bbox_ymin",
                                "bbox_xmax",
                                "bbox_ymax",
                                "tile_row",
                                "tile_col",
                            ),
                        )
                        LOGGER.info(
                            "✓ YOLO11 etiketli tespitler kaydedildi: %s",
                            _format_gpkg_target(labels_gpkg_path, labels_layer_name),
                        )
                        class_counts: Dict[str, int] = {}
                        for rec in detection_records:
                            class_name = str(rec.get("class_name", "unknown"))
                            class_counts[class_name] = class_counts.get(class_name, 0) + 1
                        if class_counts:
                            LOGGER.info("Tespit edilen sınıflar:")
                            for class_name, count in sorted(class_counts.items()):
                                LOGGER.info("  - %s: %d adet", class_name, count)
                except Exception as e:
                    LOGGER.error("YOLO11 etiketli tespitler kaydedilirken hata: %s", e)

    # Fusion (birleştirme) çalıştır (etkinse ve her iki yöntem de çalıştıysa)
    if fuse_enabled and classic_outputs is not None:
        LOGGER.info("")
        LOGGER.info("=" * 70)
        LOGGER.info(f"FUSION (BİRLEŞTİRME) BAŞLATILIYOR (alpha={config.alpha})")
        LOGGER.info("=" * 70)
        alpha = float(config.alpha)
        fuse_threshold = config.th if config.th is not None else 0.5
        
        # Per-encoder fusion (multi-encoder runs)
        if ran_multi and dl_runs:
            fuse_filter_raw = (config.fuse_encoders or "all").strip().lower()
            fuse_filter = None if fuse_filter_raw == "all" else {s.strip().lower() for s in fuse_filter_raw.split(",") if s.strip()}
            base_prefix = _output_base_path(out_prefix)
            base_prefix.parent.mkdir(parents=True, exist_ok=True)
            for enc_suffix, dl_prob_path in dl_runs:
                if fuse_filter is not None and enc_suffix.lower() not in fuse_filter:
                    continue
                _fname = build_filename_with_params(
                    base_name=base_prefix.name,
                    mode_suffix="fused",
                    arch=config.arch,
                    encoder=enc_suffix,
                    threshold=fuse_threshold,
                    tile=config.tile,
                    alpha=alpha,
                    min_area=config.min_area,
                )
                _pp = base_prefix.parent / f"{_fname}_prob.tif"
                _mp = base_prefix.parent / f"{_fname}_mask.tif"
                _tr, _crs = write_fused_probability_rasters_from_paths(
                    dl_prob_path=dl_prob_path,
                    classic_prob_path=classic_outputs.prob_path,
                    alpha=alpha,
                    threshold=fuse_threshold,
                    out_prob_path=_pp,
                    out_mask_path=_mp,
                )
                multi_fused_results.append(
                    (
                        enc_suffix,
                        FusionOutputs(
                            prob_path=_pp,
                            mask_path=_mp,
                            prob_map=None,
                            mask=None,
                            transform=_tr,
                            crs=_crs,
                            threshold=fuse_threshold,
                        ),
                    )
                )
                LOGGER.info("? Fused (%s): %s, %s", enc_suffix, _pp, _mp)
        
        dl_prob_for_fusion: Optional[Path] = None
        if outputs is not None and getattr(outputs, "prob_path", None) is not None:
            dl_prob_for_fusion = outputs.prob_path
        elif dl_runs:
            # Fallback to last encoder run
            fusion_encoder_label = fusion_encoder_label or dl_runs[-1][0]
            dl_prob_for_fusion = dl_runs[-1][1]

        base_prefix = _output_base_path(out_prefix)
        base_prefix.parent.mkdir(parents=True, exist_ok=True)
        
        # Parametreli dosya adı oluştur (fusion için)
        fused_filename = build_filename_with_params(
            base_name=base_prefix.name,
            mode_suffix="fused",
            arch=config.arch,
            encoder=fusion_encoder_label,
            threshold=fuse_threshold,
            tile=config.tile,
            alpha=alpha,
            min_area=config.min_area,
        )
        
        fused_prob_path = base_prefix.parent / f"{fused_filename}_prob.tif"
        fused_mask_path = base_prefix.parent / f"{fused_filename}_mask.tif"
        if dl_prob_for_fusion is None:
            LOGGER.warning("Fusion skipped: no DL probability raster available.")
        else:
            _tr, _crs = write_fused_probability_rasters_from_paths(
                dl_prob_path=dl_prob_for_fusion,
                classic_prob_path=classic_outputs.prob_path,
                alpha=alpha,
                threshold=fuse_threshold,
                out_prob_path=fused_prob_path,
                out_mask_path=fused_mask_path,
            )
            fusion_outputs = FusionOutputs(
                prob_path=fused_prob_path,
                mask_path=fused_mask_path,
                prob_map=None,
                mask=None,
                transform=_tr,
                crs=_crs,
                threshold=fuse_threshold,
            )
            LOGGER.info("✓ Birleştirilmiş (fused) çıktılar: %s, %s", fused_prob_path, fused_mask_path)

    # Vektörleştirme (etkinse)
    if config.vectorize or config.export_candidate_excel:
        LOGGER.info("")
        LOGGER.info("=" * 70)
        LOGGER.info("VEKTÖRLEŞTİRME BAŞLATILIYOR")
        LOGGER.info("=" * 70)
        vector_jobs: List[VectorExportJob] = []
        
        # Çoklu-encoder modunda DL çıktıları döngü sırasında vektörleştirildi; tekrar etmeyelim
        if outputs is not None and not ran_multi:
            vector_jobs.append(
                VectorExportJob(
                    label="dl_multiscale_merged" if run_multiscale else "dl",
                    mask=outputs.mask,
                    prob_map=outputs.prob_map,
                    transform=outputs.transform,
                    crs=outputs.crs,
                    out_base=_output_base_path(outputs.mask_path),
                    prob_path=outputs.prob_path,
                    mask_path=outputs.mask_path,
                    gpkg_layer_name="dl_multiscale_merged" if run_multiscale else "dl",
                )
            )
        for saved_scale in multiscale_saved_outputs:
            vector_jobs.append(
                VectorExportJob(
                    label=f"dl_{saved_scale.scale_token}",
                    mask=saved_scale.outputs.mask,
                    prob_map=saved_scale.outputs.prob_map,
                    transform=saved_scale.outputs.transform,
                    crs=saved_scale.outputs.crs,
                    out_base=_output_base_path(saved_scale.outputs.mask_path),
                    prob_path=saved_scale.outputs.prob_path,
                    mask_path=saved_scale.outputs.mask_path,
                    scale_level=saved_scale.scale_level,
                    scale_factor=saved_scale.scale_factor,
                    gpkg_layer_name=f"dl_{saved_scale.scale_token}",
                )
            )
        if yolo_outputs is not None:
            vector_jobs.append(
                VectorExportJob(
                    label="yolo11",
                    mask=yolo_outputs.mask,
                    prob_map=yolo_outputs.prob_map,
                    transform=yolo_outputs.transform,
                    crs=yolo_outputs.crs,
                    out_base=_output_base_path(yolo_outputs.mask_path),
                    prob_path=yolo_outputs.prob_path,
                    mask_path=yolo_outputs.mask_path,
                    gpkg_layer_name="yolo11_mask_polygons",
                )
            )
        if classic_outputs is not None:
            vector_jobs.append(
                VectorExportJob(
                    label="classic",
                    mask=classic_outputs.mask,
                    prob_map=classic_outputs.prob_map,
                    transform=classic_outputs.transform,
                    crs=classic_outputs.crs,
                    out_base=_output_base_path(classic_outputs.mask_path),
                    prob_path=classic_outputs.prob_path,
                    mask_path=classic_outputs.mask_path,
                    gpkg_layer_name="classic",
                )
            )
            if config.classic_save_intermediate:
                for mode_name, mode_out in classic_outputs.per_mode.items():
                    vector_jobs.append(
                        VectorExportJob(
                            label=f"classic_{mode_name}",
                            mask=mode_out.mask,
                            prob_map=mode_out.prob_map,
                            transform=classic_outputs.transform,
                            crs=classic_outputs.crs,
                            out_base=_output_base_path(mode_out.mask_path),
                            prob_path=mode_out.prob_path,
                            mask_path=mode_out.mask_path,
                            gpkg_layer_name=f"classic_{mode_name}",
                        )
                    )
        if fusion_outputs is not None:
            vector_jobs.append(
                VectorExportJob(
                    label="fused",
                    mask=fusion_outputs.mask,
                    prob_map=fusion_outputs.prob_map,
                    transform=fusion_outputs.transform,
                    crs=fusion_outputs.crs,
                    out_base=_output_base_path(fusion_outputs.mask_path),
                    prob_path=fusion_outputs.prob_path,
                    mask_path=fusion_outputs.mask_path,
                    gpkg_layer_name="fused",
                )
            )
        if multi_fused_results:
            for enc_label, fused in multi_fused_results:
                vector_jobs.append(
                    VectorExportJob(
                        label=f"fused_{enc_label}",
                        mask=fused.mask,
                        prob_map=fused.prob_map,
                        transform=fused.transform,
                        crs=fused.crs,
                        out_base=_output_base_path(fused.mask_path),
                        prob_path=fused.prob_path,
                        mask_path=fused.mask_path,
                        gpkg_layer_name=f"fused_{enc_label}",
                    )
                )
        # 'fused_multi' tekrarını engelle: zaten her encoder için 'fused_{enc_label}' eklendi

        vector_ok, vector_reason = _can_vectorize_predictions()
        if config.vectorize and not vector_ok:
            LOGGER.warning("GeoPackage export kullanilamiyor: %s. Adaylar varsa tek Excel yine yazilacak.", vector_reason)

        for job in vector_jobs:
            label = job.label
            mask_arr = job.mask
            prob_arr = job.prob_map
            if mask_arr is None or prob_arr is None:
                try:
                    mask_arr, prob_arr = load_prediction_arrays(
                        prob_path=job.prob_path,
                        mask_path=job.mask_path,
                    )
                except Exception as e:
                    LOGGER.warning("Could not load rasters for vectorization (%s): %s", label, e)
                    continue

            if config.vectorize and vector_ok:
                LOGGER.info(f"  ? {label} poligonla?t?r?l?yor...")
                try:
                    gpkg_path, gpkg_layer_name = _resolve_vector_gpkg_target(
                        out_base=job.out_base,
                        layer_name=job.gpkg_layer_name or label,
                        gpkg_mode=config.gpkg_mode,
                        single_gpkg_path=single_vector_gpkg_path,
                    )
                    vector_file = vectorize_predictions(
                        mask=mask_arr,
                        prob_map=prob_arr,
                        transform=job.transform,
                        crs=job.crs,
                        out_path=job.out_base,
                        min_area=config.min_area,
                        simplify_tol=config.simplify,
                        opening_size=config.vector_opening_size,
                        label_connectivity=config.label_connectivity,
                        export_candidate_excel=False,
                        gpkg_path=gpkg_path,
                        gpkg_layer_name=gpkg_layer_name,
                    )
                except Exception as e:
                    LOGGER.warning("Vector export failed (%s): %s", label, e)
                    vector_file = None
                if vector_file:
                    LOGGER.info(
                        "    ? Vekt?r ??kt?s? (%s): %s",
                        label,
                        _format_gpkg_target(vector_file, gpkg_layer_name),
                    )
            elif config.vectorize:
                LOGGER.warning("    Vekt?r ??kt?s? atland? (%s).", vector_reason)

            if config.export_candidate_boxes:
                try:
                    box_extra = _build_combined_candidate_extra_fields(
                        source_label=label,
                        scale_level=job.scale_level,
                        scale_factor=job.scale_factor,
                    )
                    if config.vectorize and vector_ok:
                        boxes_gpkg_path, boxes_layer_name = _resolve_vector_gpkg_target(
                            out_base=job.out_base,
                            layer_name=f"{job.gpkg_layer_name or label}_candidate_boxes",
                            gpkg_mode=config.gpkg_mode,
                            single_gpkg_path=single_vector_gpkg_path,
                        )
                        box_records = write_candidate_boxes_layer_from_prediction(
                            mask=mask_arr,
                            prob_map=prob_arr,
                            transform=job.transform,
                            crs=job.crs,
                            gpkg_path=boxes_gpkg_path,
                            layer_name=boxes_layer_name,
                            min_area=config.min_area,
                            opening_size=config.vector_opening_size,
                            label_connectivity=config.label_connectivity,
                            extra_fields=box_extra,
                        )
                        LOGGER.info(
                            "    Aday kutu katmani (%s): %s (%d kutu)",
                            label,
                            _format_gpkg_target(boxes_gpkg_path, boxes_layer_name),
                            len(box_records),
                        )
                    else:
                        box_records = _build_candidate_box_records_from_prediction(
                            mask=mask_arr,
                            prob_map=prob_arr,
                            transform=job.transform,
                            crs=job.crs,
                            min_area=config.min_area,
                            opening_size=config.vector_opening_size,
                            label_connectivity=config.label_connectivity,
                            extra_fields=box_extra,
                        )
                    if config.export_candidate_excel:
                        combined_candidate_box_rows.extend(_candidate_box_records_to_rows(box_records))
                except Exception as e:
                    LOGGER.warning("Candidate box export failed (%s): %s", label, e)

            if config.export_candidate_excel:
                if config.vectorize and vector_ok:
                    LOGGER.info(f"  ? {label} adaylar? tek Excel i?in haz?rlan?yor...")
                elif config.vectorize:
                    LOGGER.info(f"  ? {label} i?in GeoPackage atland?; adaylar tek Excel'e eklenecek...")
                else:
                    LOGGER.info(f"  ? {label} i?in adaylar tek Excel'e ekleniyor...")
                try:
                    job_rows = build_candidate_location_rows_from_prediction(
                        mask=mask_arr,
                        prob_map=prob_arr,
                        transform=job.transform,
                        crs=job.crs,
                        min_area=config.min_area,
                        opening_size=config.vector_opening_size,
                        label_connectivity=config.label_connectivity,
                        extra_fields=_build_combined_candidate_extra_fields(
                            source_label=label,
                            scale_level=job.scale_level,
                            scale_factor=job.scale_factor,
                        ),
                    )
                    combined_candidate_rows.extend(job_rows)
                    LOGGER.info("    Tek Excel listesine eklenen aday sayisi (%s): %d", label, len(job_rows))
                    combined_candidate_fallback_rows.extend(
                        _fallback_rows_for_empty_candidate_result(
                            normal_rows=job_rows,
                            enabled=config.fallback_candidates_enabled,
                            prob_map=prob_arr,
                            transform=job.transform,
                            crs=job.crs,
                            min_area=config.min_area,
                            top_k=config.fallback_candidates_top_k,
                            min_score=config.fallback_candidates_min_score,
                            extra_fields=_build_combined_candidate_extra_fields(
                                source_label=label,
                                scale_level=job.scale_level,
                                scale_factor=job.scale_factor,
                            ),
                            label=label,
                        )
                    )
                except Exception as e:
                    LOGGER.warning("Combined Excel row export failed (%s): %s", label, e)

        if config.export_candidate_excel:
            table_rows = _use_fallback_rows_if_no_candidates(
                rows=combined_candidate_rows,
                fallback_rows=combined_candidate_fallback_rows,
                top_k=config.fallback_candidates_top_k,
            )
            combined_table_path = write_combined_candidate_locations_table(
                rows=table_rows,
                candidate_box_rows=combined_candidate_box_rows if config.export_candidate_boxes else None,
                out_base=_combined_candidate_table_base(out_prefix),
            )
            if combined_table_path:
                LOGGER.info("Tek birlesik Excel aday tablosu yazildi: %s", combined_table_path)
        LOGGER.info("")
        LOGGER.info("=" * 70)
        LOGGER.info("✅ TÜM İŞLEMLER TAMAMLANDI!")
        LOGGER.info("=" * 70)


if __name__ == "__main__":
    main()

# ============================================================================
# HIZLI BAŞLANGIÇ
# ============================================================================
#
# EN KOLAY YÖNTEM: Tek config.yaml dosyası! ⭐
# --------------------------------------------
# 1. config.yaml dosyasını düzenleyin:
#    - enable_deep_learning: true/false
#    - enable_classic: true/false
#    - enable_fusion: true/false
#    - th, tile, alpha ve diğer tüm ayarlar
#
# 2. Çalıştırın:
#    python archaeo_detect.py
#
# Bu kadar! Tüm ayarlar ve detaylı açıklamalar config.yaml içinde.
#
#
# KOMUT SATIRI ÖRNEKLERİ (config.yaml'ı override eder):
# -------------------------------------------------------
#
# Temel kullanım (config.yaml'dan okur):
# python archaeo_detect.py
#
# Eşik değerini değiştir:
# python archaeo_detect.py --th 0.7
#
# Karo boyutunu ve overlap'i değiştir:
# python archaeo_detect.py --tile 512 --overlap 128
#
# Fusion karışım oranını değiştir:
# python archaeo_detect.py --alpha 0.7
#
# Eğitilmiş ağırlıklarla çalıştır:
# python archaeo_detect.py --weights unet_resnet34_arch.pth
#
#
# SENARYOLAR (config.yaml içinde):
# ----------------------------------
# Senaryo 1: Sadece DL (U-Net)
#   enable_deep_learning: true, enable_classic: false, enable_yolo: false, enable_fusion: false
#
# Senaryo 2: Sadece Klasik
#   enable_deep_learning: false, enable_classic: true, enable_yolo: false, enable_fusion: false
#
# Senaryo 3: Sadece YOLO11
#   enable_deep_learning: false, enable_classic: false, enable_yolo: true, enable_fusion: false
#
# Senaryo 4: DL + Klasik + Fusion (ÖNERİLİR)
#   enable_deep_learning: true, enable_classic: true, enable_yolo: false, enable_fusion: true
#
# Senaryo 5: Tüm Yöntemler (DL + Klasik + YOLO11)
#   enable_deep_learning: true, enable_classic: true, enable_yolo: true, enable_fusion: true
#
# YOLO11 KULLANIMI:
# -----------------
# YOLO11 için segmentasyon modelini kullanmanız önerilir:
#   yolo_weights: "yolo11n-seg.pt"  # nano segmentation model
#   yolo_weights: "yolo11s-seg.pt"  # small segmentation model
#   yolo_weights: "yolo11m-seg.pt"  # medium segmentation model
#
# Tespit (detection) modeli kullanırsanız, bounding box'lardan maske oluşturulur:
#   yolo_weights: "yolo11n.pt"      # nano detection model
#
# Özel eğitilmiş modelinizi kullanabilirsiniz:
#   yolo_weights: "path/to/your/best.pt"
#
# ============================================================================
