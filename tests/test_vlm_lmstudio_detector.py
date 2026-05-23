import json
from contextlib import ExitStack
from pathlib import Path
from zipfile import ZipFile

import numpy as np
import rasterio
from rasterio.crs import CRS
from rasterio.enums import ColorInterp
from rasterio.transform import from_origin

import vlm_lmstudio_detector as vlm


def test_auto_views_follow_available_bands():
    assert vlm._resolve_views(
        "auto",
        vlm.RasterBandLayout((1, 2, 3), None, None),
        logger=vlm.LOGGER,
    ) == ["rgb"]
    assert vlm._resolve_views(
        "auto",
        vlm.RasterBandLayout((1, 2, 3), 4, 5),
        logger=vlm.LOGGER,
    ) == ["rgb", "hillshade", "ndsm", "slope"]
    assert vlm._resolve_views(
        "auto",
        vlm.RasterBandLayout((1, 2, 3), None, 4),
        logger=vlm.LOGGER,
    ) == ["rgb", "hillshade", "slope"]
    assert vlm._resolve_views(
        "auto",
        vlm.RasterBandLayout((1, 2, 3), 4, None),
        logger=vlm.LOGGER,
    ) == ["rgb", "dsm"]


def test_auto_model_uses_loaded_lmstudio_model():
    class Model:
        id = "loaded-vision-model"

    class Models:
        data = [Model()]

    class Client:
        class models:
            @staticmethod
            def list():
                return Models()

    config = vlm.VlmLmStudioConfig(model="auto")

    assert vlm._resolve_lmstudio_model(Client(), config, logger=vlm.LOGGER) == "loaded-vision-model"


def test_base_url_without_v1_is_normalized():
    assert vlm._normalize_openai_base_url("http://127.0.0.1:8081") == "http://127.0.0.1:8081/v1"
    assert vlm._normalize_openai_base_url("http://127.0.0.1:8081/v1") == "http://127.0.0.1:8081/v1"


def test_lmstudio_native_api_base_url_is_derived_from_openai_base_url():
    assert vlm._lmstudio_native_api_base_url("http://127.0.0.1:8081") == "http://127.0.0.1:8081/api/v1"
    assert vlm._lmstudio_native_api_base_url("http://127.0.0.1:8081/v1") == "http://127.0.0.1:8081/api/v1"
    assert vlm._lmstudio_native_api_base_url("http://127.0.0.1:8081/api/v1") == "http://127.0.0.1:8081/api/v1"


def test_lmstudio_reload_unloads_then_loads_model(monkeypatch):
    calls = []

    def fake_post(config, endpoint, payload):
        calls.append((endpoint, payload))
        if endpoint == "models/load":
            return {"instance_id": "vision-model"}
        return {"instance_id": payload.get("instance_id")}

    monkeypatch.setattr(vlm, "_post_lmstudio_native_json", fake_post)
    monkeypatch.setattr(vlm.time, "sleep", lambda seconds: calls.append(("sleep", seconds)))

    config = vlm.VlmLmStudioConfig(
        base_url="http://127.0.0.1:8081/v1",
        model="vision-model",
        reload_pause_seconds=2,
    )

    assert vlm._reload_lmstudio_model(config, logger=vlm.LOGGER) == "vision-model"
    assert calls == [
        ("models/unload", {"instance_id": "vision-model"}),
        ("sleep", 2.0),
        ("models/load", {"model": "vision-model"}),
    ]


def test_prompt_includes_configured_gsd_scale():
    prompt = vlm._build_prompt(
        "rgb_only",
        ["rgb"],
        1024,
        1024,
        gsd_m=0.30,
        source_kind="rgb",
    )

    assert "0.30 m ground sampling distance" in prompt
    assert "307 m x 307 m" in prompt
    assert "nadir imagery" in prompt
    assert "coherent, plausible archaeological patterns" in prompt
    assert "mounds, rings, wall or foundation lines, enclosures, terraces" in prompt
    assert "Ignore isolated shadows, vegetation, modern or agricultural lines" in prompt
    assert "clearly organized and supported by the provided views" in prompt
    assert "Cappadocia" not in prompt
    assert "Kapadokya" not in prompt


def test_review_prompt_has_general_false_positive_gate():
    prompt = vlm._build_review_prompt(
        {
            "tile_index": 1,
            "candidate_type": "mound",
            "confidence": 0.82,
            "bbox_xyxy": [1, 1, 6, 6],
            "visual_evidence": "faint oval mark",
            "possible_false_positive": "tree shadow",
            "recommended_check": "rgb",
        },
        analysis_mode="rgb_only",
        selected_views=["rgb"],
        gsd_m=0.30,
        source_kind="rgb",
    )

    assert "Review the proposed candidate skeptically" in prompt
    assert "at least two supporting cues" in prompt
    assert "archaeological explanation clearly stronger" in prompt
    assert "natural, modern, vegetation, shadow, or image-artifact" in prompt


def test_prompt_uses_external_stage1_guidance():
    prompt = vlm._build_prompt(
        "rgb_only",
        ["rgb"],
        256,
        256,
        gsd_m=0.30,
        source_kind="rgb",
        guidance_text="CUSTOM STAGE1 GUIDANCE",
    )

    assert "CUSTOM STAGE1 GUIDANCE" in prompt
    assert "Cappadocia cultural and volcanic landscape" not in prompt
    assert "Return exactly one JSON object with this schema and no markdown" in prompt
    assert vlm.JSON_SCHEMA_TEXT in prompt


def test_review_prompt_uses_external_stage2_guidance():
    prompt = vlm._build_review_prompt(
        {
            "tile_index": 1,
            "candidate_type": "mound",
            "confidence": 0.82,
            "bbox_xyxy": [1, 1, 6, 6],
            "visual_evidence": "faint oval mark",
            "possible_false_positive": "tree shadow",
            "recommended_check": "rgb",
        },
        analysis_mode="rgb_only",
        selected_views=["rgb"],
        gsd_m=0.30,
        source_kind="rgb",
        guidance_text="CUSTOM STAGE2 REVIEW",
    )

    assert "CUSTOM STAGE2 REVIEW" in prompt
    assert "Be more skeptical than the first pass" not in prompt
    assert "First-stage proposal" in prompt
    assert vlm.REVIEW_SCHEMA_TEXT in prompt


def test_hillshade_prompt_explains_grayscale_relief_source():
    prompt = vlm._build_prompt(
        "rgb_only",
        ["rgb"],
        512,
        512,
        gsd_m=0.50,
        source_kind="hillshade",
    )

    assert "hillshade / shaded-relief visualization" in prompt
    assert "replicated grayscale hillshade source, not optical RGB" in prompt
    assert "Gray tone represents illumination of terrain relief" in prompt
    assert "Do not interpret gray tone as optical color" in prompt


def test_vlm_default_confidence_threshold_is_conservative():
    assert vlm.VlmLmStudioConfig().confidence_threshold == 0.75
    assert vlm.VlmLmStudioConfig().max_candidates_per_tile == 3


def test_confidence_tier_labels_include_very_high_bucket():
    assert vlm._confidence_tier(0.96) == "very_high_095"
    assert vlm._confidence_tier(0.92) == "strong_090"
    assert vlm._confidence_tier(0.76) == "candidate_075"
    assert vlm._confidence_tier(0.42) == "below_075"
    assert vlm._confidence_tier(0.0) == "none"


def test_prompt_requests_bounded_candidate_list():
    prompt = vlm._build_prompt(
        "rgb_only",
        ["rgb"],
        512,
        512,
        gsd_m=0.30,
        source_kind="rgb",
        max_candidates_per_tile=3,
        confidence_threshold=0.75,
    )

    assert "Return at most 3 separate candidates" in prompt
    assert "Only include candidates with confidence >= 0.75" in prompt
    assert '"candidates"' in prompt


def test_parse_model_json_repairs_common_unquoted_key_glitch():
    parsed = vlm._parse_model_json(
        '{'
        '"candidates":['
        '{"confidence":0.82,"candidate_type":"wall_trace","bbox_xyxy":[1,1,3,3]},'
        '{_candidate_type:"enclosure","confidence":0.78,"candidate_type":"enclosure","bbox_xyxy":[4,4,7,7],}'
        '],'
        '"visual_evidence":"structured lines",'
        '"possible_false_positive":"field edge",'
        '"recommended_check":"rgb",'
        '}'
    )

    assert parsed["candidates"][1]["candidate_type"] == "enclosure"


def test_parse_model_json_prefers_last_valid_json_object():
    parsed = vlm._parse_model_json(
        '{"candidates":[{"confidence":0.82,"candidate_type":"wall_trace","bbox_xyxy":[1,1,3,3]}]}\n'
        "Corrected final output:\n"
        '{"candidates":[],"visual_evidence":"none","possible_false_positive":"modern houses","recommended_check":"rgb"}'
    )

    assert parsed["candidates"] == []


def test_vlm_output_paths_include_separate_stage_gpkgs(tmp_path: Path):
    paths = vlm._build_output_paths(tmp_path / "out" / "rgb")

    assert paths.gpkg_stage1.name == "rgb_vlm_stage1_positives.gpkg"
    assert paths.gpkg_stage2.name == "rgb_vlm_stage2_verified.gpkg"
    assert paths.gpkg == paths.gpkg_stage2


def test_wgs84_google_maps_fields_are_added_to_candidate_record():
    parsed = {
        "candidate": True,
        "confidence": 0.8,
        "candidate_type": "mound",
        "bbox_xyxy": [1, 1, 3, 3],
        "visual_evidence": "oval crop mark",
        "possible_false_positive": "soil variation",
        "recommended_check": "rgb",
    }
    base = {
        "tile_index": 1,
        "tile_row": 0,
        "tile_col": 0,
        "tile_width": 8,
        "tile_height": 8,
        "used_views": ["rgb"],
        "has_rgb": True,
        "has_dsm": False,
        "has_dtm": False,
        "analysis_mode": "rgb_only",
    }

    record = vlm._model_result_to_record(
        parsed=parsed,
        raw_response="{}",
        base_record=base,
        transform=from_origin(35.0, 39.0, 0.0001, 0.0001),
        crs=CRS.from_epsg(4326),
        raster_width=8,
        raster_height=8,
    )

    assert record["gps_lon"] is not None
    assert record["gps_lat"] is not None
    assert record["google_maps_url"].startswith("https://www.google.com/maps?q=")


def test_multi_candidate_model_result_returns_distinct_candidate_records():
    parsed = {
        "candidates": [
            {
                "confidence": 0.91,
                "candidate_type": "mound",
                "bbox_xyxy": [1, 1, 4, 4],
                "visual_evidence": "first oval mark",
                "possible_false_positive": "soil variation",
                "recommended_check": "rgb",
            },
            {
                "confidence": 0.86,
                "candidate_type": "ring_ditch",
                "bbox_xyxy": [5, 5, 7, 7],
                "visual_evidence": "second ring mark",
                "possible_false_positive": "field mark",
                "recommended_check": "rgb",
            },
        ]
    }
    base = {
        "tile_index": 1,
        "tile_row": 0,
        "tile_col": 0,
        "tile_width": 8,
        "tile_height": 8,
        "used_views": ["rgb"],
        "has_rgb": True,
        "has_dsm": False,
        "has_dtm": False,
        "analysis_mode": "rgb_only",
    }

    records = vlm._model_result_to_records(
        parsed=parsed,
        raw_response="{}",
        base_record=base,
        transform=from_origin(0, 8, 1, 1),
        crs=CRS.from_epsg(3857),
        raster_width=8,
        raster_height=8,
        max_candidates_per_tile=3,
    )

    assert [record["candidate_index"] for record in records] == [1, 2]
    assert [record["candidate_count"] for record in records] == [2, 2]
    assert [record["candidate_type"] for record in records] == ["mound", "ring_ditch"]


def test_multi_candidate_model_result_dedupes_overlapping_boxes():
    parsed = {
        "candidates": [
            {
                "confidence": 0.92,
                "candidate_type": "mound",
                "bbox_xyxy": [1, 1, 6, 6],
                "visual_evidence": "strong oval mark",
                "possible_false_positive": "soil variation",
                "recommended_check": "rgb",
            },
            {
                "confidence": 0.82,
                "candidate_type": "mound",
                "bbox_xyxy": [1.2, 1.2, 6.1, 6.1],
                "visual_evidence": "same oval mark",
                "possible_false_positive": "soil variation",
                "recommended_check": "rgb",
            },
        ]
    }
    base = {
        "tile_index": 1,
        "tile_row": 0,
        "tile_col": 0,
        "tile_width": 8,
        "tile_height": 8,
        "used_views": ["rgb"],
        "has_rgb": True,
        "has_dsm": False,
        "has_dtm": False,
        "analysis_mode": "rgb_only",
    }

    records = vlm._model_result_to_records(
        parsed=parsed,
        raw_response="{}",
        base_record=base,
        transform=from_origin(0, 8, 1, 1),
        crs=CRS.from_epsg(3857),
        raster_width=8,
        raster_height=8,
        max_candidates_per_tile=3,
    )

    assert len(records) == 1
    assert records[0]["confidence"] == 0.92
    assert records[0]["candidate_count"] == 1


def test_candidate_xlsx_google_maps_url_is_clickable(tmp_path: Path):
    from openpyxl import load_workbook

    maps_url = "https://www.google.com/maps?q=39.00000000,35.00000000"
    xlsx_path = tmp_path / "vlm_candidates.xlsx"

    vlm._write_candidate_xlsx(
        xlsx_path,
        [
            {
                "tile_index": 1,
                "candidate": True,
                "confidence": 0.8,
                "candidate_type": "mound",
                "google_maps_url": maps_url,
                "status": "ok",
            }
        ],
        logger=vlm.LOGGER,
    )

    wb = load_workbook(xlsx_path)
    ws = wb["vlm_candidates"]
    google_maps_col = vlm.CSV_COLUMNS.index("google_maps_url") + 1
    cell = ws.cell(row=2, column=google_maps_col)

    assert cell.value == maps_url
    assert cell.hyperlink is not None
    assert cell.hyperlink.target == maps_url
    assert cell.style == "Hyperlink"


def test_candidate_xlsx_sorts_found_and_adds_not_found_sheet(tmp_path: Path):
    from openpyxl import load_workbook

    xlsx_path = tmp_path / "vlm_candidates.xlsx"
    low_confidence = {
        "tile_index": 2,
        "tile_row": 0,
        "tile_col": 8,
        "candidate": True,
        "confidence": 0.62,
        "review_confirmed": True,
        "review_confidence": 0.62,
        "candidate_type": "mound",
        "google_maps_url": "",
        "status": "ok",
        "geometry": {"type": "Point", "coordinates": [1, 1]},
    }
    high_confidence = {
        "tile_index": 1,
        "tile_row": 0,
        "tile_col": 0,
        "candidate": True,
        "confidence": 0.96,
        "review_confirmed": True,
        "review_confidence": 0.96,
        "candidate_type": "ring_ditch",
        "google_maps_url": "",
        "status": "ok",
        "geometry": {"type": "Point", "coordinates": [0, 0]},
    }
    no_candidate = {
        "tile_index": 3,
        "tile_row": 8,
        "tile_col": 0,
        "candidate": False,
        "confidence": 0.0,
        "candidate_type": "none",
        "google_maps_url": "",
        "status": "ok",
    }
    below_threshold = {
        "tile_index": 4,
        "tile_row": 8,
        "tile_col": 8,
        "candidate": True,
        "confidence": 0.31,
        "review_confirmed": False,
        "review_confidence": 0.0,
        "candidate_type": "unknown",
        "google_maps_url": "",
        "status": "ok",
        "geometry": {"type": "Point", "coordinates": [2, 2]},
    }

    vlm._write_candidate_xlsx(
        xlsx_path,
        [low_confidence, high_confidence],
        all_records=[low_confidence, high_confidence, no_candidate, below_threshold],
        confidence_threshold=0.5,
        logger=vlm.LOGGER,
    )

    wb = load_workbook(xlsx_path)
    found_ws = wb["vlm_candidates"]
    first_stage_ws = wb["ilk_asama_pozitifler"]
    second_stage_ws = wb["ikinci_asama_pozitifler"]
    very_high_ws = wb["kesin_095"]
    missing_ws = wb["bulunmayan_tilelar"]
    tile_col = vlm.CSV_COLUMNS.index("tile_index") + 1
    review_tier_col = vlm.CSV_COLUMNS.index("review_confidence_tier") + 1
    reason_col = vlm.NOT_FOUND_COLUMNS.index("not_found_reason") + 1

    assert [found_ws.cell(row=2, column=tile_col).value, found_ws.cell(row=3, column=tile_col).value] == [1, 2]
    assert [first_stage_ws.cell(row=2, column=tile_col).value, first_stage_ws.cell(row=3, column=tile_col).value] == [1, 2]
    assert [second_stage_ws.cell(row=2, column=tile_col).value, second_stage_ws.cell(row=3, column=tile_col).value] == [1, 2]
    assert very_high_ws.max_row == 2
    assert very_high_ws.cell(row=2, column=tile_col).value == 1
    assert very_high_ws.cell(row=2, column=review_tier_col).value == "very_high_095"
    assert [missing_ws.cell(row=2, column=tile_col).value, missing_ws.cell(row=3, column=tile_col).value] == [3, 4]
    assert [
        missing_ws.cell(row=2, column=reason_col).value,
        missing_ws.cell(row=3, column=reason_col).value,
    ] == ["no_candidate", "below_threshold"]
    assert found_ws.auto_filter.ref is None
    assert first_stage_ws.auto_filter.ref is None
    assert second_stage_ws.auto_filter.ref is None
    assert missing_ws.auto_filter.ref is None

    with ZipFile(xlsx_path) as archive:
        sheet1_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
        sheet2_xml = archive.read("xl/worksheets/sheet2.xml").decode("utf-8")
        sheet3_xml = archive.read("xl/worksheets/sheet3.xml").decode("utf-8")
        sheet4_xml = archive.read("xl/worksheets/sheet4.xml").decode("utf-8")

    assert "<autoFilter" not in sheet1_xml
    assert "<autoFilter" not in sheet2_xml
    assert "<autoFilter" not in sheet3_xml
    assert "<autoFilter" not in sheet4_xml


def test_xlsx_save_uses_alternative_name_when_excel_file_is_locked(tmp_path: Path):
    class FakeWorkbook:
        def __init__(self) -> None:
            self.save_calls: list[Path] = []

        def save(self, path: Path) -> None:
            path = Path(path)
            self.save_calls.append(path)
            if path.name in {"vlm_candidates.xlsx", "vlm_candidates_alternatif.xlsx"}:
                raise PermissionError("file is locked")
            path.write_text("saved", encoding="utf-8")

    workbook = FakeWorkbook()
    target = tmp_path / "vlm_candidates.xlsx"

    saved_path = vlm._save_workbook_with_excel_lock_fallback(workbook, target, logger=vlm.LOGGER)

    assert saved_path == tmp_path / "vlm_candidates_alternatif_2.xlsx"
    assert saved_path.read_text(encoding="utf-8") == "saved"
    assert [path.name for path in workbook.save_calls] == [
        "vlm_candidates.xlsx",
        "vlm_candidates_alternatif.xlsx",
        "vlm_candidates_alternatif_2.xlsx",
    ]


def test_gpkg_write_uses_alternative_name_when_file_is_locked(tmp_path: Path):
    calls: list[Path] = []
    target = tmp_path / "vlm_candidates.gpkg"

    def fake_writer(path: Path) -> None:
        path = Path(path)
        calls.append(path)
        if path.name in {"vlm_candidates.gpkg", "vlm_candidates_alternatif.gpkg"}:
            raise RuntimeError("database is locked")
        path.write_text("saved", encoding="utf-8")

    saved_path = vlm._write_gpkg_with_lock_fallback(target, fake_writer, logger=vlm.LOGGER)

    assert saved_path == tmp_path / "vlm_candidates_alternatif_2.gpkg"
    assert saved_path.read_text(encoding="utf-8") == "saved"
    assert [path.name for path in calls] == [
        "vlm_candidates.gpkg",
        "vlm_candidates_alternatif.gpkg",
        "vlm_candidates_alternatif_2.gpkg",
    ]


def test_candidate_outputs_write_separate_stage_gpkgs(tmp_path: Path, monkeypatch):
    paths = vlm._build_output_paths(tmp_path / "out" / "rgb")
    first_stage_record = {
        "tile_index": 1,
        "candidate": True,
        "confidence": 0.91,
        "candidate_type": "mound",
        "geometry": {"type": "Polygon", "coordinates": [[[0, 0], [1, 0], [1, 1], [0, 0]]]},
    }
    second_stage_record = {
        **first_stage_record,
        "review_confirmed": True,
        "review_confidence": 0.88,
    }
    gpkg_calls = []

    def fake_write_candidate_gpkg(path, records, crs, *, logger, main_layer_name="vlm_candidates", type_layer_prefix="vlm"):
        gpkg_calls.append(
            {
                "path": Path(path),
                "count": len(records),
                "main_layer_name": main_layer_name,
                "type_layer_prefix": type_layer_prefix,
            }
        )
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        Path(path).write_text("fake gpkg", encoding="utf-8")
        return Path(path)

    monkeypatch.setattr(vlm, "_write_candidate_gpkg", fake_write_candidate_gpkg)

    updated_paths = vlm._write_candidate_outputs(
        paths,
        [second_stage_record],
        None,
        first_stage_records=[first_stage_record],
        all_records=[first_stage_record],
        confidence_threshold=0.75,
        logger=vlm.LOGGER,
    )

    assert [call["path"].name for call in gpkg_calls] == [
        "rgb_vlm_stage1_positives.gpkg",
        "rgb_vlm_stage2_verified.gpkg",
    ]
    assert [call["count"] for call in gpkg_calls] == [1, 1]
    assert [call["main_layer_name"] for call in gpkg_calls] == [
        "all_vlm_stage1_positives",
        "all_vlm_stage2_verified",
    ]
    assert updated_paths.gpkg == updated_paths.gpkg_stage2


def test_candidate_type_layers_start_with_type_name(tmp_path: Path, monkeypatch):
    layer_calls = []

    def fake_write_layer(path, records, crs, *, layer_name):
        layer_calls.append(layer_name)

    monkeypatch.setattr(vlm, "_write_candidate_gpkg_layer", fake_write_layer)

    vlm._write_candidate_gpkg_to_path(
        tmp_path / "candidates.gpkg",
        [
            {
                "candidate_type": "mound",
                "geometry": {"type": "Polygon", "coordinates": [[[0, 0], [1, 0], [1, 1], [0, 0]]]},
            },
            {
                "candidate_type": "ring_ditch",
                "geometry": {"type": "Polygon", "coordinates": [[[2, 2], [3, 2], [3, 3], [2, 2]]]},
            },
        ],
        None,
        main_layer_name="all_vlm_stage1_positives",
        type_layer_prefix="vlm_stage1",
    )

    assert layer_calls == [
        "all_vlm_stage1_positives",
        "mound_vlm_stage1",
        "ring_ditch_vlm_stage1",
    ]


def test_rgb_alpha_raster_is_not_treated_as_dsm(tmp_path: Path):
    tif_path = tmp_path / "rgba.tif"
    data = np.ones((4, 8, 8), dtype=np.uint8) * 255
    with rasterio.open(
        tif_path,
        "w",
        driver="GTiff",
        width=8,
        height=8,
        count=4,
        dtype="uint8",
        transform=from_origin(0, 8, 1, 1),
    ) as dst:
        dst.write(data)
        dst.colorinterp = (ColorInterp.red, ColorInterp.green, ColorInterp.blue, ColorInterp.alpha)

    with rasterio.open(tif_path) as src:
        layout = vlm._detect_band_layout(src, band_indexes=(1, 2, 3, 4, 5), logger=vlm.LOGGER)

    assert layout.has_rgb is True
    assert layout.has_dsm is False
    assert layout.has_dtm is False
    assert layout.analysis_mode == "rgb_only"


def test_single_band_raster_can_be_reused_as_vlm_rgb(tmp_path: Path):
    tif_path = tmp_path / "hillshade_single_band.tif"
    data = np.arange(64, dtype=np.uint8).reshape(1, 8, 8)
    with rasterio.open(
        tif_path,
        "w",
        driver="GTiff",
        width=8,
        height=8,
        count=1,
        dtype="uint8",
        transform=from_origin(0, 8, 1, 1),
    ) as dst:
        dst.write(data)

    with rasterio.open(tif_path) as src:
        layout = vlm._detect_band_layout(src, band_indexes=(1, 1, 1), logger=vlm.LOGGER)
        arrays = vlm._read_tile_arrays(src, layout, rasterio.windows.Window(0, 0, 8, 8))

    assert layout.rgb == (1, 1, 1)
    assert layout.analysis_mode == "rgb_only"
    assert len(arrays["rgb"]) == 3
    assert np.array_equal(arrays["rgb"][0], arrays["rgb"][1])
    assert np.array_equal(arrays["rgb"][1], arrays["rgb"][2])


def test_source_kind_auto_detects_hillshade_name_for_single_band(tmp_path: Path):
    tif_path = tmp_path / "karlik_dag_dsm_hillshade.tif"
    data = np.arange(64, dtype=np.uint8).reshape(1, 8, 8)
    with rasterio.open(
        tif_path,
        "w",
        driver="GTiff",
        width=8,
        height=8,
        count=1,
        dtype="uint8",
        transform=from_origin(0, 8, 1, 1),
    ) as dst:
        dst.write(data)

    with rasterio.open(tif_path) as src:
        layout = vlm._detect_band_layout(src, band_indexes=(1, 1, 1), logger=vlm.LOGGER)
        source_kind = vlm._resolve_source_kind(
            "auto",
            input_path=tif_path,
            src=src,
            layout=layout,
            selected_views=["rgb"],
            logger=vlm.LOGGER,
        )

    assert source_kind == "hillshade"


def test_source_kind_auto_detects_relief_derivative_names(tmp_path: Path):
    for filename, expected in [
        ("area_slrm.tif", "slrm"),
        ("area_svf.tif", "svf"),
        ("area_svm.tif", "svf"),
        ("area_slope.tif", "slope"),
    ]:
        tif_path = tmp_path / filename
        data = np.arange(64, dtype=np.uint8).reshape(1, 8, 8)
        with rasterio.open(
            tif_path,
            "w",
            driver="GTiff",
            width=8,
            height=8,
            count=1,
            dtype="uint8",
            transform=from_origin(0, 8, 1, 1),
        ) as dst:
            dst.write(data)

        with rasterio.open(tif_path) as src:
            layout = vlm._detect_band_layout(src, band_indexes=(1, 1, 1), logger=vlm.LOGGER)
            source_kind = vlm._resolve_source_kind(
                "auto",
                input_path=tif_path,
                src=src,
                layout=layout,
                selected_views=["rgb"],
                logger=vlm.LOGGER,
            )

        assert source_kind == expected


def test_folder_input_discovers_rgb_and_relief_views(tmp_path: Path):
    transform = from_origin(0, 8, 1, 1)
    rgb_path = tmp_path / "area_rgb.tif"
    rgb = np.ones((3, 8, 8), dtype=np.uint8) * 100
    with rasterio.open(
        rgb_path,
        "w",
        driver="GTiff",
        width=8,
        height=8,
        count=3,
        dtype="uint8",
        transform=transform,
    ) as dst:
        dst.write(rgb)
        dst.colorinterp = (ColorInterp.red, ColorInterp.green, ColorInterp.blue)

    for name in ("area_hillshade.tif", "area_slrm.tif", "area_svm.tif", "area_slope.tif"):
        with rasterio.open(
            tmp_path / name,
            "w",
            driver="GTiff",
            width=8,
            height=8,
            count=1,
            dtype="uint8",
            transform=transform,
        ) as dst:
            dst.write(np.arange(64, dtype=np.uint8).reshape(1, 8, 8))

    plan = vlm._build_raster_input_plan(tmp_path, logger=vlm.LOGGER)
    assert plan.reference_path == rgb_path
    assert [aux.view for aux in plan.auxiliary_views] == ["hillshade", "slrm", "svf", "slope"]

    with rasterio.open(plan.reference_path) as src:
        layout = vlm._detect_band_layout(src, band_indexes=None, logger=vlm.LOGGER)
        layout = vlm.replace(layout, external_views=tuple(aux.view for aux in plan.auxiliary_views))
        assert vlm._resolve_views("auto", layout, logger=vlm.LOGGER) == ["rgb", "hillshade", "slrm", "svf", "slope"]
        assert layout.analysis_mode == "rgb_topo"


def test_folder_aux_view_with_non_overlapping_bounds_is_skipped(tmp_path: Path):
    rgb_path = tmp_path / "area_rgb.tif"
    slope_path = tmp_path / "other_area_slope.tif"
    with rasterio.open(
        rgb_path,
        "w",
        driver="GTiff",
        width=8,
        height=8,
        count=3,
        dtype="uint8",
        transform=from_origin(0, 8, 1, 1),
    ) as dst:
        dst.write(np.ones((3, 8, 8), dtype=np.uint8) * 100)
        dst.colorinterp = (ColorInterp.red, ColorInterp.green, ColorInterp.blue)

    with rasterio.open(
        slope_path,
        "w",
        driver="GTiff",
        width=8,
        height=8,
        count=1,
        dtype="uint8",
        transform=from_origin(1000, 1008, 1, 1),
    ) as dst:
        dst.write(np.arange(64, dtype=np.uint8).reshape(1, 8, 8))

    plan = vlm._build_raster_input_plan(tmp_path, logger=vlm.LOGGER)
    with ExitStack() as stack:
        src = stack.enter_context(rasterio.open(plan.reference_path))
        opened = vlm._open_auxiliary_raster_views(plan.auxiliary_views, src, stack=stack, logger=vlm.LOGGER)

    assert [aux.view for aux in plan.auxiliary_views] == ["slope"]
    assert opened == ()


def test_folder_input_sends_discovered_views_to_vlm(tmp_path: Path, monkeypatch):
    transform = from_origin(0, 8, 1, 1)
    with rasterio.open(
        tmp_path / "area_rgb.tif",
        "w",
        driver="GTiff",
        width=8,
        height=8,
        count=3,
        dtype="uint8",
        transform=transform,
    ) as dst:
        dst.write(np.ones((3, 8, 8), dtype=np.uint8) * 120)
        dst.colorinterp = (ColorInterp.red, ColorInterp.green, ColorInterp.blue)

    for offset, name in enumerate(("area_hillshade.tif", "area_slrm.tif", "area_svm.tif", "area_slope.tif"), start=1):
        data = (np.arange(64, dtype=np.uint8).reshape(1, 8, 8) + offset).astype(np.uint8)
        with rasterio.open(
            tmp_path / name,
            "w",
            driver="GTiff",
            width=8,
            height=8,
            count=1,
            dtype="uint8",
            transform=transform,
        ) as dst:
            dst.write(data)

    calls = []

    def fake_request(**kwargs):
        calls.append(
            {
                "prompt": kwargs["prompt"],
                "views": [view for view, _ in kwargs["rendered_views"]],
            }
        )
        return (
            '{"candidates":[],"visual_evidence":"none","possible_false_positive":"natural texture","recommended_check":"rgb"}',
            True,
        )

    monkeypatch.setattr(vlm, "_make_openai_client", lambda config: object())
    monkeypatch.setattr(vlm, "_resolve_lmstudio_model", lambda client, config, logger: "loaded-vision-model")
    monkeypatch.setattr(vlm, "_request_vlm_json", fake_request)

    summary = vlm.run_vlm_lmstudio_detection(
        input_path=tmp_path,
        out_prefix=tmp_path / "out" / "folder",
        config=vlm.VlmLmStudioConfig(
            tile=8,
            overlap=0,
            max_tiles=1,
            resume=False,
            featureless_std_threshold=0.0,
        ),
    )

    assert summary.processed_tiles == 1
    assert calls[0]["views"] == ["rgb", "hillshade", "slrm", "svf", "slope"]
    assert "Provided views: rgb, hillshade, slrm, svf, slope" in calls[0]["prompt"]
    assert "SLRM" in calls[0]["prompt"]


def test_empty_rgb_tile_is_skipped_before_model_call():
    rgb = [np.full((8, 8), np.nan, dtype=np.float32) for _ in range(3)]

    assert vlm._is_empty_rgb_tile(rgb) is True


def test_normalized_single_band_rgb_is_not_treated_as_empty():
    gradient = np.linspace(0.1, 0.9, 64, dtype=np.float32).reshape(8, 8)
    rgb = [gradient, gradient, gradient]

    assert vlm._is_empty_rgb_tile(rgb) is False


def test_bad_json_response_is_recorded_with_raw_response(tmp_path: Path, monkeypatch):
    tif_path = tmp_path / "rgb.tif"
    data = np.ones((3, 8, 8), dtype=np.uint8) * 100
    with rasterio.open(
        tif_path,
        "w",
        driver="GTiff",
        width=8,
        height=8,
        count=3,
        dtype="uint8",
        transform=from_origin(0, 8, 1, 1),
    ) as dst:
        dst.write(data)

    monkeypatch.setattr(vlm, "_make_openai_client", lambda config: object())
    monkeypatch.setattr(vlm, "_resolve_lmstudio_model", lambda client, config, logger: "loaded-vision-model")
    monkeypatch.setattr(vlm, "_request_vlm_json", lambda **kwargs: ("not json", True))

    summary = vlm.run_vlm_lmstudio_detection(
        input_path=tif_path,
        out_prefix=tmp_path / "out" / "rgb",
        config=vlm.VlmLmStudioConfig(tile=8, overlap=0, max_tiles=1, featureless_std_threshold=0.0),
    )

    assert summary.processed_tiles == 1
    assert summary.candidate_count == 0
    assert summary.error_count == 1
    assert summary.paths.jsonl.exists()
    assert summary.paths.csv.exists()
    assert summary.paths.xlsx.exists()
    assert summary.paths.geojson.exists()
    assert "raw_response" in summary.paths.raw_errors_jsonl.read_text(encoding="utf-8")


def test_incomplete_a1_json_response_is_retried_once(tmp_path: Path, monkeypatch):
    tif_path = tmp_path / "rgb_retry.tif"
    data = np.ones((3, 8, 8), dtype=np.uint8) * 100
    with rasterio.open(
        tif_path,
        "w",
        driver="GTiff",
        width=8,
        height=8,
        count=3,
        dtype="uint8",
        transform=from_origin(0, 8, 1, 1),
    ) as dst:
        dst.write(data)

    responses = [
        '{"candidates":[{"confidence":0.85,"candidate_type":"terrace","bbox_xyxy":[0,0,3,3],'
        '"visual_evidence":"linear terrace","possible_false_positive":"Natu',
        '{"candidates":[],"visual_evidence":"none","possible_false_positive":"natural texture","recommended_check":"rgb"}',
    ]

    def fake_request(**kwargs):
        return responses.pop(0), True

    monkeypatch.setattr(vlm, "_make_openai_client", lambda config: object())
    monkeypatch.setattr(vlm, "_resolve_lmstudio_model", lambda client, config, logger: "loaded-vision-model")
    monkeypatch.setattr(vlm, "_request_vlm_json", fake_request)

    summary = vlm.run_vlm_lmstudio_detection(
        input_path=tif_path,
        out_prefix=tmp_path / "out" / "rgb",
        config=vlm.VlmLmStudioConfig(tile=8, overlap=0, max_tiles=1, featureless_std_threshold=0.0),
    )

    assert responses == []
    assert summary.processed_tiles == 1
    assert summary.error_count == 0
    assert summary.candidate_count == 0


def test_vlm_reload_runs_between_tiles(tmp_path: Path, monkeypatch):
    tif_path = tmp_path / "rgb_two_tiles_reload.tif"
    data = np.ones((3, 8, 16), dtype=np.uint8) * 120
    with rasterio.open(
        tif_path,
        "w",
        driver="GTiff",
        width=16,
        height=8,
        count=3,
        dtype="uint8",
        transform=from_origin(0, 8, 1, 1),
    ) as dst:
        dst.write(data)

    reload_calls = []

    monkeypatch.setattr(vlm, "_make_openai_client", lambda config: object())
    monkeypatch.setattr(vlm, "_resolve_lmstudio_model", lambda client, config, logger: "loaded-vision-model")
    monkeypatch.setattr(
        vlm,
        "_request_vlm_json",
        lambda **kwargs: (
            '{"candidate":false,"confidence":0,"candidate_type":"none","bbox_xyxy":null,'
            '"visual_evidence":"","possible_false_positive":"","recommended_check":"rgb"}',
            True,
        ),
    )

    def fake_reload(config, *, logger):
        reload_calls.append(config.model)
        return config.model

    monkeypatch.setattr(vlm, "_reload_lmstudio_model", fake_reload)

    summary = vlm.run_vlm_lmstudio_detection(
        input_path=tif_path,
        out_prefix=tmp_path / "out" / "rgb",
        config=vlm.VlmLmStudioConfig(
            tile=8,
            overlap=0,
            max_tiles=2,
            export_every=0,
            resume=False,
            reload_every_tiles=1,
            reload_pause_seconds=0,
        ),
    )

    assert summary.processed_tiles == 2
    assert reload_calls == ["loaded-vision-model"]


def test_exportable_vlm_candidate_is_reviewed_before_final_export(tmp_path: Path, monkeypatch):
    from openpyxl import load_workbook

    tif_path = tmp_path / "rgb_candidate.tif"
    data = np.ones((3, 8, 8), dtype=np.uint8) * 120
    with rasterio.open(
        tif_path,
        "w",
        driver="GTiff",
        width=8,
        height=8,
        count=3,
        dtype="uint8",
        transform=from_origin(0, 8, 1, 1),
    ) as dst:
        dst.write(data)

    prompts: list[str] = []

    def fake_request(**kwargs):
        prompt = kwargs["prompt"]
        prompts.append(prompt)
        if "Second-pass review" in prompt:
            return (
                '{"confirmed":true,"review_confidence":0.88,'
                '"review_reason":"coherent oval form persists after skeptical review",'
                '"review_false_positive":"field mark","recommended_check":"field_check"}',
                True,
            )
        return (
            '{"candidate":true,"confidence":0.91,"candidate_type":"mound",'
            '"bbox_xyxy":[1,1,6,6],"visual_evidence":"oval raised mark",'
            '"possible_false_positive":"field mark","recommended_check":"hillshade"}',
            True,
        )

    monkeypatch.setattr(vlm, "_make_openai_client", lambda config: object())
    monkeypatch.setattr(vlm, "_resolve_lmstudio_model", lambda client, config, logger: "loaded-vision-model")
    monkeypatch.setattr(vlm, "_request_vlm_json", fake_request)

    summary = vlm.run_vlm_lmstudio_detection(
        input_path=tif_path,
        out_prefix=tmp_path / "out" / "rgb",
        config=vlm.VlmLmStudioConfig(tile=8, overlap=0, max_tiles=1, confidence_threshold=0.75, featureless_std_threshold=0.0),
    )

    assert len(prompts) == 2
    assert "Second-pass review" in prompts[1]
    assert summary.candidate_count == 1

    record = json.loads(summary.paths.jsonl.read_text(encoding="utf-8").strip())
    assert record["candidate"] is True
    assert record["review_confirmed"] is True
    assert record["review_confidence"] == 0.88

    wb = load_workbook(summary.paths.xlsx)
    assert wb["ilk_asama_pozitifler"].max_row == 2
    assert wb["ikinci_asama_pozitifler"].max_row == 2


def test_multi_candidate_tile_reviews_and_exports_each_candidate(tmp_path: Path, monkeypatch):
    tif_path = tmp_path / "rgb_multi_candidate.tif"
    data = np.ones((3, 8, 8), dtype=np.uint8) * 120
    with rasterio.open(
        tif_path,
        "w",
        driver="GTiff",
        width=8,
        height=8,
        count=3,
        dtype="uint8",
        transform=from_origin(0, 8, 1, 1),
    ) as dst:
        dst.write(data)

    prompts: list[str] = []

    def fake_request(**kwargs):
        prompt = kwargs["prompt"]
        prompts.append(prompt)
        if "Second-pass review" in prompt:
            return (
                '{"confirmed":true,"review_confidence":0.84,'
                '"review_reason":"confirmed distinct candidate","review_false_positive":"field mark",'
                '"recommended_check":"field_check"}',
                True,
            )
        return (
            '{"candidates":['
            '{"confidence":0.91,"candidate_type":"mound","bbox_xyxy":[1,1,3,3],'
            '"visual_evidence":"first oval mark","possible_false_positive":"soil mark","recommended_check":"rgb"},'
            '{"confidence":0.88,"candidate_type":"ring_ditch","bbox_xyxy":[5,5,7,7],'
            '"visual_evidence":"second ring mark","possible_false_positive":"field mark","recommended_check":"rgb"}'
            '],"visual_evidence":"","possible_false_positive":"","recommended_check":"rgb"}',
            True,
        )

    monkeypatch.setattr(vlm, "_make_openai_client", lambda config: object())
    monkeypatch.setattr(vlm, "_resolve_lmstudio_model", lambda client, config, logger: "loaded-vision-model")
    monkeypatch.setattr(vlm, "_request_vlm_json", fake_request)

    summary = vlm.run_vlm_lmstudio_detection(
        input_path=tif_path,
        out_prefix=tmp_path / "out" / "rgb",
        config=vlm.VlmLmStudioConfig(tile=8, overlap=0, max_tiles=1, confidence_threshold=0.75, featureless_std_threshold=0.0),
    )

    assert len(prompts) == 3
    assert summary.processed_tiles == 1
    assert summary.raw_candidate_count == 2
    assert summary.candidate_count == 2

    records = [json.loads(line) for line in summary.paths.jsonl.read_text(encoding="utf-8").splitlines()]
    assert [record["candidate_index"] for record in records] == [1, 2]
    assert [record["candidate_count"] for record in records] == [2, 2]
    assert all(record["review_confirmed"] is True for record in records)


def test_resume_reviews_unreviewed_first_stage_candidates(tmp_path: Path, monkeypatch):
    tif_path = tmp_path / "rgb_resume_candidate.tif"
    data = np.ones((3, 8, 8), dtype=np.uint8) * 120
    with rasterio.open(
        tif_path,
        "w",
        driver="GTiff",
        width=8,
        height=8,
        count=3,
        dtype="uint8",
        transform=from_origin(0, 8, 1, 1),
    ) as dst:
        dst.write(data)

    resume_path = tmp_path / "previous_vlm_candidates.jsonl"
    prompt_fingerprint = vlm._prompt_fingerprint(None, None)
    resume_record = {
        "tile_index": 1,
        "tile_row": 0,
        "tile_col": 0,
        "tile_width": 8,
        "tile_height": 8,
        "used_views": ["rgb"],
        "has_rgb": True,
        "has_dsm": False,
        "has_dtm": False,
        "analysis_mode": "rgb_only",
        "prompt_fingerprint": prompt_fingerprint,
        "candidate": True,
        "confidence": 0.91,
        "candidate_type": "mound",
        "bbox_xyxy": [1, 1, 6, 6],
        "bbox_global_xyxy": [1, 1, 6, 6],
        "bbox_crs_xyxy": [1, 2, 6, 7],
        "center_x": 3.5,
        "center_y": 4.5,
        "gps_lon": None,
        "gps_lat": None,
        "google_maps_url": "",
        "visual_evidence": "oval raised mark",
        "possible_false_positive": "field mark",
        "recommended_check": "hillshade",
        "status": "ok",
        "error_type": None,
        "error_message": None,
        "geometry": {
            "type": "Polygon",
            "coordinates": [[[1, 7], [6, 7], [6, 2], [1, 2], [1, 7]]],
        },
    }
    resume_path.write_text(json.dumps(resume_record) + "\n", encoding="utf-8")

    calls = {"count": 0}

    def fake_request(**kwargs):
        calls["count"] += 1
        assert "Second-pass review" in kwargs["prompt"]
        return (
            '{"confirmed":true,"review_confidence":0.86,'
            '"review_reason":"confirmed on review","review_false_positive":"field mark",'
            '"recommended_check":"field_check"}',
            True,
        )

    monkeypatch.setattr(vlm, "_make_openai_client", lambda config: object())
    monkeypatch.setattr(vlm, "_resolve_lmstudio_model", lambda client, config, logger: "loaded-vision-model")
    monkeypatch.setattr(vlm, "_request_vlm_json", fake_request)

    summary = vlm.run_vlm_lmstudio_detection(
        input_path=tif_path,
        out_prefix=tmp_path / "out" / "rgb",
        config=vlm.VlmLmStudioConfig(
            tile=8,
            overlap=0,
            max_tiles=1,
            resume=True,
            resume_jsonl_path=resume_path,
        ),
    )

    assert calls["count"] == 1
    assert summary.resumed_tiles == 1
    assert summary.candidate_count == 1
    updated_record = json.loads(summary.paths.jsonl.read_text(encoding="utf-8").strip().splitlines()[-1])
    assert updated_record["review_confirmed"] is True
    assert updated_record["review_confidence"] == 0.86


def test_partial_resume_continues_new_tiles_before_reviewing_old_candidates(tmp_path: Path, monkeypatch):
    tif_path = tmp_path / "rgb_two_tiles_resume_candidate.tif"
    data = np.ones((3, 8, 16), dtype=np.uint8) * 120
    with rasterio.open(
        tif_path,
        "w",
        driver="GTiff",
        width=16,
        height=8,
        count=3,
        dtype="uint8",
        transform=from_origin(0, 8, 1, 1),
    ) as dst:
        dst.write(data)

    resume_path = tmp_path / "previous_vlm_candidates.jsonl"
    prompt_fingerprint = vlm._prompt_fingerprint(None, None)
    resume_record = {
        "tile_index": 1,
        "tile_row": 0,
        "tile_col": 0,
        "tile_width": 8,
        "tile_height": 8,
        "used_views": ["rgb"],
        "has_rgb": True,
        "has_dsm": False,
        "has_dtm": False,
        "analysis_mode": "rgb_only",
        "prompt_fingerprint": prompt_fingerprint,
        "candidate": True,
        "confidence": 0.91,
        "candidate_type": "mound",
        "bbox_xyxy": [1, 1, 6, 6],
        "bbox_global_xyxy": [1, 1, 6, 6],
        "bbox_crs_xyxy": [1, 2, 6, 7],
        "center_x": 3.5,
        "center_y": 4.5,
        "gps_lon": None,
        "gps_lat": None,
        "google_maps_url": "",
        "visual_evidence": "oval raised mark",
        "possible_false_positive": "field mark",
        "recommended_check": "hillshade",
        "status": "ok",
        "error_type": None,
        "error_message": None,
        "geometry": {
            "type": "Polygon",
            "coordinates": [[[1, 7], [6, 7], [6, 2], [1, 2], [1, 7]]],
        },
    }
    resume_path.write_text(json.dumps(resume_record) + "\n", encoding="utf-8")

    prompts: list[str] = []

    def fake_request(**kwargs):
        prompt = kwargs["prompt"]
        prompts.append(prompt)
        if "Second-pass review" in prompt:
            return (
                '{"confirmed":true,"review_confidence":0.86,'
                '"review_reason":"confirmed on deferred review","review_false_positive":"field mark",'
                '"recommended_check":"field_check"}',
                True,
            )
        return (
            '{"candidate":false,"confidence":0,"candidate_type":"none","bbox_xyxy":null,'
            '"visual_evidence":"","possible_false_positive":"","recommended_check":"rgb"}',
            True,
        )

    monkeypatch.setattr(vlm, "_make_openai_client", lambda config: object())
    monkeypatch.setattr(vlm, "_resolve_lmstudio_model", lambda client, config, logger: "loaded-vision-model")
    monkeypatch.setattr(vlm, "_request_vlm_json", fake_request)

    summary = vlm.run_vlm_lmstudio_detection(
        input_path=tif_path,
        out_prefix=tmp_path / "out" / "rgb",
        config=vlm.VlmLmStudioConfig(
            tile=8,
            overlap=0,
            max_tiles=2,
            resume=True,
            resume_jsonl_path=resume_path,
            featureless_std_threshold=0.0,
        ),
    )

    assert "Task: analyze this GeoTIFF tile" in prompts[0]
    assert "Second-pass review" in prompts[1]
    assert summary.processed_tiles == 2
    assert summary.resumed_tiles == 1
    assert summary.candidate_count == 1


def test_resume_skips_tiles_already_present_in_jsonl(tmp_path: Path, monkeypatch):
    tif_path = tmp_path / "rgb_two_tiles.tif"
    data = np.ones((3, 8, 16), dtype=np.uint8) * 120
    with rasterio.open(
        tif_path,
        "w",
        driver="GTiff",
        width=16,
        height=8,
        count=3,
        dtype="uint8",
        transform=from_origin(0, 8, 1, 1),
    ) as dst:
        dst.write(data)

    resume_path = tmp_path / "previous_vlm_candidates.jsonl"
    prompt_fingerprint = vlm._prompt_fingerprint(None, None)
    resume_tile1 = {
        "tile_index": 1,
        "tile_row": 0,
        "tile_col": 0,
        "tile_width": 8,
        "tile_height": 8,
        "used_views": ["rgb"],
        "has_rgb": True,
        "has_dsm": False,
        "has_dtm": False,
        "analysis_mode": "rgb_only",
        "prompt_fingerprint": prompt_fingerprint,
        "candidate": False,
        "confidence": 0.0,
        "candidate_type": "none",
        "bbox_xyxy": None,
        "bbox_global_xyxy": None,
        "bbox_crs_xyxy": None,
        "center_x": None,
        "center_y": None,
        "gps_lon": None,
        "gps_lat": None,
        "google_maps_url": "",
        "visual_evidence": "",
        "possible_false_positive": "",
        "recommended_check": "rgb",
        "status": "ok",
        "error_type": None,
        "error_message": None,
        "geometry": None,
    }
    resume_path.write_text(json.dumps(resume_tile1) + "\n", encoding="utf-8")

    calls = {"count": 0}

    def fake_request(**kwargs):
        calls["count"] += 1
        return (
            '{"candidate":false,"confidence":0,"candidate_type":"none","bbox_xyxy":null,'
            '"visual_evidence":"","possible_false_positive":"","recommended_check":"rgb"}',
            True,
        )

    monkeypatch.setattr(vlm, "_make_openai_client", lambda config: object())
    monkeypatch.setattr(vlm, "_resolve_lmstudio_model", lambda client, config, logger: "loaded-vision-model")
    monkeypatch.setattr(vlm, "_request_vlm_json", fake_request)

    summary = vlm.run_vlm_lmstudio_detection(
        input_path=tif_path,
        out_prefix=tmp_path / "out" / "rgb",
        config=vlm.VlmLmStudioConfig(
            tile=8,
            overlap=0,
            max_tiles=2,
            export_every=1,
            resume=True,
            resume_jsonl_path=resume_path,
            featureless_std_threshold=0.0,
        ),
    )

    assert calls["count"] == 1
    assert summary.processed_tiles == 2
    assert summary.resumed_tiles == 1
    assert len(summary.paths.jsonl.read_text(encoding="utf-8").strip().splitlines()) == 2
    assert summary.paths.csv.exists()
    assert summary.paths.xlsx.exists()
