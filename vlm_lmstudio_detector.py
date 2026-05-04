"""LM Studio / OpenAI-compatible VLM candidate scanning for GeoTIFF rasters.

This module is intentionally independent from the DL/classic/YOLO pipeline. It
reads raster tiles, renders visual views as PNG images, sends them to a local
vision-language model, and writes separate candidate outputs.
"""

from __future__ import annotations

import base64
import csv
import io
import json
import logging
import math
from dataclasses import dataclass, replace
from pathlib import Path
from urllib.parse import urlparse
from typing import Any, Dict, Iterable, Iterator, List, Optional, Sequence, Tuple

import numpy as np
import rasterio
from rasterio.crs import CRS as RasterioCRS
from rasterio.transform import Affine
from rasterio.warp import transform as rasterio_transform
from rasterio.windows import Window
from tqdm import tqdm


LOGGER = logging.getLogger("vlm_lmstudio_detector")
AUTO_MODEL_TOKENS = {"auto", "active", "current", "loaded"}

VIEW_ALIASES = {
    "rgb": "rgb",
    "ortho": "rgb",
    "orthophoto": "rgb",
    "dsm": "dsm",
    "dtm": "dtm",
    "dem": "dtm",
    "ndsm": "ndsm",
    "n_dsm": "ndsm",
    "hillshade": "hillshade",
    "hill": "hillshade",
    "slope": "slope",
}

RECOMMENDED_CHECKS = {"rgb", "hillshade", "ndsm", "dsm", "dtm", "slope", "field_check"}
ALLOWED_CANDIDATE_TYPES = {
    "none",
    "mound",
    "tumulus",
    "ring_ditch",
    "wall_trace",
    "road_trace",
    "foundation",
    "enclosure",
    "terrace",
    "unknown",
}

JSON_SCHEMA_TEXT = """{
  "candidate": true,
  "confidence": 0.0,
  "candidate_type": "none | mound | tumulus | ring_ditch | wall_trace | road_trace | foundation | enclosure | terrace | unknown",
  "bbox_xyxy": [x1, y1, x2, y2],
  "visual_evidence": "...",
  "possible_false_positive": "...",
  "recommended_check": "rgb | hillshade | ndsm | dsm | dtm | slope | field_check"
}"""


class VlmConnectionError(RuntimeError):
    """Raised when the LM Studio server cannot be reached."""


@dataclass(frozen=True)
class VlmLmStudioConfig:
    base_url: str = "http://localhost:1234/v1"
    api_key: str = "lm-studio"
    model: str = "auto"
    tile: int = 1024
    overlap: int = 256
    views: str | Sequence[str] = "auto"
    gsd_m: Optional[float] = 0.30
    confidence_threshold: float = 0.75
    max_tiles: int = 0
    timeout: int = 120
    temperature: float = 0.0
    export_every: int = 50
    resume: bool = True
    resume_jsonl_path: Optional[Path | str] = None


@dataclass(frozen=True)
class RasterBandLayout:
    rgb: Tuple[int, int, int]
    dsm: Optional[int]
    dtm: Optional[int]

    @property
    def has_rgb(self) -> bool:
        return all(idx > 0 for idx in self.rgb)

    @property
    def has_dsm(self) -> bool:
        return self.dsm is not None and self.dsm > 0

    @property
    def has_dtm(self) -> bool:
        return self.dtm is not None and self.dtm > 0

    @property
    def analysis_mode(self) -> str:
        if self.has_dsm and self.has_dtm:
            return "rgb_topo"
        if self.has_dsm:
            return "rgb_dsm"
        if self.has_dtm:
            return "rgb_dtm"
        return "rgb_only"


@dataclass(frozen=True)
class VlmOutputPaths:
    jsonl: Path
    csv: Path
    xlsx: Path
    geojson: Path
    gpkg: Path
    raw_errors_jsonl: Path


@dataclass(frozen=True)
class VlmDetectionSummary:
    paths: VlmOutputPaths
    total_tiles: int
    processed_tiles: int
    resumed_tiles: int
    raw_candidate_count: int
    candidate_count: int
    error_count: int
    used_views: Tuple[str, ...]
    analysis_mode: str


def run_vlm_lmstudio_detection(
    *,
    input_path: Path | str,
    out_prefix: Path | str,
    config: VlmLmStudioConfig,
    band_indexes: Optional[Sequence[int] | str] = None,
    logger: Optional[logging.Logger] = None,
) -> VlmDetectionSummary:
    """Run a VLM tile scan and write JSONL/CSV/GeoJSON/GPKG outputs."""

    log = logger or LOGGER
    input_path = Path(input_path)
    out_prefix = _output_base_path(Path(out_prefix))
    out_prefix.parent.mkdir(parents=True, exist_ok=True)
    paths = _build_output_paths(out_prefix)

    _validate_config(config)
    _ensure_pillow_available()

    with rasterio.open(input_path) as src:
        layout = _detect_band_layout(src, band_indexes=band_indexes, logger=log)
        if not layout.has_rgb:
            raise ValueError("VLM taramasi icin en az uc RGB bandi gerekir; girdi raster RGB icermiyor.")

        selected_views = _resolve_views(config.views, layout, logger=log)
        pixel_size = _pixel_size(src.transform)
        prompt_gsd_m = _resolve_gsd_m(config.gsd_m, src.transform, src.crs, logger=log)
        total_tiles = _count_windows(src.width, src.height, config.tile, config.overlap)
        if config.max_tiles > 0:
            total_tiles = min(total_tiles, int(config.max_tiles))

        resume_source = _resolve_resume_source(config, paths)
        records, processed_tile_indexes = _load_resume_records(
            resume_source,
            total_tiles=total_tiles,
            raster_width=src.width,
            raster_height=src.height,
            tile=config.tile,
            overlap=config.overlap,
            selected_views=selected_views,
            analysis_mode=layout.analysis_mode,
            logger=log,
        )
        candidate_records: List[Dict[str, Any]] = []
        error_records: List[Dict[str, Any]] = []
        for record in records:
            if _is_exportable_candidate(record, config.confidence_threshold):
                candidate_records.append(record)
            if record.get("status") == "error":
                error_records.append(record)
        raw_candidate_count = sum(1 for record in records if bool(record.get("candidate")) and bool(record.get("geometry")))
        skipped_count = sum(1 for record in records if record.get("status") == "skipped")
        resumed_tiles = len(records)
        if resumed_tiles:
            log.info(
                "VLM resume: %d/%d tile onceki JSONL'den yuklendi: %s",
                resumed_tiles,
                total_tiles,
                resume_source,
            )
        use_response_format = True
        log.info(
            "VLM tile planı: toplam=%d, tile=%d, overlap=%d, views=%s, analysis_mode=%s, gsd=%s",
            total_tiles,
            config.tile,
            config.overlap,
            ",".join(selected_views),
            layout.analysis_mode,
            f"{prompt_gsd_m:.3f} m/px" if prompt_gsd_m is not None else "unknown",
        )
        log.info(
            "VLM resume=%s, export_every=%d",
            "on" if config.resume else "off",
            config.export_every,
        )

        if len(processed_tile_indexes) >= total_tiles:
            _write_jsonl_records(paths.jsonl, records)
            _write_jsonl_records(paths.raw_errors_jsonl, error_records)
            paths = _write_candidate_outputs(
                paths,
                candidate_records,
                src.crs,
                all_records=records,
                confidence_threshold=config.confidence_threshold,
                logger=log,
            )
            log.info("VLM resume tum hedef tile'lari kapsiyor; LM Studio'ya yeni istek gonderilmedi.")
            return VlmDetectionSummary(
                paths=paths,
                total_tiles=total_tiles,
                processed_tiles=len(records),
                resumed_tiles=resumed_tiles,
                raw_candidate_count=raw_candidate_count,
                candidate_count=len(candidate_records),
                error_count=len(error_records),
                used_views=tuple(selected_views),
                analysis_mode=layout.analysis_mode,
            )

        client = _make_openai_client(config)
        try:
            resolved_model = _resolve_lmstudio_model(client, config, logger=log)
            config = replace(config, model=resolved_model)
        except VlmConnectionError as exc:
            run_record = _run_error_record(
                error_type="connection_failed",
                message=str(exc),
                layout=layout,
                selected_views=selected_views,
            )
            _write_jsonl_records(paths.jsonl, [*records, run_record])
            _write_jsonl_records(paths.raw_errors_jsonl, [*error_records, run_record])
            paths = _write_candidate_outputs(
                paths,
                candidate_records,
                src.crs,
                all_records=[*records, run_record],
                confidence_threshold=config.confidence_threshold,
                logger=log,
            )
            raise

        log.info("VLM modeli: %s", config.model)

        with paths.jsonl.open("w", encoding="utf-8") as jsonl_fh, paths.raw_errors_jsonl.open(
            "w", encoding="utf-8"
        ) as raw_fh:
            for record in records:
                _write_jsonl_line(jsonl_fh, record)
            for record in error_records:
                _write_jsonl_line(raw_fh, record)
            if resumed_tiles:
                paths = _write_candidate_outputs(
                    paths,
                    candidate_records,
                    src.crs,
                    all_records=records,
                    confidence_threshold=config.confidence_threshold,
                    logger=log,
                )
                log.info(
                    "VLM resume ara ciktilari guncellendi: %d/%d tile, esik_ustu=%d -> %s",
                    len(records),
                    total_tiles,
                    len(candidate_records),
                    paths.xlsx,
                )

            progress_iter = _generate_windows(src.width, src.height, config.tile, config.overlap)
            with tqdm(
                total=total_tiles,
                initial=min(len(processed_tile_indexes), total_tiles),
                desc="VLM tiles",
                unit="tile",
                leave=False,
            ) as pbar:
                _set_progress_postfix(
                    pbar,
                    records=records,
                    error_records=error_records,
                    skipped_count=skipped_count,
                    raw_candidate_count=raw_candidate_count,
                    candidate_records=candidate_records,
                )
                tiles_since_export = 0
                for tile_index, (window, row, col) in enumerate(progress_iter, start=1):
                    if config.max_tiles > 0 and tile_index > config.max_tiles:
                        break
                    if tile_index in processed_tile_indexes:
                        continue

                    tile_record_base = _base_tile_record(
                        tile_index=tile_index,
                        row=row,
                        col=col,
                        window=window,
                        layout=layout,
                        selected_views=selected_views,
                    )

                    response_text: Optional[str] = None
                    try:
                        tile_arrays = _read_tile_arrays(src, layout, window)
                        if _is_empty_rgb_tile(tile_arrays["rgb"]):
                            record = _skipped_tile_record(
                                tile_record_base,
                                reason="RGB tile is empty, nodata, transparent, or effectively black.",
                            )
                        else:
                            rendered = _render_views(tile_arrays, selected_views, pixel_size=pixel_size)
                            prompt = _build_prompt(
                                layout.analysis_mode,
                                selected_views,
                                int(window.width),
                                int(window.height),
                                gsd_m=prompt_gsd_m,
                            )
                            response_text, use_response_format = _request_vlm_json(
                                client=client,
                                config=config,
                                prompt=prompt,
                                rendered_views=rendered,
                                use_response_format=use_response_format,
                                logger=log,
                            )
                            parsed = _parse_model_json(response_text)
                            record = _model_result_to_record(
                                parsed=parsed,
                                raw_response=response_text,
                                base_record=tile_record_base,
                                transform=src.transform,
                                crs=src.crs,
                                raster_width=src.width,
                                raster_height=src.height,
                            )
                    except Exception as exc:
                        error_type = _classify_exception(exc)
                        message = _friendly_exception_message(exc, config)
                        record = dict(tile_record_base)
                        record.update(
                            {
                                "candidate": False,
                                "confidence": 0.0,
                                "candidate_type": "none",
                                "bbox_xyxy": None,
                                "bbox_global_xyxy": None,
                                "bbox_crs_xyxy": None,
                                "center_x": None,
                                "center_y": None,
                                "gps_lon": None,
                                "gps_lat": None,
                                "google_maps_url": "",
                                "visual_evidence": "",
                                "possible_false_positive": "",
                                "recommended_check": "field_check",
                                "status": "error",
                                "error_type": error_type,
                                "error_message": message,
                            }
                        )
                        raw_record = dict(record)
                        if response_text is not None:
                            raw_record["raw_response"] = response_text
                        _write_jsonl_line(raw_fh, raw_record)
                        error_records.append(record)
                        if error_type == "vision_unsupported":
                            log.error(
                                "VLM modeli goruntu girdisini desteklemiyor gibi gorunuyor: %s",
                                message,
                            )
                            _write_jsonl_line(jsonl_fh, record)
                            records.append(record)
                            processed_tile_indexes.add(tile_index)
                            pbar.update(1)
                            break
                        log.warning("VLM tile %s hata ile atlandi: %s", tile_index, message)

                    _write_jsonl_line(jsonl_fh, record)
                    records.append(record)
                    processed_tile_indexes.add(tile_index)
                    if record.get("status") == "skipped":
                        skipped_count += 1
                    if bool(record.get("candidate")) and bool(record.get("geometry")):
                        raw_candidate_count += 1
                    if _is_exportable_candidate(record, config.confidence_threshold):
                        candidate_records.append(record)

                    pbar.update(1)
                    tiles_since_export += 1
                    _set_progress_postfix(
                        pbar,
                        records=records,
                        error_records=error_records,
                        skipped_count=skipped_count,
                        raw_candidate_count=raw_candidate_count,
                        candidate_records=candidate_records,
                    )
                    if config.export_every > 0 and tiles_since_export >= config.export_every:
                        paths = _write_candidate_outputs(
                            paths,
                            candidate_records,
                            src.crs,
                            all_records=records,
                            confidence_threshold=config.confidence_threshold,
                            logger=log,
                        )
                        tiles_since_export = 0
                        log.info(
                            "VLM ara ciktilar guncellendi: %d/%d tile, esik_ustu=%d -> %s",
                            len(records),
                            total_tiles,
                            len(candidate_records),
                            paths.xlsx,
                        )
                    if len(records) == 1 or len(records) % 25 == 0 or len(records) == total_tiles:
                        log.info(
                            "VLM ilerleme: %d/%d tile, model_aday=%d, esik_ustu=%d, skipped=%d, hata=%d",
                            len(records),
                            total_tiles,
                            raw_candidate_count,
                            len(candidate_records),
                            skipped_count,
                            len(error_records),
                        )

        paths = _write_candidate_outputs(
            paths,
            candidate_records,
            src.crs,
            all_records=records,
            confidence_threshold=config.confidence_threshold,
            logger=log,
        )

        return VlmDetectionSummary(
            paths=paths,
            total_tiles=total_tiles,
            processed_tiles=len(records),
            resumed_tiles=resumed_tiles,
            raw_candidate_count=raw_candidate_count,
            candidate_count=len(candidate_records),
            error_count=len(error_records),
            used_views=tuple(selected_views),
            analysis_mode=layout.analysis_mode,
        )


def _validate_config(config: VlmLmStudioConfig) -> None:
    if config.tile <= 0:
        raise ValueError("vlm_tile pozitif olmali.")
    if config.overlap < 0:
        raise ValueError("vlm_overlap negatif olamaz.")
    if config.overlap >= config.tile:
        raise ValueError("vlm_overlap, vlm_tile degerinden kucuk olmali.")
    if config.max_tiles < 0:
        raise ValueError("vlm_max_tiles negatif olamaz.")
    if config.export_every < 0:
        raise ValueError("vlm_export_every negatif olamaz.")
    if not 0.0 <= config.confidence_threshold <= 1.0:
        raise ValueError("vlm_confidence_threshold 0-1 arasinda olmali.")
    if config.timeout <= 0:
        raise ValueError("vlm_timeout pozitif olmali.")
    if config.gsd_m is not None and float(config.gsd_m) < 0:
        raise ValueError("vlm_gsd_m negatif olamaz.")
    if not str(config.base_url or "").strip():
        raise ValueError("vlm_base_url bos olamaz.")
    if not str(config.model or "").strip():
        raise ValueError("vlm_model bos olamaz.")


def _ensure_pillow_available() -> None:
    try:
        from PIL import Image  # noqa: F401
    except ImportError as exc:
        raise ImportError("VLM PNG uretimi icin pillow>=10.0.0 gereklidir.") from exc


def _output_base_path(path: Path) -> Path:
    suffixes = {
        ".tif",
        ".tiff",
        ".gpkg",
        ".xlsx",
        ".csv",
        ".geojson",
        ".json",
        ".jsonl",
    }
    return path.with_suffix("") if path.suffix.lower() in suffixes else path


def _build_output_paths(out_prefix: Path) -> VlmOutputPaths:
    base = _output_base_path(out_prefix)
    return VlmOutputPaths(
        jsonl=base.parent / f"{base.name}_vlm_candidates.jsonl",
        csv=base.parent / f"{base.name}_vlm_candidates.csv",
        xlsx=base.parent / f"{base.name}_vlm_candidates.xlsx",
        geojson=base.parent / f"{base.name}_vlm_candidates.geojson",
        gpkg=base.parent / f"{base.name}_vlm_candidates.gpkg",
        raw_errors_jsonl=base.parent / f"{base.name}_vlm_raw_errors.jsonl",
    )


def _resolve_resume_source(config: VlmLmStudioConfig, paths: VlmOutputPaths) -> Optional[Path]:
    if not config.resume:
        return None
    if config.resume_jsonl_path:
        return Path(config.resume_jsonl_path)
    if paths.jsonl.exists():
        return paths.jsonl
    return None


def _load_resume_records(
    path: Optional[Path],
    *,
    total_tiles: int,
    raster_width: int,
    raster_height: int,
    tile: int,
    overlap: int,
    selected_views: Sequence[str],
    analysis_mode: str,
    logger: logging.Logger,
) -> Tuple[List[Dict[str, Any]], set[int]]:
    if path is None or not path.exists() or not path.is_file():
        return [], set()

    expected_windows: Dict[int, Tuple[int, int, int, int]] = {}
    for tile_index, (window, row, col) in enumerate(_generate_windows(raster_width, raster_height, tile, overlap), start=1):
        if tile_index > total_tiles:
            break
        expected_windows[tile_index] = (int(row), int(col), int(window.width), int(window.height))

    expected_views = [str(view).strip().lower() for view in selected_views]
    records_by_tile: Dict[int, Dict[str, Any]] = {}
    bad_json_count = 0
    mismatch_count = 0
    duplicate_count = 0

    try:
        with path.open("r", encoding="utf-8") as fh:
            for line_no, line in enumerate(fh, start=1):
                text = line.strip()
                if not text:
                    continue
                try:
                    record = json.loads(text)
                except json.JSONDecodeError:
                    bad_json_count += 1
                    continue
                tile_index = _coerce_int(record.get("tile_index"))
                if tile_index is None or tile_index <= 0 or tile_index > total_tiles:
                    continue
                expected = expected_windows.get(tile_index)
                if expected is None or not _resume_record_matches_plan(
                    record,
                    expected=expected,
                    selected_views=expected_views,
                    analysis_mode=analysis_mode,
                ):
                    mismatch_count += 1
                    continue
                if tile_index in records_by_tile:
                    duplicate_count += 1
                records_by_tile[tile_index] = record
    except OSError as exc:
        logger.warning("VLM resume JSONL okunamadi (%s): %s", path, exc)
        return [], set()

    if bad_json_count:
        logger.warning("VLM resume JSONL icinde %d bozuk satir atlandi: %s", bad_json_count, path)
    if mismatch_count:
        logger.warning(
            "VLM resume JSONL icinde %d kayit mevcut tile/view ayarlariyla uyusmadigi icin atlandi.",
            mismatch_count,
        )
    if duplicate_count:
        logger.info("VLM resume JSONL icinde %d tekrarli tile kaydi bulundu; son kayit kullanildi.", duplicate_count)

    processed = set(records_by_tile)
    return [records_by_tile[idx] for idx in sorted(records_by_tile)], processed


def _resume_record_matches_plan(
    record: Dict[str, Any],
    *,
    expected: Tuple[int, int, int, int],
    selected_views: Sequence[str],
    analysis_mode: str,
) -> bool:
    expected_row, expected_col, expected_width, expected_height = expected
    actual = (
        _coerce_int(record.get("tile_row")),
        _coerce_int(record.get("tile_col")),
        _coerce_int(record.get("tile_width")),
        _coerce_int(record.get("tile_height")),
    )
    if actual != (expected_row, expected_col, expected_width, expected_height):
        return False

    record_mode = str(record.get("analysis_mode") or "").strip()
    if record_mode and record_mode != analysis_mode:
        return False

    record_views = _normalise_resume_views(record.get("used_views"))
    if record_views and record_views != list(selected_views):
        return False
    return True


def _normalise_resume_views(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        if text.startswith("["):
            try:
                parsed = json.loads(text)
            except json.JSONDecodeError:
                parsed = [part.strip() for part in text.split(",")]
        else:
            parsed = [part.strip() for part in text.split(",")]
    elif isinstance(value, (list, tuple)):
        parsed = list(value)
    else:
        return []
    return [str(view).strip().lower() for view in parsed if str(view).strip()]


def _coerce_int(value: Any) -> Optional[int]:
    try:
        return int(value)
    except Exception:
        return None


def _generate_windows(width: int, height: int, tile: int, overlap: int) -> Iterator[Tuple[Window, int, int]]:
    stride = tile - overlap
    for row in range(0, height, stride):
        for col in range(0, width, stride):
            win_width = min(tile, width - col)
            win_height = min(tile, height - row)
            yield Window(col, row, win_width, win_height), row, col


def _count_windows(width: int, height: int, tile: int, overlap: int) -> int:
    stride = tile - overlap
    return int(math.ceil(width / stride) * math.ceil(height / stride))


def _parse_band_indexes(band_indexes: Optional[Sequence[int] | str]) -> Optional[Tuple[int, ...]]:
    if band_indexes is None:
        return None
    if isinstance(band_indexes, str):
        parts = [part.strip() for part in band_indexes.split(",") if part.strip()]
        if not parts:
            return None
        return tuple(int(part) for part in parts)
    return tuple(int(v) for v in band_indexes)


def _detect_band_layout(
    src: rasterio.io.DatasetReader,
    *,
    band_indexes: Optional[Sequence[int] | str],
    logger: logging.Logger,
) -> RasterBandLayout:
    parsed = _parse_band_indexes(band_indexes)
    if src.count < 3:
        raise ValueError(f"VLM RGB gerektirir; raster yalnizca {src.count} band iceriyor.")

    if parsed is not None and len(parsed) >= 3:
        rgb = tuple(parsed[:3])
    else:
        rgb = (1, 2, 3)

    if any(idx <= 0 or idx > src.count for idx in rgb):
        raise ValueError(
            f"VLM RGB bandlari gecersiz: {rgb}. Raster band sayisi: {src.count}."
        )

    dsm_idx: Optional[int] = None
    dtm_idx: Optional[int] = None
    if parsed is not None:
        if len(parsed) >= 4 and 0 < parsed[3] <= src.count:
            dsm_idx = parsed[3]
        if len(parsed) >= 5 and 0 < parsed[4] <= src.count:
            dtm_idx = parsed[4]

    if dsm_idx is not None and _band_is_alpha(src, dsm_idx):
        logger.warning("VLM DSM bandi alpha/mask bandi gibi gorunuyor; DSM view atlanacak.")
        dsm_idx = None
    if dtm_idx is not None and _band_is_alpha(src, dtm_idx):
        logger.warning("VLM DTM bandi alpha/mask bandi gibi gorunuyor; topo view'lari atlanacak.")
        dtm_idx = None

    # If a four-band raster is described as terrain/DTM, prefer DTM over DSM.
    if dsm_idx is not None and dtm_idx is None and _band_seems_dtm(src, dsm_idx):
        dtm_idx = dsm_idx
        dsm_idx = None
    elif dsm_idx is not None and dtm_idx is None and src.count == 4 and not _band_seems_dsm(src, dsm_idx):
        logger.warning(
            "Dorduncu band DSM/DTM olarak etiketlenmemis; VLM bunu DSM kabul edecek. "
            "RGB+DTM icin bands: '1,2,3,0,4' kullanin."
        )

    if parsed is None:
        if src.count >= 5:
            dsm_idx = 4
            dtm_idx = 5
        elif src.count == 4:
            if _band_seems_dtm(src, 4):
                dtm_idx = 4
            elif _band_is_alpha(src, 4):
                logger.info("Dorduncu band alpha/mask olarak algilandi; VLM RGB-only calisacak.")
            else:
                dsm_idx = 4
                if not _band_seems_dsm(src, 4):
                    logger.warning(
                        "Dorduncu band DSM/DTM olarak etiketlenmemis; VLM auto bunu DSM kabul edecek. "
                        "RGB+DTM icin bands: '1,2,3,0,4' kullanin."
                    )
    else:
        # Fill conventional missing topo slots only when the parsed indexes are unavailable.
        if src.count >= 5:
            if dsm_idx is None and 4 not in rgb:
                dsm_idx = 4
            if dtm_idx is None and 5 not in rgb:
                dtm_idx = 5

    if dsm_idx in rgb:
        logger.warning("DSM bandi RGB bandlariyla cakistigi icin VLM DSM view atlanacak.")
        dsm_idx = None
    if dtm_idx in rgb:
        logger.warning("DTM bandi RGB bandlariyla cakistigi icin VLM topo view'lari atlanacak.")
        dtm_idx = None
    if dsm_idx is not None and dtm_idx is not None and dsm_idx == dtm_idx:
        logger.warning("DSM ve DTM ayni banda isaret ediyor; nDSM atlanacak, band DTM olarak kullanilacak.")
        dsm_idx = None

    return RasterBandLayout(rgb=rgb, dsm=dsm_idx, dtm=dtm_idx)


def _band_text(src: rasterio.io.DatasetReader, band_index: int) -> str:
    parts: List[str] = []
    try:
        desc = src.descriptions[band_index - 1]
        if desc:
            parts.append(str(desc))
    except Exception:
        pass
    try:
        tags = src.tags(band_index)
        for key, value in tags.items():
            parts.append(str(key))
            parts.append(str(value))
    except Exception:
        pass
    return " ".join(parts).lower()


def _band_seems_dtm(src: rasterio.io.DatasetReader, band_index: int) -> bool:
    text = _band_text(src, band_index)
    return any(token in text for token in ("dtm", "dem", "terrain", "bare earth", "bare_earth"))


def _band_seems_dsm(src: rasterio.io.DatasetReader, band_index: int) -> bool:
    text = _band_text(src, band_index)
    return any(token in text for token in ("dsm", "surface", "canopy"))


def _band_is_alpha(src: rasterio.io.DatasetReader, band_index: int) -> bool:
    try:
        interp = src.colorinterp[band_index - 1]
        return str(getattr(interp, "name", interp)).lower() == "alpha"
    except Exception:
        return False


def _resolve_views(
    requested: str | Sequence[str],
    layout: RasterBandLayout,
    *,
    logger: logging.Logger,
) -> List[str]:
    if isinstance(requested, str) and requested.strip().lower() == "auto":
        if layout.has_dsm and layout.has_dtm:
            return ["rgb", "hillshade", "ndsm", "slope"]
        if layout.has_dtm:
            return ["rgb", "hillshade", "slope"]
        if layout.has_dsm:
            return ["rgb", "dsm"]
        return ["rgb"]

    raw_views: Iterable[str]
    if isinstance(requested, str):
        raw_views = [part.strip() for part in requested.split(",") if part.strip()]
    else:
        raw_views = [str(part).strip() for part in requested if str(part).strip()]

    selected: List[str] = []
    for raw in raw_views:
        view = VIEW_ALIASES.get(raw.lower())
        if view is None:
            logger.warning("Bilinmeyen VLM view '%s' atlandi.", raw)
            continue
        missing_reason = _missing_view_reason(view, layout)
        if missing_reason:
            logger.warning("VLM view '%s' atlandi: %s", view, missing_reason)
            continue
        if view not in selected:
            selected.append(view)

    if "rgb" not in selected:
        if layout.has_rgb:
            logger.warning("VLM icin RGB zorunlu oldugundan 'rgb' view listeye eklendi.")
            selected.insert(0, "rgb")
        else:
            raise ValueError("VLM icin RGB view uretilemiyor.")

    if not selected:
        raise ValueError("VLM icin kullanilabilir view kalmadi.")
    return selected


def _missing_view_reason(view: str, layout: RasterBandLayout) -> Optional[str]:
    if view == "rgb" and not layout.has_rgb:
        return "RGB bandlari yok"
    if view == "dsm" and not layout.has_dsm:
        return "DSM bandi yok"
    if view == "dtm" and not layout.has_dtm:
        return "DTM bandi yok"
    if view in {"hillshade", "slope"} and not layout.has_dtm:
        return "DTM bandi yok"
    if view == "ndsm" and not (layout.has_dsm and layout.has_dtm):
        return "DSM ve DTM birlikte gerekli"
    return None


def _pixel_size(transform: Affine) -> Tuple[float, float]:
    px = math.hypot(transform.a, transform.b)
    py = math.hypot(transform.d, transform.e)
    return max(px, 1e-6), max(py, 1e-6)


def _resolve_gsd_m(
    configured_gsd_m: Optional[float],
    transform: Affine,
    crs: Optional[RasterioCRS],
    *,
    logger: logging.Logger,
) -> Optional[float]:
    if configured_gsd_m is not None and float(configured_gsd_m) > 0:
        return float(configured_gsd_m)

    try:
        if crs is not None and getattr(crs, "is_geographic", False):
            logger.warning("VLM GSD otomatik tahmin edilemedi: raster CRS geographic derece biriminde.")
            return None
        px, py = _pixel_size(transform)
        gsd = float((px + py) / 2.0)
        if math.isfinite(gsd) and gsd > 0:
            return gsd
    except Exception as exc:
        logger.warning("VLM GSD otomatik tahmin edilemedi: %s", exc)
    return None


def _read_band(src: rasterio.io.DatasetReader, band_index: int, window: Window) -> np.ndarray:
    arr = src.read(band_index, window=window, masked=True)
    if hasattr(arr, "filled"):
        data = np.asarray(arr.astype(np.float32).filled(np.nan), dtype=np.float32)
    else:
        data = np.asarray(arr, dtype=np.float32)
    return data


def _read_tile_arrays(
    src: rasterio.io.DatasetReader,
    layout: RasterBandLayout,
    window: Window,
) -> Dict[str, Any]:
    rgb_arrays = [_read_band(src, idx, window) for idx in layout.rgb]
    out: Dict[str, Any] = {"rgb": rgb_arrays}
    if layout.has_dsm and layout.dsm is not None:
        out["dsm"] = _read_band(src, layout.dsm, window)
    if layout.has_dtm and layout.dtm is not None:
        out["dtm"] = _read_band(src, layout.dtm, window)
    return out


def _is_empty_rgb_tile(channels: Sequence[np.ndarray]) -> bool:
    if len(channels) < 3:
        return True
    stacked = np.stack([np.asarray(ch, dtype=np.float32) for ch in channels[:3]], axis=0)
    valid = np.all(np.isfinite(stacked), axis=0)
    valid_ratio = float(np.count_nonzero(valid)) / float(valid.size or 1)
    if valid_ratio < 0.01:
        return True
    values = stacked[:, valid]
    if values.size == 0:
        return True
    finite_min = float(np.nanmin(values))
    finite_max = float(np.nanmax(values))
    if finite_max <= 1.0 and (finite_max - finite_min) <= 1.0:
        return True
    return False


def _render_views(
    arrays: Dict[str, Any],
    selected_views: Sequence[str],
    *,
    pixel_size: Tuple[float, float],
) -> List[Tuple[str, str]]:
    rendered: List[Tuple[str, str]] = []
    for view in selected_views:
        if view == "rgb":
            image = _rgb_to_uint8(arrays["rgb"])
        elif view == "dsm":
            image = _grayscale_to_rgb(_scale_to_uint8(arrays["dsm"]))
        elif view == "dtm":
            image = _grayscale_to_rgb(_scale_to_uint8(arrays["dtm"]))
        elif view == "ndsm":
            ndsm = arrays["dsm"] - arrays["dtm"]
            image = _grayscale_to_rgb(_scale_to_uint8(ndsm))
        elif view == "hillshade":
            image = _grayscale_to_rgb(_hillshade_uint8(arrays["dtm"], pixel_size=pixel_size))
        elif view == "slope":
            slope = _slope_degrees(arrays["dtm"], pixel_size=pixel_size)
            image = _grayscale_to_rgb(_scale_to_uint8(slope))
        else:
            continue
        rendered.append((view, _png_data_url(image)))
    return rendered


def _rgb_to_uint8(channels: Sequence[np.ndarray]) -> np.ndarray:
    stacked = np.stack([_scale_to_uint8(ch) for ch in channels[:3]], axis=-1)
    return stacked.astype(np.uint8, copy=False)


def _grayscale_to_rgb(arr: np.ndarray) -> np.ndarray:
    return np.repeat(arr[:, :, np.newaxis], 3, axis=2).astype(np.uint8, copy=False)


def _scale_to_uint8(arr: np.ndarray, low: float = 2.0, high: float = 98.0) -> np.ndarray:
    data = np.asarray(arr, dtype=np.float32)
    valid = np.isfinite(data)
    if not np.any(valid):
        return np.zeros(data.shape, dtype=np.uint8)
    lower = float(np.percentile(data[valid], low))
    upper = float(np.percentile(data[valid], high))
    if upper - lower <= 1e-6:
        scaled = np.zeros_like(data, dtype=np.float32)
        scaled[valid] = np.clip(data[valid] - lower, 0.0, 1.0)
    else:
        scaled = (data - lower) / (upper - lower)
    scaled = np.clip(scaled, 0.0, 1.0)
    scaled[~valid] = 0.0
    return np.rint(scaled * 255.0).astype(np.uint8)


def _fill_nodata(arr: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    data = np.asarray(arr, dtype=np.float32)
    valid = np.isfinite(data)
    if not np.any(valid):
        return np.zeros_like(data, dtype=np.float32), valid
    fill_value = float(np.nanmedian(data[valid]))
    filled = data.copy()
    filled[~valid] = fill_value
    return filled, valid


def _slope_degrees(arr: np.ndarray, *, pixel_size: Tuple[float, float]) -> np.ndarray:
    filled, valid = _fill_nodata(arr)
    px, py = pixel_size
    gy, gx = np.gradient(filled, py, px)
    slope = np.degrees(np.arctan(np.hypot(gx, gy))).astype(np.float32)
    slope[~valid] = np.nan
    return slope


def _hillshade_uint8(
    arr: np.ndarray,
    *,
    pixel_size: Tuple[float, float],
    azimuth_deg: float = 315.0,
    altitude_deg: float = 45.0,
) -> np.ndarray:
    filled, valid = _fill_nodata(arr)
    px, py = pixel_size
    gy, gx = np.gradient(filled, py, px)
    slope = np.pi / 2.0 - np.arctan(np.hypot(gx, gy))
    aspect = np.arctan2(-gx, gy)
    azimuth = math.radians(azimuth_deg)
    altitude = math.radians(altitude_deg)
    shaded = np.sin(altitude) * np.sin(slope) + np.cos(altitude) * np.cos(slope) * np.cos(
        azimuth - aspect
    )
    shaded = np.clip((shaded + 1.0) / 2.0, 0.0, 1.0)
    shaded[~valid] = 0.0
    return np.rint(shaded * 255.0).astype(np.uint8)


def _png_data_url(image: np.ndarray) -> str:
    try:
        from PIL import Image
    except ImportError as exc:
        raise ImportError("VLM PNG uretimi icin pillow>=10.0.0 gereklidir.") from exc

    buffer = io.BytesIO()
    Image.fromarray(image.astype(np.uint8, copy=False), mode="RGB").save(buffer, format="PNG")
    encoded = base64.b64encode(buffer.getvalue()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def _make_openai_client(config: VlmLmStudioConfig) -> Any:
    try:
        from openai import OpenAI
    except ImportError as exc:
        raise ImportError("VLM icin openai>=1.0.0 gereklidir: pip install openai>=1.0.0") from exc

    return OpenAI(
        base_url=_normalize_openai_base_url(config.base_url),
        api_key=str(config.api_key or "lm-studio"),
        timeout=float(config.timeout),
    )


def _normalize_openai_base_url(base_url: str) -> str:
    text = str(base_url or "").strip().rstrip("/")
    if not text:
        return text
    parsed = urlparse(text)
    if parsed.scheme and parsed.netloc and parsed.path in ("", "/"):
        return f"{text}/v1"
    return text


def _resolve_lmstudio_model(client: Any, config: VlmLmStudioConfig, *, logger: logging.Logger) -> str:
    requested_model = str(config.model or "").strip()
    auto_model = requested_model.lower() in AUTO_MODEL_TOKENS
    try:
        models = client.models.list()
    except Exception as exc:
        if _looks_like_connection_error(exc):
            raise VlmConnectionError(
                f"LM Studio sunucusuna baglanilamadi ({_normalize_openai_base_url(config.base_url)}). "
                "LM Studio'da Local Server'i baslatin ve base URL'yi kontrol edin."
            ) from exc
        if auto_model:
            raise RuntimeError(
                f"vlm_model={requested_model!r} icin LM Studio model listesi okunamadi: {exc}"
            ) from exc
        logger.warning("LM Studio model listesi okunamadi, verilen modelle devam edilecek: %s", exc)
        return requested_model

    try:
        model_ids = [str(item.id) for item in models.data if str(getattr(item, "id", "")).strip()]
    except Exception:
        model_ids = []

    if auto_model:
        if not model_ids:
            raise RuntimeError(
                "vlm_model='auto' icin LM Studio'da yuklu model bulunamadi. "
                "Once vision destekli bir modeli Load Model ile yukleyin."
            )
        if len(model_ids) > 1:
            logger.warning(
                "LM Studio birden fazla yuklu model dondurdu; ilk model kullanilacak: %s. Tum modeller: %s",
                model_ids[0],
                ", ".join(model_ids),
            )
        else:
            logger.info("LM Studio aktif/yuklu modeli otomatik secildi: %s", model_ids[0])
        return model_ids[0]

    if model_ids and requested_model not in model_ids:
        logger.warning(
            "VLM modeli LM Studio model listesinde gorunmedi: %s. Yuklu modeller: %s",
            requested_model,
            ", ".join(model_ids),
        )
    return requested_model


def _request_vlm_json(
    *,
    client: Any,
    config: VlmLmStudioConfig,
    prompt: str,
    rendered_views: Sequence[Tuple[str, str]],
    use_response_format: bool,
    logger: logging.Logger,
) -> Tuple[str, bool]:
    content: List[Dict[str, Any]] = [{"type": "text", "text": prompt}]
    for view_name, data_url in rendered_views:
        content.append({"type": "text", "text": f"View: {view_name}"})
        content.append({"type": "image_url", "image_url": {"url": data_url}})

    kwargs: Dict[str, Any] = {
        "model": config.model,
        "messages": [
            {
                "role": "system",
                "content": (
                    "You are a conservative archaeological remote-sensing analyst reviewing aerial and terrain-derived imagery. "
                    "Your job is to flag only clear, review-worthy archaeological candidates. "
                    "Prefer candidate=false over speculative candidates whenever the evidence is weak, single-cue, modern-looking, "
                    "natural-looking, or explainable as agriculture, infrastructure, vegetation, shadow, or imagery artifact. "
                    "Return only strict JSON. Do not include markdown or commentary."
                ),
            },
            {"role": "user", "content": content},
        ],
        "temperature": float(config.temperature),
    }
    if use_response_format:
        kwargs["response_format"] = {"type": "json_object"}

    try:
        response = client.chat.completions.create(**kwargs)
    except Exception as exc:
        if use_response_format and _looks_like_response_format_error(exc):
            logger.warning("LM Studio response_format=json_object desteklemedi; prompt-only JSON ile tekrar deneniyor.")
            kwargs.pop("response_format", None)
            response = client.chat.completions.create(**kwargs)
            use_response_format = False
        else:
            raise

    message = response.choices[0].message
    return _message_content_to_text(message.content), use_response_format


def _message_content_to_text(content: Any) -> str:
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts: List[str] = []
        for item in content:
            if isinstance(item, dict) and item.get("type") == "text":
                parts.append(str(item.get("text", "")))
            elif hasattr(item, "text"):
                parts.append(str(item.text))
            else:
                parts.append(str(item))
        return "\n".join(parts)
    return str(content)


def _build_prompt(
    analysis_mode: str,
    selected_views: Sequence[str],
    width: int,
    height: int,
    *,
    gsd_m: Optional[float],
) -> str:
    if gsd_m is not None and gsd_m > 0:
        tile_width_m = width * float(gsd_m)
        tile_height_m = height * float(gsd_m)
        scale_text = (
            f"Spatial scale: this is nadir imagery with approximately {float(gsd_m):.2f} m ground sampling distance (GSD), "
            f"so one pixel is about {float(gsd_m):.2f} m on the ground and this tile covers about "
            f"{tile_width_m:.0f} m x {tile_height_m:.0f} m. "
            "Use this scale when deciding whether an anomaly has a plausible archaeological size. "
            "Do not flag tiny isolated objects or modern-scale clutter as archaeological features. "
        )
    else:
        scale_text = (
            "Spatial scale: exact GSD is unknown. Avoid strong size-based claims and rely on visible pattern coherence. "
        )
    common = "".join(
        [
            "Task: analyze this GeoTIFF tile for possible archaeological features. ",
            f"The tile size is {width}x{height} pixels. ",
            scale_text,
            f"Provided views: {', '.join(selected_views)}. ",
            "Decision gate: set candidate=true only when the anomaly shows clear archaeological morphology and at least two supporting cues, "
            "such as coherent circular/oval/rectilinear geometry, plausible archaeological scale, contrast with the surrounding terrain, "
            "spatial organization, and support from more than one relevant view or visual cue. ",
            "Before marking candidate=true, actively test alternative explanations: agriculture, field parcel edges, ploughing, irrigation, "
            "drainage, roads or tracks, modern buildings or walls, vehicle marks, vegetation rows or individual tree crowns, natural gullies, "
            "erosion, geology, shadows, seams, compression artifacts, and tile-edge artifacts. ",
            "If any non-archaeological explanation is as plausible as archaeology, set candidate=false. ",
            "Do not mark isolated color changes, vague texture, random stone/soil patterns, single shadows, single vegetation differences, "
            "straight modern boundaries, road curves, or regular agricultural traces as archaeological candidates. ",
            "Look for repeated or regular geometry, circular or oval forms, rectilinear traces, banks, ditches, terraces, mounds, tumuli, "
            "old route alignments, enclosures, foundation traces, and anomalies that are spatially organized rather than random. ",
            "Candidate-type rules: mound/tumulus requires a coherent raised or soil/vegetation signature with plausible size; "
            "ring_ditch requires a continuous or strongly implied circular/oval ditch pattern; wall_trace/foundation/enclosure requires "
            "intentional rectilinear or enclosed geometry; road_trace requires an old alignment that is not a modern road, track, or field edge. ",
            "If the evidence is weak, ambiguous, modern-looking, only natural texture, or only one cue, set candidate=false. ",
            "If there is a candidate, return bbox_xyxy in tile pixel coordinates [x1,y1,x2,y2], not normalized coordinates. ",
            "The bbox should tightly cover the visible anomaly but is approximate and must not be treated as a final archaeological boundary. ",
            "Confidence calibration: below 0.75 means not strong enough for export, 0.75 means clear multi-cue evidence, "
            "0.90 means very strong evidence with archaeology more likely than common false positives. ",
            "Use candidate_type=\"unknown\" when the pattern looks archaeological but the type is unclear. ",
            "If there is no candidate, return candidate=false, confidence=0, candidate_type=\"none\", bbox_xyxy=null. ",
            "Keep visual_evidence short and concrete; mention the specific views that support the decision. ",
            "In possible_false_positive, name the strongest non-archaeological explanation considered, even for candidate=true. ",
            "In recommended_check, choose the single most useful next check from the allowed values. ",
            "Return exactly one JSON object with this schema and no markdown:\n",
            JSON_SCHEMA_TEXT,
        ]
    )
    if analysis_mode == "rgb_only":
        mode_text = (
            "Data mode: RGB orthophoto only. No DSM, DTM, hillshade, nDSM, or slope information is available. "
            "Base your judgment on crop marks, soil marks, color/vegetation differences, surface texture, regular geometry, "
            "structure stains, and visible surface anomalies. "
            "Because there is no topographic confirmation, be stricter: candidate=true should require very coherent geometry and plausible scale. "
            "Do not make strong micro-topographic claims in RGB-only mode."
        )
    elif analysis_mode == "rgb_dsm":
        mode_text = (
            "Data mode: RGB plus DSM/elevation surface. DTM-derived terrain derivatives are not available. "
            "Use RGB for crop/soil/surface traces and DSM for elevated surface anomalies. "
            "Be careful: DSM can include vegetation, buildings, and other modern surface objects, so do not treat every height anomaly "
            "as terrain archaeology. Reject height patterns that align with trees, roofs, roads, field edges, or other modern surface objects."
        )
    else:
        mode_text = (
            "Data mode: RGB plus topographic derivative views. "
            "Use RGB for crop/soil/surface traces and hillshade/nDSM/DSM/DTM/slope views for micro-topographic evidence. "
            "Prioritize anomalies that appear consistently across relevant views, such as mounds, tumuli, ring ditches, banks, walls, "
            "terraces, foundation traces, old road traces, enclosures, and regular terrain forms. "
            "Reject anomalies that appear in only one noisy view or are better explained by natural drainage, vegetation, or modern land use."
        )
    return f"{mode_text}\n\n{common}"


def _parse_model_json(text: str) -> Dict[str, Any]:
    raw = str(text or "").strip()
    json_text = _extract_json_object(raw)
    parsed = json.loads(json_text)
    if not isinstance(parsed, dict):
        raise ValueError("Model JSON object dondurmedi.")
    return parsed


def _extract_json_object(text: str) -> str:
    stripped = text.strip()
    if stripped.startswith("```"):
        lines = stripped.splitlines()
        if lines and lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip().startswith("```"):
            lines = lines[:-1]
        stripped = "\n".join(lines).strip()

    start = stripped.find("{")
    if start < 0:
        raise ValueError(f"Model JSON dondurmedi. Ham cevap: {stripped[:300]}")

    depth = 0
    in_string = False
    escaped = False
    for idx in range(start, len(stripped)):
        ch = stripped[idx]
        if in_string:
            if escaped:
                escaped = False
            elif ch == "\\":
                escaped = True
            elif ch == '"':
                in_string = False
            continue
        if ch == '"':
            in_string = True
        elif ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return stripped[start : idx + 1]
    raise ValueError(f"Model JSON objesi tamamlanmamis. Ham cevap: {stripped[:300]}")


def _model_result_to_record(
    *,
    parsed: Dict[str, Any],
    raw_response: str,
    base_record: Dict[str, Any],
    transform: Affine,
    crs: Optional[RasterioCRS],
    raster_width: int,
    raster_height: int,
) -> Dict[str, Any]:
    record = dict(base_record)
    candidate = bool(parsed.get("candidate", False))
    confidence = _coerce_float(parsed.get("confidence", 0.0), default=0.0)
    candidate_type = str(parsed.get("candidate_type", "none")).strip().lower() or "none"
    if candidate_type not in ALLOWED_CANDIDATE_TYPES:
        candidate_type = "unknown" if candidate else "none"
    recommended_check = str(parsed.get("recommended_check", "field_check")).strip().lower() or "field_check"
    if recommended_check not in RECOMMENDED_CHECKS:
        recommended_check = "field_check"

    bbox = parsed.get("bbox_xyxy")
    bbox_local: Optional[List[float]] = None
    bbox_global: Optional[List[float]] = None
    bbox_crs: Optional[List[float]] = None
    geometry: Optional[Dict[str, Any]] = None
    center_x: Optional[float] = None
    center_y: Optional[float] = None
    gps_lon: Optional[float] = None
    gps_lat: Optional[float] = None
    google_maps_url: str = ""
    parse_warning: Optional[str] = None

    if candidate:
        bbox_local = _validate_bbox(
            bbox,
            tile_width=int(record["tile_width"]),
            tile_height=int(record["tile_height"]),
        )
        if bbox_local is None:
            parse_warning = "candidate=true ancak bbox_xyxy gecersiz veya eksik; aday geometriye aktarilmadi"
            candidate = False
            confidence = 0.0
            candidate_type = "none"
        else:
            x1, y1, x2, y2 = bbox_local
            col = float(record["tile_col"])
            row = float(record["tile_row"])
            gx1 = min(max(col + x1, 0.0), float(raster_width))
            gy1 = min(max(row + y1, 0.0), float(raster_height))
            gx2 = min(max(col + x2, 0.0), float(raster_width))
            gy2 = min(max(row + y2, 0.0), float(raster_height))
            bbox_global = [gx1, gy1, gx2, gy2]
            geometry, bbox_crs, center_x, center_y = _bbox_global_to_geometry(transform, bbox_global)
            gps_lon, gps_lat = _point_to_wgs84(center_x, center_y, crs)
            if gps_lon is not None and gps_lat is not None:
                google_maps_url = f"https://www.google.com/maps?q={gps_lat:.8f},{gps_lon:.8f}"
    else:
        confidence = 0.0
        candidate_type = "none"
        bbox_local = None

    record.update(
        {
            "candidate": candidate,
            "confidence": float(np.clip(confidence, 0.0, 1.0)),
            "candidate_type": candidate_type,
            "bbox_xyxy": bbox_local,
            "bbox_global_xyxy": bbox_global,
            "bbox_crs_xyxy": bbox_crs,
            "center_x": center_x,
            "center_y": center_y,
            "gps_lon": gps_lon,
            "gps_lat": gps_lat,
            "google_maps_url": google_maps_url,
            "visual_evidence": str(parsed.get("visual_evidence", "") or ""),
            "possible_false_positive": str(parsed.get("possible_false_positive", "") or ""),
            "recommended_check": recommended_check,
            "status": "ok" if parse_warning is None else "warning",
            "error_type": None,
            "error_message": parse_warning,
            "geometry": geometry,
        }
    )
    if parse_warning:
        record["raw_response_excerpt"] = raw_response[:500]
    return record


def _coerce_float(value: Any, *, default: float) -> float:
    try:
        return float(value)
    except Exception:
        return default


def _validate_bbox(value: Any, *, tile_width: int, tile_height: int) -> Optional[List[float]]:
    if not isinstance(value, (list, tuple)) or len(value) != 4:
        return None
    try:
        x1, y1, x2, y2 = [float(v) for v in value]
    except Exception:
        return None
    if not all(math.isfinite(v) for v in (x1, y1, x2, y2)):
        return None
    x1 = min(max(x1, 0.0), float(tile_width))
    x2 = min(max(x2, 0.0), float(tile_width))
    y1 = min(max(y1, 0.0), float(tile_height))
    y2 = min(max(y2, 0.0), float(tile_height))
    if x2 <= x1 or y2 <= y1:
        return None
    return [x1, y1, x2, y2]


def _bbox_global_to_geometry(
    transform: Affine,
    bbox: Sequence[float],
) -> Tuple[Dict[str, Any], List[float], float, float]:
    x1, y1, x2, y2 = bbox
    corners_px = [(x1, y1), (x2, y1), (x2, y2), (x1, y2), (x1, y1)]
    coords = [tuple(transform * xy) for xy in corners_px]
    xs = [float(x) for x, _ in coords[:-1]]
    ys = [float(y) for _, y in coords[:-1]]
    cx, cy = transform * ((x1 + x2) / 2.0, (y1 + y2) / 2.0)
    geometry = {"type": "Polygon", "coordinates": [[list(coord) for coord in coords]]}
    return geometry, [min(xs), min(ys), max(xs), max(ys)], float(cx), float(cy)


def _point_to_wgs84(
    x: Optional[float],
    y: Optional[float],
    crs: Optional[RasterioCRS],
) -> Tuple[Optional[float], Optional[float]]:
    if x is None or y is None or crs is None:
        return None, None
    try:
        if crs.to_epsg() == 4326:
            return float(x), float(y)
        lon_vals, lat_vals = rasterio_transform(crs, "EPSG:4326", [float(x)], [float(y)])
        return float(lon_vals[0]), float(lat_vals[0])
    except Exception:
        return None, None


def _base_tile_record(
    *,
    tile_index: int,
    row: int,
    col: int,
    window: Window,
    layout: RasterBandLayout,
    selected_views: Sequence[str],
) -> Dict[str, Any]:
    return {
        "tile_index": int(tile_index),
        "tile_row": int(row),
        "tile_col": int(col),
        "tile_width": int(window.width),
        "tile_height": int(window.height),
        "used_views": list(selected_views),
        "has_rgb": layout.has_rgb,
        "has_dsm": layout.has_dsm,
        "has_dtm": layout.has_dtm,
        "analysis_mode": layout.analysis_mode,
    }


def _run_error_record(
    *,
    error_type: str,
    message: str,
    layout: RasterBandLayout,
    selected_views: Sequence[str],
) -> Dict[str, Any]:
    record = _base_tile_record(
        tile_index=0,
        row=0,
        col=0,
        window=Window(0, 0, 0, 0),
        layout=layout,
        selected_views=selected_views,
    )
    record.update(
        {
            "candidate": False,
            "confidence": 0.0,
            "candidate_type": "none",
            "bbox_xyxy": None,
            "bbox_global_xyxy": None,
            "bbox_crs_xyxy": None,
            "center_x": None,
            "center_y": None,
            "gps_lon": None,
            "gps_lat": None,
            "google_maps_url": "",
            "visual_evidence": "",
            "possible_false_positive": "",
            "recommended_check": "field_check",
            "status": "error",
            "error_type": error_type,
            "error_message": message,
        }
    )
    return record


def _skipped_tile_record(base_record: Dict[str, Any], *, reason: str) -> Dict[str, Any]:
    record = dict(base_record)
    record.update(
        {
            "candidate": False,
            "confidence": 0.0,
            "candidate_type": "none",
            "bbox_xyxy": None,
            "bbox_global_xyxy": None,
            "bbox_crs_xyxy": None,
            "center_x": None,
            "center_y": None,
            "gps_lon": None,
            "gps_lat": None,
            "google_maps_url": "",
            "visual_evidence": "",
            "possible_false_positive": "",
            "recommended_check": "rgb",
            "status": "skipped",
            "error_type": "empty_tile",
            "error_message": reason,
            "geometry": None,
        }
    )
    return record


def _is_exportable_candidate(record: Dict[str, Any], threshold: float) -> bool:
    return bool(record.get("candidate")) and float(record.get("confidence") or 0.0) >= threshold and bool(
        record.get("geometry")
    )


def _classify_exception(exc: Exception) -> str:
    if isinstance(exc, json.JSONDecodeError) or "json" in str(exc).lower():
        return "invalid_json"
    text = str(exc).lower()
    if _looks_like_connection_error(exc):
        return "connection_failed"
    if "model" in text and ("not found" in text or "does not exist" in text):
        return "model_not_found"
    if any(token in text for token in ("vision", "image_url", "image input", "does not support image", "multimodal")):
        return "vision_unsupported"
    return "api_error"


def _friendly_exception_message(exc: Exception, config: VlmLmStudioConfig) -> str:
    error_type = _classify_exception(exc)
    if error_type == "connection_failed":
        return (
            f"LM Studio sunucusuna baglanilamadi ({config.base_url}). "
            "Local Server acik mi ve port dogru mu kontrol edin."
        )
    if error_type == "vision_unsupported":
        return (
            f"Model goruntu girdisini desteklemiyor olabilir ({config.model}). "
            "LM Studio'da vision/multimodal destekli bir model yukleyin."
        )
    if error_type == "model_not_found":
        return f"Model bulunamadi: {config.model}. LM Studio'da yuklu model adini kontrol edin."
    if error_type == "invalid_json":
        return f"Model strict JSON dondurmedi: {str(exc)[:500]}"
    return str(exc)[:1000]


def _looks_like_connection_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return any(
        token in text
        for token in (
            "connection",
            "connect",
            "refused",
            "timed out",
            "timeout",
            "winerror 10061",
            "server disconnected",
        )
    )


def _looks_like_response_format_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return "response_format" in text or "json_object" in text or "json schema" in text


def _write_jsonl_line(fh: Any, record: Dict[str, Any]) -> None:
    fh.write(json.dumps(record, ensure_ascii=False) + "\n")
    fh.flush()


def _write_jsonl_records(path: Path, records: Sequence[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        for record in records:
            _write_jsonl_line(fh, record)


def _write_candidate_outputs(
    paths: VlmOutputPaths,
    records: Sequence[Dict[str, Any]],
    crs: Optional[RasterioCRS],
    *,
    all_records: Optional[Sequence[Dict[str, Any]]] = None,
    confidence_threshold: float = 0.0,
    logger: logging.Logger,
) -> VlmOutputPaths:
    sorted_records = _sort_candidate_records(records)
    _write_candidate_csv(paths.csv, sorted_records)
    xlsx_path = _write_candidate_xlsx(
        paths.xlsx,
        sorted_records,
        all_records=all_records,
        confidence_threshold=confidence_threshold,
        logger=logger,
    )
    _write_candidate_geojson(paths.geojson, sorted_records, crs)
    _write_candidate_gpkg(paths.gpkg, sorted_records, crs, logger=logger)
    return replace(paths, xlsx=xlsx_path)


def _set_progress_postfix(
    pbar: Any,
    *,
    records: Sequence[Dict[str, Any]],
    error_records: Sequence[Dict[str, Any]],
    skipped_count: int,
    raw_candidate_count: int,
    candidate_records: Sequence[Dict[str, Any]],
) -> None:
    pbar.set_postfix(
        {
            "ok": len(records) - len(error_records) - skipped_count,
            "skip": skipped_count,
            "err": len(error_records),
            "cand": raw_candidate_count,
            "out": len(candidate_records),
        }
    )


CSV_COLUMNS = [
    "tile_index",
    "tile_row",
    "tile_col",
    "tile_width",
    "tile_height",
    "candidate",
    "confidence",
    "candidate_type",
    "bbox_xyxy",
    "bbox_global_xyxy",
    "bbox_crs_xyxy",
    "center_x",
    "center_y",
    "gps_lon",
    "gps_lat",
    "google_maps_url",
    "visual_evidence",
    "possible_false_positive",
    "recommended_check",
    "used_views",
    "has_rgb",
    "has_dsm",
    "has_dtm",
    "analysis_mode",
    "status",
    "error_type",
    "error_message",
]

NOT_FOUND_COLUMNS = [*CSV_COLUMNS, "not_found_reason"]


def _record_confidence(record: Dict[str, Any]) -> float:
    try:
        return float(record.get("confidence") or 0.0)
    except Exception:
        return 0.0


def _record_tile_index(record: Dict[str, Any]) -> int:
    try:
        return int(record.get("tile_index") or 0)
    except Exception:
        return 0


def _sort_candidate_records(records: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return sorted(records, key=lambda record: (-_record_confidence(record), _record_tile_index(record)))


def _not_found_reason(record: Dict[str, Any], threshold: float) -> str:
    status = str(record.get("status") or "").strip().lower()
    if status in {"skipped", "error", "warning"}:
        return status
    if bool(record.get("candidate")):
        if not record.get("geometry"):
            return "missing_geometry"
        if _record_confidence(record) < threshold:
            return "below_threshold"
    return "no_candidate"


def _not_found_records(
    records: Sequence[Dict[str, Any]],
    *,
    confidence_threshold: float,
) -> List[Dict[str, Any]]:
    not_found: List[Dict[str, Any]] = []
    for record in records:
        if _is_exportable_candidate(record, confidence_threshold):
            continue
        row = dict(record)
        row["not_found_reason"] = _not_found_reason(record, confidence_threshold)
        not_found.append(row)
    return not_found


def _write_candidate_csv(path: Path, records: Sequence[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for record in records:
            writer.writerow(_flatten_record(record, CSV_COLUMNS))


def _write_candidate_xlsx(
    path: Path,
    records: Sequence[Dict[str, Any]],
    *,
    all_records: Optional[Sequence[Dict[str, Any]]] = None,
    confidence_threshold: float = 0.0,
    logger: logging.Logger,
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        logger.warning("VLM Excel yazimi atlandi; openpyxl kurulu degil. CSV ciktisi kullanilabilir: %s", path.with_suffix(".csv"))
        return path

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active

    def write_sheet(
        sheet: Any,
        *,
        title: str,
        columns: Sequence[str],
        sheet_records: Sequence[Dict[str, Any]],
        table_name: str,
    ) -> None:
        sheet.title = title
        sheet.append(list(columns))
        for record in sheet_records:
            flat = _flatten_record(record, columns)
            sheet.append([flat.get(col, "") for col in columns])

        google_maps_col = columns.index("google_maps_url") + 1
        for row_idx in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=google_maps_col)
            maps_url = str(cell.value or "").strip()
            if maps_url:
                cell.hyperlink = maps_url
                cell.style = "Hyperlink"

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        sheet.freeze_panes = "A2"
        if sheet.max_row >= 2:
            table = Table(displayName=table_name, ref=sheet.dimensions)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            sheet.add_table(table)

        for col_idx, column_name in enumerate(columns, start=1):
            max_len = len(column_name)
            for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                value = row[0].value
                if value is not None:
                    max_len = max(max_len, min(len(str(value)), 80))
            sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = min(
                max(max_len + 2, 10),
                60,
            )

    write_sheet(
        ws,
        title="vlm_candidates",
        columns=CSV_COLUMNS,
        sheet_records=_sort_candidate_records(records),
        table_name="VlmCandidates",
    )
    missing_records = _not_found_records(
        list(all_records) if all_records is not None else [],
        confidence_threshold=confidence_threshold,
    )
    write_sheet(
        wb.create_sheet("bulunmayan_tilelar"),
        title="bulunmayan_tilelar",
        columns=NOT_FOUND_COLUMNS,
        sheet_records=missing_records,
        table_name="VlmNotFoundTiles",
    )

    return _save_workbook_with_excel_lock_fallback(wb, path, logger=logger)


def _is_excel_lock_error(exc: OSError) -> bool:
    return (
        isinstance(exc, PermissionError)
        or getattr(exc, "winerror", None) in {32, 33}
        or getattr(exc, "errno", None) == 13
    )


def _xlsx_alternative_path(path: Path, index: int) -> Path:
    suffix = "_alternatif" if index <= 1 else f"_alternatif_{index}"
    return path.with_name(f"{path.stem}{suffix}{path.suffix}")


def _save_workbook_with_excel_lock_fallback(
    workbook: Any,
    path: Path,
    *,
    logger: logging.Logger,
) -> Path:
    try:
        workbook.save(path)
        return path
    except OSError as exc:
        if not _is_excel_lock_error(exc):
            raise
        original_exc = exc

    for index in range(1, 101):
        alternative_path = _xlsx_alternative_path(path, index)
        try:
            workbook.save(alternative_path)
        except OSError as exc:
            if _is_excel_lock_error(exc):
                continue
            raise
        logger.warning(
            "VLM Excel dosyasi acik/kilitli gorunuyor; cikti alternatif adla kaydedildi: %s",
            alternative_path,
        )
        return alternative_path

    raise original_exc


def _write_candidate_geojson(
    path: Path,
    records: Sequence[Dict[str, Any]],
    crs: Optional[RasterioCRS],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    features = []
    for record in records:
        geometry = record.get("geometry")
        if not geometry:
            continue
        props = _flatten_record(record, CSV_COLUMNS)
        features.append({"type": "Feature", "geometry": geometry, "properties": props})
    payload: Dict[str, Any] = {"type": "FeatureCollection", "features": features}
    if crs is not None:
        payload["name"] = path.stem
        payload["crs_wkt"] = crs.to_wkt()
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_candidate_gpkg(
    path: Path,
    records: Sequence[Dict[str, Any]],
    crs: Optional[RasterioCRS],
    *,
    logger: logging.Logger,
) -> None:
    try:
        import fiona
    except ImportError:
        logger.warning("VLM GPKG yazimi atlandi; fiona kurulu degil.")
        return

    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        path.unlink()
    _write_candidate_gpkg_layer(path, records, crs, layer_name="vlm_candidates")

    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for record in records:
        candidate_type = str(record.get("candidate_type") or "unknown").strip().lower() or "unknown"
        if candidate_type == "none":
            continue
        grouped.setdefault(candidate_type, []).append(record)
    for candidate_type, group_records in sorted(grouped.items()):
        _write_candidate_gpkg_layer(
            path,
            group_records,
            crs,
            layer_name=f"vlm_{_safe_layer_token(candidate_type)}",
        )


def _safe_layer_token(value: str) -> str:
    token = "".join(ch if ch.isalnum() or ch == "_" else "_" for ch in str(value).strip().lower())
    token = "_".join(part for part in token.split("_") if part)
    return (token or "unknown")[:48]


def _write_candidate_gpkg_layer(
    path: Path,
    records: Sequence[Dict[str, Any]],
    crs: Optional[RasterioCRS],
    *,
    layer_name: str,
) -> None:
    import fiona

    schema = {
        "geometry": "Polygon",
        "properties": {
            "tile_index": "int",
            "tile_row": "int",
            "tile_col": "int",
            "confidence": "float",
            "cand_type": "str",
            "bbox_px": "str",
            "bbox_crs": "str",
            "center_x": "float",
            "center_y": "float",
            "gps_lon": "float",
            "gps_lat": "float",
            "gmaps": "str",
            "evidence": "str",
            "false_pos": "str",
            "check": "str",
            "views": "str",
            "has_rgb": "int",
            "has_dsm": "int",
            "has_dtm": "int",
            "mode": "str",
            "status": "str",
        },
    }
    open_kwargs: Dict[str, Any] = {
        "driver": "GPKG",
        "layer": layer_name,
        "schema": schema,
    }
    if crs is not None:
        open_kwargs["crs_wkt"] = crs.to_wkt()
    with fiona.open(path, "w", **open_kwargs) as dst:
        for record in records:
            geometry = record.get("geometry")
            if not geometry:
                continue
            dst.write(
                {
                    "geometry": geometry,
                    "properties": {
                        "tile_index": int(record.get("tile_index") or 0),
                        "tile_row": int(record.get("tile_row") or 0),
                        "tile_col": int(record.get("tile_col") or 0),
                        "confidence": float(record.get("confidence") or 0.0),
                        "cand_type": str(record.get("candidate_type") or ""),
                        "bbox_px": _jsonish(record.get("bbox_xyxy")),
                        "bbox_crs": _jsonish(record.get("bbox_crs_xyxy")),
                        "center_x": _optional_float(record.get("center_x")),
                        "center_y": _optional_float(record.get("center_y")),
                        "gps_lon": _optional_float(record.get("gps_lon")),
                        "gps_lat": _optional_float(record.get("gps_lat")),
                        "gmaps": str(record.get("google_maps_url") or ""),
                        "evidence": str(record.get("visual_evidence") or "")[:500],
                        "false_pos": str(record.get("possible_false_positive") or "")[:500],
                        "check": str(record.get("recommended_check") or ""),
                        "views": ",".join(str(v) for v in record.get("used_views") or []),
                        "has_rgb": int(bool(record.get("has_rgb"))),
                        "has_dsm": int(bool(record.get("has_dsm"))),
                        "has_dtm": int(bool(record.get("has_dtm"))),
                        "mode": str(record.get("analysis_mode") or ""),
                        "status": str(record.get("status") or ""),
                    },
                }
            )


def _flatten_record(record: Dict[str, Any], columns: Sequence[str]) -> Dict[str, Any]:
    flat: Dict[str, Any] = {}
    for col in columns:
        value = record.get(col)
        if isinstance(value, (list, tuple, dict)):
            flat[col] = json.dumps(value, ensure_ascii=False)
        elif isinstance(value, bool):
            flat[col] = int(value)
        elif value is None:
            flat[col] = ""
        else:
            flat[col] = value
    return flat


def _jsonish(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False)


def _optional_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except Exception:
        return None
